# THON SAFE JSON V58.19 - atomic json protection

# Hooks antigos de lote desativados.
# O lote de caça agora tem UMA fonte oficial: sync_api_qualificados_para_lote().
print("[LOTE API DIRECT] ativo")

try:
    import thon_safe_json
    thon_safe_json.activate()
except Exception as _thon_safe_json_err:
    print("[SAFE_JSON] falha ao ativar:", _thon_safe_json_err)

"""
THON Toolkit v58.11 - ENGINE SELECT SAFE
- PROSPECTOR API: YouTube Data API só para descobrir canais; depois dispara verificação DLP automática
- PROSPECTOR DLP: yt-dlp faz descoberta + verificação completa, com Auto Hunt
- Múltiplas chaves lidas de api_keys.txt, uma por linha
- Salva TODOS os canais brutos (antes de qualquer filtro) em canais_brutos_api.json
- Atualiza status: bruto -> qualificado -> aprovado / reprovado
- Fila persistente de leads (não perde lote ao reiniciar)
- Todas as funcionalidades: CRM, projetos, trabalhos, downloader, metas, etc.
"""

from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
import openpyxl
from io import BytesIO
import subprocess, threading, os, time, json, re, sys, uuid, shutil, platform, random, logging, shlex, tempfile, html
from datetime import datetime, timedelta

try:
    from query_factory.query_factory import get_next_queries as query_factory_get_next_queries
except Exception as _query_factory_import_error:
    query_factory_get_next_queries = None
    print(f"[QUERY FACTORY] indisponivel, usando fallback antigo: {_query_factory_import_error}")

try:
    from pytubefix import YouTube as _PytubeYouTube
    import xml.etree.ElementTree as _ET
except Exception as _pytubefix_import_error:
    _PytubeYouTube = None
    _ET = None
    print(f"[PYTUBEFIX] indisponivel ({_pytubefix_import_error}); rode: pip install pytubefix --break-system-packages")

APP_DIR = os.path.dirname(os.path.abspath(__file__))
try:
    os.chdir(APP_DIR)
except Exception:
    pass
app = Flask(__name__, static_folder=".")
CORS(app)

# V-SaaS: Ativa sistema de login via variavel de ambiente
try:
    from thon_auth import setup_auth_middleware
    setup_auth_middleware(app)
    print("[AUTH] Sistema de autenticacao carregado. Defina THON_AUTH_PASSWORD no ambiente para ativar.")
except Exception as _e_auth:
    print(f"[AUTH] Aviso: thon_auth nao carregado: {_e_auth}")
logging.getLogger("werkzeug").setLevel(logging.ERROR)
APP_VERSION = "v58.22-crm-api-collector-v2"
ENGINE_VERSION = "PROSPECTORS_API_DISCOVERY_THEN_DLP_OR_DLP_FULL"
RUNTIME_LOG_FILE = os.environ.get("THON_RUNTIME_LOG", "thon_backend_runtime.log")

# ===== ARQUIVOS =====
MEMORY_FILE = "winchester_vistos.json"
REPROVADOS_FILE = "winchester_reprovados.json"
APROVADOS_FILE = "winchester_aprovados.json"
API_KEYS_FILE = os.environ.get("YOUTUBE_API_KEYS_FILE", "api_keys.txt")
CANAL_BRUTO_FILE = "canais_brutos_api.json"   # <-- depósito de todos os brutos
CANAL_STATUS_FILE = "canais_brutos_status.json"  # <-- status por ID (opcional, podemos integrar no mesmo)
LOTE_ATIVO_FILE = "fila_pendente_api.json"  # fila/lote aguardando aprovação; nunca apagar em ciclo novo
MODE_FILE = "thon_engine_mode.json"  # api ou dlp
DLP_VERIFY_QUEUE_FILE = "dlp_verification_queue.json"  # canais achados pela API aguardando verificação DLP

MIN_SUBS = 10_000
MAX_SUBS = 200_000
LOTE_PADRAO = 60
SEARCH_LIMIT = 80
QUERY_LIMIT_PADRAO = 40
QUERY_LIMIT_MAX = 600
PRESETS = {
    "api_multi_source_fast": {
        "candidate_target": 5000, "pages_per_query": 2, "score_min": 65,
        "min_avg_views": 500, "videos_per_channel": 3, "query_workers": 6,
        "workers": 8, "max_last_days": 180, "modo_brasil": True,
        "parada_inteligente": True,
        "source_mix": {
            "video_search_long": 0.45, "channel_search": 0.20,
            "playlist_search": 0.15, "uploads_expansion": 0.10,
            "subscriptions_seed": 0.05, "activities_seed": 0.05,
        }
    }
}

# ============================================================
# CHAVES DA API (lidas de api_keys.txt, uma por linha)
# ============================================================
_API_KEYS = []
_API_KEY_INDEX = 0
_API_KEY_LOCK = threading.Lock()
_API_KEY_FAILED = {}

def _api_key_id(key):
    key = str(key or "").strip()
    if len(key) <= 14:
        return key
    return f"{key[:8]}...{key[-6:]}"

def _api_marcar_chave_falha(key, motivo="quota/403"):
    """Marca uma chave como ruim só para o dia atual, deixando o rodízio pular para as próximas."""
    kid = _api_key_id(key)
    if not kid:
        return
    _API_KEY_FAILED[kid] = {"date": _api_today(), "motivo": str(motivo or ""), "at": str(datetime.now())}

def _api_chave_disponivel(key):
    falha = _API_KEY_FAILED.get(_api_key_id(key))
    return not falha or falha.get("date") != _api_today()

def _api_carregar_chaves():
    """Carrega múltiplas chaves de api_keys.txt. Também aceita env YOUTUBE_API_KEY como primeira chave."""
    chaves = []
    chave_env = os.environ.get("YOUTUBE_API_KEY", "").strip()
    if chave_env:
        chaves.append(chave_env)
        print("[API] chave carregada de env YOUTUBE_API_KEY")
    try:
        if os.path.exists(API_KEYS_FILE):
            with open(API_KEYS_FILE, "r", encoding="utf-8") as f:
                for linha in f:
                    linha = linha.strip()
                    if linha and not linha.startswith("#") and linha not in chaves:
                        chaves.append(linha)
    except Exception as e:
        print(f"[API] erro lendo {API_KEYS_FILE}: {e}")
    # Compatibilidade: se api_keys.txt não existir, lê youtube_api_key.txt antigo.
    if not chaves and os.path.exists("youtube_api_key.txt"):
        try:
            with open("youtube_api_key.txt", "r", encoding="utf-8") as f:
                for linha in f:
                    linha = linha.strip()
                    if linha and not linha.startswith("#") and linha not in chaves:
                        chaves.append(linha)
            if chaves:
                print("[API] usando youtube_api_key.txt por compatibilidade")
        except Exception as e:
            print(f"[API] erro lendo youtube_api_key.txt: {e}")
    if not chaves:
        print(f"[API] AVISO: nenhuma chave encontrada em {API_KEYS_FILE}")
    return chaves

def _api_proxima_chave():
    global _API_KEY_INDEX
    with _API_KEY_LOCK:
        if not _API_KEYS:
            raise RuntimeError(f"Nenhuma chave API carregada. Crie {API_KEYS_FILE} na pasta do app.")
        for _ in range(len(_API_KEYS)):
            chave = _API_KEYS[_API_KEY_INDEX % len(_API_KEYS)]
            _API_KEY_INDEX += 1
            if _api_chave_disponivel(chave):
                return chave
        # Se todas falharam hoje, libera uma última tentativa circular para não travar eternamente.
        chave = _API_KEYS[_API_KEY_INDEX % len(_API_KEYS)]
        _API_KEY_INDEX += 1
        return chave

_API_KEYS = _api_carregar_chaves()
print(f"[API] {len(_API_KEYS)} chave(s) carregada(s) de {API_KEYS_FILE}.")

# ============================================================
# GESTÃO DO DEPÓSITO DE BRUTOS
# ============================================================
def _carregar_brutos():
    try:
        if os.path.exists(CANAL_BRUTO_FILE):
            with open(CANAL_BRUTO_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, list):
                    return data
                if isinstance(data, dict) and "canais" in data:
                    return data.get("canais", [])
    except Exception as e:
        print(f"[bruto] erro carregando: {e}")
    return []

def _salvar_brutos(brutos):
    try:
        with open(CANAL_BRUTO_FILE, "w", encoding="utf-8") as f:
            json.dump({"canais": brutos, "total": len(brutos), "updated_at": str(datetime.now())}, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"[bruto] erro salvando: {e}")

def _atualizar_status_bruto(channel_id, novo_status, meta=None):
    """Atualiza o status de um canal no depósito de brutos."""
    brutos = _carregar_brutos()
    encontrado = False
    for c in brutos:
        if c.get("channel_id") == channel_id or c.get("id") == channel_id:
            c["status"] = novo_status
            c["updated_at"] = str(datetime.now())
            if meta:
                c.update(meta)
            encontrado = True
            break
    if encontrado:
        _salvar_brutos(brutos)
    return encontrado

def _adicionar_bruto(canal_data):
    """Adiciona um canal bruto (se ainda não existir)."""
    brutos = _carregar_brutos()
    cid = canal_data.get("channel_id") or canal_data.get("id")
    if not cid:
        return False
    # Verifica se já existe
    for c in brutos:
        if c.get("channel_id") == cid or c.get("id") == cid:
            return False
    canal_data["status"] = "bruto"
    canal_data["salvo_em"] = str(datetime.now())
    canal_data["engine"] = ENGINE_VERSION
    brutos.append(canal_data)
    # Limita a 20.000 registros (para não crescer infinitamente)
    if len(brutos) > 20000:
        brutos = brutos[-20000:]
    _salvar_brutos(brutos)
    return True

def _adicionar_brutos_lista(lista_canais):
    """Adiciona vários canais brutos de uma vez."""
    adicionados = 0
    for c in lista_canais:
        if _adicionar_bruto(c):
            adicionados += 1
    return adicionados


# ============================================================
# FILA PENDENTE / LOTE ATIVO — PERSISTÊNCIA REAL
# ============================================================
def _fmt_numero_fila(n):
    try:
        n = int(n or 0)
        if n >= 1_000_000:
            return f"{n/1e6:.1f}M"
        if n >= 1_000:
            return f"{n/1e3:.0f}K"
        return str(n)
    except Exception:
        return "N/A"

def _eh_url_video_youtube(valor):
    s = str(valor or "").strip().lower()
    return bool(s and ("watch?v=" in s or "youtu.be/" in s or "/shorts/" in s or "/video/" in s))

def _extrair_channel_id_de_valor(valor):
    s = str(valor or "").strip()
    if not s:
        return ""
    if _eh_url_video_youtube(s):
        return ""
    # V58.25: regex mais flexivel (aceita 18+ chars apos UC)
    m = re.search(r"(UC[0-9A-Za-z_-]{18,})", s)
    if m:
        return m.group(1)
    if s.startswith("UC") and len(s) >= 18:
        return s
    return ""

def _lead_channel_id_seguro(lead):
    lead = lead or {}
    for key in ("channel_id", "canal_id", "youtube_channel_id", "channelId", "id"):
        cid = _extrair_channel_id_de_valor(lead.get(key))
        if cid:
            return cid
    for key in ("channel_url", "canal_url", "url_canal", "url"):
        cid = _extrair_channel_id_de_valor(lead.get(key))
        if cid:
            return cid
    return ""

def _lead_channel_url_segura(lead, cid=""):
    lead = lead or {}
    for key in ("channel_url", "canal_url", "url_canal", "url"):
        u = str(lead.get(key) or "").strip()
        if u and not _eh_url_video_youtube(u):
            return u
    return f"https://youtube.com/channel/{cid}" if cid else ""

def _normalizar_canal_lote(c):
    if not isinstance(c, dict):
        return None
    cid = _lead_channel_id_seguro(c)
    if not cid:
        return None
    nome = c.get("nome") or c.get("title") or c.get("search_title") or "Canal"
    url = _lead_channel_url_segura(c, cid)
    subs = c.get("subs") if c.get("subs") is not None else c.get("subscriber_count")
    try:
        subs_int = int(subs or 0)
    except Exception:
        subs_int = 0
    views = c.get("views", c.get("recent_avg_views", c.get("avg_views", 0)))
    score = int(c.get("score", c.get("api_score", 0)) or 0)
    return {
        **c,
        "id": cid,
        "channel_id": cid,
        "nome": nome,
        "title": c.get("title") or nome,
        "url": url,
        "subs": subs_int,
        "inscritos": int(c.get("inscritos", subs_int) or 0),
        "subs_fmt": c.get("subs_fmt") or _fmt_numero_fila(subs_int),
        "score": score,
        "api_score": int(c.get("api_score", score) or 0),
        "views": int(views or 0),
        "shorts": int(c.get("shorts", 0) or 0),
        "longos": int(c.get("longos", 0) or 0),
        "status": c.get("status") or "pendente",
        "status_fila": c.get("status_fila") or "pendente",
        "fila_salvo_em": c.get("fila_salvo_em") or str(datetime.now()),
        "engine": c.get("engine") or ENGINE_VERSION,
    }

def _extrair_lista_canais(data):
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        for key in ("lote", "fila", "pendentes", "qualificados", "canais", "items"):
            if isinstance(data.get(key), list):
                return data.get(key) or []
    return []

def _dedupe_lote(lista):
    mapa = {}
    for item in lista or []:
        c = _normalizar_canal_lote(item)
        if not c:
            continue
        cid = c.get("id")
        antigo = mapa.get(cid, {})
        # conserva campos antigos, mas permite campos novos/score atualizarem
        mapa[cid] = {**antigo, **c}
    return list(mapa.values())

def carregar_lote_ativo():
    """Carrega somente a fonte oficial do lote: fila_pendente_api.json."""
    itens = []
    try:
        if os.path.exists(LOTE_ATIVO_FILE):
            with open(LOTE_ATIVO_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            itens = _extrair_lista_canais(data)
    except Exception as e:
        print(f"[fila] erro lendo {LOTE_ATIVO_FILE}: {e}")
    lote = _dedupe_lote(itens)
    ids_ok = ids_aprovados()
    lote = [c for c in lote if c.get("id") not in ids_ok]
    return lote

def salvar_lote_ativo(lote):
    try:
        lote = _dedupe_lote(lote or [])
        ids_ok = ids_aprovados()
        lote = [c for c in lote if c.get("id") not in ids_ok]
        if os.path.exists(LOTE_ATIVO_FILE):
            try:
                shutil.copy2(LOTE_ATIVO_FILE, LOTE_ATIVO_FILE + ".bak")
            except Exception:
                pass
        tmp = LOTE_ATIVO_FILE + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(lote, f, ensure_ascii=False, indent=2)
            f.write("\n")
        os.replace(tmp, LOTE_ATIVO_FILE)
        return lote
    except Exception as e:
        print(f"[fila] erro salvando {LOTE_ATIVO_FILE}: {e}")
        return lote or []

def adicionar_lote_ativo(novos, base=None):
    base = list(base or carregar_lote_ativo())
    lote = _dedupe_lote(base + list(novos or []))
    return salvar_lote_ativo(lote)

def sync_api_qualificados_para_lote(qualificados, source="api_direct"):
    """Única função oficial para alimentar o Lote de caça com qualificados reais."""
    qualificados = [q for q in (qualificados or []) if isinstance(q, dict)]
    atual = carregar_lote_ativo()
    antes = len(atual)
    aprovados_ids = ids_aprovados()
    mapa = {c.get("channel_id") or c.get("id"): dict(c) for c in atual if isinstance(c, dict) and (c.get("channel_id") or c.get("id"))}
    novos_adicionados = 0
    ja_existiam = 0
    aprovados_ignorados = 0
    agora = datetime.now().isoformat(timespec="seconds")

    for raw in qualificados:
        c = _normalizar_canal_lote(raw)
        if not c:
            continue
        cid = c.get("channel_id") or c.get("id")
        if cid in aprovados_ids:
            aprovados_ignorados += 1
            continue
        c.update({
            "source": source,
            "qualified_by": source,
            "api_last_batch": True,
            "qualified_at": c.get("qualified_at") or agora,
            "status": "pendente",
            "status_fila": "pendente",
        })
        if cid in mapa:
            antigo = mapa[cid]
            mapa[cid] = {**antigo, **c}
            ja_existiam += 1
        else:
            mapa[cid] = c
            novos_adicionados += 1

    lote_final = salvar_lote_ativo(list(mapa.values()))
    report = {
        "antes": antes,
        "input": len(qualificados),
        "novos_adicionados": novos_adicionados,
        "ja_existiam": ja_existiam,
        "aprovados_ignorados": aprovados_ignorados,
        "depois": len(lote_final),
        "file": LOTE_ATIVO_FILE,
    }
    with lock:
        estado["lote"] = lote_final
        estado["qualificados_lista"] = lote_final
    print(f"[API DIRECT] qualificados_api={report['input']} | novos_adicionados={report['novos_adicionados']} | ja_existiam={report['ja_existiam']} | fila_total={report['depois']}")
    return report

def remover_lote_ativo_ids(ids):
    ids = {str(x) for x in (ids or []) if str(x).strip()}
    lote = [c for c in carregar_lote_ativo() if c.get("id") not in ids and c.get("channel_id") not in ids]
    return salvar_lote_ativo(lote)

def reconstruir_lote_dos_brutos(limit=500):
    """Tenta reconstruir fila a partir de canais_brutos_api.json com status qualificado."""
    brutos = _carregar_brutos()
    candidatos = []
    for b in brutos:
        st = str(b.get("status") or "").lower()
        if st not in {"qualificado", "pendente", "aguardando", "aprovacao", "aprovação"}:
            continue
        cid = b.get("id") or b.get("channel_id")
        if not cid:
            continue
        candidatos.append({
            "id": cid,
            "channel_id": cid,
            "nome": b.get("nome") or b.get("title") or b.get("search_title") or "Canal",
            "title": b.get("title") or b.get("nome") or b.get("search_title") or "Canal",
            "url": b.get("url") or f"https://youtube.com/channel/{cid}",
            "subs": int(b.get("subs") or b.get("subscriber_count") or 0),
            "subs_fmt": b.get("subs_fmt") or _fmt_numero_fila(b.get("subs") or b.get("subscriber_count") or 0),
            "score": int(b.get("score") or 0),
            "nicho": b.get("nicho", ""),
            "query": b.get("query") or b.get("found_query") or "",
            "found_query": b.get("found_query") or b.get("query") or "",
            "description": b.get("description") or b.get("descricao") or "",
            "source": b.get("source") or "reconstruido_dos_brutos",
            "engine": b.get("engine") or ENGINE_VERSION,
            "status_fila": "pendente",
        })
        if len(candidatos) >= limit:
            break
    return adicionar_lote_ativo(candidatos)

# ============================================================
# FILA DE VERIFICAÇÃO DLP — canais achados pela API
# ============================================================
def carregar_fila_dlp_verificacao():
    try:
        if os.path.exists(DLP_VERIFY_QUEUE_FILE):
            with open(DLP_VERIFY_QUEUE_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            itens = _extrair_lista_canais(data)
        else:
            itens = []
    except Exception as e:
        print(f"[dlp queue] erro lendo {DLP_VERIFY_QUEUE_FILE}: {e}")
        itens = []
    ids_finalizados = ids_aprovados() | ids_reprovados()
    return [c for c in _dedupe_lote(itens) if c.get("id") not in ids_finalizados]

def salvar_fila_dlp_verificacao(lista):
    try:
        lista = _dedupe_lote(lista or [])
        ids_finalizados = ids_aprovados() | ids_reprovados()
        lista = [c for c in lista if c.get("id") not in ids_finalizados]
        with open(DLP_VERIFY_QUEUE_FILE, "w", encoding="utf-8") as f:
            json.dump({
                "fila": lista,
                "total": len(lista),
                "updated_at": str(datetime.now()),
                "engine": ENGINE_VERSION,
                "nota": "Fila de canais descobertos pela API que ainda precisam passar na verificação DLP.",
            }, f, ensure_ascii=False, indent=2)
        return lista
    except Exception as e:
        print(f"[dlp queue] erro salvando {DLP_VERIFY_QUEUE_FILE}: {e}")
        return lista or []

def adicionar_fila_dlp_verificacao(novos):
    fila = carregar_fila_dlp_verificacao()
    fila = _dedupe_lote(fila + list(novos or []))
    return salvar_fila_dlp_verificacao(fila)

def remover_fila_dlp_verificacao_ids(ids):
    ids = {str(x) for x in (ids or []) if str(x).strip()}
    fila = [c for c in carregar_fila_dlp_verificacao() if c.get("id") not in ids and c.get("channel_id") not in ids]
    return salvar_fila_dlp_verificacao(fila)

def garantir_arquivos_persistentes():
    """Garante que os arquivos centrais existam sem apagar dados antigos."""
    defaults = {
        CANAL_BRUTO_FILE: {"canais": [], "total": 0, "created_at": str(datetime.now()), "engine": ENGINE_VERSION},
        LOTE_ATIVO_FILE: [],
        DLP_VERIFY_QUEUE_FILE: {"fila": [], "total": 0, "created_at": str(datetime.now()), "engine": ENGINE_VERSION},
        API_KEYS_FILE: "",
    }
    for path, default in defaults.items():
        try:
            if os.path.exists(path):
                continue
            with open(path, "w", encoding="utf-8") as f:
                if isinstance(default, str):
                    f.write(default)
                else:
                    json.dump(default, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"[persistência] não consegui criar {path}: {e}")

# ===== FUNÇÕES DE ARQUIVO (persistência) =====
def carregar_memoria():
    if os.path.exists(MEMORY_FILE):
        try:
            with open(MEMORY_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return set(data.get('vistos', []))
        except:
            return set()
    return set()

def salvar_memoria(vistos):
    try:
        with open(MEMORY_FILE, 'w', encoding='utf-8') as f:
            json.dump({'vistos': list(vistos), 'ultima_atualizacao': str(datetime.now())}, f, ensure_ascii=False)
    except:
        pass

def carregar_aprovados():
    """Carrega aprovados com log, fallback e sem engolir JSON quebrado em silêncio."""
    paths = [APROVADOS_FILE, APROVADOS_FILE + ".bak"]
    for path in paths:
        if not os.path.exists(path):
            continue
        try:
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            if isinstance(data, dict) and isinstance(data.get('aprovados'), list):
                return data.get('aprovados') or []
            if isinstance(data, list):
                return data
            print(f"[CRM/APROVADOS] formato inesperado em {path}: {type(data).__name__}")
        except Exception as e:
            print(f"[CRM/APROVADOS] erro lendo {path}: {e}")
    return []

def salvar_aprovados(aprovados):
    """Salva aprovados de forma atômica e cria backup para o CRM não sumir."""
    try:
        if os.path.exists(APROVADOS_FILE):
            try:
                shutil.copy2(APROVADOS_FILE, APROVADOS_FILE + ".bak")
            except Exception as be:
                print(f"[CRM/APROVADOS] aviso backup: {be}")
        tmp = APROVADOS_FILE + ".tmp"
        with open(tmp, 'w', encoding='utf-8') as f:
            json.dump({'aprovados': aprovados or [], 'ultima_atualizacao': str(datetime.now()), 'total': len(aprovados or [])}, f, ensure_ascii=False, indent=2)
            f.write("\n")
        os.replace(tmp, APROVADOS_FILE)
    except Exception as e:
        print(f"[CRM/APROVADOS] ERRO salvar_aprovados: {e}")

def carregar_reprovados():
    if os.path.exists(REPROVADOS_FILE):
        try:
            with open(REPROVADOS_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return data.get('reprovados', [])
        except:
            return []
    return []

# V58.33: TTL de 60 dias na blacklist de reprovados (canais antigos podem ser reprocessados)
REPROVADOS_TTL_DIAS = 60

def _data_reprovado_valida(data_str):
    """Tenta parsear data do reprovado. Retorna datetime ou None."""
    if not data_str:
        return None
    s = str(data_str).strip()
    # tenta vários formatos comuns
    for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s[:26] if '.' in s else s[:19], fmt)
        except Exception:
            continue
    return None

def limpar_reprovados_antigos(dias_ttl=None, dry_run=False):
    """V58.33: remove reprovados com mais de N dias da blacklist.

    Returns: dict com {total_antes, total_depois, removidos, mantidos}
    """
    dias_ttl = dias_ttl or REPROVADOS_TTL_DIAS
    limite = datetime.now() - timedelta(days=dias_ttl)
    reprovados = carregar_reprovados()
    total_antes = len(reprovados)
    mantidos = []
    removidos = 0
    for r in reprovados:
        if not isinstance(r, dict):
            continue
        data_r = _data_reprovado_valida(r.get('data'))
        if data_r is None:
            # sem data = mantém (conservador, pode ser reprovado manual recente)
            mantidos.append(r)
            continue
        if data_r >= limite:
            mantidos.append(r)
        else:
            removidos += 1
    if not dry_run and removidos > 0:
        try:
            with open(REPROVADOS_FILE, 'w', encoding='utf-8') as f:
                json.dump({'reprovados': mantidos, 'ultima_atualizacao': str(datetime.now())}, f, ensure_ascii=False)
        except Exception as e:
            print(f"[BLACKLIST TTL] erro salvando: {e}")
            return {"total_antes": total_antes, "total_depois": total_antes, "removidos": 0, "mantidos": total_antes, "erro": str(e)}
    return {
        "total_antes": total_antes,
        "total_depois": len(mantidos),
        "removidos": removidos,
        "mantidos": len(mantidos),
        "dias_ttl": dias_ttl,
        "dry_run": dry_run,
    }

def salvar_reprovado(canal):
    reprovados = carregar_reprovados()
    if not any(r['id'] == canal['id'] for r in reprovados):
        reprovados.append({
            'id': canal['id'],
            'nome': canal['nome'],
            'url': canal.get('url', f"https://youtube.com/channel/{canal['id']}"),
            'nicho': canal.get('nicho', ''),
            'score': canal.get('score', 0),
            'subs_fmt': canal.get('subs_fmt', ''),
            'motivo': canal.get('motivo', 'Nao qualificado'),
            'data': str(datetime.now())
        })
        try:
            with open(REPROVADOS_FILE, 'w', encoding='utf-8') as f:
                json.dump({'reprovados': reprovados, 'ultima_atualizacao': str(datetime.now())}, f, ensure_ascii=False)
        except:
            pass
    return reprovados

def salvar_reprovados_lista(canais):
    reprovados = carregar_reprovados()
    ids_existentes = {r['id'] for r in reprovados}
    for canal in canais:
        if canal['id'] not in ids_existentes:
            reprovados.append({
                'id': canal['id'],
                'nome': canal['nome'],
                'url': canal.get('url', f"https://youtube.com/channel/{canal['id']}"),
                'nicho': canal.get('nicho', ''),
                'score': canal.get('score', 0),
                'subs_fmt': canal.get('subs_fmt', ''),
                'motivo': canal.get('motivo', 'Nao qualificado'),
                'data': str(datetime.now())
            })
    try:
        with open(REPROVADOS_FILE, 'w', encoding='utf-8') as f:
            json.dump({'reprovados': reprovados, 'ultima_atualizacao': str(datetime.now())}, f, ensure_ascii=False)
    except:
        pass
    return reprovados

def ids_aprovados():
    return {c.get('id') for c in carregar_aprovados() if c.get('id')}

def ids_reprovados():
    return {c.get('id') for c in carregar_reprovados() if c.get('id')}

def _ler_json_lista(arquivo, chave):
    if not os.path.exists(arquivo):
        return []
    try:
        with open(arquivo, 'r', encoding='utf-8') as f:
            data = json.load(f)
        if isinstance(data, dict):
            return data.get(chave, []) or []
        if isinstance(data, list):
            return data
    except Exception as e:
        print(f"⚠️ Erro lendo {arquivo}: {e}")
    return []

def _id_canal(item):
    if isinstance(item, str):
        return item.strip()
    if isinstance(item, dict):
        return str(item.get('id') or item.get('channel_id') or '').strip()
    return ""

def unir_por_id(lista_a, lista_b):
    mapa = {}
    for item in (lista_a or []) + (lista_b or []):
        cid = _id_canal(item)
        if not cid:
            continue
        if isinstance(item, dict):
            mapa[cid] = {**mapa.get(cid, {}), **item, "id": cid}
        else:
            mapa[cid] = {"id": cid}
    return list(mapa.values())

def sincronizar_banco_dados(salvar=True):
    vistos_novo = set(_ler_json_lista(MEMORY_FILE, 'vistos'))
    vistos_antigo = set(_ler_json_lista("canais_vistos.json", 'vistos'))
    aprovados_novo = carregar_aprovados()
    aprovados_antigo = _ler_json_lista("canais_aprovados.json", 'aprovados')
    reprovados_novo = carregar_reprovados()
    reprovados_antigo = _ler_json_lista("canais_reprovados.json", 'reprovados')
    aprovados_unidos = unir_por_id(aprovados_novo, aprovados_antigo)
    reprovados_unidos = unir_por_id(reprovados_novo, reprovados_antigo)
    ids_aprovados = {_id_canal(c) for c in aprovados_unidos if _id_canal(c)}
    ids_reprovados = {_id_canal(c) for c in reprovados_unidos if _id_canal(c)}
    vistos_unidos = {v for v in (vistos_novo | vistos_antigo | ids_aprovados | ids_reprovados) if v}
    if salvar:
        salvar_memoria(vistos_unidos)
        salvar_aprovados(aprovados_unidos)
        with open(REPROVADOS_FILE, 'w', encoding='utf-8') as f:
            json.dump({'reprovados': reprovados_unidos, 'ultima_atualizacao': str(datetime.now()), 'total': len(reprovados_unidos)}, f, ensure_ascii=False, indent=2)
    return {
        "vistos": vistos_unidos,
        "aprovados": aprovados_unidos,
        "reprovados": reprovados_unidos,
        "ids_aprovados": ids_aprovados,
        "ids_reprovados": ids_reprovados,
        "blacklist_total": len(vistos_unidos),
        "vistos_arquivo": len(vistos_novo),
        "vistos_antigo": len(vistos_antigo),
        "aprovados_total": len(aprovados_unidos),
        "reprovados_total": len(reprovados_unidos),
    }

def ids_banco_dados():
    ids = set(estado.get('vistos', set())) | {c.get('id') for c in estado.get('aprovados', []) if c.get('id')} | {c.get('id') for c in estado.get('reprovados', []) if c.get('id')}
    try:
        ids |= _ids_de_lista_arquivo(LOTE_ATIVO_FILE)
        ids |= _ids_de_lista_arquivo(DLP_VERIFY_QUEUE_FILE, keys=("fila", "items", "candidatos"))
        ids |= _ids_de_lista_arquivo(CANAL_BRUTO_FILE, keys=("canais", "items"))
        ids |= _ids_de_lista_arquivo("qualificados_lista.json")
        ids |= _ids_de_lista_arquivo("thon_api_direct_qualified_last.json")
    except Exception:
        pass
    return {str(x) for x in ids if x}

def limpar_memoria():
    if os.path.exists(MEMORY_FILE):
        os.remove(MEMORY_FILE)
    if os.path.exists(APROVADOS_FILE):
        os.remove(APROVADOS_FILE)
    if os.path.exists(REPROVADOS_FILE):
        os.remove(REPROVADOS_FILE)
    return set(), [], []

# ===== ESTADO GLOBAL =====
garantir_arquivos_persistentes()
sync_inicial = sincronizar_banco_dados(salvar=True)
memoria_vistos = sync_inicial["vistos"]
aprovados_salvos = sync_inicial["aprovados"]
reprovados_salvos = sync_inicial["reprovados"]
lote_ativo_salvo = carregar_lote_ativo()

estado = {
    "rodando": False,
    "lote": lote_ativo_salvo,
    "aprovados": aprovados_salvos,
    "reprovados": reprovados_salvos,
    "vistos": memoria_vistos,
    "verificados": 0,
    "qualificados": 0,
    "status": "idle",
    "msg": "",
    "pausa": False,
    "queries_processadas": 0,
    "candidatos_encontrados": 0,
    "pipeline_stage": "idle",
    "qualificados_lista": list(lote_ativo_salvo),
    "motivos_reprovacao": {},
    "auto_mode": False,
    "api_quota_used": 0,
    "api_quota_budget": 9000,
    "api_engine": ENGINE_VERSION,
    "api_last_error": "",
    "api_channels_detailed": 0,
    "api_videos_collected": 0,
    "api_rejected": 0,
    "ytdlp_verified": 0,
    "ytdlp_approved": 0,
    "ytdlp_verify_max": 160,
    "engine_mode": "api",
    "api_discovery_only": True,
}
lock = threading.RLock()

def _sleep_interrompivel(total, step=0.1):
    """Dorme em fatias pequenas para /parar e /prospector/stop responderem rápido."""
    fim = time.time() + max(0, float(total or 0))
    while time.time() < fim:
        if not estado.get("rodando"):
            return False
        time.sleep(min(step, max(0, fim - time.time())))
    return True

def _sleep_auto_hunt(total, step=0.5):
    """Intervalo entre ciclos: interrompe pelo flag do Auto Hunt, não por estado['rodando']."""
    fim = time.time() + max(0, float(total or 0))
    while time.time() < fim:
        if not auto_hunt.get("enabled"):
            return False
        time.sleep(min(step, max(0, fim - time.time())))
    return True

print(f"Banco sincronizado: {sync_inicial['blacklist_total']} canais conhecidos/blacklist")
print(f"Memoria carregada: {len(memoria_vistos)} canais ja vistos")
print(f"Aprovados carregados: {len(aprovados_salvos)} canais")
print(f"Reprovados carregados: {len(reprovados_salvos)} canais")
print(f"Fila pendente carregada: {len(lote_ativo_salvo)} canais")

# ===== PALAVRAS NEGATIVAS =====
PALAVRAS_NEGATIVAS = [
    "cortes", "melhores momentos", "highlights", "resumo",
    "kids", "infantil", "crianca", "baby", "gameplay", "live", "twitch"
]

# ===== NICHOS + FORMATOS GERAIS =====
FORMATOS_CONTEUDO = [
    "podcast", "entrevista", "talkhead", "canal de cortes", "shorts", "reels",
    "conteudo vertical", "vlog", "documentario", "video aula", "conteudo longo",
    "criador de conteudo", "influenciador", "youtube channel"
]

FORMATOS_POR_MODO = {
    "longform": ["podcast", "entrevista", "talkhead", "vlog", "documentario", "video aula", "conteudo longo", "youtube channel"],
    "shortform": ["shorts", "reels", "conteudo vertical", "canal de cortes", "criador de conteudo", "influenciador"],
    "ambos": FORMATOS_CONTEUDO,
}

QUERIES_BASE = {
    "empreendedorismo": ["empreendedorismo brasil", "startup brasil", "ceo brasil", "fundador brasil", "empresario brasil", "lideranca empreendedora", "gestao de empresas", "cases de sucesso"],
    "marketing digital": ["marketing digital brasil", "trafego pago brasil", "vendas online", "social media", "ecommerce brasil", "seo brasil", "copywriting", "funil de vendas"],
    "negocios": ["negocios brasil", "gestao empresarial", "lideranca brasil", "estrategia de negocios", "administracao", "inovacao negocios", "produtividade"],
    "tecnologia": ["tecnologia brasil", "programacao", "dev portugues", "inteligencia artificial", "saas brasil", "startup tech", "software development", "data science"],
    "financas": ["financas brasil", "investimentos", "educacao financeira", "bolsa de valores", "financas pessoais", "economia brasil", "criptomoedas", "fundos imobiliarios"],
    "saude": ["saude brasil", "medicina", "bem estar", "nutricao", "fitness brasil", "saude mental", "psicologia", "vida saudavel"],
    "esportes": ["esportes brasil", "futebol", "fitness", "treinamento", "mma brasil", "corrida", "musculacao", "performance esportiva"],
    "relacionamento": ["relacionamento brasil", "relacionamento amoroso", "casamento", "autoconhecimento", "terapia casal", "vida a dois"],
    "religiao": ["cristao brasil", "igreja", "fe", "lideranca crista", "catolico", "biblia estudo", "pastor", "padre"],
    "humor": ["humor brasil", "comedia", "stand up brasil", "humoristas", "piadas", "comedia brasileira"],
    "educacao": ["educacao brasil", "professores", "pedagogia", "ensino", "aprendizagem", "universidade", "aula online"],
    "politica": ["politica brasil", "governo", "eleicoes", "camara", "senado", "analise politica", "politicos"]
}

def gerar_queries_expandidas(nicho, modo="ambos"):
    base = QUERIES_BASE.get(nicho, [f"{nicho} brasil"])
    formatos = FORMATOS_POR_MODO.get(modo, FORMATOS_CONTEUDO)
    queries = []
    for tema in base:
        queries.append(f"{tema} criador de conteudo")
        for formato in formatos:
            queries.append(f"{formato} {tema} brasil")
    vistas = set()
    limpas = []
    for q in queries:
        q = re.sub(r'\s+', ' ', q).strip()
        if q not in vistas:
            vistas.add(q)
            limpas.append(q)
    return limpas

def limitar_queries(todas_queries, query_limit):
    try:
        query_limit = int(query_limit)
    except:
        query_limit = QUERY_LIMIT_PADRAO
    query_limit = max(1, min(QUERY_LIMIT_MAX, query_limit))
    return todas_queries[:query_limit]

# ===== YT-DLP (verificação) =====
def _thon_ytdlp_base_cmd():
    cmd = [sys.executable, "-m", "yt_dlp", "--no-warnings", "--quiet", "--no-cache-dir"]
    cookie_browser = os.environ.get("THON_YTDLP_COOKIE_BROWSER", "none").strip().lower()
    if cookie_browser and cookie_browser not in {"none", "off", "0", "false"}:
        cmd += ["--cookies-from-browser", cookie_browser]
    return cmd

_403_state = {"count": 0, "last": 0.0}

def _registrar_403(stderr=""):
    now = time.time()
    if now - _403_state.get("last", 0) > 60:
        _403_state["count"] = 0
    _403_state["count"] += 1
    _403_state["last"] = now
    print(f"  [yt-dlp 403] pulado sem retry | seguidos={_403_state['count']}")
    _sleep_interrompivel(min(8.0, 1.5 + _403_state["count"] * 0.8))

def run_cmd(args, timeout=60):
    try:
        r = subprocess.run(_thon_ytdlp_base_cmd() + args,
                           capture_output=True, text=True, timeout=timeout)
        stderr = (r.stderr or "")
        if r.returncode != 0:
            if "403" in stderr or "Forbidden" in stderr:
                _registrar_403(stderr)
            elif stderr.strip():
                print(f"  [yt-dlp erro] {stderr.strip()[:160]}")
            return ""
        out = (r.stdout or "").strip()
        if out:
            _403_state["count"] = 0
        return out
    except subprocess.TimeoutExpired:
        print("  [yt-dlp timeout] pulado")
        return ""
    except Exception as e:
        print(f"  [yt-dlp exception] {e}")
        return ""

def run_lines(args, timeout=60):
    out = run_cmd(args, timeout=timeout)
    return [l for l in out.splitlines() if l.strip()]

def pegar_videos_canal(channel_id, variant=None):
    """Busca vídeos recentes do canal via yt-dlp. Parâmetro variant ignorado (compatibilidade)."""
    url = f"https://www.youtube.com/channel/{channel_id}"
    linhas = run_lines(["--flat-playlist", "--print", "%(duration)s\t%(view_count)s\t%(upload_date)s",
                        "--playlist-end", "15", url], timeout=50)
    if not linhas:
        return None
    longos = medios = shorts = total_views = total_videos = 0
    datas = []
    for linha in linhas:
        partes = linha.split("\t")
        if len(partes) >= 2:
            try:
                dur = int(float(partes[0])) if partes[0] else 0
                views = int(float(partes[1])) if len(partes) > 1 and partes[1] else 0
                data = partes[2] if len(partes) > 2 else ""
                if dur > 600:
                    longos += 1
                    if views > 0:
                        total_views += views
                        total_videos += 1
                elif dur > 180:
                    medios += 1
                    if views > 0:
                        total_views += views
                        total_videos += 1
                elif 0 < dur <= 60:
                    shorts += 1
                if data and len(data) >= 8:
                    datas.append(data[:8])
            except:
                continue
    avg_views = total_views // total_videos if total_videos > 0 else 0
    dias_ultimo = 999
    if datas:
        datas.sort(reverse=True)
        try:
            dt = datetime.strptime(datas[0], "%Y%m%d")
            dias_ultimo = (datetime.now() - dt).days
        except:
            pass
    return {"longos": longos, "shorts": shorts, "avg_views": avg_views, "dias_ultimo": dias_ultimo, "variant": variant or "default"}

def fmt_numero(n):
    try:
        n = int(n)
        if n >= 1_000_000:
            return f"{n/1e6:.1f}M"
        if n >= 1_000:
            return f"{n/1e3:.0f}K"
        return str(n)
    except:
        return "N/A"

def formatar_dias(dias):
    if dias == 0: return "hoje"
    if dias == 1: return "ontem"
    if dias < 30: return f"{dias}d"
    if dias < 365: return f"{dias//30}m"
    return f"{dias//365}a"

# ============================================================
# FUNÇÕES DA API (com rodízio e salvamento de brutos)
# ============================================================
API_DAILY_BUDGET = max(9000, len(_API_KEYS) * 9500)
API_SEARCH_COST = 100
API_OTHER_COST = 1
API_QUERY_LIMIT_MAX = QUERY_LIMIT_MAX
API_VIDEOS_PER_CHANNEL = 5
API_PAGES_PER_QUERY = 2
API_STATE_FILE = "youtube_api_quota_state.json"
QUERY_STATS_FILE = "query_stats.json"
API_DISCOVERY_SEEN_FILE = "api_discovery_seen.json"
QUERY_PERFORMANCE_FILE = "query_performance.json"
QUERY_MAX_PAGE = 8
QUERY_MIN_RODADAS_PARA_JULGAR = 2
QUERY_TAXA_SATURADA = 0.08
QUERY_TAXA_LIXO = 0.05
QUERY_COOLDOWN_SATURADA_DIAS = 7
QUERY_COOLDOWN_LIXO_DIAS = 15
API_SEEN_COOLDOWN_DIAS = 15
REPROVADOS_COOLDOWN_DIAS = 45

VETOS_PRE_FILTRO = [
    "cortes","highlights","melhores momentos","shorts canal","ministério","ministerio","secretaria","prefeitura",
    "câmara municipal","camara municipal","senado","tribunal","governo federal","governo estadual","governo municipal",
    "universidade","faculdade","instituto federal","senac","senai","fiap","sebrae","concurso público","concurso publico",
    "vestibular","enem","tecconcursos","qconcursos","estratégia concursos","culto","louvor","sermão","sermao","diocese",
    "suporte técnico","suporte tecnico","assistência técnica","assistencia tecnica","review produto","unboxing","gameplay",
    "lets play","canal oficial empresa","institucional","mitsubishi","samsung oficial","microsoft brasil oficial",
]
SINAIS_POSITIVOS_PRE = [
    "podcast","videocast","entrevista","bate papo","bate-papo","conversa","talk show","mesa redonda","episódio","episodio",
    "host","convidado","convidada","papo cast","cast",
]

def _qperf_load():
    data = _json_load_safe(QUERY_PERFORMANCE_FILE, {})
    return data if isinstance(data, dict) else {}

def _qperf_save(data):
    _json_write_atomic(QUERY_PERFORMANCE_FILE, data if isinstance(data, dict) else {})

def _qperf_key(query, source_type):
    return f"{source_type}::{query}"

def _qperf_get(data, query, source_type):
    key = _qperf_key(query, source_type)
    if key not in data or not isinstance(data.get(key), dict):
        data[key] = {"query": query, "source_type": source_type, "runs": 0, "last_run_at": None, "channels_found": 0, "channels_new": 0, "channels_repeated": 0, "pre_rejected": 0, "qualified": 0, "rejected": 0, "new_rate": 0, "qualification_rate": 0, "quota_used": 0, "status": "active", "cooldown_until": None}
    return data[key]

def _qperf_set_cooldown(data, query, source_type, dias):
    row = _qperf_get(data, query, source_type)
    row["status"] = "cooldown"
    row["cooldown_until"] = datetime.fromtimestamp(time.time() + int(dias) * 86400).isoformat(timespec="seconds")

def _qperf_em_cooldown(data, query, source_type):
    row = _qperf_get(data, query, source_type)
    dt = _parse_dt_any(row.get("cooldown_until"))
    return bool(dt and dt > datetime.now())

def _qperf_update(data, query, source_type, novos, repetidos, pre_reprovados, qualificados, quota_usada):
    row = _qperf_get(data, query, source_type)
    row["runs"] = int(row.get("runs", 0) or 0) + 1
    row["last_run_at"] = _dt_now_iso()
    row["channels_found"] = int(row.get("channels_found", 0) or 0) + int(novos or 0) + int(repetidos or 0) + int(pre_reprovados or 0)
    row["channels_new"] = int(row.get("channels_new", 0) or 0) + int(novos or 0)
    row["channels_repeated"] = int(row.get("channels_repeated", 0) or 0) + int(repetidos or 0)
    row["pre_rejected"] = int(row.get("pre_rejected", 0) or 0) + int(pre_reprovados or 0)
    row["qualified"] = int(row.get("qualified", 0) or 0) + int(qualificados or 0)
    row["quota_used"] = int(row.get("quota_used", 0) or 0) + int(quota_usada or 0)
    total = max(1, int(row["channels_found"] or 0))
    row["new_rate"] = round(row["channels_new"] / total, 4)
    row["qualification_rate"] = round(row["qualified"] / total, 4)
    row["status"] = "active"
    if row["new_rate"] == 0 and row["runs"] >= 1:
        _qperf_set_cooldown(data, query, source_type, 3)
    elif row["new_rate"] < 0.05 and row["runs"] >= 2:
        _qperf_set_cooldown(data, query, source_type, 7)
    elif row["qualification_rate"] == 0 and row["runs"] >= 3:
        _qperf_set_cooldown(data, query, source_type, 15)
    return row

def pre_filtro_comercial(ch, config):
    text = " ".join(str(ch.get(k, "")) for k in ("search_title", "search_desc", "source_video_title", "source_query", "nome", "title")).lower()
    for veto in VETOS_PRE_FILTRO:
        if veto in text:
            return False, f"veto:{veto}"
    perfis = set(config.get("perfis") or [])
    exige_podcast = bool(config.get("modo_brasil")) and (not perfis or perfis == {"podcast"} or "podcast" in perfis)
    if exige_podcast and not any(s in text for s in SINAIS_POSITIVOS_PRE):
        return True, "sem_sinal_podcast"
    return True, "ok"

def _multi_candidate(channel_id, source_type, query="", nicho="", **extra):
    now = _dt_now_iso()
    return {"id": channel_id, "channel_id": channel_id, "source_type": source_type, "source_query": query, "nicho": nicho, "first_seen_at": now, "last_seen_at": now, "seen_count": 1, **extra}

def _add_candidate_multi(out, cand, seen, banco, config):
    cid = cand.get("channel_id") or cand.get("id")
    if not cid or cid in seen or cid in banco:
        return "repetido"
    ok, motivo = pre_filtro_comercial(cand, config)
    if not ok:
        return motivo
    seen.add(cid)
    out.append(cand)
    _adicionar_bruto({**cand, "source": cand.get("source_type")})
    return "novo"

def _api_today():
    return datetime.now().strftime("%Y-%m-%d")

def _json_load_safe(path, default):
    try:
        if not os.path.exists(path):
            return default
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        try:
            shutil.move(path, f"{path}.quebrado_{int(time.time())}")
        except Exception:
            pass
        print(f"[json safe] {path} quebrado, iniciando vazio: {e}")
        return default

def _json_write_atomic(path, data):
    try:
        if os.path.exists(path):
            shutil.copy2(path, path + ".bak")
    except Exception:
        pass
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")
    os.replace(tmp, path)

def _dt_now_iso():
    return datetime.now().isoformat(timespec="seconds")

def _parse_dt_any(value):
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value).replace("Z", ""))
    except Exception:
        try:
            return datetime.strptime(str(value)[:19], "%Y-%m-%d %H:%M:%S")
        except Exception:
            return None

def _days_since_any(value, default=999999):
    dt = _parse_dt_any(value)
    if not dt:
        return default
    return max(0, (datetime.now() - dt).days)

def carregar_query_stats():
    data = _json_load_safe(QUERY_STATS_FILE, {})
    return data if isinstance(data, dict) else {}

def salvar_query_stats(stats):
    _json_write_atomic(QUERY_STATS_FILE, stats if isinstance(stats, dict) else {})

def carregar_api_seen():
    data = _json_load_safe(API_DISCOVERY_SEEN_FILE, {})
    return data if isinstance(data, dict) else {}

def salvar_api_seen(seen):
    _json_write_atomic(API_DISCOVERY_SEEN_FILE, seen if isinstance(seen, dict) else {})

def _ids_de_lista_arquivo(path, keys=("lote", "qualificados", "canais", "items")):
    data = _json_load_safe(path, [] if path.endswith(".json") else [])
    itens = data if isinstance(data, list) else []
    if isinstance(data, dict):
        for k in keys:
            if isinstance(data.get(k), list):
                itens = data.get(k)
                break
    return {str(x.get("id") or x.get("channel_id")) for x in itens if isinstance(x, dict) and (x.get("id") or x.get("channel_id"))}

def _reprovado_em_cooldown(cid):
    for r in carregar_reprovados():
        if not isinstance(r, dict) or str(r.get("id") or r.get("channel_id")) != str(cid):
            continue
        return _days_since_any(r.get("data") or r.get("updated_at") or r.get("created_at"), 999999) < REPROVADOS_COOLDOWN_DIAS
    return False

def discovery_channel_ja_visto(channel_id, contexto=None):
    cid = str(channel_id or "")
    if not cid:
        return "sem_id"
    if cid in _ids_de_lista_arquivo(LOTE_ATIVO_FILE):
        return "lote"
    if cid in ids_aprovados():
        return "aprovados"
    if cid in _ids_de_lista_arquivo("qualificados_lista.json"):
        return "qualificados_lista"
    if cid in _ids_de_lista_arquivo("thon_api_direct_qualified_last.json"):
        return "ultimo_qualificado"
    seen = carregar_api_seen()
    if cid in seen and _days_since_any((seen.get(cid) or {}).get("last_seen_at"), 999999) < API_SEEN_COOLDOWN_DIAS:
        return "api_seen_cooldown"
    if _reprovado_em_cooldown(cid):
        return "reprovado_cooldown"
    return ""

def registrar_api_seen(channel_id, query, pagina):
    cid = str(channel_id or "")
    if not cid:
        return
    seen = carregar_api_seen()
    now = _dt_now_iso()
    row = seen.get(cid) if isinstance(seen.get(cid), dict) else {}
    seen[cid] = {
        "first_seen_at": row.get("first_seen_at") or now,
        "last_seen_at": now,
        "count": int(row.get("count", 0) or 0) + 1,
        "last_query": query,
        "last_page": pagina,
    }
    salvar_api_seen(seen)

def query_decide_pagina(query, stats):
    qst = stats.get(query) if isinstance(stats.get(query), dict) else {}
    cooldown = qst.get("cooldown_ate")
    if cooldown and _parse_dt_any(cooldown) and _parse_dt_any(cooldown) > datetime.now():
        return None, None
    paginas = qst.get("paginas") if isinstance(qst.get("paginas"), dict) else {}
    for pagina in range(1, QUERY_MAX_PAGE + 1):
        pst = paginas.get(str(pagina)) if isinstance(paginas.get(str(pagina)), dict) else {}
        if not pst:
            prev = paginas.get(str(pagina - 1), {}) if pagina > 1 else {}
            return pagina, prev.get("ultimo_token")
        rodadas = int(pst.get("rodadas", 0) or 0)
        taxa = float(pst.get("taxa_novos", 1) or 0)
        if rodadas < QUERY_MIN_RODADAS_PARA_JULGAR or taxa >= QUERY_TAXA_SATURADA:
            prev = paginas.get(str(pagina - 1), {}) if pagina > 1 else {}
            return pagina, None if pagina == 1 else prev.get("ultimo_token")
    qst["cooldown_ate"] = (datetime.now().timestamp() + QUERY_COOLDOWN_SATURADA_DIAS * 86400)
    qst["cooldown_ate"] = datetime.fromtimestamp(qst["cooldown_ate"]).isoformat(timespec="seconds")
    stats[query] = qst
    salvar_query_stats(stats)
    return None, None

def query_registrar_resultado(query, pagina, novos, repetidos, qualificados, next_token):
    stats = carregar_query_stats()
    qst = stats.setdefault(query, {"paginas": {}, "cooldown_ate": None, "prioridade": 0, "total_qualificados": 0})
    paginas = qst.setdefault("paginas", {})
    pst = paginas.setdefault(str(pagina), {"rodadas": 0, "novos": 0, "repetidos": 0, "qualificados": 0})
    teve_busca = int(novos or 0) + int(repetidos or 0) > 0
    if teve_busca:
        pst["rodadas"] = int(pst.get("rodadas", 0) or 0) + 1
        pst["novos"] = int(pst.get("novos", 0) or 0) + int(novos or 0)
        pst["repetidos"] = int(pst.get("repetidos", 0) or 0) + int(repetidos or 0)
        pst["ultimo_token"] = next_token
        pst["ultima_rodada"] = _dt_now_iso()
    pst["qualificados"] = int(pst.get("qualificados", 0) or 0) + int(qualificados or 0)
    total = max(1, pst["novos"] + pst["repetidos"])
    pst["taxa_novos"] = round(pst["novos"] / total, 4)
    pst["taxa_qualificacao"] = round(pst["qualificados"] / total, 4)
    if teve_busca and pst["rodadas"] >= QUERY_MIN_RODADAS_PARA_JULGAR and pst["taxa_novos"] < QUERY_TAXA_LIXO:
        qst["cooldown_ate"] = datetime.fromtimestamp(time.time() + QUERY_COOLDOWN_LIXO_DIAS * 86400).isoformat(timespec="seconds")
    qst["total_qualificados"] = sum(int((p or {}).get("qualificados", 0) or 0) for p in paginas.values() if isinstance(p, dict))
    qst["prioridade"] = round(max((float((p or {}).get("taxa_novos", 0) or 0) + float((p or {}).get("taxa_qualificacao", 0) or 0) * 2) for p in paginas.values() if isinstance(p, dict)), 4) if paginas else 0
    salvar_query_stats(stats)
    return pst

def query_atualizar_qualificados_por_origem(qualificados):
    grupos = {}
    for c in qualificados or []:
        q = c.get("_discovery_query") or c.get("found_query") or c.get("query")
        p = int(c.get("_discovery_page") or 1)
        if q:
            grupos[(q, p)] = grupos.get((q, p), 0) + 1
    for (q, p), qtd in grupos.items():
        query_registrar_resultado(q, p, 0, 0, qtd, None)
    if grupos:
        print(f"[QUERY STATS] qualificados registrados | queries={len(grupos)} | total={sum(grupos.values())}")

def ordenar_queries_inteligente(queries, stats):
    now = datetime.now()
    def score(item):
        q = item[0]
        st = stats.get(q) if isinstance(stats.get(q), dict) else {}
        cd = _parse_dt_any(st.get("cooldown_ate"))
        if cd and cd > now:
            return -999
        return float(st.get("prioridade", 0) or 0)
    return sorted(queries, key=score, reverse=True)

def _api_load_state():
    try:
        if os.path.exists(API_STATE_FILE):
            with open(API_STATE_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            if data.get("date") == _api_today():
                return data
    except Exception:
        pass
    return {"date": _api_today(), "used": 0, "calls": 0, "history": []}

def _api_save_state(data):
    try:
        data["date"] = _api_today()
        data["updated_at"] = str(datetime.now())
        hist = data.get("history") or []
        data["history"] = hist[-80:]
        with open(API_STATE_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"[api quota] erro salvando estado: {e}")

def _api_quota_used():
    st = _api_load_state()
    return int(st.get("used", 0) or 0)

def _api_consume_quota(cost, endpoint):
    st = _api_load_state()
    used = int(st.get("used", 0) or 0)
    if used + cost > API_DAILY_BUDGET:
        return False, used
    st["used"] = used + cost
    st["calls"] = int(st.get("calls", 0) or 0) + 1
    hist = st.setdefault("history", [])
    hist.append({"time": str(datetime.now()), "endpoint": endpoint, "cost": cost, "used": st["used"]})
    _api_save_state(st)
    return True, st["used"]

def youtube_api_get(endpoint, params, cost=None, quiet_404=False, max_tentativas=None):
    if cost is None:
        cost = API_SEARCH_COST if endpoint == "search" else API_OTHER_COST
    ok, used = _api_consume_quota(cost, endpoint)
    if not ok:
        raise RuntimeError(f"Limite interno da API atingido: {used}/{API_DAILY_BUDGET}")

    total_chaves = max(1, len(_API_KEYS))
    tentativas = total_chaves if max_tentativas is None else max(1, min(int(max_tentativas or 1), total_chaves))
    for tentativa in range(tentativas):
        key = _api_proxima_chave()
        if not key:
            break
        params_copy = dict(params or {})
        params_copy["key"] = key
        import urllib.parse
        url = "https://www.googleapis.com/youtube/v3/" + endpoint + "?" + urllib.parse.urlencode(params_copy)
        curl_bin = shutil.which("curl") or "/usr/bin/curl"
        try:
            out = subprocess.check_output(
                [curl_bin, "-sS", "-L", "-w", "\n__HTTP_STATUS__:%{http_code}", url],
                stderr=subprocess.STDOUT,
                timeout=45,
            ).decode("utf-8", errors="replace")
            marker = "\n__HTTP_STATUS__:"
            if marker in out:
                body, status_raw = out.rsplit(marker, 1)
                try:
                    status = int(status_raw.strip())
                except Exception:
                    status = 0
            else:
                body, status = out, 0
            try:
                data = json.loads(body) if body.strip() else {}
            except Exception:
                data = {"raw": body[:1000]}
            if status >= 400:
                reason = ""
                try:
                    reason = data.get("error", {}).get("errors", [{}])[0].get("reason", "")
                except Exception:
                    reason = ""
                msg = data.get("error", {}).get("message", "") if isinstance(data, dict) else ""
                if status == 404 and quiet_404:
                    return {"items": [], "_http_status": 404, "_error": msg or "404"}
                # V58.20: separar erro de quota/chave de erro normal da fonte.
                # subscriptionForbidden/forbidden/private/notFound não podem matar a API key.
                quota_reasons = {"quotaExceeded", "dailyLimitExceeded", "rateLimitExceeded", "keyInvalid", "accessNotConfigured"}
                benign_reasons = {"subscriptionForbidden", "forbidden", "notFound", "playlistNotFound", "videoNotFound", "channelNotFound"}
                if reason in quota_reasons:
                    _api_marcar_chave_falha(key, reason or f"HTTP {status}")
                    print(f"[API] chave {_api_key_id(key)} quota/key, tentando próxima ({tentativa+1}/{tentativas}) | reason={reason}")
                    continue
                if reason in benign_reasons or status in {400, 401, 404}:
                    if not quiet_404:
                        print(f"[api aviso] {endpoint} HTTP {status} | {reason} | {str(msg)[:180]}")
                    return {"items": [], "_http_status": status, "_error": msg or reason or f"HTTP {status}", "_reason": reason}
                if status == 403:
                    # 403 genérico: não marque todas as chaves como mortas sem reason de quota.
                    print(f"[api aviso] {endpoint} HTTP 403 não-quota | {reason} | {str(msg)[:180]}")
                    return {"items": [], "_http_status": status, "_error": msg or reason or f"HTTP {status}", "_reason": reason}
                print(f"[api erro] {endpoint} HTTP {status} | {reason} | {str(msg)[:180]}")
                return {"items": [], "_http_status": status, "_error": msg or reason or f"HTTP {status}", "_reason": reason}
            return data
        except subprocess.TimeoutExpired:
            print(f"[api timeout] {endpoint}")
            continue
        except Exception as e:
            print(f"[api exception] {endpoint}: {e}")
            continue
    raise RuntimeError("Todas as chaves da API falharam (quota/403).")

def _api_num(x):
    try:
        return int(x)
    except Exception:
        return None

def _api_parse_duration_seconds(iso):
    if not iso:
        return 0
    m = re.match(r"PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?", str(iso))
    if not m:
        return 0
    h = int(m.group(1) or 0)
    mi = int(m.group(2) or 0)
    s = int(m.group(3) or 0)
    return h * 3600 + mi * 60 + s

def _api_days_since(date_str):
    if not date_str:
        return None
    try:
        dt = datetime.strptime(date_str[:19], "%Y-%m-%dT%H:%M:%S")
        return max(0, (datetime.utcnow() - dt).days)
    except Exception:
        return None

def _api_clean_query(q):
    q = re.sub(r"\s+", " ", str(q or "").strip().lower())
    lixo = {"tier5", "aprendizado", "persistente"}
    palavras = [p for p in q.split() if p not in lixo]
    out = []
    for p in palavras:
        if out and out[-1] == p:
            continue
        out.append(p)
    q = " ".join(out).strip()
    return q

API_QUERY_TEMPLATES = [
    "{nicho} podcast", "podcast {nicho} brasil", "{nicho} podcast brasileiro", "entrevista {nicho}",
    "{nicho} entrevista podcast", "bate papo {nicho}", "cast {nicho}", "canal {nicho} podcast",
    "programa {nicho} youtube", "talk show {nicho} brasil", "conversa {nicho}", "papo {nicho}",
    "{nicho} ao vivo youtube", "mesa redonda {nicho}", "podcast de {nicho}", "especialista {nicho} podcast",
    "{nicho} com entrevistados", "voices {nicho}", "youtube {nicho} brasil", "{nicho} criador de conteudo",
]

API_NICHO_QUERIES = {
    "empreendedorismo": ["como abrir empresa podcast brasil","founder podcast brasil","ceo entrevista youtube","case de sucesso empreendedor","empreendedor de sucesso podcast","startup founder brasil podcast","scale up brasil podcast","empreendedorismo digital podcast","negocio do zero podcast","empresario de sucesso entrevista"],
    "marketing digital": ["trafego pago podcast brasil","growth hacking podcast","vendas digitais podcast","funil de vendas entrevista","copy podcast brasil","email marketing podcast","performance digital podcast","social media podcast brasileiro","influencer marketing podcast","lançamento digital podcast"],
    "negocios": ["gestao empresarial podcast","lideranca corporativa podcast","rh podcast brasil","vendas b2b podcast","cultura organizacional podcast","estrategia empresarial entrevista","gestao de pessoas podcast","business brasil podcast","corporativo podcast","operacoes empresariais podcast"],
    "tecnologia": ["dev podcast brasil","programador podcast","inteligencia artificial podcast","machine learning brasil podcast","saas podcast brasil","dados podcast","cloud computing podcast","ciberseguranca podcast","dev tools podcast","tech talks brasil"],
    "financas": ["educacao financeira podcast","investimentos podcast brasil","bolsa de valores podcast","fundos imobiliarios podcast","renda fixa podcast","criptomoedas podcast brasil","finanças pessoais podcast","planejamento financeiro podcast","independencia financeira podcast","dinheiro podcast brasil"],
    "politica": ["politica brasil podcast","podcast politica brasileira","analise politica podcast brasil","politica entrevista podcast","debate politico podcast brasil","geopolitica podcast brasil","eleicoes brasil podcast","congresso nacional podcast","deputado podcast brasil","senado podcast brasil","cientista politico podcast","jornalismo politico podcast","seguranca publica podcast brasil","reforma tributaria podcast brasil","historia politica podcast brasil"],
    "saude": ["saude podcast brasil","medicina podcast brasileiro","medico podcast brasil","saude mental podcast brasil","psicologia podcast brasil","nutricao podcast brasil","fitness podcast brasil","medicina preventiva podcast brasil","cardiologia podcast brasil","psiquiatria podcast brasil","longevidade podcast brasil","biohacking podcast brasil","dentista podcast brasil","entrevista medico brasil"],
    "esportes": ["esportes podcast brasil","futebol podcast brasil","jogador futebol podcast","basquete podcast brasil","mma podcast brasil","jiu jitsu podcast brasil","crossfit podcast brasil","fisiculturismo podcast brasil","atleta entrevista podcast brasil","performance esportiva podcast","marketing esportivo podcast","gestao esportiva podcast brasil","e-sports podcast brasil","mercado esportivo podcast brasil"],
    "relacionamento": ["relacionamento podcast brasil","relacionamento amoroso podcast","casal podcast brasil","casamento podcast brasil","terapia casal podcast brasil","familia podcast brasil","autoconhecimento podcast brasil","inteligencia emocional podcast","sexualidade podcast brasil","comportamento humano podcast brasil","entrevista casal podcast brasil"],
    "religiao": ["religiao podcast brasil","fe podcast brasil","espiritualidade podcast brasil","crista podcast brasil","biblia podcast brasil","pastor podcast brasil","padre podcast brasil","catolicismo podcast brasil","evangelico podcast brasil","teologia podcast brasil","lideranca crista podcast brasil","entrevista pastor youtube","entrevista lider religioso podcast"],
    "humor": ["humor podcast brasil","comedia podcast brasileiro","stand up podcast brasil","comediante podcast brasil","humorista podcast brasil","entretenimento podcast brasil","youtuber humor podcast brasil","celebridade podcast brasil","talk show podcast brasil","podcast descontraido brasil","humor inteligente podcast brasil"],
    "educacao": ["educacao podcast brasil","professor podcast brasil","ensino podcast brasil","pedagogia podcast brasil","universidade podcast brasil","pesquisa academica podcast brasil","concurso publico podcast brasil","edtech podcast brasil","ensino online podcast brasil","youtube educativo brasil","entrevista professor youtube brasil"],
    "gamer": ["gamer podcast brasil","games podcast brasileiro","videogame podcast brasil","jogos podcast brasil","game design podcast brasil","desenvolvimento games podcast brasil","rpg podcast brasil","valorant podcast brasil","minecraft podcast brasil","esport podcast brasil","streamer podcast brasil","criador conteudo gamer podcast","mercado jogos podcast brasil","geek nerd podcast brasil"],
}

API_PERFIL_QUERIES = {
    "medico": ["medico podcast brasil","doutor podcast brasil","medico empreendedor podcast","healthtech podcast brasil"],
    "advogado": ["advogado podcast brasil","juridico podcast brasil","direito podcast brasil","legaltech podcast brasil"],
    "engenheiro": ["engenheiro podcast brasil","engenharia podcast brasil","construcao civil podcast brasil","engenheiro empreendedor podcast"],
    "corretor": ["corretor imoveis podcast brasil","mercado imobiliario podcast","investimento imobiliario podcast","corretor empreendedor podcast"],
    "personal": ["personal trainer podcast brasil","treinamento funcional podcast","negocio fitness podcast","academia podcast brasil"],
    "consultor": ["consultor podcast brasil","consultoria podcast brasil","mentor negocios podcast brasil","consultor digital podcast brasil"],
    "psicologo": ["psicologo podcast brasil","terapeuta podcast brasil","psicanalista podcast brasil","terapia online podcast brasil"],
    "nutricionista": ["nutricionista podcast brasil","nutricao podcast brasil","nutricao esportiva podcast","nutricionista empreendedora podcast"],
    "dentista": ["dentista podcast brasil","odontologia podcast brasil","clinica odontologica podcast","marketing odontologia podcast"],
    "arquiteto": ["arquiteto podcast brasil","arquitetura podcast brasil","design interiores podcast brasil","bim podcast brasil"],
    "contador": ["contador podcast brasil","contabilidade podcast brasil","tributario podcast brasil","contabilidade digital podcast"],
    "gamer": API_NICHO_QUERIES["gamer"],
}

def gerar_queries_api_first(nichos, modo="ambos", extras=None, limit=80, perfis=None):
    templates = API_QUERY_TEMPLATES
    queries = []
    for q in (extras or []):
        q = _api_clean_query(q)
        if q:
            queries.append((q, "manual"))
    for nicho in nichos:
        nicho_limpo = str(nicho or "").strip().lower()
        if not nicho_limpo:
            continue
        for tpl in templates:
            q = _api_clean_query(tpl.format(nicho=nicho_limpo))
            if q:
                queries.append((q, nicho_limpo))
        for q in API_NICHO_QUERIES.get(nicho_limpo, []):
            q = _api_clean_query(q)
            if q:
                queries.append((q, nicho_limpo))
    for perfil in (perfis or []):
        perfil_limpo = str(perfil or "").strip().lower()
        for q in API_PERFIL_QUERIES.get(perfil_limpo, []):
            q = _api_clean_query(q)
            if q:
                queries.append((q, perfil_limpo))
    seen, out = set(), []
    for q, nicho in queries:
        if q in seen:
            continue
        seen.add(q)
        out.append((q, nicho))
        if len(out) >= limit:
            break
    return out

def obter_queries_para_engine(config=None, limit=None, formato="tuplas", fallback_kind="api"):
    """Adaptador único: engines continuam recebendo o formato antigo.

    A Query Factory é a fonte principal. As queries hardcoded antigas ficam
    apenas como fallback de segurança, sem mudar pipeline/API/DLP.
    """
    config = dict(config or {})
    extras = [_api_clean_query(q) for q in (config.get("queries_extras") or []) if _api_clean_query(q)]
    nichos = config.get("nichos") or ["empreendedorismo"]
    perfis = config.get("perfis") or []
    try:
        limit = int(limit or config.get("query_limit") or QUERY_LIMIT_PADRAO)
    except Exception:
        limit = QUERY_LIMIT_PADRAO
    limit = max(1, min(int(limit), max(QUERY_LIMIT_MAX, 1000)))

    queries = []
    seen = set()
    for q in extras:
        if q and q not in seen:
            seen.add(q)
            queries.append((q, "manual"))

    factory_limit = max(1, limit - len(queries))
    if query_factory_get_next_queries:
        try:
            factory_queries = query_factory_get_next_queries(config=config, limit=factory_limit, as_tuples=True)
            for q, nicho in factory_queries:
                q = _api_clean_query(q)
                nicho = str(nicho or "factory").strip().lower()
                if q and q not in seen:
                    seen.add(q)
                    queries.append((q, nicho))
        except Exception as e:
            print(f"[QUERY FACTORY] falhou, fallback antigo ativo: {e}")

    if len(queries) < limit:
        if fallback_kind == "dlp":
            legacy = []
            for nicho in nichos or ["empreendedorismo"]:
                n = str(nicho or "").strip().lower()
                for tpl in DLP_EXTRA_TEMPLATES if "DLP_EXTRA_TEMPLATES" in globals() else []:
                    legacy.append((_api_clean_query(tpl.format(nicho=n)), n))
                if "empreendedor" in n or "negocio" in n or "negócio" in n:
                    for q in DLP_GENERAL_EXPANSIONS if "DLP_GENERAL_EXPANSIONS" in globals() else []:
                        legacy.append((_api_clean_query(q), n))
        else:
            legacy = gerar_queries_api_first(nichos, extras=[], limit=limit, perfis=perfis)
        for q, nicho in legacy:
            if q and q not in seen:
                seen.add(q)
                queries.append((q, nicho))
            if len(queries) >= limit:
                break

    print(f"[QUERY FACTORY] entregues={len(queries)} | manuais={len(extras)} | formato={formato} | fallback={fallback_kind}")
    if formato == "strings":
        return [q for q, _n in queries]
    return queries

NEGATIVOS_FORTES_API = [
    # Original
    "cortes", "corte ", "clips", "clip ", "melhores momentos", "highlights", "resumo",
    "notícias", "noticias", "news", "rádio", "radio", "tv ", "ao vivo", "live", "gameplay",
    # V58.26: vlog/lifestyle sem orcamento
    "vlog", "rotina", "dia a dia", "mukbang", "estilo de vida", "lifestyle",
    "minha rotina", "o que faco no dia", "rotina matinal",
    # React / low effort
    "react", "reaction", "reagindo", "reacao",
    # Gamers que nao pagam
    "free fire gameplay", "minecraft gameplay", "fortnite gameplay", "gta gameplay",
    "lets play", "let's play", "jogando",
    # Musicais (selos)
    "clip oficial", "clipe oficial", " videoclipe", "music video", "music video oficial",
    # Tutoriais genéricos sem $$ (DIY/beleza)
    "tutorial maquiagem", "diy", "como fazer ", "passo a passo maquiagem",
    # Vendedores de curso (nao sao leads - sao concorrentes de venda)
    "curso gratuito", "aula gratuita", "mentor gratuito", "mentoria gratuita",
    "clique no link", "link na bio", "baixe agora",
]

GRINGO_FORTE_API = [
    "daily marketing tips", "marketing school", "with a twist", "chisme", "constructora", "escuela", "negocios a la", "negocios à portuguesa",
    "product hackers", "podcast en español", "en español", "emprendimiento", "mercado latino",
    "portuguese", "portuguesa", "à portuguesa", "a la", "latino", "mexico", "argentina", "colombia"
]


# ============================================================
# REGIOES / IDIOMAS — V58.23-TERRITORIOS
# Mapa de territorio -> (regionCode YouTube, relevanceLanguage BCP-47, label)
# Garante que Portugal (PT) use pt-PT e filtre portugues europeu (nao ingles)
# ============================================================
REGION_LANG_MAP = {
    "BR":     {"regionCode": "BR", "relevanceLanguage": "pt-BR", "lang_query_suffix": "brasil",         "label": "Brasil"},
    "PT":     {"regionCode": "PT", "relevanceLanguage": "pt-PT", "lang_query_suffix": "portugal",       "label": "Portugal"},
    "US":     {"regionCode": "US", "relevanceLanguage": "en",    "lang_query_suffix": "",               "label": "EUA"},
    "MX":     {"regionCode": "MX", "relevanceLanguage": "es-MX", "lang_query_suffix": "mexico",         "label": "Mexico"},
    "AR":     {"regionCode": "AR", "relevanceLanguage": "es-AR", "lang_query_suffix": "argentina",      "label": "Argentina"},
    "ES":     {"regionCode": "ES", "relevanceLanguage": "es-ES", "lang_query_suffix": "espana",         "label": "Espanha"},
    "CO":     {"regionCode": "CO", "relevanceLanguage": "es-CO", "lang_query_suffix": "colombia",       "label": "Colombia"},
    "CL":     {"regionCode": "CL", "relevanceLanguage": "es-CL", "lang_query_suffix": "chile",          "label": "Chile"},
    "LATAM":  {"regionCode": None, "relevanceLanguage": "es",    "lang_query_suffix": "latinoamerica",  "label": "LATAM"},
    "GLOBAL": {"regionCode": None, "relevanceLanguage": None,    "lang_query_suffix": "",               "label": "Global"},
}

def _region_lang_params_list(config):
    """Retorna lista de dicts {regionCode, relevanceLanguage, lang_query_suffix, label} para cada territorio marcado.

    Se config['regioes'] estiver vazio, usa [BR] como default (comportamento legado).
    Garante que Portugal (PT) use pt-PT e filtre videos em portugues europeu.
    """
    regioes = config.get("regioes") or []
    if isinstance(regioes, str):
        regioes = [r.strip() for r in regioes.split(",") if r.strip()]
    regioes = [str(r).strip().upper() for r in regioes if str(r).strip()]
    if not regioes:
        regioes = ["BR"]
    out, seen = [], set()
    for r in regioes:
        r = r.upper()
        if r in seen:
            continue
        seen.add(r)
        cfg = REGION_LANG_MAP.get(r)
        if cfg:
            out.append(dict(cfg))
        else:
            out.append(dict(REGION_LANG_MAP["BR"]))
    return out

def _apply_region_to_params(params, region_cfg):
    """Aplica regionCode/relevanceLanguage de region_cfg em params (mutavel)."""
    if not region_cfg:
        return params
    rc = region_cfg.get("regionCode")
    rl = region_cfg.get("relevanceLanguage")
    if rc:
        params["regionCode"] = rc
    elif "regionCode" in params:
        del params["regionCode"]
    if rl:
        params["relevanceLanguage"] = rl
    elif "relevanceLanguage" in params:
        del params["relevanceLanguage"]
    return params



# ============================================================
# FILTRO ANTI-GRINGO V58.24
# Heurística de detecção de idioma baseada em título + descrição + nome.
# O yt-dlp (modo DLP) ignora regionCode/relevanceLanguage da YouTube API,
# então precisamos filtrar pós-busca os canais que não são do idioma alvo.
# ============================================================

# Sinais fortes de português (Brasil + Portugal)
SINAIS_PT_FORTES = [
    # Brasileirismos
    "brasil", "brasileiro", "brasileira", "são paulo", "rio de janeiro", "minas gerais",
    "bahia", "curitiba", "brasília", "brasilia", "portugues", "português",
    # Conectores e palavras functionais PT
    "para que", "gente", "cara", "negócio", "negócios", "negocios", "empresa",
    "empreendedor", "empreendedora", "mercado", "vendas", "vender", "cliente",
    "sucesso", "sucesso", "dicas", "canal", "conteúdo", "conteudo", "inscreva",
    "inscrição", "video", "vídeo", "episódio", "episodio", "entrevista",
    "convidado", "convidada", "papo", "conversa", "bate papo", "bate-papo",
    # Portugal
    "portugal", "português europeu", "tuga", "tuíte", "cá", "estou em", "olá pessoal",
    "boas", "fixe", "giro", "bué", "mas que", "estádio", "metro", "telemóvel",
    # Expressões
    "se inscreva", "deixa o like", "deixe o like", "comenta ai", "comente aqui",
    "deixa nos comentários", "até o próximo", "até o proximo", "um abraço",
    "fala pessoal", "e ai pessoal", "e ai galera",
]

# Sinais fortes de inglês (gringo)
SINAIS_EN_FORTES = [
    "subscribe", "subscribers", "sub to", "like and subscribe", "hit the bell",
    "comment down below", "in this video", "today we", "today i", "hey guys",
    "what's up", "whats up", "welcome back", "welcome to my channel",
    "lets talk about", "let\'s talk", "let\'s dive", "let\'s get",
    "podcast", "podcasts", "show", "episode", "episodes", "interview",
    "interviews", "guest", "host", "business", "entrepreneur", "marketing",
    "finance", "investing", "tech", "technology", "money", "income",
    "startup", "founder", "ceo", "growth", "leadership", "mindset",
    "success", "tips", "how to", "tutorial", "guide", "review",
    "watch", "watching", "watch now", "check out", "check this out",
    "join me", "follow me", "instagram", "twitter", "facebook",
    "morning", "evening", "weekly", "daily", "monthly",
    "the best", "the most", "in the world", "in america",
    "amazing", "awesome", "incredible", "fantastic", "guys",
    "english", "en español", "español", "suscríbete", "suscribete",
]

# Sinais de espanhol (LATAM/Espanhol) — só conta como gringo se territorio != LATAM/ES/MX/AR/CO/CL
SINAIS_ES_FORTES = [
    "hola", "buenos días", "buenas tardes", "buenas noches", "bienvenido",
    "bienvenida", "suscríbete", "suscribete", "me gusta", "comenta",
    "empresario", "emprendedor", "negocio", "mercado", "cliente",
    "éxito", "exito", "consejos", "canal", "contenido", "vídeo",
    "video", "episodio", "entrevista", "invitado", "invitada",
    "charla", "conversación", "conversacion", "plática", "platica",
    "en español", "mexicano", "mexicana", "argentino", "argentina",
    "colombiano", "colombiana", "chileno", "chilena", "español",
]

# Lista de territórios que aceitam português
TERRITORIOS_PT = {"BR", "PT"}
# Territórios que aceitam espanhol
TERRITORIOS_ES = {"MX", "AR", "ES", "CO", "CL", "LATAM"}
# Territórios que aceitam inglês
TERRITORIOS_EN = {"US", "GLOBAL"}

def _detectar_idioma_canal(nome, descricao, uploader=""):
    """Detecta idioma predominante do canal baseado em heurística.

    Retorna: {'pt': int, 'en': int, 'es': int, 'dominante': 'pt'|'en'|'es'|'neutro'}
    """
    texto = " ".join([str(nome or ""), str(descricao or ""), str(uploader or "")]).lower()
    if not texto.strip():
        return {"pt": 0, "en": 0, "es": 0, "dominante": "neutro"}

    pt = sum(1 for s in SINAIS_PT_FORTES if s in texto)
    en = sum(1 for s in SINAIS_EN_FORTES if s in texto)
    es = sum(1 for s in SINAIS_ES_FORTES if s in texto)

    # Bonus: palavras acentuadas PT sao forte indicio
    pt_acentuadas = len(re.findall(r"[àáâãéêíóôõúçÀÁÂÃÉÊÍÓÔÕÚÇ]", texto))
    pt += min(pt_acentuadas // 3, 5)  # cap em 5

    # Determina dominante
    scores = {"pt": pt, "en": en, "es": es}
    dominante = max(scores, key=scores.get)
    if scores[dominante] == 0:
        return {"pt": pt, "en": en, "es": es, "dominante": "neutro"}
    # Só considera dominante se tem pelo menos 2x o segundo
    sorted_scores = sorted(scores.values(), reverse=True)
    if sorted_scores[0] < 2 or (sorted_scores[0] < sorted_scores[1] * 2 and sorted_scores[1] > 0):
        # Ambiguo: pega quem tiver mais
        pass
    return {"pt": pt, "en": en, "es": es, "dominante": dominante}

def _filtrar_gringo_por_territorio(canal, info, config):
    """Filtra canais gringos baseado nos territorios marcados.

    Returns: (reprovado: bool, motivo: str, idioma: str)
    """
    regioes = config.get("regioes") or ["BR"]
    if isinstance(regioes, str):
        regioes = [r.strip() for r in regioes.split(",") if r.strip()]
    regioes = [str(r).strip().upper() for r in regioes if str(r).strip()]
    if not regioes:
        regioes = ["BR"]

    # Se GLOBAL está marcado, nao filtra idioma
    if "GLOBAL" in regioes:
        return False, "", "neutro"

    nome = canal.get("nome") or canal.get("title") or ""
    descricao = info.get("descricao") or info.get("description") or canal.get("description") or ""
    uploader = info.get("uploader") or canal.get("uploader") or ""

    det = _detectar_idioma_canal(nome, descricao, uploader)
    idioma = det["dominante"]

    aceita_pt = bool(set(regioes) & TERRITORIOS_PT)
    aceita_es = bool(set(regioes) & TERRITORIOS_ES)
    aceita_en = bool(set(regioes) & TERRITORIOS_EN)

    # Se aceita todos (PT + ES + EN), nao reprova por idioma
    if aceita_pt and aceita_es and aceita_en:
        return False, "", idioma

    # Se o canal é neutro (sem sinais claros), deixa passar (nao punir canais sem descricao)
    if idioma == "neutro":
        return False, "", idioma

    # Logica de reprovacao
    motivos = []

    # Canal em ingles mas territorio nao aceita ingles
    if idioma == "en" and not aceita_en:
        if det["en"] >= 3 and det["pt"] == 0 and det["es"] == 0:
            return True, f"gringo_en (sinais={det['en']})", idioma
        # Se tem mais ingles que portugues/espanhol por margem grande
        if det["en"] >= 4 and det["en"] >= max(det["pt"], det["es"]) * 2:
            return True, f"gringo_en_dominante (en={det['en']} pt={det['pt']} es={det['es']})", idioma

    # Canal em espanhol mas territorio nao aceita espanhol
    if idioma == "es" and not aceita_es:
        if det["es"] >= 3 and det["pt"] == 0:
            return True, f"gringo_es (sinais={det['es']})", idioma
        # V58.24: se tem muito mais ES que PT e territorio é PT-only, reprova tambem
        if det["es"] >= 4 and det["es"] >= det["pt"] * 2:
            return True, f"gringo_es_dominante (es={det['es']} pt={det['pt']})", idioma

    # Canal em portugues mas territorio é so ES/EN (raro)
    if idioma == "pt" and not aceita_pt and not aceita_en:
        return True, f"idioma_pt_inesperado (sinais={det['pt']})", idioma

    # Penalizacao leve (nao reprova, só reduz score) se tem mistura
    return False, "", idioma



# ============================================================
# V58.26-LEADS-UP: helpers novos
# ============================================================

# Sinais de monetizacao real (canal que fatura = tem budget p/ edicao)
SINAIS_MONETIZACAO = [
    "patrocinado", "patrocinadora", "sponsor", "sponsored", "apoiado por",
    "parceria com", "em parceria", "use o codigo", "codigo promocional",
    "cupom de desconto", "compre agora", "link na descricao",
    "loja oficial", "merchandising", "merch", "shop now",
    "afiliado", "afiliados", "link de afiliado",
    "nordvpn", "skillshare", "squarespace", "expressvpn", "blue microphones",
    "de outline", "desconto exclusivo", "parceiro oficial",
]

# Padrões de créditos de edicao em descrição de vídeo
PADROES_EDITOR_CREDITO = [
    r"edi[çc][ãa]o[:\s]+[@\w._-]+",
    r"edited by[:\s]+[@\w._-]+",
    r"editor[:\s]+[@\w._-]+",
    r"montagem[:\s]+[@\w._-]+",
    r"finaliza[çc][ãa]o[:\s]+[@\w._-]+",
    r"pos-produ[çc][ãa]o[:\s]+[@\w._-]+",
    r"video editor[:\s]+[@\w._-]+",
    r"creditos[:\s\n]+edi[çc][ãa]o",
    r"edicao de video[:\s]+[@\w._-]+",
    r"cortes por[:\s]+[@\w._-]+",
]

# Sinais de canal profissional (producao serio)
SINAIS_PROFISSIONAL = [
    "site:", "website:", "http", "www.", ".com", ".com.br", ".net",
    "instagram.com/", "facebook.com/", "linkedin.com/",
    "contato@", "business@", "parceria@", "commercial@",
    "para parcerias", "para contato", "business inquiries",
    "empresa", "agencia", "estudio", "estúdio", "producoes", "produções",
]

# Arquivo de blacklist de concorrentes (canais para NUNCA trazer)
CONCORRENTES_BLACKLIST_FILE = "thon_concorrentes_blacklist.json"

def _carregar_concorrentes_blacklist():
    """Carrega lista de IDs de canais concorrentes para nunca trazer."""
    try:
        if os.path.exists(CONCORRENTES_BLACKLIST_FILE):
            with open(CONCORRENTES_BLACKLIST_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, list):
                return set(str(x).strip() for x in data if str(x).strip())
            if isinstance(data, dict):
                ids = set()
                for c in (data.get("canais") or data.get("ids") or []):
                    cid = str(c).strip() if not isinstance(c, dict) else str(c.get("id") or c.get("channel_id") or "").strip()
                    if cid:
                        ids.add(cid)
                return ids
    except Exception as e:
        print(f"[BLACKLIST CONCORRENTES] erro carregando: {e}")
    return set()

def _salvar_concorrentes_blacklist(ids_set):
    """Salva lista de concorrentes."""
    try:
        ids_list = sorted(set(str(x).strip() for x in ids_set if str(x).strip()))
        with open(CONCORRENTES_BLACKLIST_FILE, "w", encoding="utf-8") as f:
            json.dump({"canais": ids_list, "total": len(ids_list), "updated_at": str(datetime.now())}, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"[BLACKLIST CONCORRENTES] erro salvando: {e}")
        return False

def _detectar_editor_nos_creditos(recent_videos):
    """Verifica se algum video recente cita um editor nos creditos.

    Returns: (tem_editor: bool, editor_detectado: str)
    """
    if not recent_videos:
        return False, ""
    for v in recent_videos[:5]:  # checa so os 5 mais recentes
        desc = str(v.get("description") or v.get("title") or "").lower()
        if not desc:
            continue
        for padrao in PADROES_EDITOR_CREDITO:
            m = re.search(padrao, desc, flags=re.I)
            if m:
                # Nao conta se for o PROPRIO canal mencionando (ex: "edição própria")
                match_text = m.group(0).lower()
                if any(skip in match_text for skip in ["propria", "própria", "proprio", "próprio", "eu mesmo", "self"]):
                    continue
                return True, m.group(0)
    return False, ""

def _detectar_monetizacao(ch, recent_videos=None):
    """Detecta sinais de que o canal fatura (tem budget para edicao).

    Returns: (tem_monetizacao: bool, sinais: list[str])
    """
    text = _api_text_channel(ch).lower()
    sinais = []
    # checa descricao do canal + titulos de videos recentes
    if recent_videos:
        for v in recent_videos[:5]:
            text += " " + str(v.get("title") or v.get("description") or "").lower()
    for sinal in SINAIS_MONETIZACAO:
        if sinal in text:
            sinais.append(sinal)
    # checa se tem descricao longa (>200 chars = canal serio)
    desc_len = len(str(ch.get("description") or ""))
    if desc_len > 200:
        sinais.append("descricao_longa")
    # checa links externos (sinal de canal profissional)
    has_links = any(s in text for s in SINAIS_PROFISSIONAL[:7])
    if has_links:
        sinais.append("tem_links_externos")
    return len(sinais) > 0, sinais

def _detectar_canal_profissional(ch, recent_videos=None):
    """Detecta se canal é profissional (tem site, IG business, descricao longa).

    Returns: (profissional: bool, sinais: list)
    """
    text = _api_text_channel(ch).lower()
    sinais = []
    for sinal in SINAIS_PROFISSIONAL:
        if sinal in text:
            sinais.append(sinal)
    desc_len = len(str(ch.get("description") or ""))
    if desc_len > 100:
        sinais.append("descricao_detalhada")
    return len(sinais) >= 2, sinais

def _enrich_leads_rapido(canais_list, max_per_call=8):
    """V58.26: enrich rapido de leads NO LOTE (nao no CRM).

    Para cada canal, busca descricao do canal + about + extrai email/IG.
    Mais leve que o _crmv2_enrich_card (nao baixa videos, so o canal).

    Returns: lista de canais com campos 'email', 'instagram_handle', 'tem_contato' adicionados.
    """
    if not canais_list:
        return canais_list
    if not _API_KEYS:
        # sem chaves, pula enrich (nao trava o ciclo)
        return canais_list

    print(f"[ENRICH LOTE] enriquecendo {min(len(canais_list), max_per_call)}/{len(canais_list)} canais (contato rapido)")
    enriquecidos = 0
    for ch in canais_list[:max_per_call]:
        try:
            cid = ch.get("id") or ch.get("channel_id")
            if not cid:
                continue
            # se ja tem contato, pula
            if ch.get("email") or ch.get("instagram_handle"):
                continue
            # busca dados do canal via API
            data = youtube_api_get("channels", {"part": "snippet,brandingSettings", "id": cid, "maxResults": 1}, cost=1, quiet_404=True)
            items = data.get("items") or []
            if not items:
                continue
            item = items[0]
            sn = item.get("snippet") or {}
            desc = sn.get("description") or ""
            branding = item.get("brandingSettings") or {}
            channel_desc = (branding.get("channel") or {}).get("description") or ""
            all_text = desc + "\n" + channel_desc
            # extrai contatos
            contacts = _crmv2_contacts(all_text)
            if contacts.get("emails"):
                ch["email"] = contacts["emails"][0]
                ch["emails_detectados"] = contacts["emails"]
                enriquecidos += 1
            ig = _crmv2_pick_instagram(contacts)
            if ig:
                ch["instagram_handle"] = ig
                ch["instagram_link"] = f"https://instagram.com/{ig.lstrip('@')}"
                enriquecidos += 1
            if contacts.get("emails") or ig:
                ch["tem_contato"] = True
            else:
                ch["tem_contato"] = False
            # atualiza descricao se estava vazia
            if not ch.get("description"):
                ch["description"] = desc[:500]
            # sem sleep - API ja tem rate control interno
        except Exception as e:
            # enrich falhou? nao trava o ciclo, so pula
            print(f"[ENRICH LOTE] aviso em {ch.get('nome','?')[:30]}: {e}")
            continue
    print(f"[ENRICH LOTE] finalizado | enriquecidos={enriquecidos}/{min(len(canais_list), max_per_call)}")
    return canais_list


SINAIS_PODCAST_API = ["podcast", "pod cast", "cast", "entrevista", "bate papo", "conversa", "talk", "talks"]
SINAIS_NICHO_API = ["empreendedor", "negócios", "negocios", "marketing", "finanças", "financas", "investimento", "tecnologia", "carreira", "vendas", "empresa", "startup", "gestão", "gestao"]

def _api_text_channel(ch):
    return (str(ch.get("title") or "") + " " + str(ch.get("description") or "") + " " + str(ch.get("custom_url") or "")).lower()

# ============================================================
# V58.32: FILTRO DE PAIS ANTES DO SCORE
# ============================================================
# Paises aceitos por territorio marcado
PAISES_POR_TERRITORIO = {
    "BR":     {"BR"},                    # Brasil so aceita BR
    "PT":     {"PT", "BR"},              # Portugal aceita PT e BR (idioma igual)
    "US":     {"US", None, ""},          # EUA aceita US e sem pais (global)
    "MX":     {"MX"},
    "AR":     {"AR"},
    "ES":     {"ES"},
    "CO":     {"CO"},
    "CL":     {"CL"},
    "LATAM":  {"MX", "AR", "CO", "CL", "ES", "BR"},  # LATAM aceita todos latinos
    "GLOBAL": None,                       # Global = nao filtra (None = sem filtro)
}

def _pais_aceito_pelo_territorio(country_do_canal, territorios_marcados):
    """V58.32: verifica se o pais do canal e aceito pelos territorios marcados.

    Returns: (aceito: bool, motivo_reprovacao: str)
    """
    if not territorios_marcados:
        territorios_marcados = ["BR"]

    # Se GLOBAL esta marcado, nao filtra por pais
    if "GLOBAL" in territorios_marcados:
        return True, ""

    country = (str(country_do_canal or "").strip().upper() or None)

    # Para cada territorio marcado, pega conjunto de paises aceitos
    paises_aceitos = set()
    for terr in territorios_marcados:
        aceitos = PAISES_POR_TERRITORIO.get(terr)
        if aceitos is None:
            # GLOBAL = sem filtro
            return True, ""
        paises_aceitos |= aceitos

    # Se pais do canal e vazio/None e territorio nao e BR/PT, aceita (nao punir cego)
    # V58.32: se e BR/PT marcado e canal sem pais, aceita mas marca pra revisao
    if not country:
        if "BR" in territorios_marcados or "PT" in territorios_marcados:
            return True, ""  # deixa passar, filtro de idioma cuida depois
        return True, ""

    # Se pais do canal esta na lista de aceitos, passa
    if country in paises_aceitos:
        return True, ""

    # Se nao, reprova
    return False, f"pais_incompativel_{country}_territorios_{','.join(territorios_marcados)}"


def _enrich_country_canais(canais_list, max_per_call=15):
    """V58.32: busca snippet.country da YouTube API para canais do lote.

    Mais leve que o enrich completo — so pega o country (1 unidade de quota por canal).
    Atualiza canal['country'] no objeto.
    """
    if not canais_list:
        return canais_list
    if not _API_KEYS:
        return canais_list  # sem chaves, pula

    print(f"[ENRICH COUNTRY] buscando country de {min(len(canais_list), max_per_call)}/{len(canais_list)} canais")
    enriquecidos = 0
    for ch in canais_list[:max_per_call]:
        try:
            cid = ch.get("id") or ch.get("channel_id")
            if not cid:
                continue
            # se ja tem country, pula
            if ch.get("country"):
                continue
            data = youtube_api_get("channels", {"part": "snippet", "id": cid, "maxResults": 1}, cost=1, quiet_404=True)
            items = data.get("items") or []
            if not items:
                continue
            sn = items[0].get("snippet") or {}
            country = sn.get("country") or ""
            ch["country"] = country
            if country:
                enriquecidos += 1
        except Exception as e:
            print(f"[ENRICH COUNTRY] aviso em {ch.get('nome','?')[:30]}: {e}")
            continue
    print(f"[ENRICH COUNTRY] finalizado | {enriquecidos} canais com country preenchido")
    return canais_list


def _filtrar_canais_por_pais(canais_list, config, source_label="filtro_pais"):
    """V58.32: filtra lista de canais por pais ANTES do score.

    Returns: (canais_aceitos: list, canais_reprovados: list)
    """
    territorios = config.get("regioes") or ["BR"]
    if isinstance(territorios, str):
        territorios = [t.strip() for t in territorios.split(",") if t.strip()]
    territorios = [t.strip().upper() for t in territorios if t.strip()]

    # Se GLOBAL marcado, nao filtra
    if "GLOBAL" in territorios:
        return canais_list, []

    aceitos = []
    reprovados = []
    for c in canais_list:
        country = c.get("country") or ""
        ok, motivo = _pais_aceito_pelo_territorio(country, territorios)
        if ok:
            aceitos.append(c)
        else:
            c["motivo"] = motivo
            c["score"] = 0
            reprovados.append(c)
            print(f"  [FILTRO PAIS] REPROVADO {c.get('nome','')[:35]} (country={country or 'vazio'} | {motivo})")

    if reprovados:
        print(f"[FILTRO PAIS] {len(reprovados)} reprovados por pais | {len(aceitos)} aceitos")
    return aceitos, reprovados


def _api_rejeicao_previa(ch, config):
    cid = ch.get("id") or ch.get("channel_id")
    text = _api_text_channel(ch)
    subs = ch.get("subscriber_count") or ch.get("subs")
    if not cid:
        return "sem id"
    if any(k in text for k in NEGATIVOS_FORTES_API):
        return "termo negativo forte"
    if any(k in text for k in GRINGO_FORTE_API):
        return "provavel gringo/espanhol"
    # V58.24: filtro anti-gringo por territorio (modo API)
    info_api = {"descricao": ch.get("description") or "", "uploader": ch.get("title") or ""}
    gringo_reprovado, gringo_motivo, _idioma = _filtrar_gringo_por_territorio(ch, info_api, config)
    if gringo_reprovado:
        return gringo_motivo
    if subs is None:
        return "inscritos ocultos"
    if subs < config["min_subs"]:
        return f"inscritos abaixo do minimo ({fmt_numero(subs)})"
    if subs > config["max_subs"]:
        return f"inscritos acima do maximo ({fmt_numero(subs)})"
    return ""

def api_search_channels(config):
    queries = obter_queries_para_engine(config, limit=config.get("query_limit"), formato="tuplas", fallback_kind="api")
    refs = []
    seen = set()
    banco = ids_banco_dados()
    target = config["candidate_target"]
    print(f"\n[API] BUSCA API-FIRST | meta={target} canais | queries={len(queries)} | budget={_api_quota_used()}/{API_DAILY_BUDGET}")

    for qi, (q, nicho) in enumerate(queries, start=1):
        if not estado.get("rodando") or len(refs) >= target:
            break
        page_token = None
        with lock:
            estado["queries_processadas"] = qi
            estado["pipeline_stage"] = "api_search"
            estado["msg"] = f"API search {qi}/{len(queries)}: {q[:45]}"

        print(f"  [api search] {qi}/{len(queries)} {q!r}")
        for page in range(config.get("pages_per_query", API_PAGES_PER_QUERY)):
            if not estado.get("rodando") or len(refs) >= target:
                break
            # V58.23: itera por todos os territorios marcados (nao so BR)
            region_list = _region_lang_params_list(config)
            data = None
            for _rcfg in region_list:
                if not estado.get("rodando") or len(refs) >= target:
                    break
                params = {
                    "part": "snippet",
                    "q": q,
                    "type": "channel",
                    "maxResults": 50,
                    "safeSearch": "none",
                }
                _apply_region_to_params(params, _rcfg)
                if page_token:
                    params["pageToken"] = page_token
                try:
                    data = youtube_api_get("search", params, cost=API_SEARCH_COST)
                except Exception as _re:
                    print(f"  [api search] regiao={_rcfg.get('label')} erro: {_re}")
                    continue
                before = len(refs)
                for item in data.get("items", []) or []:
                    sn = item.get("snippet") or {}
                    cid = sn.get("channelId")
                    if not cid or cid in seen or cid in banco:
                        continue
                    canal_bruto = {
                        "channel_id": cid,
                        "found_query": q,
                        "nicho": nicho,
                        "search_title": sn.get("title", ""),
                        "search_desc": (sn.get("description") or "").replace("\n", " ").strip(),
                        "source": "youtube_api_search",
                        "salvo_em": str(datetime.now()),
                    }
                    # Salva no depósito de brutos
                    _adicionar_bruto(canal_bruto)
                    seen.add(cid)
                    refs.append({
                        "channel_id": cid,
                        "found_query": q,
                        "nicho": nicho,
                        "search_title": sn.get("title", ""),
                        "search_desc": (sn.get("description") or "").replace("\n", " ").strip(),
                    })
                    if len(refs) >= target:
                        break
                added = len(refs) - before
                print(f"    pagina {page+1} reg={_rcfg.get('label','BR')}: +{added} | total={len(refs)} | quota={_api_quota_used()}/{API_DAILY_BUDGET}")
                if data and data.get("nextPageToken"):
                    page_token = data.get("nextPageToken")
                with lock:
                    estado["candidatos_encontrados"] = len(refs)
                if not page_token:
                    break
                time.sleep(0.15)
            if not page_token:
                break
        time.sleep(0.25)
    return refs

def api_channels_details(refs):
    ref_map = {r["channel_id"]: r for r in refs}
    ids = [r["channel_id"] for r in refs]
    out = []
    print(f"\n[API] channels.list em batch | canais={len(ids)}")
    for bi, batch in enumerate([ids[i:i+50] for i in range(0, len(ids), 50)], start=1):
        if not estado.get("rodando"):
            break
        data = youtube_api_get("channels", {
            "part": "snippet,statistics,contentDetails",
            "id": ",".join(batch),
            "maxResults": 50,
        }, cost=API_OTHER_COST)
        for item in data.get("items", []) or []:
            cid = item.get("id")
            sn = item.get("snippet") or {}
            st = item.get("statistics") or {}
            uploads = ((item.get("contentDetails") or {}).get("relatedPlaylists") or {}).get("uploads", "")
            hidden = bool(st.get("hiddenSubscriberCount", False))
            subs = None if hidden else _api_num(st.get("subscriberCount"))
            ref = ref_map.get(cid, {})
            canal = {
                "id": cid,
                "channel_id": cid,
                "nome": sn.get("title", ref.get("search_title", "")),
                "title": sn.get("title", ref.get("search_title", "")),
                "url": f"https://youtube.com/channel/{cid}",
                "custom_url": sn.get("customUrl", ""),
                "description": (sn.get("description") or ref.get("search_desc", "")).replace("\n", " ").strip(),
                "country": sn.get("country", ""),
                "published_at": sn.get("publishedAt", ""),
                "subscriber_count": subs,
                "subs": subs or 0,
                "hidden_subs": hidden,
                "video_count": _api_num(st.get("videoCount")),
                "view_count": _api_num(st.get("viewCount")),
                "uploads_playlist": uploads,
                "found_query": ref.get("found_query", ""),
                "query": ref.get("found_query", ""),
                "origem_query": ref.get("found_query", ""),
                "nicho": ref.get("nicho", ""),
                "source": "youtube_data_api",
                "source_type": ref.get("source_type", ref.get("source", "")),
                "source_query": ref.get("source_query") or ref.get("found_query", ""),
                "source_video_id": ref.get("source_video_id", ""),
                "source_video_title": ref.get("source_video_title", ""),
                "source_playlist_id": ref.get("source_playlist_id", ""),
                "source_channel_seed": ref.get("source_channel_seed", ""),
                "_discovery_query": ref.get("_discovery_query") or ref.get("source_query") or ref.get("found_query", ""),
                "_discovery_page": ref.get("_discovery_page") or 1,
                "recent_videos": [],
            }
            # Atualiza o depósito de brutos com mais informações
            _atualizar_status_bruto(cid, "bruto", {"nome": canal["nome"], "subs": subs, "descricao": canal["description"][:500]})
            out.append(canal)
        print(f"  batch {bi}: {len(batch)} ok | quota={_api_quota_used()}/{API_DAILY_BUDGET}")
        with lock:
            estado["api_channels_detailed"] = len(out)
            estado["msg"] = f"Detalhando canais API: {len(out)}/{len(ids)}"
        time.sleep(0.15)
    return out

def filter_videos(videos, config):
    filtro = config.get("video_filter") or {}
    if not filtro:
        return videos
    min_dur = int(filtro.get("min_duration") or 0)
    max_dur = int(filtro.get("max_duration") or 999999)
    min_last_days = int(filtro.get("min_last_days") or 0)
    max_last_days = int(filtro.get("max_last_days") or 365)
    keywords = [str(k).strip().lower() for k in (filtro.get("keywords_title") or []) if str(k).strip()]
    ignore_shorts = bool(filtro.get("ignore_shorts"))
    require_keyword = bool(filtro.get("require_keyword"))

    filtered = []
    for v in videos or []:
        dur = int(v.get("duration_seconds") or v.get("duration") or 0)
        title = str(v.get("title") or "").lower()
        if dur < min_dur or dur > max_dur:
            continue
        if ignore_shorts and dur <= 60:
            continue
        if keywords:
            has_kw = any(kw in title for kw in keywords)
            if require_keyword and not has_kw:
                continue
        pub = v.get("published_at") or v.get("upload_date") or ""
        if pub:
            try:
                days = _api_days_since(pub) if len(pub) > 8 else None
            except Exception:
                days = None
            if days is None:
                pub_date = datetime.strptime(pub[:10], "%Y-%m-%d")
                days = (datetime.now() - pub_date).days
            if days is not None:
                if days < min_last_days or days > max_last_days:
                    continue
        filtered.append(v)
    return filtered

def api_collect_videos(channels, config):
    video_filter = (config or {}).get("video_filter") or {}
    videos_per_channel = config.get("videos_per_channel", API_VIDEOS_PER_CHANNEL)
    if isinstance(video_filter, dict) and video_filter.get("ignore_shorts"):
        hard_limit = max(int(videos_per_channel or 1) * 4, 25)
    else:
        hard_limit = max(int(videos_per_channel or 1) * 2, 10)
    hard_limit = min(hard_limit, 50)
    channels_videos = _api_collect_videos_impl(channels, config, hard_limit)
    for c in channels_videos:
        c["recent_videos"] = filter_videos(c.get("recent_videos"), config)
    return channels_videos

def _api_collect_videos_impl(channels, config, hard_limit):
    video_ids = []
    owner = {}
    videos_per_channel = config.get("videos_per_channel", API_VIDEOS_PER_CHANNEL)
    print(f"\n[API] playlistItems.list | canais para verificar={len(channels)} | videos/canal={videos_per_channel}")
    for idx, ch in enumerate(channels, start=1):
        if not estado.get("rodando"):
            break
        uploads = ch.get("uploads_playlist")
        if not uploads:
            ch["video_error"] = "sem uploads playlist"
            continue
        data = youtube_api_get("playlistItems", {
            "part": "snippet,contentDetails",
            "playlistId": uploads,
            "maxResults": videos_per_channel,
        }, cost=API_OTHER_COST, quiet_404=True)
        if data.get("_http_status") == 404:
            ch["video_error"] = "uploads playlist 404"
            continue
        for item in data.get("items", []) or []:
            vid = ((item.get("contentDetails") or {}).get("videoId") or "").strip()
            sn = item.get("snippet") or {}
            if not vid:
                continue
            owner[vid] = ch["id"]
            video_ids.append(vid)
            ch["recent_videos"].append({
                "video_id": vid,
                "title": sn.get("title", ""),
                "published_at": sn.get("publishedAt", ""),
                "url": f"https://youtube.com/watch?v={vid}",
            })
        if idx % 25 == 0 or idx == len(channels):
            print(f"  playlistItems: {idx}/{len(channels)} canais | videos={len(video_ids)} | quota={_api_quota_used()}/{API_DAILY_BUDGET}")
        with lock:
            estado["api_videos_collected"] = len(video_ids)
            estado["msg"] = f"Coletando videos API: {idx}/{len(channels)}"
        time.sleep(0.05)

    print(f"\n[API] videos.list em batch | videos={len(video_ids)}")
    ch_map = {c["id"]: c for c in channels}
    for bi, batch in enumerate([video_ids[i:i+50] for i in range(0, len(video_ids), 50)], start=1):
        if not estado.get("rodando"):
            break
        data = youtube_api_get("videos", {
            "part": "snippet,contentDetails,statistics",
            "id": ",".join(batch),
            "maxResults": 50,
        }, cost=API_OTHER_COST)
        for v in data.get("items", []) or []:
            vid = v.get("id")
            cid = owner.get(vid)
            if not cid or cid not in ch_map:
                continue
            st = v.get("statistics") or {}
            cd = v.get("contentDetails") or {}
            sn = v.get("snippet") or {}
            seconds = _api_parse_duration_seconds(cd.get("duration", ""))
            for rv in ch_map[cid].get("recent_videos", []):
                if rv.get("video_id") == vid:
                    rv.update({
                        "title": sn.get("title", rv.get("title", "")),
                        "published_at": sn.get("publishedAt", rv.get("published_at", "")),
                        "duration": cd.get("duration", ""),
                        "duration_seconds": seconds,
                        "view_count": _api_num(st.get("viewCount")),
                        "like_count": _api_num(st.get("likeCount")),
                        "comment_count": _api_num(st.get("commentCount")),
                    })
                    break
        print(f"  videos batch {bi}: {len(batch)} ok | quota={_api_quota_used()}/{API_DAILY_BUDGET}")
        time.sleep(0.12)
    return channels

def api_score_channel(ch, config):
    text = _api_text_channel(ch)
    subs = ch.get("subscriber_count") or ch.get("subs")
    videos = ch.get("recent_videos") or []
    view_vals = [v.get("view_count") for v in videos if isinstance(v.get("view_count"), int)]
    durations = [int(v.get("duration_seconds") or 0) for v in videos]
    dates = [v.get("published_at") for v in videos if v.get("published_at")]

    avg_views = round(sum(view_vals) / len(view_vals)) if view_vals else None
    max_views = max(view_vals) if view_vals else None
    longos = sum(1 for s in durations if s >= 20 * 60)
    medios = sum(1 for s in durations if 3 * 60 <= s < 20 * 60)
    shorts = sum(1 for s in durations if 0 < s <= 90)
    latest = sorted(dates, reverse=True)[0] if dates else ""
    last_days = _api_days_since(latest) if latest else None

    score = 0
    reasons = []

    if subs is not None and config["min_subs"] <= subs <= config["max_subs"]:
        score += 35; reasons.append("subs_ok")
    else:
        score -= 30; reasons.append("subs_fora")

    if any(k in text for k in SINAIS_PODCAST_API):
        score += 18; reasons.append("sinal_podcast")
    if any(k in text for k in SINAIS_NICHO_API):
        score += 12; reasons.append("nicho_ok")

    if longos >= 3:
        score += 25; reasons.append("longos_3+")
    elif longos >= 2:
        score += 20; reasons.append("longos_2")
    elif longos == 1:
        score += 10; reasons.append("longo_1")
    else:
        score -= 20; reasons.append("sem_longo_recente")

    # shorts nao penalizam mais; canal shorts-only ainda reprova por min_longos

    if last_days is not None:
        if last_days <= 45:
            score += 10; reasons.append("recente")
        elif last_days <= 120:
            score += 5; reasons.append("meio_recente")
        elif last_days > config.get("max_last_days", 365):
            score -= 14; reasons.append("parado")
    else:
        score -= 8; reasons.append("sem_data_video")

    ratio = None
    if subs and avg_views is not None:
        ratio = round(avg_views / max(subs, 1), 4)
        if ratio >= 0.03:
            score += 10; reasons.append("views_ratio_ok")
        elif ratio >= 0.01:
            score += 5; reasons.append("views_ratio_medio")
        elif ratio < 0.005:
            score -= 12; reasons.append("views_mortas")
    if avg_views is not None:
        if config.get("engine_mode") == "api_multi_source_fast":
            if avg_views < 200:
                score -= 40; reasons.append("views_mortas")
            elif avg_views < 500:
                score -= 15; reasons.append("views_baixas")
            elif avg_views >= 2000:
                score += 12; reasons.append("avg_2k+")
            elif avg_views >= 1000:
                score += 8; reasons.append("avg_1k+")
            elif avg_views >= 500:
                score += 4; reasons.append("avg_500+")
        else:
            if avg_views >= 1000:
                score += 8; reasons.append("avg_1k+")
            elif avg_views >= 300:
                score += 4; reasons.append("avg_300+")

    if any(k in text for k in GRINGO_FORTE_API):
        score -= 30; reasons.append("gringo_es")
    if any(k in text for k in NEGATIVOS_FORTES_API):
        score -= 25; reasons.append("negativo")

    # V58.26 #9: filtro de audiencia comprada (threshold mais agressivo)
    if subs and avg_views is not None and subs > 5000:
        ratio_val = avg_views / max(subs, 1)
        if ratio_val < 0.005:
            score -= 30; reasons.append("audiencia_comprada")
        elif ratio_val < 0.01:
            score -= 15; reasons.append("views_mortas_v26")
        elif ratio_val >= 0.08:
            score += 15; reasons.append("audiencia_ouro")

    # V58.26 #8: deteccao de monetizacao (canal que fatura = bom lead)
    tem_monet, sinais_monet = _detectar_monetizacao(ch, videos)
    if tem_monet:
        bonus_monet = min(len(sinais_monet) * 3, 12)
        score += bonus_monet; reasons.append(f"monetizado(+{bonus_monet})")
        ch["sinais_monetizacao"] = sinais_monet

    # V58.26: deteccao de canal profissional
    prof, sinais_prof = _detectar_canal_profissional(ch, videos)
    if prof:
        score += 8; reasons.append("profissional")
        ch["sinais_profissional"] = sinais_prof

    # V58.26 #5: deteccao de editor nos creditos (reprova automatico)
    tem_editor, editor_str = _detectar_editor_nos_creditos(videos)
    if tem_editor:
        score -= 40; reasons.append("ja_tem_editor")
        ch["ja_tem_editor"] = True
        ch["editor_detectado"] = editor_str
    else:
        ch["ja_tem_editor"] = False

    # V58.26 #4: blacklist de concorrentes (reprova direto)
    concorrentes_ids = _carregar_concorrentes_blacklist()
    cid_ch = str(ch.get("id") or ch.get("channel_id") or "").strip()
    if cid_ch and cid_ch in concorrentes_ids:
        score = 0; reasons.append("concorrente_blacklist")
        ch["concorrente_blacklist"] = True

    score = max(0, min(100, int(score)))
    ch["score"] = score
    ch["score_reasons"] = ",".join(reasons)
    ch["score_tags_v26"] = reasons  # V58.26 #13: tags explicativas no card
    ch["recent_avg_views"] = avg_views
    ch["recent_max_views"] = max_views
    ch["views_subs_ratio"] = ratio
    ch["longos"] = longos
    ch["medios"] = medios
    ch["shorts"] = shorts
    ch["ultimo"] = formatar_dias(last_days) if last_days is not None else "N/A"
    ch["last_video_at"] = latest
    ch["last_video_days"] = last_days
    ch["subs"] = subs or 0
    ch["subs_fmt"] = fmt_numero(subs or 0)
    ch["tem_shorts"] = shorts > 0

    motivos = []
    if score < config["score_min"]:
        motivos.append(f"score baixo ({score})")
    if longos < config.get("min_longos", 1):
        motivos.append(f"poucos longos recentes ({longos})")
    if last_days is not None and last_days > config.get("max_last_days", 365):
        motivos.append(f"canal parado ({last_days}d)")
    # V58.26: reprova se ja tem editor (nao é lead)
    if ch.get("ja_tem_editor"):
        motivos.append(f"ja tem editor ({ch.get('editor_detectado','')[:30]})")
    # V58.26: reprova se é concorrente blacklist
    if ch.get("concorrente_blacklist"):
        motivos.append("concorrente blacklist")
    ch["motivo_reprovacao"] = "; ".join(motivos) if motivos else ""
    ch["criterio_status"] = "qualificado" if not motivos else "reprovado"
    return ch

def api_verificar_canais(channels, config):
    pre_reprovados = []
    para_video = []
    # V58.32: enrich de country ANTES do pre-filtro (precisa do country pra filtrar por pais)
    if channels and _API_KEYS:
        try:
            _enrich_country_canais(channels, max_per_call=20)
        except Exception as e:
            print(f"[API VERIFY] aviso enrich country: {e}")
    for ch in channels:
        if config.get("engine_mode") == "api_multi_source_fast":
            ok_pre, motivo_pre = pre_filtro_comercial(ch, config)
            if not ok_pre:
                ch["score"] = 0
                ch["motivo"] = motivo_pre
                pre_reprovados.append(ch)
                print(f"[PREFILTER] {motivo_pre} | channels.list | {ch.get('nome') or ch.get('title')}")
                continue
        # V58.32: FILTRO DE PAIS ANTES DO SCORE
        territorios = config.get("regioes") or ["BR"]
        if isinstance(territorios, str):
            territorios = [t.strip() for t in territorios.split(",") if t.strip()]
        territorios = [t.strip().upper() for t in territorios if t.strip()]
        if "GLOBAL" not in territorios:
            pais_ok, pais_motivo = _pais_aceito_pelo_territorio(ch.get("country", ""), territorios)
            if not pais_ok:
                ch["score"] = 0
                ch["motivo"] = pais_motivo
                pre_reprovados.append(ch)
                print(f"[FILTRO PAIS] {pais_motivo} | {ch.get('nome') or ch.get('title')}")
                continue
        motivo = _api_rejeicao_previa(ch, config)
        if motivo:
            ch["score"] = 0
            ch["motivo"] = motivo
            pre_reprovados.append(ch)
        else:
            para_video.append(ch)

    print(f"\n[API] pre-filtro | verificar videos={len(para_video)} | reprovados_pre={len(pre_reprovados)}")
    with lock:
        estado["api_rejected"] = len(pre_reprovados)
        estado["msg"] = f"Pré-filtro API: {len(para_video)} para verificar"
        estado["pipeline_stage"] = "api_prefiltro"

    api_collect_videos(para_video, config)

    qualificados = []
    reprovados = []
    for i, ch in enumerate(para_video, start=1):
        if not estado.get("rodando"):
            break
        api_score_channel(ch, config)
        if ch.get("criterio_status") == "qualificado":
            qualificados.append(ch)
            # Atualiza status no depósito de brutos
            _atualizar_status_bruto(ch["id"], "qualificado", {"score": ch["score"], "longos": ch["longos"], "subs": ch["subs"]})
            print(f"  [API QUALIFICADO] score={ch['score']} | {ch.get('subs_fmt')} | {ch.get('nome')} | longos={ch.get('longos')} avg={fmt_numero(ch.get('recent_avg_views') or 0)}")
        else:
            ch["motivo"] = ch.get("motivo_reprovacao") or "não atingiu critérios"
            reprovados.append(ch)
            _atualizar_status_bruto(ch["id"], "reprovado_api", {"motivo": ch["motivo"], "score": ch["score"]})
        with lock:
            estado["verificados"] = i
            estado["qualificados"] = len(qualificados)
            estado["msg"] = f"Score API {i}/{len(para_video)}: {ch.get('nome','')[:35]}"
            estado["pipeline_stage"] = "api_score"

    for ch in pre_reprovados:
        reprovados.append(ch)
    # V58.26: ordena por score (maior primeiro) - #11
    qualificados.sort(key=lambda c: (int(c.get("score", 0) or 0), int(c.get("subs", 0) or 0)), reverse=True)
    # V58.26 #1: enrich automatico de contato (email + Instagram) nos top qualificados
    if qualificados:
        try:
            _enrich_leads_rapido(qualificados, max_per_call=10)
        except Exception as e:
            print(f"[API VERIFY] aviso enrich: {e}")
    return qualificados, reprovados

def _canal_api_para_lote(ch):
    cid = ch.get("id") or ch.get("channel_id")
    return {
        "id": cid,
        "nome": ch.get("nome") or ch.get("title") or "",
        "url": ch.get("url") or f"https://youtube.com/channel/{cid}",
        "subs": int(ch.get("subs") or ch.get("subscriber_count") or 0),
        "subs_fmt": ch.get("subs_fmt") or fmt_numero(ch.get("subscriber_count") or 0),
        "nicho": ch.get("nicho", ""),
        "longos": int(ch.get("longos") or 0),
        "shorts": int(ch.get("shorts") or 0),
        "ultimo": ch.get("ultimo", "N/A"),
        "score": int(ch.get("score") or 0),
        "tem_shorts": bool(ch.get("tem_shorts")),
        "query": ch.get("query") or ch.get("found_query") or "",
        "origem_query": ch.get("origem_query") or ch.get("found_query") or "",
        "found_query": ch.get("found_query", ""),
        "_discovery_query": ch.get("_discovery_query") or ch.get("found_query") or ch.get("query") or "",
        "_discovery_page": ch.get("_discovery_page") or 1,
        "source_type": ch.get("source_type", ""),
        "source_query": ch.get("source_query", ""),
        "source_video_id": ch.get("source_video_id", ""),
        "source_video_title": ch.get("source_video_title", ""),
        "source_playlist_id": ch.get("source_playlist_id", ""),
        "source_channel_seed": ch.get("source_channel_seed", ""),
        "source": "youtube_data_api",
        "engine": ENGINE_VERSION,
        "description": ch.get("description", ""),
        "recent_avg_views": ch.get("recent_avg_views"),
        "recent_max_views": ch.get("recent_max_views"),
        "views_subs_ratio": ch.get("views_subs_ratio"),
        "last_video_at": ch.get("last_video_at", ""),
        "last_video_days": ch.get("last_video_days"),
        "score_reasons": ch.get("score_reasons", ""),
        "recent_videos": ch.get("recent_videos", []),
        "criterio_status": "qualificado",
        # V58.32: country (para frontend mostrar e filtro de pais funcionar)
        "country": ch.get("country") or ch.get("snippet_country") or "",
        # V58.26: campos novos para qualidade de leads
        "score_tags_v26": ch.get("score_tags_v26", []),
        "sinais_monetizacao": ch.get("sinais_monetizacao", []),
        "sinais_profissional": ch.get("sinais_profissional", []),
        "ja_tem_editor": ch.get("ja_tem_editor", False),
        "editor_detectado": ch.get("editor_detectado", ""),
        "concorrente_blacklist": ch.get("concorrente_blacklist", False),
        "email": ch.get("email", ""),
        "instagram_handle": ch.get("instagram_handle", ""),
        "instagram_link": ch.get("instagram_link", ""),
        "tem_contato": ch.get("tem_contato", False),
    }

def _canal_api_para_reprovado(ch):
    cid = ch.get("id") or ch.get("channel_id")
    return {
        "id": cid,
        "nome": ch.get("nome") or ch.get("title") or "",
        "url": ch.get("url") or f"https://youtube.com/channel/{cid}",
        "nicho": ch.get("nicho", ""),
        "score": int(ch.get("score") or 0),
        "subs_fmt": ch.get("subs_fmt") or fmt_numero(ch.get("subscriber_count") or ch.get("subs") or 0),
        "motivo": ch.get("motivo") or ch.get("motivo_reprovacao") or "não qualificado API",
        "query": ch.get("query") or ch.get("found_query") or "",
        "origem_query": ch.get("origem_query") or ch.get("found_query") or "",
        "source": "youtube_data_api",
        "engine": ENGINE_VERSION,
        "data": str(datetime.now()),
    }


# ============================================================
# MODO DE TRABALHO: API OU DLP
# ============================================================
def _normalizar_engine_mode(value):
    v = str(value or "").strip().lower()
    aliases = {
        "api": "api", "youtube_api": "api", "api_dlp": "api", "api_then_dlp": "api", "api->dlp": "api",
        "api_multi_source_fast": "api_multi_source_fast", "multi_source": "api_multi_source_fast",
        "dlp": "dlp", "yt-dlp": "dlp", "ytdlp": "dlp", "yt_dlp": "dlp", "dlp_only": "dlp", "100_dlp": "dlp", "100% dlp": "dlp",
    }
    return aliases.get(v, "api")

def carregar_engine_mode():
    try:
        if os.path.exists(MODE_FILE):
            with open(MODE_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            return _normalizar_engine_mode(data.get("mode") or data.get("engine_mode"))
    except Exception as e:
        print(f"[modo] erro lendo {MODE_FILE}: {e}")
    return "api"

def salvar_engine_mode(mode):
    mode = _normalizar_engine_mode(mode)
    try:
        with open(MODE_FILE, "w", encoding="utf-8") as f:
            json.dump({"mode": mode, "updated_at": str(datetime.now()), "descricao": "api = YouTube API descobre, pontua e joga direto no lote; dlp = yt-dlp descobre e verifica com Auto Hunt"}, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"[modo] erro salvando {MODE_FILE}: {e}")
    return mode

def _modo_label(mode):
    mode = _normalizar_engine_mode(mode)
    if mode == "dlp":
        return "DLP completo — yt-dlp descobre + verifica; Auto Hunt só aqui"
    return "API direto — YouTube API descobre + pontua + joga no lote; DLP desligado"

# ============================================================
# API DISCOVERY-ONLY: acha o máximo de canais, sem verificação
# ============================================================
def api_search_channels_discovery_only(config):
    """Busca API smart multi-page. Respeita pages_per_query no mesmo ciclo."""
    queries = obter_queries_para_engine(config, limit=config.get("query_limit"), formato="tuplas", fallback_kind="api")
    stats = carregar_query_stats()
    queries = ordenar_queries_inteligente(queries, stats)

    refs = []
    seen = set()
    target = int(config.get("candidate_target") or 1000)
    max_pages_cycle = int(config.get("pages_per_query") or API_PAGES_PER_QUERY)
    max_pages_cycle = max(1, min(max_pages_cycle, QUERY_MAX_PAGE))

    print(f"\n[API DISCOVERY SMART MULTIPAGE] queries={len(queries)} | target={target} | pages/query={max_pages_cycle} | seen_cooldown={API_SEEN_COOLDOWN_DIAS}d | budget={_api_quota_used()}/{API_DAILY_BUDGET}")

    chamadas_ciclo = set()
    for qi, (q, nicho) in enumerate(queries, start=1):
        if not estado.get("rodando") or len(refs) >= target:
            break

        paginas_rodadas_query = 0
        page_token_local = None

        while estado.get("rodando") and len(refs) < target and paginas_rodadas_query < max_pages_cycle:
            stats = carregar_query_stats()
            pagina = paginas_rodadas_query + 1
            if paginas_rodadas_query == 0:
                pagina_decidida, token_decidido = query_decide_pagina(q, stats)
                if pagina_decidida is None:
                    print(f"[query skip] {q!r} | cooldown/saturada")
                    break
                page_token_local = token_decidido if pagina_decidida > 1 else None
                pagina = int(pagina_decidida or 1)
            page_token = page_token_local

            if pagina > 1 and not page_token:
                print(f"[query stop] {q!r} | pág={pagina} sem page_token anterior")
                break

            call_key = ("channel_search", q, page_token or "__FIRST__")
            if call_key in chamadas_ciclo:
                print(f"[query guard] {q!r} pág={pagina} ignorada: chamada duplicada no mesmo ciclo")
                break
            chamadas_ciclo.add(call_key)

            with lock:
                estado["queries_processadas"] = qi
                estado["pipeline_stage"] = "api_discovery_search"
                estado["msg"] = f"API smart {qi}/{len(queries)} pág {pagina}: {q[:45]}"

            # V58.23: itera por todos os territorios marcados (BR, PT, US, etc)
            region_list = _region_lang_params_list(config)
            _any_region_ok = False
            _last_err = None
            data = {"items": []}
            for _rcfg in region_list:
                if not estado.get("rodando") or len(refs) >= target:
                    break
                params = {
                    "part": "snippet",
                    "q": q,
                    "type": "channel",
                    "maxResults": 50,
                    "safeSearch": "none",
                }
                _apply_region_to_params(params, _rcfg)
                if page_token:
                    params["pageToken"] = page_token

                try:
                    _data_r = youtube_api_get("search", params, cost=API_SEARCH_COST)
                    _any_region_ok = True
                    if _data_r and _data_r.get("items"):
                        data["items"].extend(_data_r["items"])
                    if _data_r and _data_r.get("nextPageToken") and not data.get("nextPageToken"):
                        data["nextPageToken"] = _data_r["nextPageToken"]
                except Exception as e:
                    _last_err = str(e)
                    print(f"[API DISCOVERY] regiao={_rcfg.get('label')} erro em {q!r}: {_last_err}")
                    continue

            if not _any_region_ok and _last_err:
                print(f"[API DISCOVERY SMART MULTIPAGE] busca parou em todas as regioes: {_last_err}. Encontrados={len(refs)}.")
                with lock:
                    estado["api_last_error"] = _last_err
                    estado["msg"] = f"API parou: {_last_err[:120]}"
                return refs

            novos_pagina = 0
            repetidos_pagina = 0
            ignorados = {"seen": 0, "lote": 0, "aprovados": 0, "reprovados_cooldown": 0}

            for item in data.get("items", []) or []:
                sn = item.get("snippet") or {}
                cid = sn.get("channelId")

                if not cid or cid in seen:
                    repetidos_pagina += 1
                    continue

                motivo = discovery_channel_ja_visto(cid, {"query": q, "pagina": pagina})
                if motivo:
                    repetidos_pagina += 1
                    if motivo == "api_seen_cooldown":
                        ignorados["seen"] += 1
                    elif motivo == "lote":
                        ignorados["lote"] += 1
                    elif motivo == "aprovados":
                        ignorados["aprovados"] += 1
                    elif motivo == "reprovado_cooldown":
                        ignorados["reprovados_cooldown"] += 1
                    continue

                seen.add(cid)

                ref = {
                    "id": cid,
                    "channel_id": cid,
                    "nome": sn.get("title", ""),
                    "title": sn.get("title", ""),
                    "url": f"https://youtube.com/channel/{cid}",
                    "found_query": q,
                    "query": q,
                    "nicho": nicho,
                    "search_title": sn.get("title", ""),
                    "search_desc": (sn.get("description") or "").replace("\n", " ").strip(),
                    "source": "api_discovery_to_dlp",
                    "_discovery_query": q,
                    "_discovery_page": pagina,
                }

                _adicionar_bruto({
                    "channel_id": cid,
                    "found_query": q,
                    "query": q,
                    "nicho": nicho,
                    "search_title": ref["search_title"],
                    "title": ref["search_title"],
                    "search_desc": ref["search_desc"],
                    "description": ref["search_desc"],
                    "source": "youtube_api_discovery_to_dlp",
                    "_discovery_query": q,
                    "_discovery_page": pagina,
                })

                registrar_api_seen(cid, q, pagina)
                refs.append(ref)
                novos_pagina += 1

                if len(refs) >= target:
                    break

            next_token = data.get("nextPageToken")
            query_registrar_resultado(q, pagina, novos_pagina, repetidos_pagina, 0, next_token)
            # V58.29: registra tambem no query_factory V2 (saturacao inteligente)
            try:
                if query_factory_get_next_queries and hasattr(query_factory_get_next_queries, '__module__'):
                    from query_factory.query_factory import registrar_resultado_query as _v29_registrar
                    sat = _v29_registrar(q, novos_pagina, repetidos_pagina, 0)
                    if sat:
                        print(f"[SATURACAO V2] {q!r} foi pro cooldown: {sat['motivo']} ({sat['cooldown_seg']}s)")
            except Exception as _e_sat:
                pass

            taxa = int((novos_pagina / max(1, novos_pagina + repetidos_pagina)) * 100)
            print(f"[query {qi:02d}/{len(queries)}] {q!r} pág={pagina} | +{novos_pagina} novos | {repetidos_pagina} repetidos | taxa_novos={taxa}% | ign={ignorados} | total={len(refs)}")

            paginas_rodadas_query += 1

            with lock:
                estado["candidatos_encontrados"] = len(refs)

            if not next_token:
                break
            if config.get("parada_inteligente", True) and taxa < 10:
                print(f"[query stop] {q!r} | taxa_novos={taxa}% < 10%")
                break
            page_token_local = next_token

            if not _sleep_interrompivel(0.15):
                break

        if not _sleep_interrompivel(0.10):
            break

    return refs


def api_discovery_only_process(config):
    """API direta para lote. Nao chama DLP.

    Fluxo:
    1. search.list acha canais.
    2. channels.list pega inscritos/descricao/uploads.
    3. playlistItems/videos.list pega videos recentes.
    4. api_score_channel aplica o score normal ja existente.
    5. qualificados entram direto no lote de caca.
    """
    refs = api_search_channels_discovery_only(config)
    details = []
    if refs:
        try:
            details = api_channels_details(refs)
        except Exception as e:
            print(f"[API DIRECT] channels.list falhou: {e}. Usando refs basicos.")
            details = []
    if not details:
        details = refs

    print(f"[API DIRECT] API vai qualificar direto | canais detalhados={len(details)} | DLP desligado")

    qualificados_api, reprovados = api_verificar_canais(details, config)
    lote_api = []
    for ch in qualificados_api:
        cid = ch.get("id") or ch.get("channel_id")
        item = _canal_api_para_lote(ch)
        if not item:
            continue
        item.update({
            "source": "api_direct",
            "qualified_by": "api_direct",
            "engine_mode": "api_direct_no_dlp",
            "status": "pendente",
            "status_fila": "pendente",
            "lote_status": "pendente_aprovacao",
            "qualificado": True,
            "api_last_batch": True,
            "qualified_at": datetime.now().isoformat(timespec="seconds"),
            "aviso": "Qualificado pela YouTube API direta. DLP/yt-dlp nao foi usado neste modo.",
        })
        lote_api.append(item)
        if cid:
            _atualizar_status_bruto(cid, "qualificado_api_direct_lote", {"score": item.get("score"), "subs": item.get("subs"), "longos": item.get("longos"), "query": item.get("query")})

    # Como o modo API agora e direto, limpa fila DLP para nao sobrar lixo antigo/verificacao aleatoria.
    try:
        salvar_fila_dlp_verificacao([])
    except Exception as e:
        print(f"[API DIRECT] aviso: nao consegui limpar fila DLP: {e}")

    # Snapshot do último lote qualificado pela API direta.
    try:
        with open("thon_api_direct_qualified_last.json", "w", encoding="utf-8") as _f_apiq:
            json.dump({
                "engine": APP_VERSION,
                "created_at": str(datetime.now()),
                "qualificados": lote_api,
                "total": len(lote_api),
            }, _f_apiq, ensure_ascii=False, indent=2)
        print(f"[API DIRECT] snapshot qualificados salvo | total={len(lote_api)}")
    except Exception as e:
        print("[API DIRECT] aviso: snapshot falhou:", e)

    sync_report = sync_api_qualificados_para_lote(lote_api, source="api_direct")
    query_atualizar_qualificados_por_origem(lote_api)

    with lock:
        estado["pipeline_stage"] = "api_direct_lote"
        estado["msg"] = f"API direta finalizada: {len(lote_api)} qualificados | +{sync_report['novos_adicionados']} novos no lote | {len(reprovados)} reprovados | DLP desligado"
        estado["qualificados"] = len(lote_api)
        estado["dlp_queue_count"] = len(carregar_fila_dlp_verificacao())

    print(f"[API DIRECT] finalizada | qualificados_api={len(lote_api)} | novos_adicionados={sync_report['novos_adicionados']} | ja_existiam={sync_report['ja_existiam']} | fila_total={sync_report['depois']} | reprovados={len(reprovados)} | fila_dlp=0")
    return lote_api, reprovados

def _api_call_once(call_cache, source_type, query, page_token, endpoint, params, cost, quiet_404=False):
    key = (source_type, str(query or ""), str(page_token or "__FIRST__"), endpoint)
    if key in call_cache:
        print(f"[DEDUP CENTRAL] chamada duplicada bloqueada | {source_type} | {query!r} | token={page_token or 'FIRST'}")
        return None
    call_cache.add(key)
    return youtube_api_get(endpoint, params, cost=cost, quiet_404=quiet_404)

def _registrar_resultado_multi(out, cand, seen, banco, config):
    r = _add_candidate_multi(out, cand, seen, banco, config)
    if r == "novo":
        print(f"[DEDUP CENTRAL] novo | {cand.get('source_type')} | {cand.get('channel_id')} | {cand.get('search_title') or cand.get('source_query')}")
        return "novo"
    if r == "repetido":
        return "repetido"
    print(f"[PREFILTER] {r} | {cand.get('source_type')} | {cand.get('search_title') or cand.get('source_video_title') or cand.get('source_query')}")
    return "pre"

def _discovery_search_paginado(source_type, queries, config, seen, banco, qperf, call_cache, search_type, extra_params=None, max_results=50, on_batch=None, batch_size=None):
    out = []
    target = int(config.get("candidate_target", 5000) or 5000)
    max_pages = max(1, min(int(config.get("pages_per_query") or API_PAGES_PER_QUERY), 10))
    chunked = bool(config.get("chunked_pipeline") or on_batch)
    batch_size = int(batch_size or config.get("discovery_batch_size") or 0)
    def total_descoberto():
        if not chunked:
            return len(out)
        try:
            return int(estado.get("multi_candidates_discovered_total", estado.get("candidatos_encontrados", 0)) or 0)
        except Exception:
            return len(out)
    def limite_atingido():
        return total_descoberto() >= target
    def flush_if_needed(force=False):
        if not on_batch or not batch_size:
            return True
        if out and (force or len(out) >= batch_size):
            bloco = list(out)
            out.clear()
            return bool(on_batch(bloco, source_type))
        return True
    for qi, (query, nicho) in enumerate(queries, start=1):
        if limite_atingido():
            break
        if _qperf_em_cooldown(qperf, query, source_type):
            print(f"[DISCOVERY][{source_type}] skip cooldown | {query!r}")
            continue
        page_token = None
        for page in range(1, max_pages + 1):
            if limite_atingido() or not estado.get("rodando"):
                break
            # V58.23: itera por todos os territorios marcados (multi-source)
            region_list = _region_lang_params_list(config)
            _any_ok = False
            data = None
            for _rcfg in region_list:
                if not estado.get("rodando") or limite_atingido():
                    break
                params = {
                    "part": "snippet",
                    "q": query,
                    "type": search_type,
                    "maxResults": max_results,
                    "safeSearch": "none",
                }
                _apply_region_to_params(params, _rcfg)
                if extra_params:
                    params.update(extra_params)
                if page_token:
                    params["pageToken"] = page_token
                with lock:
                    estado["pipeline_stage"] = f"multi_{source_type}"
                    estado["queries_processadas"] = qi
                    estado["msg"] = f"{source_type} {qi}/{len(queries)} pág {page} reg={_rcfg.get('label','BR')}: {query[:45]}"
                print(f"[DISCOVERY][{source_type}] query={query!r} pág={page} reg={_rcfg.get('label','BR')} token={page_token or 'FIRST'}")
                try:
                    _data_r = _api_call_once(call_cache, source_type + ":" + (_rcfg.get("regionCode") or "ALL"), query, page_token, "search", params, API_SEARCH_COST)
                    if _data_r is not None:
                        data = _data_r
                        _any_ok = True
                        break  # achou nesta regiao, nao precisa tentar outras
                except Exception as e:
                    print(f"[DISCOVERY][{source_type}] erro reg={_rcfg.get('label','BR')} | {query!r}: {e}")
                    continue
            if not _any_ok and data is None:
                break
            novos = repetidos = pre = 0
            playlist_ids = []
            for item in data.get("items", []) or []:
                sn = item.get("snippet") or {}
                if search_type == "video":
                    cid = sn.get("channelId")
                    vid = (item.get("id") or {}).get("videoId")
                    cand = _multi_candidate(cid, source_type, query, nicho, source_video_id=vid, source_video_title=sn.get("title", ""), search_title=sn.get("channelTitle", ""), search_desc=sn.get("description", ""))
                    r = _registrar_resultado_multi(out, cand, seen, banco, config)
                elif search_type == "channel":
                    cid = sn.get("channelId")
                    cand = _multi_candidate(cid, source_type, query, nicho, search_title=sn.get("title", ""), search_desc=sn.get("description", ""))
                    r = _registrar_resultado_multi(out, cand, seen, banco, config)
                else:
                    pid = (item.get("id") or {}).get("playlistId")
                    if pid:
                        playlist_ids.append(pid)
                    continue
                if r == "novo": novos += 1
                elif r == "repetido": repetidos += 1
                else: pre += 1
            if search_type == "playlist":
                for pid in playlist_ids:
                    try:
                        items = _api_call_once(call_cache, "playlistItems", pid, None, "playlistItems", {"part": "snippet", "playlistId": pid, "maxResults": 50}, API_OTHER_COST, quiet_404=True)
                    except Exception as e:
                        print(f"[DISCOVERY][playlist_search] playlistItems falhou {pid}: {e}")
                        continue
                    if items is None:
                        continue
                    for it in items.get("items", []) or []:
                        sn = it.get("snippet") or {}
                        cid = sn.get("videoOwnerChannelId") or sn.get("channelId")
                        cand = _multi_candidate(cid, source_type, query, nicho, source_playlist_id=pid, source_video_title=sn.get("title", ""), search_title=sn.get("videoOwnerChannelTitle", ""), search_desc=sn.get("description", ""))
                        r = _registrar_resultado_multi(out, cand, seen, banco, config)
                        if r == "novo": novos += 1
                        elif r == "repetido": repetidos += 1
                        else: pre += 1
            _qperf_update(qperf, query, source_type, novos, repetidos, pre, 0, API_SEARCH_COST)
            total = novos + repetidos + pre
            taxa = novos / max(1, total)
            with lock:
                by_source = estado.setdefault("multi_candidates_by_source", {})
                if source_type not in by_source:
                    by_source[source_type] = int(by_source.get(source_type, 0) or 0)
                # _add_candidate_multi já colocou os novos em out; aqui somamos só a novidade da página.
                estado["multi_candidates_discovered_total"] = int(estado.get("multi_candidates_discovered_total", 0) or 0) + int(novos or 0)
                by_source[source_type] = int(by_source.get(source_type, 0) or 0) + int(novos or 0)
                estado["candidatos_encontrados"] = int(estado.get("multi_candidates_discovered_total", 0) or 0)
                estado["fonte_atual"] = source_type
                estado["query_atual"] = query
                estado["pagina_atual"] = page
                estado["repetidos_total"] = int(estado.get("repetidos_total", 0) or 0) + int(repetidos or 0)
                estado["pre_reprovados_total"] = int(estado.get("pre_reprovados_total", 0) or 0) + int(pre or 0)
                estado["ultimo_evento"] = f"{source_type}: +{novos} novos, {repetidos} repetidos, {pre} pré-reprovados"
            print(f"[DISCOVERY][{source_type}] result | {query!r} pág={page} +{novos} rep={repetidos} pre={pre} taxa={taxa:.0%} total_fonte={estado.get('multi_candidates_by_source',{}).get(source_type, len(out))}")
            if not flush_if_needed(False):
                return out
            page_token = data.get("nextPageToken")
            if not page_token:
                print(f"[DISCOVERY][{source_type}] sem nextPageToken | {query!r}")
                break
            if config.get("parada_inteligente", True) and total and taxa < 0.10:
                print(f"[DISCOVERY][{source_type}] parada inteligente | {query!r} | taxa_novos={taxa:.0%}")
                break
    flush_if_needed(True)
    return out

def discovery_video_search(queries, config, seen, banco, qperf, call_cache=None, on_batch=None, batch_size=None):
    return _discovery_search_paginado("video_search_long", queries, config, seen, banco, qperf, call_cache or set(), "video", {"videoDuration": "long", "order": "relevance"}, 50, on_batch=on_batch, batch_size=batch_size)

def discovery_channel_search(queries, config, seen, banco, qperf, call_cache=None, on_batch=None, batch_size=None):
    return _discovery_search_paginado("channel_search", queries, config, seen, banco, qperf, call_cache or set(), "channel", None, 50, on_batch=on_batch, batch_size=batch_size)

def discovery_playlist_search(queries, config, seen, banco, qperf, call_cache=None, on_batch=None, batch_size=None):
    return _discovery_search_paginado("playlist_search", queries, config, seen, banco, qperf, call_cache or set(), "playlist", None, 25, on_batch=on_batch, batch_size=batch_size)

def _seed_ids(limit=20):
    ids = []
    for c in (estado.get("aprovados") or carregar_aprovados() or [])[:limit]:
        cid = c.get("id") or c.get("channel_id")
        if cid:
            ids.append(cid)
    return ids

def discovery_uploads_expansion(seed_channels, seen, banco, config=None):
    out = []
    call_cache = set()
    config = config or {}
    for sid in seed_channels[:10]:
        if not sid or sid in seen or sid in banco:
            continue
        try:
            data = youtube_api_get("channels", {"part": "snippet,contentDetails", "id": sid, "maxResults": 1}, cost=API_OTHER_COST)
            item = (data.get("items") or [None])[0] or {}
            sn = item.get("snippet") or {}
            uploads = (((item.get("contentDetails") or {}).get("relatedPlaylists") or {}).get("uploads") or "")
            if uploads:
                _api_call_once(call_cache, "uploads_expansion_items", uploads, None, "playlistItems", {"part": "snippet", "playlistId": uploads, "maxResults": 5}, API_OTHER_COST, quiet_404=True)
            cand = _multi_candidate(sid, "uploads_expansion", "", "", source_channel_seed=sid, search_title=sn.get("title", ""), search_desc=sn.get("description", ""))
            _registrar_resultado_multi(out, cand, seen, banco, config)
        except Exception as e:
            print(f"[DISCOVERY][uploads_expansion] falhou {sid}: {e}")
    print(f"[DISCOVERY][uploads_expansion] result | novos={len(out)}")
    return out

def discovery_subscriptions_seed(seed_channels, seen, banco, config=None):
    out = []
    for sid in seed_channels[:10]:
        try:
            data = youtube_api_get("subscriptions", {"part": "snippet", "channelId": sid, "maxResults": 50, "order": "relevance"}, cost=API_OTHER_COST)
        except Exception as e:
            print(f"[MULTI subscriptions] privadas/falhou {sid}: {e}")
            continue
        for item in data.get("items", []) or []:
            rid = ((item.get("snippet") or {}).get("resourceId") or {}).get("channelId")
            _registrar_resultado_multi(out, _multi_candidate(rid, "subscriptions_seed", "", "", source_channel_seed=sid, search_title=(item.get("snippet") or {}).get("title", ""), search_desc=(item.get("snippet") or {}).get("description", "")), seen, banco, config or {})
    print(f"[DISCOVERY][subscriptions_seed] result | novos={len(out)}")
    return out

def discovery_activities_seed(seed_channels, seen, banco, config=None):
    out = []
    config = config or {}
    for sid in seed_channels[:10]:
        try:
            data = youtube_api_get("activities", {"part": "snippet,contentDetails", "channelId": sid, "maxResults": 20}, cost=API_OTHER_COST, quiet_404=True)
        except Exception as e:
            print(f"[DISCOVERY][activities_seed] falhou {sid}: {e}")
            continue
        for item in data.get("items", []) or []:
            sn = item.get("snippet") or {}
            cid = sn.get("channelId")
            cand = _multi_candidate(cid, "activities_seed", "", "", source_channel_seed=sid, search_title=sn.get("channelTitle", ""), search_desc=sn.get("description", ""))
            _registrar_resultado_multi(out, cand, seen, banco, config)
    print(f"[DISCOVERY][activities_seed] result | novos={len(out)}")
    return out

def _estado_multi_reset(config):
    target = int(config.get("candidate_target", 5000) or 5000)
    bloco = int(config.get("discovery_batch_size", 500) or 500)
    with lock:
        estado.update({
            "target_candidatos": target,
            "api_candidate_target": target,
            "discovery_batch_size": bloco,
            "api_verify_batch_size": int(config.get("api_verify_batch_size", 50) or 50),
            "approval_goal": int(config.get("approval_goal", 0) or 0),
            "multi_candidates_discovered_total": 0,
            "multi_candidates_by_source": {},
            "multi_chunks_processed": 0,
            "multi_chunks_estimated": max(1, (target + bloco - 1) // max(1, bloco)),
            "bloco_atual": 0,
            "bloco_tamanho": 0,
            "bloco_processado": 0,
            "fonte_atual": "",
            "query_atual": "",
            "pagina_atual": 0,
            "repetidos_total": 0,
            "pre_reprovados_total": 0,
            "channels_detailed_total": 0,
            "verificados_total": 0,
            "qualificados_total": 0,
            "reprovados_total": 0,
            "ultimo_evento": "Iniciando pipeline em blocos",
        })

def _dedupe_refs_por_id(refs):
    mapa = {}
    for r in refs or []:
        if not isinstance(r, dict):
            continue
        cid = str(r.get("channel_id") or r.get("id") or "").strip()
        if not cid:
            continue
        mapa[cid] = {**mapa.get(cid, {}), **r, "id": cid, "channel_id": cid}
    return list(mapa.values())

def _multi_source_processar_bloco(refs, config, bloco_num, source_hint="multi_source"):
    refs = _dedupe_refs_por_id(refs)
    if not refs or not estado.get("rodando"):
        return [], [], True
    target = int(config.get("candidate_target", 5000) or 5000)
    bloco_total = len(refs)
    print(f"\n[MULTI SOURCE][BLOCO {bloco_num}] processando {bloco_total} candidatos | source={source_hint} | quota={_api_quota_used()}/{API_DAILY_BUDGET}")
    with lock:
        estado["pipeline_stage"] = "multi_block_details"
        estado["stage"] = "multi_block_details"
        estado["stage_label"] = f"Detalhando bloco {bloco_num}"
        estado["bloco_atual"] = bloco_num
        estado["bloco_tamanho"] = bloco_total
        estado["bloco_processado"] = 0
        estado["msg"] = f"Bloco {bloco_num}: detalhando {bloco_total} canais"
        estado["ultimo_evento"] = f"Bloco {bloco_num} enviado para channels.list"
    try:
        details = api_channels_details(refs) if refs else []
    except Exception as e:
        erro = str(e)
        print(f"[MULTI SOURCE][BLOCO {bloco_num}] channels.list falhou: {erro}")
        with lock:
            estado["api_last_error"] = erro
            estado["ultimo_evento"] = f"Erro no bloco {bloco_num}: {erro[:120]}"
            estado["msg"] = f"Bloco {bloco_num} falhou em channels.list: {erro[:120]}"
            estado["pipeline_stage"] = "multi_block_error"
            estado["stage_label"] = "Erro no bloco"
            estado["rodando"] = False
        # Não apaga progresso anterior. Para o ciclo de forma limpa se for quota real.
        return [], [], False
    details = details or refs
    with lock:
        estado["channels_detailed_total"] = int(estado.get("channels_detailed_total", 0) or 0) + len(details)
        estado["api_channels_detailed"] = int(estado.get("channels_detailed_total", 0) or 0)
        estado["pipeline_stage"] = "multi_block_verify"
        estado["stage_label"] = f"Verificando bloco {bloco_num}"
        estado["msg"] = f"Bloco {bloco_num}: verificando {len(details)} canais"
        estado["ultimo_evento"] = f"Bloco {bloco_num}: {len(details)} canais detalhados"
    qualificados_api, reprovados = api_verificar_canais(details, config)
    lote_api = []
    for ch in qualificados_api:
        item = _canal_api_para_lote(ch)
        if item:
            item.update({
                "source": "api_multi_source_fast",
                "qualified_by": "api_multi_source_fast",
                "status": "pendente",
                "status_fila": "pendente",
                "api_last_batch": True,
                "multi_source_block": bloco_num,
                "qualified_at": datetime.now().isoformat(timespec="seconds"),
            })
            lote_api.append(item)
    sync_report = sync_api_qualificados_para_lote(lote_api, source="api_multi_source_fast")
    if reprovados:
        try:
            salvar_reprovados_lista([_canal_api_para_reprovado(r) if not r.get("motivo") else r for r in reprovados])
        except Exception as e:
            print(f"[MULTI SOURCE][BLOCO {bloco_num}] aviso: falha salvando reprovados: {e}")
    query_atualizar_qualificados_por_origem(lote_api)
    with lock:
        estado["multi_chunks_processed"] = int(estado.get("multi_chunks_processed", 0) or 0) + 1
        estado["verificados_total"] = int(estado.get("verificados_total", 0) or 0) + len(details)
        estado["verificados"] = int(estado.get("verificados_total", 0) or 0)
        estado["qualificados_total"] = int(estado.get("qualificados_total", 0) or 0) + len(lote_api)
        estado["qualificados"] = int(estado.get("qualificados_total", 0) or 0)
        estado["reprovados_total"] = int(estado.get("reprovados_total", 0) or 0) + len(reprovados or [])
        estado["api_rejected"] = int(estado.get("reprovados_total", 0) or 0)
        estado["bloco_processado"] = bloco_total
        estado["api_quota_used"] = _api_quota_used()
        estado["api_quota_budget"] = API_DAILY_BUDGET
        estado["api_quota_pct"] = round((_api_quota_used() / max(1, API_DAILY_BUDGET)) * 100, 1)
        estado["pipeline_stage"] = "multi_block_saved"
        estado["stage_label"] = f"Bloco {bloco_num} salvo"
        estado["msg"] = f"Bloco {bloco_num} salvo: +{len(lote_api)} qualificados | {len(reprovados)} reprovados | {estado.get('multi_candidates_discovered_total', 0)}/{target} candidatos"
        estado["ultimo_evento"] = f"Bloco {bloco_num}: +{len(lote_api)} no lote, {len(reprovados)} reprovados"
    print(f"[MULTI SOURCE][BLOCO {bloco_num}] salvo | qualificados={len(lote_api)} | novos_lote={sync_report['novos_adicionados']} | fila_total={sync_report['depois']} | reprovados={len(reprovados)}")
    return lote_api, reprovados, True

def worker_api_multi_source(config, auto=False):
    # V58.20: pipeline em blocos. Não espera juntar 5000 para só depois verificar.
    config = dict(config or {})
    config["chunked_pipeline"] = True
    discovery_batch_size = int(config.get("discovery_batch_size") or 500)
    api_verify_batch_size = int(config.get("api_verify_batch_size") or 50)
    approval_goal = int(config.get("approval_goal") or 0)
    config["discovery_batch_size"] = discovery_batch_size
    config["api_verify_batch_size"] = api_verify_batch_size
    config["approval_goal"] = approval_goal
    print(f"[MULTI SOURCE] start CHUNKED | target={config.get('candidate_target')} bloco={discovery_batch_size} api_batch={api_verify_batch_size} pages={config.get('pages_per_query')} qworkers={config.get('query_workers')} workers={config.get('workers')} | sem DLP")
    _estado_multi_reset(config)
    queries = obter_queries_para_engine(config, limit=config.get("query_limit"), formato="tuplas", fallback_kind="api")
    qperf = _qperf_load()
    seen = set()
    banco = ids_banco_dados()
    call_cache = set()
    pending = []
    lote_total = []
    reprovados_total = []
    stats_fonte = {}
    bloco_num = 0

    def processar_pending(force=False, source_hint="multi_source"):
        nonlocal pending, bloco_num, lote_total, reprovados_total
        ok = True
        while pending and (force or len(pending) >= discovery_batch_size):
            bloco = pending[:discovery_batch_size]
            pending = pending[discovery_batch_size:]
            bloco_num += 1
            lote, reps, ok = _multi_source_processar_bloco(bloco, config, bloco_num, source_hint=source_hint)
            lote_total.extend(lote or [])
            reprovados_total.extend(reps or [])
            if approval_goal and len(lote_total) >= approval_goal:
                print(f"[MULTI SOURCE] approval_goal atingido | {len(lote_total)}/{approval_goal}")
                return False
            if not ok:
                return False
            if not estado.get("rodando"):
                return False
            if not force:
                break
        return ok

    def on_batch(bloco_refs, source_type):
        nonlocal pending
        pending.extend(_dedupe_refs_por_id(bloco_refs))
        with lock:
            estado["bloco_tamanho"] = len(pending)
            estado["bloco_processado"] = min(len(pending), discovery_batch_size)
            estado["fonte_atual"] = source_type
            estado["msg"] = f"Buffer do bloco: {min(len(pending), discovery_batch_size)}/{discovery_batch_size} | fonte {source_type}"
        return processar_pending(False, source_hint=source_type)

    for source_name, fn, subset in [
        ("video_search_long", discovery_video_search, queries),
        ("channel_search", discovery_channel_search, queries),
        ("playlist_search", discovery_playlist_search, queries[:max(1, len(queries)//3)]),
    ]:
        if not estado.get("rodando"):
            break
        antes_total = int(estado.get("multi_candidates_discovered_total", 0) or 0)
        leftovers = fn(subset, config, seen, banco, qperf, call_cache, on_batch=on_batch, batch_size=discovery_batch_size)
        if leftovers:
            pending.extend(_dedupe_refs_por_id(leftovers))
        stats_fonte[source_name] = int(estado.get("multi_candidates_discovered_total", 0) or 0) - antes_total
        print(f"[DISCOVERY][{source_name}] acumulado | +{stats_fonte[source_name]} | total={estado.get('multi_candidates_discovered_total', 0)} | pending={len(pending)}")
        # Fecha bloco parcial ao trocar de fonte se ele já está razoável.
        if len(pending) >= max(100, discovery_batch_size // 2):
            if not processar_pending(True, source_hint=source_name):
                break
        if approval_goal and len(lote_total) >= approval_goal:
            break
        if int(estado.get("multi_candidates_discovered_total", 0) or 0) >= int(config.get("candidate_target", 5000) or 5000):
            break

    # Fontes best-effort. Não deixam o ciclo morrer.
    if estado.get("rodando") and (not approval_goal or len(lote_total) < approval_goal) and int(estado.get("multi_candidates_discovered_total", 0) or 0) < int(config.get("candidate_target", 5000) or 5000):
        seeds = _seed_ids()
        for source_name, fn in [
            ("uploads_expansion", discovery_uploads_expansion),
            ("subscriptions_seed", discovery_subscriptions_seed),
            ("activities_seed", discovery_activities_seed),
        ]:
            if not estado.get("rodando"):
                break
            antes_total = int(estado.get("multi_candidates_discovered_total", 0) or 0)
            try:
                novos = fn(seeds, seen, banco, config)
            except Exception as e:
                print(f"[DISCOVERY][{source_name}] best-effort falhou: {e}")
                novos = []
            if novos:
                with lock:
                    estado["multi_candidates_discovered_total"] = int(estado.get("multi_candidates_discovered_total", 0) or 0) + len(novos)
                    estado["candidatos_encontrados"] = int(estado.get("multi_candidates_discovered_total", 0) or 0)
                    by = estado.setdefault("multi_candidates_by_source", {})
                    by[source_name] = int(by.get(source_name, 0) or 0) + len(novos)
                pending.extend(_dedupe_refs_por_id(novos))
                processar_pending(False, source_hint=source_name)
            stats_fonte[source_name] = int(estado.get("multi_candidates_discovered_total", 0) or 0) - antes_total

    _qperf_save(qperf)
    if pending and estado.get("rodando"):
        processar_pending(True, source_hint="final_flush")
    with lock:
        estado["api_quota_used"] = _api_quota_used()
        estado["api_quota_budget"] = API_DAILY_BUDGET
        estado["api_quota_pct"] = round((_api_quota_used() / max(1, API_DAILY_BUDGET)) * 100, 1)
        estado["stage_label"] = "Multi-source finalizado"
        estado["ultimo_evento"] = f"Final: {len(lote_total)} qualificados, {len(reprovados_total)} reprovados"
    print(f"[MULTI SOURCE] final CHUNKED | candidatos={estado.get('multi_candidates_discovered_total', 0)} | blocos={bloco_num} | qualificados={len(lote_total)} | reprovados={len(reprovados_total)} | chamadas_unicas={len(call_cache)} | fontes={stats_fonte}")
    return lote_total, reprovados_total

def _dlp_buscar_canais(config):
    queries = obter_queries_para_engine(config, limit=config.get("query_limit"), formato="tuplas", fallback_kind="api")
    target = int(config.get("candidate_target") or 100)
    per_query = max(5, min(50, int(config.get("dlp_search_per_query") or 50)))
    banco = ids_banco_dados()
    seen = set()
    candidatos = []
    print(f"\n[DLP] BUSCA 100% yt-dlp | meta={target} canais | queries={len(queries)}")
    for qi, (q, nicho) in enumerate(queries, start=1):
        if not estado.get("rodando") or len(candidatos) >= target:
            break
        with lock:
            estado["queries_processadas"] = qi
            estado["pipeline_stage"] = "dlp_search"
            estado["msg"] = f"DLP buscando {qi}/{len(queries)}: {q[:45]}"
        print(f"  [dlp search] {qi}/{len(queries)} {q!r}")
        linhas = run_lines(["--flat-playlist", "--print", "%(channel_id)s\t%(uploader)s", f"ytsearch{per_query}:{q}"], timeout=55)
        before = len(candidatos)
        for linha in linhas:
            partes = linha.split("\t")
            if len(partes) < 2:
                continue
            cid = (partes[0] or "").strip()
            nome = (partes[1] or "").strip()
            if not cid or cid == "NA" or len(cid) < 8:
                continue
            if cid in seen or cid in banco:
                continue
            text = nome.lower()
            if any(k in text for k in NEGATIVOS_FORTES_API):
                continue
            seen.add(cid)
            item = {"id": cid, "channel_id": cid, "nome": nome, "title": nome, "url": f"https://youtube.com/channel/{cid}", "nicho": nicho, "query": q, "found_query": q, "source": "yt_dlp_search"}
            candidatos.append(item)
            _adicionar_bruto({**item, "status": "bruto_dlp"})
            if len(candidatos) >= target:
                break
        print(f"    +{len(candidatos)-before} | total={len(candidatos)}")
        with lock:
            estado["candidatos_encontrados"] = len(candidatos)
        if not _sleep_interrompivel(0.6):
            break
    return candidatos

def pegar_info_canal_dlp(channel_id, variant=None):
    """Busca inscritos e descrição do canal via yt-dlp. Parâmetro variant ignorado (compatibilidade)."""
    try:
        url = f"https://www.youtube.com/channel/{channel_id}"
        result = run_cmd(["--dump-single-json", "--flat-playlist", "--playlist-end", "1", url], timeout=50)
        if result:
            data = json.loads(result)
            return {
                "subs": data.get("channel_follower_count") or data.get("subscriber_count") or 0,
                "descricao": data.get("description") or "",
                "uploader": data.get("uploader") or data.get("channel") or "",
                "variant": variant or "default",
            }
    except Exception as e:
        print(f"  [dlp info erro] {channel_id}: {e}")
    return {"subs": 0, "descricao": "", "uploader": "", "variant": variant or "default"}

def _dlp_score(c, info, videos, config):
    subs = int(info.get("subs") or 0)
    longos = int((videos or {}).get("longos") or 0)
    shorts = int((videos or {}).get("shorts") or 0)
    avg_views = int((videos or {}).get("avg_views") or 0)
    dias = int((videos or {}).get("dias_ultimo") or 999)
    score = 0
    motivos = []
    if config["min_subs"] <= subs <= config["max_subs"]:
        score += 35
    else:
        motivos.append(f"subs fora ({fmt_numero(subs)})")
    if longos >= 3:
        score += 30
    elif longos >= 1:
        score += 18
    else:
        motivos.append(f"poucos longos ({longos})")
    if dias <= 45:
        score += 12
    elif dias <= 120:
        score += 6
    elif dias > config.get("max_last_days", 365):
        motivos.append(f"canal parado ({dias}d)")
    if avg_views >= 1000:
        score += 12
    elif avg_views >= 300:
        score += 6
    # shorts nao penalizam mais no DLP; canal shorts-only ainda reprova por min_longos
    text = (c.get("nome", "") + " " + info.get("descricao", "")).lower()
    if any(k in text for k in SINAIS_PODCAST_API):
        score += 8
    if any(k in text for k in NEGATIVOS_FORTES_API):
        score -= 20
        motivos.append("termo negativo")
    # V58.24: filtro anti-gringo por territorio
    gringo_reprovado, gringo_motivo, idioma = _filtrar_gringo_por_territorio(c, info, config)
    if gringo_reprovado:
        motivos.append(gringo_motivo)
        return 0, motivos
    # Penalizacao leve se canal é neutro mas tem alguns sinais EN (nao reprova)
    if idioma == "en" and not any(r in (config.get("regioes") or ["BR"]) for r in ["US", "GLOBAL"]):
        en_count = _detectar_idioma_canal(c.get("nome",""), info.get("descricao",""), info.get("uploader","")).get("en", 0)
        if en_count > 0:
            score -= min(en_count * 3, 15)
            if en_count >= 2:
                motivos.append(f"tendencia_gringo_en ({en_count})")

    # V58.26 #9: filtro de audiencia comprada
    if subs > 5000 and avg_views > 0:
        ratio_val = avg_views / max(subs, 1)
        if ratio_val < 0.005:
            score -= 30; motivos.append("audiencia_comprada")
        elif ratio_val < 0.01:
            score -= 15; motivos.append("views_mortas_v26")
        elif ratio_val > 0.10:
            score += 15

    # V58.26 #8: deteccao de monetizacao
    tem_monet, sinais_monet = _detectar_monetizacao(c)
    if tem_monet:
        bonus_monet = min(len(sinais_monet) * 3, 12)
        score += bonus_monet
        c["sinais_monetizacao"] = sinais_monet

    # V58.26 #5: deteccao de editor nos creditos (videos recentes)
    # No DLP, videos = pegar_videos_canal retorna so metricas (sem descricao).
    # Para deteccao de editor no DLP, checamos a descricao do canal.
    desc_canal = info.get("descricao") or ""
    for padrao in PADROES_EDITOR_CREDITO:
        if re.search(padrao, desc_canal, flags=re.I):
            match_text = re.search(padrao, desc_canal, flags=re.I).group(0).lower()
            if not any(s in match_text for s in ["propria", "própria", "proprio", "próprio", "eu mesmo"]):
                score -= 40; motivos.append("ja_tem_editor")
                c["ja_tem_editor"] = True
                break

    # V58.26 #4: blacklist de concorrentes
    concorrentes_ids = _carregar_concorrentes_blacklist()
    cid_c = str(c.get("id") or c.get("channel_id") or "").strip()
    if cid_c and cid_c in concorrentes_ids:
        score = 0; motivos.append("concorrente_blacklist")
        c["concorrente_blacklist"] = True

    score = max(0, min(100, int(score)))
    c["score_tags_v26"] = ["dlp_mode"]  # DLP nao tem tantos sinais quanto API
    if score < config.get("score_min", 55):
        motivos.append(f"score baixo ({score})")
    if longos < config.get("min_longos", 1):
        motivos.append(f"min_longos ({longos})")
    return score, motivos

def dlp_verify_candidates(candidatos, config, source_label="api_then_dlp"):
    """Verifica uma lista de canais já descobertos usando somente yt-dlp."""
    verify_max = int(config.get("ytdlp_verify_max") or len(candidatos) or 0)
    if bool(config.get("verify_until_empty", False)):
        verify_max = len(candidatos)
    candidatos = list(candidatos or [])[:max(0, verify_max)]
    lote = []
    reprovados = []
    print(f"\n[DLP VERIFY] verificação de fila | source={source_label} | canais={len(candidatos)}")
    for i, c in enumerate(candidatos, start=1):
        if not estado.get("rodando"):
            break
        cid = c.get("id") or c.get("channel_id")
        if not cid:
            continue
        with lock:
            estado["pipeline_stage"] = "dlp_verify_queue"
            estado["msg"] = f"DLP verificando fila {i}/{len(candidatos)}: {c.get('nome','')[:35]}"
            estado["verificados"] = i
            estado["ytdlp_verified"] = int(estado.get("ytdlp_verified", 0) or 0) + 1
        info = pegar_info_canal_dlp(cid)
        videos = pegar_videos_canal(cid) or {"longos": 0, "shorts": 0, "avg_views": 0, "dias_ultimo": 999}
        score, motivos = _dlp_score(c, info, videos, config)
        c.update({"subs": int(info.get("subs") or 0), "subscriber_count": int(info.get("subs") or 0), "subs_fmt": fmt_numero(info.get("subs") or 0), "description": info.get("descricao", ""), "longos": videos.get("longos", 0), "shorts": videos.get("shorts", 0), "recent_avg_views": videos.get("avg_views", 0), "ultimo": formatar_dias(videos.get("dias_ultimo", 999)), "last_video_days": videos.get("dias_ultimo", 999), "score": score, "score_reasons": source_label, "engine_mode": source_label, "source": source_label})
        if not motivos:
            item = _canal_api_para_lote(c)
            item["source"] = source_label
            item["engine_mode"] = source_label
            item["status_fila"] = "pendente"
            lote.append(item)
            _atualizar_status_bruto(cid, "qualificado_dlp", {"score": score, "subs": c.get("subs"), "longos": c.get("longos")})
            print(f"  [DLP QUALIFICADO] score={score} | {c.get('subs_fmt')} | {c.get('nome')} | longos={c.get('longos')} avg={fmt_numero(c.get('recent_avg_views') or 0)}")
        else:
            c["motivo"] = "; ".join(motivos)
            reprovados.append(_canal_api_para_reprovado(c))
            _atualizar_status_bruto(cid, "reprovado_dlp", {"motivo": c["motivo"], "score": score})
        with lock:
            estado["qualificados"] = len(lote)
        if not _sleep_interrompivel(0.25):
            break
    # V58.26 #11: ordenar lote DLP por score (maior primeiro)
    lote.sort(key=lambda c: (int(c.get("score", 0) or 0), int(c.get("subs", 0) or 0)), reverse=True)
    # V58.26 #1: enrich automatico nos top DLP (email + Instagram)
    if lote:
        try:
            _enrich_leads_rapido(lote, max_per_call=10)
        except Exception as e:
            print(f"[DLP VERIFY] aviso enrich: {e}")
    return lote, reprovados

def dlp_verify_queue_until_empty(config):
    """Processa a fila DLP persistente até acabar, ou até o usuário parar."""
    lote_total = []
    reprovados_total = []
    while estado.get("rodando"):
        fila = carregar_fila_dlp_verificacao()
        if not fila:
            break
        with lock:
            estado["pipeline_stage"] = "dlp_queue_start"
            estado["msg"] = f"DLP iniciando verificação automática da fila: {len(fila)} canais"
        lote, reprovados = dlp_verify_candidates(fila, {**config, "ytdlp_verify_max": len(fila), "verify_until_empty": True}, source_label="api_then_dlp")
        ids_processados = {c.get("id") for c in (lote + reprovados) if c.get("id")}
        remover_fila_dlp_verificacao_ids(ids_processados)
        lote_total.extend(lote)
        reprovados_total.extend(reprovados)
        if not ids_processados:
            break
    return lote, reprovados


    # ============================================================
    # WORKER API-FIRST (com depósito de brutos)
    # ============================================================
def normalizar_config_api(data):
    data = data or {}
    if isinstance(data, dict) and isinstance(data.get("config"), dict):
        merged = dict(data.get("config") or {})
        for k, v in data.items():
            if k != "config" and k not in merged:
                merged[k] = v
        data = merged

    perf = str(data.get("performance") or data.get("performance_mode") or "normal").lower()
    perf_defaults = {"baixo": 8, "normal": 15, "alto": 25, "extremo": 40}
    query_limit = _safe_int(data.get("query_limit", data.get("queryLimit", perf_defaults.get(perf, 15))), perf_defaults.get(perf, 15), 1, API_QUERY_LIMIT_MAX)

    nichos = data.get("nichos") or []
    if isinstance(nichos, str):
        nichos = [n.strip() for n in nichos.split(",") if n.strip()]
    nichos = [str(n).strip().lower() for n in nichos if str(n).strip()]
    if not nichos:
        nichos = ["empreendedorismo", "marketing digital", "negocios", "financas", "tecnologia"]
    perfis = data.get("perfis") or data.get("profiles") or []
    if isinstance(perfis, str):
        perfis = [p.strip() for p in perfis.split(",") if p.strip()]
    perfis = [str(p).strip().lower() for p in perfis if str(p).strip()]
    objetivos = data.get("objetivos") or data.get("objetivo") or []
    if isinstance(objetivos, str):
        objetivos = [p.strip() for p in objetivos.split(",") if p.strip()]
    objetivos = [str(p).strip() for p in objetivos if str(p).strip()]
    formatos = data.get("formatos") or data.get("formato_video") or data.get("formato") or []
    if isinstance(formatos, str):
        formatos = [p.strip() for p in formatos.split(",") if p.strip()]
    formatos = [str(p).strip() for p in formatos if str(p).strip()]
    regioes = data.get("regioes") or data.get("regiao") or data.get("localizacoes") or data.get("localizacao") or []
    if isinstance(regioes, str):
        regioes = [p.strip() for p in regioes.split(",") if p.strip()]
    regioes = [str(p).strip() for p in regioes if str(p).strip()]

    query_personalizada = str(data.get("query_personalizada") or data.get("query") or "").strip()
    queries_extras = [query_personalizada] if query_personalizada else []

    min_subs = _safe_int(data.get("min_subs", data.get("minSubs", MIN_SUBS)), MIN_SUBS, 1, 10_000_000)
    max_subs = _safe_int(data.get("max_subs", data.get("maxSubs", MAX_SUBS)), MAX_SUBS, 1, 10_000_000)
    if max_subs < min_subs:
        min_subs, max_subs = max_subs, min_subs

    engine_mode = _normalizar_engine_mode(data.get("engine_mode") or data.get("modo_coleta") or data.get("collect_mode") or carregar_engine_mode())
    lote_size = _safe_int(data.get("lote_size", data.get("loteSize", LOTE_PADRAO)), LOTE_PADRAO, 1, 1000)
    score_min = _safe_int(data.get("score_min", data.get("scoreMin", 55)), 55, 0, 100)
    api_like = engine_mode in {"api", "api_multi_source_fast"}
    candidate_target_max = 50000 if api_like else 1000
    candidate_target = _safe_int(data.get("candidate_target", data.get("candidateTarget", data.get("meta_candidatos", 1000 if api_like else 120))), 1000 if api_like else 120, 20, candidate_target_max)
    pages_per_query = _safe_int(data.get("pages_per_query", data.get("pagesPerQuery", API_PAGES_PER_QUERY)), API_PAGES_PER_QUERY, 1, 10)
    videos_per_channel = _safe_int(data.get("videos_per_channel", data.get("videosPerChannel", API_VIDEOS_PER_CHANNEL)), API_VIDEOS_PER_CHANNEL, 1, 15)
    min_longos = _safe_int(data.get("min_longos", data.get("minLongos", 1)), 1, 0, 5)
    max_last_days = _safe_int(data.get("max_last_days", data.get("maxLastDays", 365)), 365, 30, 5000)
    ytdlp_verify_max = _safe_int(data.get("ytdlp_verify_max", data.get("ytdlpVerifyMax", candidate_target)), candidate_target, 1, 1000)
    default_workers = PRESETS.get("api_multi_source_fast", {}).get("workers", 1) if engine_mode == "api_multi_source_fast" else 1
    default_query_workers = PRESETS.get("api_multi_source_fast", {}).get("query_workers", 1) if engine_mode == "api_multi_source_fast" else 1
    workers = _safe_int(data.get("workers", data.get("worker_count", default_workers)), default_workers, 1, 32)
    query_workers = _safe_int(data.get("query_workers", data.get("queryWorkers", default_query_workers)), default_query_workers, 1, 32)

    discovery_batch_size = _safe_int(data.get("discovery_batch_size", data.get("discoveryBatchSize", 500)), 500, 50, 2000)
    api_verify_batch_size = _safe_int(data.get("api_verify_batch_size", data.get("apiVerifyBatchSize", 50)), 50, 10, 200)
    dlp_verify_batch_size = _safe_int(data.get("dlp_verify_batch_size", data.get("dlpVerifyBatchSize", 75)), 75, 10, 200)
    approval_goal = _safe_int(data.get("approval_goal", data.get("approvalGoal", 0)), 0, 0, 100000)
    dlp_queries_per_cycle = _safe_int(data.get("dlp_queries_per_cycle", data.get("dlpQueriesPerCycle", 60)), 60, 5, 200)
    max_total_queries = _safe_int(data.get("max_total_queries", data.get("maxTotalQueries", 500)), 500, 100, 2000)
    video_filter = data.get("video_filter") or {}
    if isinstance(video_filter, dict):
        min_dur = _safe_int(video_filter.get("min_duration", video_filter.get("minDuration", 0)), 0, 0, 86400)
        max_dur = _safe_int(video_filter.get("max_duration", video_filter.get("maxDuration", 0)), 0, 0, 86400)
        min_last_days = _safe_int(video_filter.get("min_last_days", video_filter.get("minLastDays", 0)), 0, 0, 3650)
        max_last_days = _safe_int(video_filter.get("max_last_days", video_filter.get("maxLastDays", 365)), 365, 0, 3650)
        keywords = video_filter.get("keywords_title") or video_filter.get("keywordsTitle") or []
        if isinstance(keywords, str):
            keywords = [k.strip() for k in keywords.split(",") if k.strip()]
        ignore_shorts = bool(video_filter.get("ignore_shorts", video_filter.get("ignoreShorts", False)))
        require_keyword = bool(video_filter.get("require_keyword", video_filter.get("requireKeyword", False)))
    else:
        min_dur = max_dur = min_last_days = 0
        keywords = []
        ignore_shorts = False
        require_keyword = False

    config_out = {
        "nichos": nichos,
        "perfis": perfis,
        "objetivos": objetivos,
        "formatos": formatos,
        "regioes": regioes,
        "lote_size": lote_size,
        "score_min": score_min,
        "query_limit": query_limit,
        "candidate_target": candidate_target,
        "pages_per_query": pages_per_query,
        "videos_per_channel": videos_per_channel,
        "min_subs": min_subs,
        "max_subs": max_subs,
        "min_longos": min_longos,
        "max_last_days": max_last_days,
        "queries_extras": queries_extras,
        "performance": perf,
        "engine": ENGINE_VERSION,
        "engine_mode": engine_mode,
        "ytdlp_verify_max": ytdlp_verify_max,
        "api_discovery_only": engine_mode == "api",
        "workers": workers,
        "query_workers": query_workers,
        "discovery_batch_size": discovery_batch_size,
        "api_verify_batch_size": api_verify_batch_size,
        "dlp_verify_batch_size": dlp_verify_batch_size,
        "approval_goal": approval_goal,
        "dlp_queries_per_cycle": dlp_queries_per_cycle,
        "max_total_queries": max_total_queries,
        "video_filter": {
            "min_duration": min_dur,
            "max_duration": max_dur if max_dur else 999999,
            "min_last_days": min_last_days,
            "max_last_days": max_last_days,
            "keywords_title": [str(k).strip().lower() for k in keywords if str(k).strip()],
            "ignore_shorts": ignore_shorts,
            "require_keyword": require_keyword,
        }
    }
    if engine_mode == "api_multi_source_fast":
        preset = dict(PRESETS["api_multi_source_fast"])
        preset.update(config_out)
        preset["engine_mode"] = "api_multi_source_fast"
        preset["api_discovery_only"] = True
        config_out = preset
    return config_out

def _safe_int(value, default, minimo=None, maximo=None):
    try:
        value = int(value)
    except Exception:
        value = default
    if minimo is not None:
        value = max(minimo, value)
    if maximo is not None:
        value = min(maximo, value)
    return value

def worker_api_first(config, auto=False):
    mode = _normalizar_engine_mode(config.get("engine_mode") or carregar_engine_mode())
    salvar_engine_mode(mode)
    with lock:
        estado["status"] = "buscando"
        estado["rodando"] = True
        estado["pausa"] = False
        estado["auto_mode"] = bool(auto)
        estado["pipeline_stage"] = f"{mode}_start"
        estado["msg"] = f"Iniciando {_modo_label(mode)}..."
        estado["verificados"] = 0
        estado["qualificados"] = 0
        estado["queries_processadas"] = 0
        estado["candidatos_encontrados"] = 0
        estado["api_channels_detailed"] = 0
        estado["api_videos_collected"] = 0
        estado["api_rejected"] = 0
        estado["api_last_error"] = ""
        estado["api_engine"] = ENGINE_VERSION
        estado["engine_mode"] = mode
        estado["api_discovery_only"] = mode == "api"
        estado["dlp_queue_count"] = len(carregar_fila_dlp_verificacao())

    print(f"\n[PROSPECTOR] ===== CICLO INICIADO | modo={mode} | auto={auto} | config={config} =====")
    lote_api = []
    reprovados_api = []
    # V58.33: limpa reprovados antigos automaticamente a cada 24h
    try:
        last_cleanup_file = ".thon_blacklist_last_cleanup"
        need_cleanup = True
        if os.path.exists(last_cleanup_file):
            try:
                last_ts = float(open(last_cleanup_file).read().strip() or 0)
                if (time.time() - last_ts) < 86400:  # menos de 24h
                    need_cleanup = False
            except Exception:
                pass
        if need_cleanup:
            r = limpar_reprovados_antigos(dias_ttl=REPROVADOS_TTL_DIAS, dry_run=False)
            if r.get("removidos", 0) > 0:
                print(f"[BLACKLIST TTL] limpeza automática: removidos={r['removidos']} | blacklist {r['total_antes']} → {r['total_depois']}")
            try:
                with open(last_cleanup_file, "w") as f:
                    f.write(str(time.time()))
            except Exception:
                pass
    except Exception as _e_cleanup:
        print(f"[BLACKLIST TTL] aviso limpeza: {_e_cleanup}")
    try:
        if mode == "dlp":
            lote_api, reprovados_api = dlp_full_process(config)
            final_label = "Prospector DLP"
        elif mode == "api_multi_source_fast":
            lote_api, reprovados_api = worker_api_multi_source(config, auto=auto)
            final_label = "Prospector API Multi Source"
        else:
            # Modo API direto. A API qualifica com score normal e joga no lote.
            # Nao chama DLP/yt-dlp/verify_only neste modo.
            lote_api, reprovados_api = api_discovery_only_process(config)
            final_label = "Prospector API Direto"
        if reprovados_api:
            salvar_reprovados_lista(reprovados_api)
            for c in reprovados_api:
                if c.get("id"):
                    _atualizar_status_bruto(c["id"], f"reprovado_{mode}", {"motivo": c.get("motivo", "")})
        # FIX BUG 2: garantir que todo item do lote tem 'id' antes de salvar
        for item in lote_api:
            if not item.get("id") and item.get("channel_id"):
                item["id"] = item["channel_id"]
            elif not item.get("channel_id") and item.get("id"):
                item["channel_id"] = item["id"]
        lote_api = [c for c in lote_api if c.get("id")]
        # V58.25: SEMPRE adiciona qualificados ao lote ativo (camada extra de seguranca)
        # Modos api/api_multi_source_fast ja chamam sync_api_qualificados_para_lote internamente,
        # mas isso garante que se algo falhar la dentro, ainda salvamos aqui.
        if lote_api:
            try:
                # Para modos api/api_multi_source_fast, sync ja foi feito internamente.
                # Para modo DLP, sync foi feito no dlp_full_process.
                # Aqui apenas garantimos que o estado["lote"] esteja sincronizado com o arquivo.
                estado_lote_atual = adicionar_lote_ativo(lote_api if mode == "dlp" else [], base=estado.get("lote", []))
            except Exception as e:
                print(f"[PROSPECTOR] aviso: falha adicionar_lote_ativo no final: {e}")
        with lock:
            ids_processados = {c.get("id") for c in lote_api + reprovados_api if c.get("id")}
            estado["vistos"].update(ids_processados)
            estado["reprovados"] = carregar_reprovados()
            salvar_memoria(estado["vistos"])
            # V58.25: SEMPRE recarrega do arquivo (fonte unica de verdade)
            estado["lote"] = carregar_lote_ativo()
            print(f"[PROSPECTOR] FINAL lote_count={len(estado['lote'])} | arquivo={LOTE_ATIVO_FILE}")
            estado["status"] = "aguardando" if estado.get("lote") else "fim"
            estado["pausa"] = bool(estado.get("lote"))
            estado["qualificados"] = len(lote_api)
            estado["pipeline_stage"] = "aguardando_revisao" if lote_api else "concluido"
            estado["dlp_queue_count"] = len(carregar_fila_dlp_verificacao())
            estado["msg"] = f"{final_label} finalizado: +{len(lote_api)} no lote | {len(reprovados_api)} reprovados | fila aprovação {len(estado.get('lote', []) or [])} | fila DLP {estado['dlp_queue_count']}"
            qlista = estado.setdefault("qualificados_lista", [])
            ja = {c.get("id") for c in qlista if c.get("id")}
            qlista.extend([c for c in lote_api if c.get("id") not in ja])
        print(f"[PROSPECTOR] FINAL | modo={mode} | novos_lote={len(lote_api)} | reprovados={len(reprovados_api)} | fila_total={len(estado.get('lote', []) or [])} | fila_dlp={len(carregar_fila_dlp_verificacao())} | quota={_api_quota_used()}/{API_DAILY_BUDGET}")
    except Exception as e:
        erro = str(e)
        print(f"[PROSPECTOR ERRO] {erro}")
        with lock:
            estado["api_last_error"] = erro
            estado["status"] = "erro"
            estado["pipeline_stage"] = "erro"
            estado["msg"] = f"Erro {mode}: {erro[:180]}"
    finally:
        with lock:
            estado["rodando"] = False
            estado["auto_mode"] = False
            salvar_memoria(estado["vistos"])
            salvar_aprovados(estado["aprovados"])
            salvar_lote_ativo(estado.get("lote", []))
        if auto:
            auto_hunt["cycles"] = int(auto_hunt.get("cycles", 0) or 0) + 1
            auto_hunt["last_finished"] = str(datetime.now())


@app.route("/api/engine_mode", methods=["GET", "POST"])
def api_engine_mode_route():
    if request.method == "GET":
        mode = carregar_engine_mode()
        with lock:
            estado["engine_mode"] = mode
            estado["api_discovery_only"] = mode == "api"
        return jsonify({"ok": True, "mode": mode, "label": _modo_label(mode), "file": MODE_FILE})
    data = request.json or {}
    mode = salvar_engine_mode(data.get("mode") or data.get("engine_mode") or data.get("modo_coleta"))
    with lock:
        estado["engine_mode"] = mode
        estado["api_discovery_only"] = mode == "api"
        estado["msg"] = f"Modo de coleta definido: {_modo_label(mode)}"
    return jsonify({"ok": True, "mode": mode, "label": _modo_label(mode), "file": MODE_FILE})

def iniciar_busca_api(config, auto=False):
    with lock:
        if estado.get("rodando"):
            return False
        estado["rodando"] = True
    t = threading.Thread(target=worker_api_first, args=(config, auto), daemon=True, name="garimpo-api-first")
    t.start()
    return True

@app.route("/modo-coleta")
def modo_coleta_page():
    return "", 302, {"Location": "/prospector?engine_mode=api"}

@app.route("/api/reset_estado", methods=["POST"])
def reset_estado():
    with lock:
        estado["rodando"] = False
        estado["pausa"] = False
        estado["status"] = "idle"
        estado["pipeline_stage"] = "idle"
        estado["msg"] = ""
        estado["api_last_error"] = ""
    auto_hunt["enabled"] = False
    return jsonify({"ok": True})


@app.route("/thon_mode_widget.js")
def thon_mode_widget_js():
    return "/* widget antigo desativado na V58.11 */", 200, {"Content-Type": "application/javascript; charset=utf-8"}

@app.route("/iniciar", methods=["POST"])
def iniciar_api_route():
    try:
        data = dict(request.json or {})
        ref = (request.headers.get("Referer") or "").lower()
        # Sem mexer no HTML: se o Prospector foi aberto como /prospector?engine_mode=api ou /prospector?engine_mode=dlp,
        # o backend força o modo correto pela URL de origem.
        if "engine_mode=api" in ref or "/prospector-api" in ref:
            data["engine_mode"] = "api"
        elif "engine_mode=dlp" in ref or "/prospector-dlp" in ref:
            data["engine_mode"] = "dlp"
        config = normalizar_config_api(data)
    except Exception as exc:
        return jsonify({"ok": False, "erro": str(exc)}), 400
    started = iniciar_busca_api(config, auto=False)
    if not started:
        return jsonify({"ok": False, "erro": "Busca ja rodando"}), 400
    return jsonify({"ok": True, "version": APP_VERSION, "engine": ENGINE_VERSION, "config": config, "modo": config.get("engine_mode")})

@app.route("/reset_memoria", methods=["POST"])
def reset_memoria():
    with lock:
        novos_vistos, novos_aprovados, novos_reprovados = limpar_memoria()
        estado['vistos'] = novos_vistos
        estado['aprovados'] = novos_aprovados
        estado['reprovados'] = novos_reprovados
    return jsonify({"ok": True})

@app.route("/reprovados")
def get_reprovados():
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 100))
    with lock:
        total = len(estado['reprovados'])
        start = (page - 1) * per_page
        end = start + per_page
        return jsonify({
            "reprovados": estado['reprovados'][start:end],
            "total": total,
            "page": page,
            "per_page": per_page,
            "total_pages": (total + per_page - 1) // per_page
        })

@app.route("/sincronizar_banco", methods=["POST"])
def sincronizar_banco():
    with lock:
        dados = sincronizar_banco_dados(salvar=True)
        estado['vistos'] = dados['vistos']
        estado['aprovados'] = dados['aprovados']
        estado['reprovados'] = dados['reprovados']
    return jsonify({"ok": True, **diagnostico_banco()})

def diagnostico_banco():
    banco = ids_banco_dados()
    return {
        "vistos_memoria": len(estado.get("vistos", set())),
        "aprovados": len(estado.get("aprovados", [])),
        "reprovados": len(estado.get("reprovados", [])),
        "banco_total": len(banco),
        "arquivos": {
            MEMORY_FILE: os.path.exists(MEMORY_FILE),
            APROVADOS_FILE: os.path.exists(APROVADOS_FILE),
            REPROVADOS_FILE: os.path.exists(REPROVADOS_FILE),
            "canais_vistos.json": os.path.exists("canais_vistos.json"),
            "canais_aprovados.json": os.path.exists("canais_aprovados.json"),
            "canais_reprovados.json": os.path.exists("canais_reprovados.json"),
        }
    }

@app.route("/diagnostico")
def diagnostico():
    with lock:
        return jsonify(diagnostico_banco())

@app.route("/estado")
def get_estado():
    with lock:
        estado["lote"] = carregar_lote_ativo()
        return jsonify({
            "status": estado["status"],
            "msg": estado["msg"],
            "rodando": bool(estado.get("rodando")),
            "pausa": bool(estado.get("pausa")),
            "pipeline_stage": estado["pipeline_stage"],
            "lote": estado["lote"],
            "lote_count": len(estado.get("lote", []) or []),
            "fila_pendente_file": LOTE_ATIVO_FILE,
            "verificados": estado["verificados"],
            "qualificados": estado["qualificados"],
            "aprovados": len(estado["aprovados"]),
            "reprovados": len(estado["reprovados"]),
            "vistos_total": len(ids_banco_dados()),
            "banco_total": len(ids_banco_dados()),
            "vistos_memoria": len(estado["vistos"]),
            "version": APP_VERSION,
            "engine": ENGINE_VERSION,
            "engine_mode": estado.get("engine_mode", carregar_engine_mode()),
            "engine_mode_label": _modo_label(estado.get("engine_mode", carregar_engine_mode())),
            "api_discovery_only": bool(estado.get("api_discovery_only", True)),
            "api_quota_used": _api_quota_used(),
            "api_quota_budget": API_DAILY_BUDGET,
            "api_channels_detailed": estado.get("api_channels_detailed", 0),
            "api_videos_collected": estado.get("api_videos_collected", 0),
            "ytdlp_verified": estado.get("ytdlp_verified", 0),
            "candidatos_encontrados": estado.get("candidatos_encontrados", 0),
            "brutos_total": len(_carregar_brutos()),
            "dlp_queue_count": len(carregar_fila_dlp_verificacao()),
            "dlp_queue_file": DLP_VERIFY_QUEUE_FILE,
            "auto_hunt": dict(auto_hunt),
        })

@app.route("/proximo", methods=["POST"])
def proximo():
    # Compatibilidade com o frontend antigo: antes essa rota limpava o lote.
    # Agora ela preserva a fila para não sumir candidato sem aprovação/descartar.
    data = request.json or {}
    clear = bool(data.get("clear") or data.get("limpar"))
    with lock:
        if clear:
            estado["lote"] = salvar_lote_ativo([])
            estado["pausa"] = False
            estado["status"] = "idle"
            return jsonify({"ok": True, "cleared": True, "lote_count": 0})
        estado["lote"] = carregar_lote_ativo()
        estado["status"] = "aguardando" if estado["lote"] else "idle"
        estado["pausa"] = bool(estado["lote"])
    return jsonify({"ok": True, "preservado": True, "lote_count": len(estado.get("lote", []) or [])})

@app.route("/aprovar", methods=["POST"])
def aprovar():
    ids = {str(x) for x in ((request.json or {}).get("ids", []) or []) if str(x).strip()}
    if not ids:
        return jsonify({"erro": "Nenhum canal selecionado"}), 400
    crm_total = 0
    with lock:
        lote_oficial = carregar_lote_ativo()
        estado["lote"] = lote_oficial
        novos = [c for c in lote_oficial if str(c.get("id") or "") in ids or str(c.get("channel_id") or "") in ids]
        aprovados_ids = {str(a.get("id") or a.get("channel_id") or "") for a in estado.get("aprovados", []) or []}
        for c in novos:
            cid = c.get("id") or c.get("channel_id")
            if cid and str(cid) not in aprovados_ids:
                c["id"] = cid
                c["channel_id"] = cid
                estado.setdefault("aprovados", []).append(c)
                aprovados_ids.add(str(cid))
                _atualizar_status_bruto(cid, "aprovado", {"aprovado_em": str(datetime.now())})
        estado["vistos"].update(ids)
        estado["lote"] = [c for c in lote_oficial if (str(c.get("id") or "") not in ids and str(c.get("channel_id") or "") not in ids)]
        salvar_lote_ativo(estado["lote"])
        salvar_memoria(estado['vistos'])
        salvar_aprovados(estado['aprovados'])
    try:
        if "sincronizar_crm_leads_incremental" in globals():
            crm_total = sincronizar_crm_leads_incremental(novos)
        else:
            crm_total = len(sincronizar_crm_com_aprovados())
    except Exception as e:
        print(f"[CRM] sync apos aprovar falhou: {e}")
    return jsonify({"ok": True, "aprovados_novos": len(novos), "total": len(estado.get("aprovados", []) or []), "crm_total": crm_total, "lote_count": len(estado.get("lote", []) or []), "crm_sync": True})

@app.route("/aprovados")
def get_aprovados():
    with lock:
        return jsonify({"canais": estado["aprovados"], "total": len(estado["aprovados"])})

@app.route("/remover_aprovado", methods=["POST"])
def remover_aprovado():
    ids = set(request.json.get("ids", []))
    if not ids:
        return jsonify({"erro": "Nenhum canal selecionado"}), 400
    with lock:
        estado["aprovados"] = [c for c in estado["aprovados"] if c["id"] not in ids]
        estado["vistos"].update(ids)
        salvar_aprovados(estado["aprovados"])
        salvar_memoria(estado["vistos"])
    return jsonify({"ok": True, "total": len(estado["aprovados"])})

@app.route("/descartar", methods=["POST"])
def descartar():
    data = request.json or {}
    ids = {str(x) for x in (data.get("ids", []) or []) if str(x).strip()}
    motivo = data.get("motivo", "Descartado manualmente")
    if not ids:
        return jsonify({"erro": "Nenhum canal selecionado"}), 400
    descartados = []
    with lock:
        lote_oficial = carregar_lote_ativo()
        estado["lote"] = lote_oficial
        for canal in lote_oficial:
            cid = str(canal.get("id") or canal.get("channel_id") or "")
            if cid in ids:
                descartados.append({
                    "id": cid,
                    "nome": canal.get('nome') or canal.get("title") or "Canal",
                    "url": canal.get('url') or (f"https://youtube.com/channel/{cid}" if cid else ""),
                    "nicho": canal.get('nicho', ''),
                    "score": canal.get('score', 0),
                    "subs_fmt": canal.get('subs_fmt', ''),
                    "motivo": motivo,
                    "data": str(datetime.now())
                })
                estado["vistos"].add(cid)
                _atualizar_status_bruto(cid, "descartado", {"motivo": motivo})
        if descartados:
            salvar_reprovados_lista(descartados)
            estado['reprovados'] = carregar_reprovados()
        estado["lote"] = [c for c in lote_oficial if str(c.get("id") or c.get("channel_id") or "") not in ids]
        salvar_lote_ativo(estado["lote"])
        salvar_memoria(estado['vistos'])
    return jsonify({"ok": True, "descartados": len(descartados), "total": len(estado["reprovados"]), "lote_count": len(estado.get("lote", []) or [])})

@app.route("/parar", methods=["POST"])
def parar():
    auto_hunt["enabled"] = False
    with lock:
        estado["rodando"] = False
        estado["pausa"] = False
        estado["status"] = "idle"
        estado["pipeline_stage"] = "parado"
        estado["msg"] = "Parado pelo usuário"
        salvar_lote_ativo(estado.get("lote", []))
        salvar_memoria(estado['vistos'])
        salvar_aprovados(estado['aprovados'])
    return jsonify({"ok": True, "auto_hunt": auto_hunt})

# ============================================================
# AUTO HUNT
# ============================================================
auto_hunt = {
    "enabled": False,
    "cycles": 0,
    "last_config": None,
    "last_started": None,
    "last_finished": None,
    "thread_alive": False,
}
_auto_thread = None
_auto_lock = threading.RLock()

def _auto_loop_api():
    auto_hunt["thread_alive"] = True
    try:
        while auto_hunt.get("enabled"):
            config = dict(auto_hunt.get("last_config") or {})
            if not config:
                print("[AUTO HUNT] sem config salva, encerrando loop")
                break

            mode = str(config.get("engine_mode") or "api").lower()
            config["engine_mode"] = mode

            # Só dispara se não está rodando
            if estado.get("rodando"):
                # Já tem ciclo ativo (ex: usuário iniciou manualmente), só espera
                time.sleep(1.0)
                continue

            ciclo_num = int(auto_hunt.get('cycles', 0) or 0) + 1
            auto_hunt["last_started"] = str(datetime.now())
            print(f"[AUTO HUNT] disparando ciclo {ciclo_num} | mode={mode}")
            started = iniciar_busca_api(config, auto=True)
            if not started:
                print(f"[AUTO HUNT] ciclo {ciclo_num} não iniciou (já rodando), aguardando 3s...")
                time.sleep(3.0)
                continue

            # Aguarda a thread subir e setar rodando=True (race condition fix)
            for _ in range(30):  # até 3s de espera
                if estado.get("rodando") or not auto_hunt.get("enabled"):
                    break
                time.sleep(0.1)

            # Aguarda o ciclo terminar
            while auto_hunt.get("enabled"):
                if not estado.get("rodando"):
                    break
                time.sleep(1.0)

            if not auto_hunt.get("enabled"):
                break

            auto_hunt["last_finished"] = str(datetime.now())
            ciclos_feitos = int(auto_hunt.get('cycles', 0) or 0)
            lote_atual = len(estado.get('lote') or [])
            print(f"[AUTO HUNT] ciclo {ciclos_feitos} finalizado | lote_acumulado={lote_atual} | aguardando próximo...")

            delay = _safe_int(os.environ.get("THON_AUTOHUNT_DELAY", 15), 15, 5, 24*3600)
            if not _sleep_auto_hunt(delay):
                break

            # Recarrega config para capturar mudanças feitas pelo usuário entre ciclos
            config_atualizada = dict(auto_hunt.get("last_config") or config)
            config_atualizada["engine_mode"] = str(config_atualizada.get("engine_mode") or mode).lower()
            auto_hunt["last_config"] = config_atualizada

    finally:
        auto_hunt["thread_alive"] = False
        print("[AUTO HUNT] loop encerrado")

@app.route("/auto_hunt", methods=["GET", "POST"])
def auto_hunt_route():
    global _auto_thread
    if request.method == "GET":
        return jsonify({"ok": True, "version": APP_VERSION, "engine": ENGINE_VERSION, "auto_hunt": auto_hunt, "nota": "Auto Hunt repete a engine/config selecionada."})

    data = request.json or {}
    enabled = bool(data.get("enabled", False))
    if not enabled:
        auto_hunt["enabled"] = False
        with lock:
            estado["rodando"] = False
            estado["pausa"] = False
            estado["status"] = "idle"
            estado["pipeline_stage"] = "auto_off"
            estado["msg"] = "Auto Hunt desligado"
        return jsonify({"ok": True, "enabled": False, "auto_hunt": auto_hunt})

    try:
        payload = dict(data)
        if isinstance(payload.get("config"), dict):
            payload = {**payload.get("config", {}), **{k: v for k, v in payload.items() if k != "config"}}
        config = normalizar_config_api(payload)
    except Exception as exc:
        return jsonify({"ok": False, "erro": str(exc)}), 400

    auto_hunt["enabled"] = True
    auto_hunt["last_config"] = config
    with _auto_lock:
        if not _auto_thread or not _auto_thread.is_alive():
            _auto_thread = threading.Thread(target=_auto_loop_api, daemon=True, name="auto-hunt")
            _auto_thread.start()
    print(f"[AUTO HUNT] enabled=True | mode={config.get('engine_mode')} | config={config.get('performance')}/{config.get('candidate_target')}")
    return jsonify({"ok": True, "enabled": True, "started": True, "version": APP_VERSION, "engine": ENGINE_VERSION, "config": config, "auto_hunt": auto_hunt, "nota": "Auto Hunt repete a engine/config selecionada."})

# ============================================================
# FUNCIONALIDADES ADICIONAIS (CRM, PROJETOS, TRABALHOS, DOWNLOADER)
# ============================================================
DATA_DIR = "thon_data"
os.makedirs(DATA_DIR, exist_ok=True)
CRM_FILE = os.path.join(DATA_DIR, "crm.json")
CRM_DELETED_FILE = os.path.join(DATA_DIR, "crm_apagados.json")
PROJECTS_FILE = os.path.join(DATA_DIR, "projetos.json")
GOALS_FILE = os.path.join(DATA_DIR, "metas.json")
JOBS_FILE = os.path.join(DATA_DIR, "trabalhos.json")
DOWNLOADS_FILE = os.path.join(DATA_DIR, "downloads.json")
DOWNLOADER_CONFIG_FILE = os.path.join(DATA_DIR, "downloader_config.json")

CRM_COLUMNS_FILE = os.path.join(DATA_DIR, "crm_columns.json")
CRM_SCHEMA_DEFAULT = {
    "columns": ["aprovado", "coletar_dados", "contato_encontrado", "sem_contato", "dm_enviada", "respondeu", "reuniao", "proposta", "fechado", "cliente_ativo", "perdido"],
    "colors": {
        "aprovado":"#35f28f",
        "coletar_dados":"#00c6ff",
        "contato_encontrado":"#2effa0",
        "sem_contato":"#ffd166",
        "dm_enviada":"#12a8ff",
        "respondeu":"#ffd166",
        "reuniao":"#a78bfa",
        "proposta":"#fb923c", "fechado":"#22c55e",
        "cliente_ativo":"#06d6a0",
        "perdido":"#ff5b5b"
    },
    "labels": {
        "aprovado":"Aprovado",
        "coletar_dados":"Coletar dados",
        "contato_encontrado":"Contato encontrado",
        "sem_contato":"Sem contato",
        "dm_enviada":"DM enviada",
        "respondeu":"Respondeu",
        "reuniao":"Reunião",
        "proposta":"Proposta",
        "fechado":"Fechado",
        "cliente_ativo":"Cliente ativo",
        "perdido":"Perdido"
    },
}
def carregar_crm_schema():
    try:
        if os.path.exists(CRM_COLUMNS_FILE):
            with open(CRM_COLUMNS_FILE, "r", encoding="utf-8") as f:
                schema = json.load(f)
            if isinstance(schema, dict) and isinstance(schema.get("columns"), list) and schema["columns"]:
                out = dict(CRM_SCHEMA_DEFAULT)
                cols = [str(c).strip().lower().replace(" ", "_") for c in schema["columns"] if str(c).strip()]
                cols = [c for c in cols if c]
                if not cols:
                    cols = list(CRM_SCHEMA_DEFAULT["columns"])
                out["columns"] = cols
                labels = schema.get("labels") or {}
                out["labels"] = {str(k).strip().lower(): str(v or k) for k, v in labels.items()}
                colors = schema.get("colors") or {}
                out["colors"] = {str(k).strip().lower(): str(v or "#2effa0") for k, v in colors.items()}
                return out
    except Exception as e:
        print(f"[CRM] erro carregando {CRM_COLUMNS_FILE}: {e}")
    return dict(CRM_SCHEMA_DEFAULT)

def salvar_crm_schema(schema):
    if not isinstance(schema, dict) or not isinstance(schema.get("columns"), list) or not schema["columns"]:
        return False
    try:
        with open(CRM_COLUMNS_FILE, "w", encoding="utf-8") as f:
            json.dump(schema, f, ensure_ascii=False, indent=2)
            f.write("\n")
        return True
    except Exception as e:
        print(f"[CRM] erro salvando {CRM_COLUMNS_FILE}: {e}")

def _read_json(path, default):
    try:
        if os.path.exists(path):
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception as e:
        print(f"Erro lendo {path}: {e}")
    return default

def _write_json(path, data):
    os.makedirs(os.path.dirname(path) or '.', exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def carregar_crm_raw():
    data = _read_json(CRM_FILE, {"cards": [], "ultima_atualizacao": str(datetime.now())})
    if not isinstance(data, dict):
        data = {"cards": []}
    data.setdefault("cards", [])
    return data

def salvar_crm(cards):
    _write_json(CRM_FILE, {"cards": cards or [], "ultima_atualizacao": str(datetime.now()), "total": len(cards or [])})

def carregar_crm_apagados():
    data = _read_json(CRM_DELETED_FILE, {"cards": [], "ids": [], "ultima_atualizacao": str(datetime.now())})
    if isinstance(data, list):
        data = {"cards": data, "ids": [_crm_lead_id(c) for c in data if isinstance(c, dict)]}
    if not isinstance(data, dict):
        data = {"cards": [], "ids": []}
    data.setdefault("cards", [])
    data.setdefault("ids", [])
    return data

def salvar_crm_apagados(cards, ids=None):
    ids_final = sorted({str(x) for x in (ids or []) if str(x).strip()} | {_crm_lead_id(c) for c in (cards or []) if isinstance(c, dict) and _crm_lead_id(c)})
    _write_json(CRM_DELETED_FILE, {"cards": cards or [], "ids": ids_final, "ultima_atualizacao": str(datetime.now()), "total": len(cards or [])})

def ids_crm_apagados():
    data = carregar_crm_apagados()
    return {str(x) for x in (data.get("ids") or []) if str(x).strip()} | {_crm_lead_id(c) for c in (data.get("cards") or []) if isinstance(c, dict) and _crm_lead_id(c)}

def _crm_lead_id(lead):
    return _lead_channel_id_seguro(lead)

def _crm_norm_status(status):
    schema = carregar_crm_schema()
    cols = {str(c).strip().lower().replace(" ", "_") for c in schema.get("columns", [])}
    status = str(status or "aprovado").strip().lower().replace(" ", "_")
    return status if status in cols else "aprovado"

def _crm_source(lead):
    src = str((lead or {}).get("source") or (lead or {}).get("origem") or "youtube").strip().lower()
    if not src:
        src = "youtube"
    if "insta" in src:
        return "instagram"
    if "yt" in src or "youtube" in src or "api" in src or "dlp" in src:
        return "youtube"
    return src

def _crm_card_from_lead(lead, existing=None):
    existing = existing or {}
    cid = _crm_lead_id(lead)
    nome = lead.get("nome") or lead.get("title") or lead.get("channel_title") or existing.get("nome") or "Lead"
    url = _lead_channel_url_segura(lead, cid) or existing.get("url") or (f"https://youtube.com/channel/{cid}" if cid else "")
    now = str(datetime.now())
    merged = {
        **existing,
        "id": cid,
        "channel_id": cid,
        "nome": nome,
        "url": url,
        "nicho": lead.get("nicho") or lead.get("source_query") or existing.get("nicho", ""),
        "subs_fmt": lead.get("subs_fmt") or lead.get("inscritos_fmt") or existing.get("subs_fmt", ""),
        "score": int(lead.get("score") or lead.get("api_score") or existing.get("score") or 0),
        "status": _crm_norm_status(existing.get("status") or lead.get("status") or "aprovado"),
        "source": existing.get("source") or _crm_source(lead),
        "valor_mensal": existing.get("valor_mensal", 0),
        "qtd_videos": existing.get("qtd_videos", 0),
        "preco_por_video": existing.get("preco_por_video", 0),
        "responsavel": existing.get("responsavel", ""),
        "observacoes": existing.get("observacoes", lead.get("observacoes", "")),
        "mensagem_dm": existing.get("mensagem_dm", lead.get("mensagem_dm", lead.get("msg", ""))),
        "demo_url": existing.get("demo_url", lead.get("demo_url", "")),
        "instagram_handle": existing.get("instagram_handle", lead.get("instagram_handle", lead.get("host_instagram", ""))),
        "avatar_url": existing.get("avatar_url", lead.get("avatar_url", lead.get("thumbnail", lead.get("thumb", "")))),
        "ultimo_episodio_titulo": existing.get("ultimo_episodio_titulo", lead.get("ultimo_episodio_titulo", lead.get("source_video_title", lead.get("video_title", "")))),
        "transcricao_status": existing.get("transcricao_status", lead.get("transcricao_status", lead.get("trans", ""))),
        "tem_transcricao_full": bool(existing.get("tem_transcricao_full") or lead.get("tem_transcricao_full") or lead.get("tem_transcricao")),
        "tem_shorts": bool(existing.get("tem_shorts") or lead.get("tem_shorts") or lead.get("shorts")),
        "archived": bool(existing.get("archived", False)),
        "cor": existing.get("cor") or gerar_cor_lead(cid),
        "created_at": existing.get("created_at") or now,
        "updated_at": now,
        "historico": existing.get("historico") or [{"data": now, "evento": "Importado dos aprovados do Prospector"}],
    }
    return merged

def sincronizar_crm_com_aprovados():
    """Sincroniza CRM a partir do disco + estado, para não depender só de estado['aprovados']."""
    data = carregar_crm_raw()
    cards = data.get("cards", []) or []
    apagados_ids = ids_crm_apagados()
    if apagados_ids:
        cards_filtrados = [c for c in cards if _crm_lead_id(c) not in apagados_ids]
        if len(cards_filtrados) != len(cards):
            cards = cards_filtrados
            data["cards"] = cards
            salvar_crm(cards)
    by_id = {_crm_lead_id(c): c for c in cards if _crm_lead_id(c)}

    aprovados_disco = carregar_aprovados()
    with lock:
        aprovados_estado = list(estado.get("aprovados", []) or [])
        if len(aprovados_disco) > len(aprovados_estado):
            estado["aprovados"] = aprovados_disco
        elif len(aprovados_estado) > len(aprovados_disco):
            # preserva memória atual se disco ficou atrasado
            aprovados_disco = aprovados_estado

    mudou = False
    for lead in aprovados_disco:
        if not isinstance(lead, dict):
            continue
        cid = _crm_lead_id(lead)
        if not cid:
            continue
        if cid in apagados_ids:
            continue
        existing = by_id.get(cid)
        if existing:
            # Só completa campos ausentes; não sobrescreve status/observações do CRM.
            card = _crm_card_from_lead(lead, existing)
            if card != existing:
                existing.update(card)
                mudou = True
        else:
            card = _crm_card_from_lead(lead)
            cards.append(card)
            by_id[cid] = card
            mudou = True
    if mudou:
        salvar_crm(cards)
    return cards

def sincronizar_crm_leads_incremental(leads):
    """Importa só os leads recém-aprovados no CRM, sem varrer todos os aprovados antigos."""
    leads = [l for l in (leads or []) if isinstance(l, dict)]
    if not leads:
        return len((carregar_crm_raw().get("cards", []) or []))
    data = carregar_crm_raw()
    cards = data.get("cards", []) or []
    apagados_ids = ids_crm_apagados()
    by_id = {_crm_lead_id(c): c for c in cards if _crm_lead_id(c)}
    mudou = False
    for lead in leads:
        cid = _crm_lead_id(lead)
        if not cid or cid in apagados_ids:
            continue
        existing = by_id.get(cid)
        if existing:
            card = _crm_card_from_lead(lead, existing)
            if card != existing:
                existing.update(card)
                mudou = True
        else:
            card = _crm_card_from_lead(lead)
            cards.append(card)
            by_id[cid] = card
            mudou = True
    if mudou:
        salvar_crm(cards)
    return len(cards)

def gerar_cor_lead(seed):
    pal = ["#35f28f", "#12a8ff", "#ffd166", "#fb923c", "#a78bfa", "#06d6a0", "#f472b6", "#38bdf8", "#84cc16", "#f97316"]
    return pal[sum(ord(c) for c in str(seed)) % len(pal)]

def _crm_has_transcript(card):
    return bool(card.get("tem_transcricao_full") or card.get("transcricao_full") or card.get("transcricao_preview") or str(card.get("transcricao_status") or "").lower() in {"ok", "cache", "full", "parcial"})

def _crm_export_val(card, keys, default="NAO_ENCONTRADO"):
    if isinstance(keys, str):
        keys = [keys]
    for k in keys:
        v = (card or {}).get(k)
        if v is None:
            continue
        if isinstance(v, (list, dict)):
            if v:
                return json.dumps(v, ensure_ascii=False)
            continue
        s = str(v).strip()
        if s:
            return s
    return default

def _crm_export_slug(text, max_len=70):
    text = str(text or "lead").strip().lower()
    text = re.sub(r"[áàãâä]", "a", text)
    text = re.sub(r"[éèêë]", "e", text)
    text = re.sub(r"[íìîï]", "i", text)
    text = re.sub(r"[óòõôö]", "o", text)
    text = re.sub(r"[úùûü]", "u", text)
    text = re.sub(r"ç", "c", text)
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return (text or "lead")[:max_len].strip("-") or "lead"

def _crm_export_duration_seconds(card):
    raw = card.get("ultimo_episodio_duration_seconds") or card.get("episodio_duration_seconds") or card.get("duration_seconds")
    try:
        sec = int(float(raw or 0))
    except Exception:
        sec = 0
    fmt = str(card.get("ultimo_episodio_duracao_fmt") or card.get("episodio_duracao") or card.get("duration") or "").strip()
    if fmt and ":" in fmt:
        try:
            parts = [int(x) for x in fmt.split(":")]
            parsed = 0
            for p in parts:
                parsed = parsed * 60 + p
            sec = max(sec, parsed)
        except Exception:
            pass
    return sec

def _crm_export_fmt_duration(seconds, fallback="NAO_ENCONTRADO"):
    try:
        seconds = int(seconds or 0)
    except Exception:
        seconds = 0
    if seconds <= 0:
        return fallback
    h, rem = divmod(seconds, 3600)
    m, s = divmod(rem, 60)
    return f"{h}:{m:02d}:{s:02d}" if h else f"{m}:{s:02d}"

def _crm_export_parse_video_seconds(v):
    raw = (v or {}).get("duration_seconds") or (v or {}).get("duracao_seconds") or 0
    try:
        sec = int(float(raw or 0))
    except Exception:
        sec = 0
    iso = str((v or {}).get("duration") or (v or {}).get("duracao") or "").strip()
    if iso.startswith("PT"):
        m = re.match(r"PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?", iso)
        if m:
            h = int(m.group(1) or 0); mi = int(m.group(2) or 0); s = int(m.group(3) or 0)
            sec = max(sec, h * 3600 + mi * 60 + s)
    elif ":" in iso:
        try:
            parsed = 0
            for p in [int(x) for x in iso.split(":")]:
                parsed = parsed * 60 + p
            sec = max(sec, parsed)
        except Exception:
            pass
    return sec

def _crm_export_pick_long_video(card):
    videos = card.get("ultimos_videos") or card.get("recent_videos") or []
    if not isinstance(videos, list):
        return None
    cortes = re.compile(r"\b(shorts?|corte|clip|highlights?|trecho|react|review|unboxing|gameplay)\b", re.I)
    podcast = re.compile(r"\b(podcast|videocast|entrevista|conversa|bate papo|talk show|mesa redonda|epis[oó]dio|host|convidad)\b", re.I)
    longos = []
    for v in videos:
        if not isinstance(v, dict):
            continue
        title = str(v.get("title") or v.get("titulo") or "")
        url = str(v.get("url") or "")
        sec = _crm_export_parse_video_seconds(v)
        if not url or sec < 30 * 60:
            continue
        if "shorts/" in url.lower():
            continue
        if cortes.search(title):
            continue
        longos.append((0 if podcast.search(title) else 1, v))
    if not longos:
        return None
    longos.sort(key=lambda x: x[0])
    return longos[0][1]

def _crm_export_episode_status(card):
    title = _crm_export_val(card, ["ultimo_episodio_titulo", "episodio_titulo", "ultimo_video_titulo", "video_title"], "")
    url = _crm_export_val(card, ["ultimo_episodio_url", "episodio_url", "ultimo_video_url", "video_url", "source_video_url"], "")
    seconds = _crm_export_duration_seconds(card)
    obs = []
    if not _crm_has_transcript(card):
        picked = _crm_export_pick_long_video(card)
        if picked:
            title = str(picked.get("title") or picked.get("titulo") or title or "").strip()
            url = str(picked.get("url") or url or "").strip()
            seconds = max(seconds, _crm_export_parse_video_seconds(picked))
            obs.append("episodio_longo_escolhido_de_ultimos_videos_salvos")
    if not title or title == "NAO_ENCONTRADO":
        obs.append("episodio_titulo_nao_encontrado")
    if not url or url == "NAO_ENCONTRADO":
        obs.append("episodio_url_nao_encontrado")
    if seconds and seconds < 30 * 60:
        obs.append("ATENCAO: episodio abaixo de 30 minutos; exportado por ser o melhor dado salvo disponivel")
    if re.search(r"\b(shorts?|corte|clip|highlights?|trecho)\b", title or "", re.I):
        obs.append("ATENCAO: titulo tem sinal de short/corte/clipe; revisar antes de usar")
    return title or "NAO_ENCONTRADO", url or "NAO_ENCONTRADO", seconds, obs

def _crm_export_transcript(card):
    tr = _crm_export_val(card, ["transcricao_full", "transcricao", "transcript_full", "transcricao_preview"], "")
    if not tr or tr == "NAO_ENCONTRADO":
        return "NAO_ENCONTRADO\n[transcricao nao encontrada para o episodio exportado]"
    return tr

def _crm_export_host(card):
    host = _crm_export_val(card, ["host_nome", "host_name"], "")
    if host and host != "NAO_ENCONTRADO":
        return host, _crm_export_val(card, ["host_confidence", "host_confidence_level"], "medio")
    nome = _crm_export_val(card, "nome")
    return nome, "baixo"

def _crm_export_observacoes(card, episode_notes):
    parts = []
    obs = str(card.get("observacoes") or "").strip()
    if obs:
        parts.append(obs)
    motivo = _crm_export_val(card, ["motivo_aprovado", "motivo", "aviso", "ultimo_episodio_motivo"], "")
    if motivo and motivo != "NAO_ENCONTRADO":
        parts.append("Motivo/sinal salvo: " + motivo)
    if episode_notes:
        parts.extend(episode_notes)
    if not str(card.get("host_nome") or "").strip():
        parts.append("Host nao encontrado com confianca; HOST_NOME usa nome do canal com baixa confianca.")
    if not _crm_has_transcript(card):
        parts.append("Transcricao completa nao encontrada; exportacao mantida para nao perder o lead.")
    return " | ".join(parts) if parts else "NAO_ENCONTRADO"

def _crm_lead_txt(card):
    title, url, seconds, episode_notes = _crm_export_episode_status(card)
    host, host_conf = _crm_export_host(card)
    transcript = _crm_export_transcript(card)
    motivo = _crm_export_val(card, ["motivo_aprovado", "motivo", "aviso", "ultimo_episodio_motivo"], "NAO_ENCONTRADO")
    lines = [
        f"NOME_CANAL: {_crm_export_val(card, ['nome','title','channel_title'])}",
        f"URL_CANAL: {_crm_export_val(card, ['url','channel_url'])}",
        f"INSCRITOS: {_crm_export_val(card, ['subs_fmt','subs','inscritos','subscriber_count'])}",
        f"NICHO: {_crm_export_val(card, ['nicho','source_query','query'])}",
        f"SCORE: {_crm_export_val(card, ['score','api_score'])}",
        f"MOTIVO_APROVADO: {motivo}",
        "",
        f"HOST_NOME: {host}",
        f"HOST_CONFIDENCE: {host_conf}",
        f"INSTAGRAM_CANAL: {_crm_export_val(card, ['instagram_handle','instagram_url','instagram_links_detectados'])}",
        f"INSTAGRAM_HOST: {_crm_export_val(card, ['host_instagram'])}",
        f"EMAIL: {_crm_export_val(card, ['email_comercial','email','emails_detectados'])}",
        "",
        f"EPISODIO_TITULO: {title}",
        f"EPISODIO_URL: {url}",
        f"EPISODIO_DURACAO: {_crm_export_fmt_duration(seconds, _crm_export_val(card, ['ultimo_episodio_duracao_fmt','episodio_duracao','duration'], 'NAO_ENCONTRADO'))}",
        f"DATA_EPISODIO: {_crm_export_val(card, ['ultimo_episodio_data','episodio_data','published_at','data_episodio'])}",
        "",
        f"OBSERVACOES_PROSPECTOR: {_crm_export_observacoes(card, episode_notes)}",
        f"LINK_DEMO: {_crm_export_val(card, ['demo_url','link_demo'], '')}",
        "MODO: DEMO_PRONTA",
        "",
        "TRANSCRICAO:",
        transcript,
        "",
    ]
    return "\n".join(lines)

def _crm_lead_txt_filename(card):
    score = re.sub(r"[^0-9]", "", str(card.get("score") or card.get("api_score") or 0)) or "0"
    nome = _crm_export_slug(card.get("nome") or card.get("title") or card.get("id") or "lead")
    return f"lead_{score}_{nome}.txt"

def _zip_unique_name(used, name):
    base, ext = os.path.splitext(name)
    candidate = name
    i = 2
    while candidate in used:
        candidate = f"{base}-{i}{ext}"
        i += 1
    used.add(candidate)
    return candidate

def crm_stats(cards):
    schema = carregar_crm_schema()
    cols = schema["columns"]
    by = {c: 0 for c in cols}
    mrr = 0
    com_transcricao = sem_instagram = prontos_editar = 0
    for card in cards or []:
        if card.get("archived"):
            continue
        st = _crm_norm_status(card.get("status"))
        by[st] = by.get(st, 0) + 1
        if _crm_has_transcript(card):
            com_transcricao += 1
        if not (card.get("instagram_handle") or card.get("host_instagram")):
            sem_instagram += 1
        if _crm_has_transcript(card) and (card.get("instagram_handle") or card.get("host_instagram") or card.get("url")):
            prontos_editar += 1
        if st in ["fechado", "cliente_ativo"]:
            try: mrr += float(card.get("valor_mensal") or 0)
            except Exception: pass
    dms = by.get("dm_enviada",0)+by.get("respondeu",0)+by.get("reuniao",0)+by.get("proposta",0)+by.get("fechado",0)+by.get("cliente_ativo",0)+by.get("perdido",0)
    respostas = by.get("respondeu",0)+by.get("reuniao",0)+by.get("proposta",0)+by.get("fechado",0)+by.get("cliente_ativo",0)
    fechados = by.get("fechado",0)+by.get("cliente_ativo",0)
    taxa_resposta = round((respostas/dms)*100,1) if dms else 0
    taxa_fechamento = round((fechados/max(dms,1))*100,1) if dms else 0
    return {"por_status": by, "mrr": mrr, "dms": dms, "respostas": respostas, "fechados": fechados, "taxa_resposta": taxa_resposta, "taxa_fechamento": taxa_fechamento, "com_transcricao": com_transcricao, "sem_instagram": sem_instagram, "prontos_editar": prontos_editar, "total": len([c for c in cards or [] if not c.get("archived")])}

def _crm_filtered_cards(cards):
    src = str(request.args.get("source", "all") or "all").lower()
    archived = str(request.args.get("archived", "0")).lower() in {"1", "true", "yes", "sim"}
    filt = str(request.args.get("filter", "all") or "all")
    q = str(request.args.get("q", "") or "").strip().lower()
    out = []
    for c in cards or []:
        if bool(c.get("archived")) != archived:
            continue
        if src not in {"all", "archived"} and _crm_source(c) != src:
            continue
        if q:
            hay = " ".join(str(c.get(k, "")) for k in ["nome", "nicho", "url", "instagram_handle", "observacoes", "ultimo_episodio_titulo"]).lower()
            if q not in hay:
                continue
        if filt == "com_transcricao" and not _crm_has_transcript(c):
            continue
        if filt == "sem_transcricao" and _crm_has_transcript(c):
            continue
        if filt == "sem_instagram" and (c.get("instagram_handle") or c.get("host_instagram")):
            continue
        if filt == "score80" and int(c.get("score") or 0) < 80:
            continue
        if filt == "prontos_editar" and not (_crm_has_transcript(c) and (c.get("instagram_handle") or c.get("host_instagram") or c.get("url"))):
            continue
        out.append(c)
    return out

@app.route("/api/crm")
def api_crm():
    with lock:
        cards_all = sincronizar_crm_com_aprovados()
    cards = _crm_filtered_cards(cards_all)
    try: page = max(1, int(request.args.get("page", 1) or 1))
    except Exception: page = 1
    try: per_page = max(1, min(500, int(request.args.get("per_page", request.args.get("perPage", 200)) or 200)))
    except Exception: per_page = 200
    total = len(cards)
    start = (page - 1) * per_page
    end = start + per_page
    page_cards = cards[start:end]
    schema = carregar_crm_schema()
    return jsonify({"ok": True, "columns": schema["columns"], "labels": schema.get("labels", {}), "colors": schema.get("colors", {}), "cards": page_cards, "stats": crm_stats(cards_all), "total": total, "page": page, "per_page": per_page, "total_pages": max(1, (total + per_page - 1)//per_page), "all_count": len(cards_all)})


@app.route("/api/crm/columns", methods=["GET"])
def api_crm_columns_get():
    schema = carregar_crm_schema()
    return jsonify({"columns": schema.get("columns", []), "labels": schema.get("labels", {}), "colors": schema.get("colors", {})})

@app.route("/api/crm/columns", methods=["POST"])
def api_crm_columns_save():
    data = request.json or {}
    novas = data.get("columns") or []
    if not isinstance(novas, list) or not novas:
        return jsonify({"erro": "formato inválido"}), 400
    salvar_crm_schema({"columns": novas})
    schema = carregar_crm_schema()
    return jsonify({"ok": True, "columns": schema.get("columns", []), "labels": schema.get("labels", {}), "colors": schema.get("colors", {})})

@app.route("/api/crm/sync", methods=["POST", "GET"])
def api_crm_sync():
    with lock:
        estado["aprovados"] = carregar_aprovados()
        cards = sincronizar_crm_com_aprovados()
    return jsonify({"ok": True, "total": len(cards), "aprovados": len(estado.get("aprovados", []) or []), "file": CRM_FILE})

@app.route("/api/debug/crm_sync")
def debug_crm_sync():
    data = carregar_crm_raw()
    aprovados = carregar_aprovados()
    with lock:
        estado_count = len(estado.get("aprovados", []) or [])
    return jsonify({
        "ok": True,
        "aprovados_count_file": len(aprovados),
        "aprovados_count_estado": estado_count,
        "aprovados_amostra": aprovados[:5],
        "crm_cards_count": len(data.get("cards", []) or []),
        "winchester_exists": os.path.exists(APROVADOS_FILE),
        "winchester_size": os.path.getsize(APROVADOS_FILE) if os.path.exists(APROVADOS_FILE) else 0,
        "crm_file": CRM_FILE,
    })

@app.route("/api/crm/update", methods=["POST"])
def api_crm_update():
    data = request.json or {}
    cid = str(data.get("id") or "").strip()
    if not cid: return jsonify({"erro":"id ausente"}), 400
    cards = sincronizar_crm_com_aprovados()
    found = None
    src = str(data.get("source") or "").lower().strip()
    for c in cards:
        if str(c.get("id")) == cid and (not src or _crm_source(c) == src or str(c.get("source", "")).lower() == src):
            found = c; break
    if not found: return jsonify({"erro":"card nao encontrado"}), 404
    old_status = found.get("status")
    allowed = {"status","valor_mensal","qtd_videos","preco_por_video","responsavel","observacoes","nome","nicho","url","cor","source","archived","mensagem_dm","demo_url","instagram_handle","avatar_url","proxima_acao","proxima_acao_data","notas_editor"}
    if "whatsapp_handle" in data:
        found["whatsapp_handle"] = data["whatsapp_handle"]
        allowed.add("whatsapp_handle")
    if "email" in data:
        found["email"] = data["email"]
        allowed.add("email")
    for k,v in data.items():
        if k in allowed:
            found[k] = v
    found["status"] = _crm_norm_status(found.get("status"))
    found["updated_at"] = str(datetime.now())
    if old_status != found.get("status"):
        schema = carregar_crm_schema()
        labels = schema.get("labels") or {}
        labels_map = {str(k).strip().lower(): str(v or k) for k, v in labels.items()}
        nome_antigo = labels_map.get(str(old_status or "aprovado").strip().lower().replace(" ", "_"), old_status or "aprovado")
        nome_novo = labels_map.get(found.get("status"), found.get("status"))
        found.setdefault("historico", []).append({"data": str(datetime.now()), "evento": f"Status: {nome_antigo} → {nome_novo}"})
    salvar_crm(cards)
    return jsonify({"ok": True, "card": found, "stats": crm_stats(cards)})

def _crm_br_score(card):
    text = " ".join(str(card.get(k, "")) for k in [
        "nome", "url", "nicho", "descricao_canal", "description", "observacoes",
        "ultimo_episodio_titulo", "custom_url", "country", "query", "source_query"
    ]).lower()
    score = 0
    motivos = []
    country = str(card.get("country") or "").upper()
    if country == "BR":
        score += 4; motivos.append("country_BR")
    if re.search(r"[áàâãéêíóôõúç]", text):
        score += 2; motivos.append("acentos_pt")
    termos = ["brasil", "brasileiro", "brasileira", "português", "portugues", "são paulo", "rio de janeiro", "brasília", "brasilia", "goiânia", "goiania", "curitiba", "pernambuco", "bahia", "minas", "ceará", "ceara", "empreendedorismo", "negócios", "negocios", "imobiliário", "imobiliario", "corretor", "advogado", "contador"]
    hits = [t for t in termos if t in text]
    if hits:
        score += min(4, len(hits)); motivos.extend(hits[:6])
    label = "BR provável" if score >= 3 else ("BR incerto" if score >= 1 else "NAO_ENCONTRADO")
    return score, label, motivos

@app.route("/api/crm/verificar_br", methods=["POST"])
@app.route("/api/verificar_br", methods=["POST"])
def api_crm_verificar_br():
    """V58.32: verifica BR no LOTE (nao so CRM). Se ids vazios, verifica lote atual."""
    data = request.json or {}
    ids = {str(x) for x in (data.get("ids") or []) if str(x).strip()}
    atualizados = 0
    br = incerto = fora = 0
    verificados_info = []

    # V58.32: se ids vazios OU se vem do lote, verifica lote atual
    lote_atual = carregar_lote_ativo()
    if not ids:
        # sem ids = verifica lote atual inteiro
        alvos = lote_atual
        print(f"[VERIFICAR BR V58.32] verificando lote atual: {len(alvos)} canais")
    else:
        # com ids = pega do lote + CRM
        alvos = [c for c in lote_atual if str(c.get("id") or c.get("channel_id") or "") in ids]
        cards_crm = sincronizar_crm_com_aprovados()
        alvos_crm = [c for c in cards_crm if _crm_lead_id(c) in ids]
        alvos.extend(alvos_crm)
        print(f"[VERIFICAR BR V58.32] verificando {len(alvos)} canais (ids especificos)")

    # V58.32: enrich de country antes de pontuar
    if alvos and _API_KEYS:
        try:
            _enrich_country_canais(alvos, max_per_call=30)
        except Exception as e:
            print(f"[VERIFICAR BR] aviso enrich: {e}")

    for c in alvos:
        cid = str(c.get("id") or c.get("channel_id") or "")
        score, label, motivos = _crm_br_score(c)
        # V58.32: bonus se country == BR (da API)
        country = str(c.get("country") or "").upper()
        if country == "BR":
            score += 4
            motivos.append("country_BR_api")
            label = "BR confirmado (API)"
        elif country:
            score = max(0, score - 2)
            motivos.append(f"country_{country}_api")
            if score < 3:
                label = f"NAO_BR (country={country})"
        c["pais_score"] = score
        c["pais_label"] = label
        c["pais_status"] = "br" if score >= 3 else ("incerto" if score >= 1 else "nao_encontrado")
        c["pais_motivos"] = motivos
        c["pais_verificado_em"] = str(datetime.now())
        atualizados += 1
        if score >= 3: br += 1
        elif score >= 1: incerto += 1
        else: fora += 1
        verificados_info.append({
            "id": cid,
            "nome": c.get("nome") or c.get("title") or "",
            "country": country,
            "score": score,
            "label": label,
            "status": c["pais_status"],
        })

    # Salva lote atualizado (com country)
    if alvos and not ids:
        salvar_lote_ativo(lote_atual)
    # Se veio do CRM, salva CRM
    if ids:
        cards_crm = sincronizar_crm_com_aprovados()
        salvar_crm(cards_crm)

    return jsonify({
        "ok": True,
        "atualizados": atualizados,
        "br": br,
        "incerto": incerto,
        "nao_encontrado": fora,
        "verificados": verificados_info[:50],  # limita pra nao estourar response
        "stats": crm_stats(sincronizar_crm_com_aprovados()),
    })

@app.route("/api/crm/delete", methods=["POST"])
def api_crm_delete():
    data = request.json or {}
    ids = {str(x) for x in (data.get("ids") or []) if str(x).strip()}
    if not ids:
        return jsonify({"erro": "ids ausentes"}), 400
    cards = sincronizar_crm_com_aprovados()
    apagados_data = carregar_crm_apagados()
    apagados_cards = apagados_data.get("cards", []) or []
    apagados_ids = set(apagados_data.get("ids") or [])
    removidos = []
    restantes = []
    for c in cards:
        cid = _crm_lead_id(c)
        if cid in ids:
            c = dict(c)
            c["deleted_at"] = str(datetime.now())
            c["deleted_reason"] = data.get("motivo") or "apagado_no_crm"
            c.setdefault("historico", []).append({"data": str(datetime.now()), "evento": "Apagado do CRM"})
            removidos.append(c)
            apagados_ids.add(cid)
        else:
            restantes.append(c)
    if removidos:
        by_id = {_crm_lead_id(c): c for c in apagados_cards if isinstance(c, dict) and _crm_lead_id(c)}
        for c in removidos:
            by_id[_crm_lead_id(c)] = c
        apagados_cards = list(by_id.values())
        salvar_crm(restantes)
        salvar_crm_apagados(apagados_cards, apagados_ids)
    return jsonify({"ok": True, "removidos": len(removidos), "total": len(restantes), "apagados_file": CRM_DELETED_FILE})

@app.route("/api/crm/create", methods=["POST"])
def api_crm_create():
    data = request.json or {}
    cards = sincronizar_crm_com_aprovados()
    cid = data.get("id") or ("manual_" + str(int(time.time()*1000)))
    card = _crm_card_from_lead({**data, "id": cid, "source": data.get("source") or "youtube"})
    card["historico"] = [{"data":str(datetime.now()),"evento":"Criado manualmente"}]
    cards.append(card); salvar_crm(cards)
    return jsonify({"ok": True, "card": card})

@app.route("/api/crm/export", methods=["POST"])
def api_crm_export():
    import io, csv, zipfile
    data = request.json or {}
    ids = {str(x) for x in (data.get("ids") or []) if str(x).strip()}
    cards = sincronizar_crm_com_aprovados()
    if ids:
        cards = [c for c in cards if str(c.get("id")) in ids]
    mem = io.BytesIO()
    with zipfile.ZipFile(mem, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("crm_leads.json", json.dumps(cards, ensure_ascii=False, indent=2))
        s = io.StringIO()
        w = csv.writer(s)
        w.writerow(["nome","url","nicho","subs","score","status","instagram","demo","observacoes"])
        used_txt_names = set()
        for c in cards:
            w.writerow([c.get("nome",""), c.get("url",""), c.get("nicho",""), c.get("subs_fmt",""), c.get("score",0), c.get("status",""), c.get("instagram_handle",""), c.get("demo_url",""), c.get("observacoes","")])
            txt_name = _zip_unique_name(used_txt_names, _crm_lead_txt_filename(c))
            z.writestr("leads_txt/" + txt_name, _crm_lead_txt(c))
        z.writestr("crm_leads.csv", "\ufeff" + s.getvalue())
    mem.seek(0)
    from flask import send_file
    return send_file(mem, mimetype="application/zip", as_attachment=True, download_name=f"crm_leads_{datetime.now().strftime('%Y%m%d_%H%M')}.zip")

@app.route("/exportar_leads_completos_txt", methods=["GET", "POST"])
def exportar_leads_completos_txt():
    import io, zipfile
    data = request.json or {} if request.method == "POST" else {}
    ids = {str(x) for x in (data.get("ids") or request.args.get("ids", "").split(",")) if str(x).strip()}
    cards = sincronizar_crm_com_aprovados()
    if ids:
        cards = [c for c in cards if str(c.get("id")) in ids]
    mem = io.BytesIO()
    with zipfile.ZipFile(mem, "w", zipfile.ZIP_DEFLATED) as z:
        used_txt_names = set()
        for c in cards:
            txt_name = _zip_unique_name(used_txt_names, _crm_lead_txt_filename(c))
            z.writestr("leads_txt/" + txt_name, _crm_lead_txt(c))
    mem.seek(0)
    from flask import send_file
    return send_file(mem, mimetype="application/zip", as_attachment=True, download_name=f"leads_txt_{datetime.now().strftime('%Y%m%d_%H%M')}.zip")



# ============================================================
# THON CRM API COLLECTOR V2 - Coleta no CRM com console real
# ============================================================
CRM_ENRICH_CONSOLE = []
CRM_ENRICH_LAST = {"running": False, "started_at": None, "finished_at": None, "processed": 0, "ok": 0, "fail": 0}

def _crmv2_now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def _crmv2_log(msg, level="info", extra=None):
    row = {"ts": _crmv2_now(), "level": level, "msg": str(msg)}
    if extra is not None:
        row["extra"] = extra
    CRM_ENRICH_CONSOLE.append(row)
    del CRM_ENRICH_CONSOLE[:-300]
    try:
        print(f"[CRM API V2] {level.upper()} {msg}")
    except Exception:
        pass
    return row

def _crmv2_keys():
    keys = []
    try:
        with open(API_KEYS_FILE, "r", encoding="utf-8", errors="replace") as f:
            for line in f:
                k = line.strip()
                if k and not k.startswith("#"):
                    keys.append(k)
    except Exception as e:
        _crmv2_log(f"Falha lendo {API_KEYS_FILE}: {e}", "error")
    return keys

def _crmv2_mask_key(k):
    k = str(k or "")
    return k[:6] + "..." + k[-4:] if len(k) > 12 else k

def _crmv2_curl_json(endpoint, params, key=None):
    import urllib.parse as _up
    keys = [key] if key else _crmv2_keys()
    if not keys:
        raise RuntimeError("api_keys.txt sem chave válida")
    last = None
    for idx, k in enumerate(keys, 1):
        q = dict(params or {})
        q["key"] = k
        url = "https://www.googleapis.com/youtube/v3/" + endpoint + "?" + _up.urlencode(q)
        try:
            r = subprocess.run(["curl", "-sS", "-L", url], capture_output=True, text=True, timeout=45)
            if r.returncode != 0:
                last = r.stderr.strip() or f"curl saiu {r.returncode}"
                _crmv2_log(f"curl falhou em {endpoint} usando chave #{idx}: {last[:180]}", "warn")
                continue
            data = json.loads(r.stdout or "{}")
            if data.get("error"):
                err = data.get("error") or {}
                msg = err.get("message") or str(err)
                reason = ""
                try:
                    reason = err.get("errors", [{}])[0].get("reason", "")
                except Exception:
                    pass
                last = f"{err.get('code')} {reason} {msg}"
                _crmv2_log(f"API erro em {endpoint} chave #{idx}: {last[:220]}", "warn")
                continue
            return data, idx
        except Exception as e:
            last = str(e)
            _crmv2_log(f"Exceção API em {endpoint}: {last[:220]}", "warn")
            continue
    raise RuntimeError("Todas as chaves falharam. Último erro: " + str(last))

def _crmv2_contacts(text):
    text = text or ""
    emails = sorted(set(re.findall(r"[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+", text)))
    ig_links = sorted(set(re.findall(r"(?:https?://)?(?:www\.)?instagram\.com/[A-Za-z0-9_.]+/?", text, flags=re.I)))
    handles = sorted(set(re.findall(r"(?<!\w)@([A-Za-z0-9_.]{3,30})", text)))
    links = sorted(set(re.findall(r"https?://[^\s\)\]\}<>\"']+", text)))
    return {"emails": emails[:30], "instagram_links": ig_links[:30], "instagram_handles_possiveis": ["@" + h for h in handles[:40]], "links": links[:100]}

def _crmv2_fmt_num(n):
    try:
        n = int(n or 0)
    except Exception:
        return ""
    if n >= 1000000:
        return f"{n/1000000:.1f}M"
    if n >= 1000:
        return f"{n/1000:.0f}K" if n >= 10000 else f"{n/1000:.1f}K"
    return str(n)

def _crmv2_resolve_channel_id(card, key=None):
    cid = str((card or {}).get("channel_id") or (card or {}).get("id") or "").strip()
    if cid.startswith("UC") and len(cid) > 20:
        return cid
    url = str((card or {}).get("url") or "").strip()
    m = re.search(r"/channel/(UC[a-zA-Z0-9_-]+)", url)
    if m:
        return m.group(1)
    # fallback via yt-dlp só para resolver ID. Coleta continua sendo API.
    if url:
        try:
            r = subprocess.run([sys.executable, "-m", "yt_dlp", "--no-warnings", "--quiet", "--print", "%(channel_id)s", "--playlist-end", "1", url], capture_output=True, text=True, timeout=45)
            for line in (r.stdout or "").splitlines():
                line = line.strip()
                if line.startswith("UC"):
                    return line
        except Exception as e:
            _crmv2_log(f"yt-dlp não resolveu channel_id: {e}", "warn")
    name = str((card or {}).get("nome") or (card or {}).get("custom_url") or url or "").strip()
    if name:
        data, _ = _crmv2_curl_json("search", {"part": "snippet", "type": "channel", "maxResults": 1, "q": name}, key=key)
        items = data.get("items") or []
        if items:
            found = items[0].get("snippet", {}).get("channelId")
            if found:
                return found
    return ""

def _crmv2_pick_instagram(contacts):
    contacts = contacts or {}
    links = contacts.get("instagram_links") or []
    handles = contacts.get("instagram_handles_possiveis") or []
    if links:
        return links[0]
    if handles:
        return handles[0]
    return ""

def _crmv2_xml_captions_to_text(xml_str):
    """Faz o parsing do XML de legenda do pytubefix (caption.xml_captions) e devolve a
    lista de trechos [{start, text}], na ordem do vídeo. Diferente do .vtt do yt-dlp,
    o XML do InnerTube já vem em trechos únicos (sem duplicação de "rolagem"), então
    não precisa de nenhuma lógica de dedup."""
    if not xml_str or _ET is None:
        return []
    try:
        root = _ET.fromstring(xml_str)
    except Exception:
        return []
    trechos = []
    for elem in root.findall(".//text"):
        try:
            start = float(elem.get("start", 0) or 0)
        except Exception:
            start = 0.0
        texto = (elem.text or "").strip()
        if not texto:
            continue
        texto = html.unescape(texto).replace("\n", " ").strip()
        if texto:
            trechos.append({"start": start, "text": texto})
    return trechos

_TRANSCRIPT_QUEUE_LOCK = threading.Lock()
_TRANSCRIPT_QUEUE_STATE = {"em_fila": 0, "processando_video_id": None}

def _crmv2_melhor_faixa_legenda(yt, idiomas_preferidos):
    """Escolhe, entre TODAS as faixas de legenda disponíveis (não só a primeira que bater
    o idioma preferido), a que cobre o vídeo mais completamente. Um vídeo pode ter várias
    faixas (manual pt, automática pt, automática en...) e nem sempre a preferida por
    idioma é a mais completa — então mede cobertura real (timestamp da última legenda /
    duração do vídeo) e só para de testar quando acha uma que cobre quase o vídeo todo."""
    captions = list(yt.captions)
    try:
        duracao = int(yt.length or 0)
    except Exception:
        duracao = 0
    ordenados = sorted(captions, key=lambda c: (idiomas_preferidos.index(c.code) if c.code in idiomas_preferidos else len(idiomas_preferidos)))
    melhor = None
    for c in ordenados:
        try:
            xml_str = c.xml_captions
        except Exception:
            continue
        trechos = _crmv2_xml_captions_to_text(xml_str)
        if not trechos:
            continue
        ultima_ts = trechos[-1]["start"]
        cobertura = (ultima_ts / duracao) if duracao > 0 else 1.0
        cobertura = min(cobertura, 1.0)
        if melhor is None or cobertura > melhor["cobertura"]:
            melhor = {"caption": c, "trechos": trechos, "cobertura": cobertura, "ultima_ts": ultima_ts, "idioma": c.code}
        if cobertura >= 0.92:
            break  # já cobre quase o vídeo inteiro, não precisa testar as outras faixas
    return melhor, duracao

def _crmv2_fetch_transcript_pytube(video_id, idiomas_preferidos=None):
    """Busca a transcrição REAL e COMPLETA (100% do vídeo) via pytubefix (InnerTube).
    Substitui o yt-dlp como fonte de transcrição porque o yt-dlp estava tomando 429
    (rate limit) com frequência e, mesmo funcionando, tinha bug de dedup que cortava
    parte do conteúdo em vídeo longo.

    Duas garantias novas:
      1) COBERTURA: em vez de aceitar a primeira faixa de legenda que bater o idioma
         preferido, testa a cobertura real dela (último timestamp / duração do vídeo)
         e, se não cobrir quase o vídeo inteiro, tenta as outras faixas disponíveis até
         achar a mais completa (ou fica com a melhor encontrada, marcando status parcial
         se nenhuma cobrir tudo).
      2) FILA: todo pedido de transcrição passa por um lock global, então mesmo que
         várias requisições cheguem ao mesmo tempo (lote + card avulso, por exemplo),
         elas são processadas uma vídeo por vez, em fila, nunca em paralelo — o que
         evita corte por rate limit e deixa o processamento previsível.
    """
    if not video_id:
        return {"status": "sem_video_id", "preview": "Nenhum ultimo_video_id disponível para buscar transcrição.", "full": "", "video_id": ""}
    if _PytubeYouTube is None:
        return {"status": "erro_pytube", "preview": "pytubefix não está instalado neste ambiente. Rode: pip install pytubefix --break-system-packages", "full": "", "video_id": video_id}
    idiomas_preferidos = idiomas_preferidos or ["pt-BR", "pt", "a.pt-BR", "a.pt"]

    _TRANSCRIPT_QUEUE_STATE["em_fila"] += 1
    posicao = _TRANSCRIPT_QUEUE_STATE["em_fila"]
    if not _TRANSCRIPT_QUEUE_LOCK.acquire(blocking=False):
        _crmv2_log(f"Transcrição de {video_id} entrou na fila (posição {posicao}) - aguardando o vídeo anterior terminar...", "info")
        _TRANSCRIPT_QUEUE_LOCK.acquire(blocking=True)
    _TRANSCRIPT_QUEUE_STATE["em_fila"] -= 1
    _TRANSCRIPT_QUEUE_STATE["processando_video_id"] = video_id
    try:
        url = f"https://www.youtube.com/watch?v={video_id}"
        yt = _PytubeYouTube(url)
        titulo = ""
        try:
            titulo = yt.title or ""
        except Exception:
            titulo = ""
        if not yt.captions:
            return {"status": "nao_encontrada_pytube", "preview": "Nenhuma legenda disponível para este vídeo via pytubefix.", "full": "", "video_id": video_id, "titulo": titulo}
        melhor, duracao = _crmv2_melhor_faixa_legenda(yt, idiomas_preferidos)
        if not melhor:
            return {"status": "legenda_vazia_pytube", "preview": "Legenda encontrada mas não foi possível extrair texto dela.", "full": "", "video_id": video_id, "titulo": titulo}
        trechos = melhor["trechos"]
        cobertura = melhor["cobertura"]
        texto_full = " ".join(t["text"] for t in trechos).strip()
        status = "ok_pytube" if cobertura >= 0.92 or duracao == 0 else "parcial_pytube"
        preview = texto_full[:400]
        if status == "parcial_pytube":
            preview = f"[cobertura ~{int(cobertura*100)}% do vídeo, melhor faixa disponível] " + preview
            _crmv2_log(f"Transcrição de {video_id} ficou parcial (~{int(cobertura*100)}% do vídeo, faixa {melhor['idioma']}) - nenhuma faixa disponível cobria o vídeo inteiro", "warn")
        return {
            "status": status,
            "preview": preview,
            "full": texto_full,
            "video_id": video_id,
            "titulo": titulo,
            "idioma": melhor["idioma"],
            "total_trechos": len(trechos),
            "cobertura_pct": round(cobertura * 100, 1),
            "duracao_segundos": duracao,
            "fonte": "pytubefix",
        }
    except Exception as e:
        return {"status": "erro_pytube", "preview": f"Erro ao buscar transcrição via pytubefix: {e}", "full": "", "video_id": video_id}
    finally:
        _TRANSCRIPT_QUEUE_STATE["processando_video_id"] = None
        _TRANSCRIPT_QUEUE_LOCK.release()
        time.sleep(random.uniform(1.0, 2.0))  # respiro antes do próximo da fila

def _crmv2_transcript_stub(card, video_id, engine):
    """A YouTube Data API não entrega transcrição (só metadados de vídeo/canal). A coleta
    de transcrição usa pytubefix (InnerTube) em vez de yt-dlp — independente do engine de
    DESCOBERTA escolhido (api ou dlp), a transcrição em si sempre vem do pytubefix, sempre
    em fila (um vídeo por vez) e sempre tentando cobrir 100% do vídeo antes de aceitar."""
    return _crmv2_fetch_transcript_pytube(video_id)

def _crmv2_video_id_para_transcricao(updates, card):
    """O id do último vídeo pode não estar em 'ultimo_video_id' se a busca fresca de
    vídeos falhou/veio vazia nessa rodada (quota, uploads playlist vazia, etc). Mas o
    card pode já ter um vídeo salvo de uma coleta anterior — só que em outro campo
    (ex.: 'ultimo_episodio_url' com a URL completa, de coletas mais antigas). Tenta
    todas as fontes possíveis antes de desistir, pra não perder a chance de transcrição
    só porque o campo novo não foi preenchido nessa passada."""
    direto = (updates.get("ultimo_video_id") or card.get("ultimo_video_id") or card.get("video_id") or "").strip()
    if direto:
        return direto
    candidatos = [
        updates.get("ultimo_episodio_url"), card.get("ultimo_episodio_url"),
        card.get("episodio_url"), card.get("ultimo_video_url"), card.get("video_url"),
        card.get("source_video_url"),
    ]
    for c in candidatos:
        if not c:
            continue
        m = re.search(r'(?:v=|youtu\.be/|shorts/|embed/)([A-Za-z0-9_-]{11})', str(c))
        if m:
            return m.group(1)
    return ""

def _crmv2_enrich_card(card, options=None, engine="api", max_videos=10):
    options = options or {}
    engine = str(engine or "api").lower()
    max_videos = max(1, min(25, int(max_videos or 10)))
    cid = _crmv2_resolve_channel_id(card)
    if not cid:
        raise RuntimeError("channel_id não encontrado no card")
    _crmv2_log(f"Coletando {card.get('nome') or cid} via API", "info", {"id": cid})

    ch, key_idx = _crmv2_curl_json("channels", {"part": "snippet,statistics,contentDetails", "id": cid, "maxResults": 1})
    items = ch.get("items") or []
    if not items:
        raise RuntimeError("API não retornou canal")
    item = items[0]
    sn = item.get("snippet") or {}
    st = item.get("statistics") or {}
    uploads = ((item.get("contentDetails") or {}).get("relatedPlaylists") or {}).get("uploads") or ""
    thumbs = sn.get("thumbnails") or {}
    avatar = (thumbs.get("high") or thumbs.get("medium") or thumbs.get("default") or {}).get("url") or ""
    desc = sn.get("description") or ""
    all_text = desc
    updates = {
        "channel_id": cid,
        "id": card.get("id") or cid,
        "crm_enriched": True,
        "crm_enrich_source": "api_curl",
        "crm_enriched_at": str(datetime.now()),
        "crm_enrich_key_index": key_idx,
    }

    if options.get("channel", True):
        updates.update({
            "nome": sn.get("title") or card.get("nome") or cid,
            "descricao_canal": desc,
            "country": sn.get("country") or "",
            "custom_url": sn.get("customUrl") or "",
            "published_at": sn.get("publishedAt") or "",
            "url": card.get("url") or (f"https://www.youtube.com/{sn.get('customUrl')}" if sn.get("customUrl") else f"https://www.youtube.com/channel/{cid}"),
            "uploads_playlist": uploads,
            "subs": st.get("subscriberCount") or "",
            "subs_fmt": _crmv2_fmt_num(st.get("subscriberCount")),
            "views_total": st.get("viewCount") or "",
            "video_count": st.get("videoCount") or "",
        })
    if options.get("avatar", True) and avatar:
        updates["avatar_url"] = avatar

    videos = []
    if options.get("videos", True) and uploads:
        pl, _ = _crmv2_curl_json("playlistItems", {"part": "snippet,contentDetails", "playlistId": uploads, "maxResults": max_videos})
        ids = []
        for it in pl.get("items") or []:
            vid = ((it.get("contentDetails") or {}).get("videoId"))
            sni = it.get("snippet") or {}
            if vid:
                ids.append(vid)
            all_text += "\n" + (sni.get("title") or "")
            all_text += "\n" + (sni.get("description") or "")
        if not ids:
            _crmv2_log(f"playlistItems de {card.get('nome') or cid} veio sem nenhum videoId (uploads={uploads!r}, itens brutos={len(pl.get('items') or [])})", "warn")
        if ids:
            vd, _ = _crmv2_curl_json("videos", {"part": "snippet,statistics,contentDetails", "id": ",".join(ids)})
            by_vid = {}
            for v in vd.get("items") or []:
                vs = v.get("snippet") or {}
                stats = v.get("statistics") or {}
                all_text += "\n" + (vs.get("description") or "")
                by_vid[v.get("id")] = {
                    "id": v.get("id"),
                    "title": vs.get("title"),
                    "publishedAt": vs.get("publishedAt"),
                    "views": stats.get("viewCount"),
                    "likes": stats.get("likeCount"),
                    "duration": (v.get("contentDetails") or {}).get("duration"),
                }
            # IMPORTANTE: o endpoint videos.list NÃO garante devolver os itens na mesma
            # ordem dos ids passados em "id=a,b,c". Reconstrói na ordem original da
            # playlist de uploads (que já vem mais recente -> mais antigo) e ainda
            # reordena por publishedAt como segurança extra (ex.: premieres agendadas
            # que a playlist às vezes lista fora de ordem cronológica real).
            videos = [by_vid[i] for i in ids if i in by_vid]
            def _crmv2_parse_published(v):
                try:
                    return datetime.strptime(v.get("publishedAt") or "", "%Y-%m-%dT%H:%M:%SZ")
                except Exception:
                    return datetime.min
            videos.sort(key=_crmv2_parse_published, reverse=True)
            if not videos:
                _crmv2_log(f"videos.list de {card.get('nome') or cid} não devolveu nenhum item pros ids buscados ({len(ids)} ids enviados)", "warn")
        updates["ultimos_videos"] = videos
        if videos:
            updates["ultimo_episodio_titulo"] = videos[0].get("title") or ""
            updates["ultimo_video_id"] = videos[0].get("id") or ""
            nums = []
            for v in videos:
                try:
                    nums.append(int(v.get("views") or 0))
                except Exception:
                    pass
            updates["media_views_recentes"] = int(sum(nums)/len(nums)) if nums else 0
    elif options.get("videos", True) and not uploads:
        _crmv2_log(f"{card.get('nome') or cid} sem uploads playlist (contentDetails.relatedPlaylists.uploads vazio na resposta do channels.list)", "warn")

    # ===== TRANSCRIÇÃO AUTOMÁTICA (sempre que houver vídeos) =====
    # Agora a transcrição é puxada automaticamente, independente da opção 'transcript'.
    # Se o usuário quiser pular, pode desmarcar no frontend, mas por padrão é sempre buscada.
    video_id_transcricao = _crmv2_video_id_para_transcricao(updates, card)
    if video_id_transcricao and options.get("transcript", True):
        tr = _crmv2_transcript_stub(card, video_id_transcricao, engine)
        updates["transcricao_status"] = tr.get("status")
        updates["transcricao_preview"] = tr.get("preview")
        updates["transcricao_full"] = tr.get("full", "")
        updates["transcricao_fonte"] = tr.get("fonte", "")
        updates["tem_transcricao_full"] = bool(tr.get("full"))
        if not updates.get("ultimo_video_id") and video_id_transcricao:
            updates["ultimo_video_id"] = video_id_transcricao

    if options.get("contacts", True):
        ct = _crmv2_contacts(all_text)
        updates["contatos_detectados"] = ct
        updates["emails_detectados"] = ct.get("emails") or []
        updates["instagram_links_detectados"] = ct.get("instagram_links") or []
        updates["links_detectados"] = ct.get("links") or []
        if ct.get("emails"):
            updates["email_comercial"] = ct["emails"][0]
        ig = _crmv2_pick_instagram(ct)
        if ig:
            updates["instagram_handle"] = ig
            updates["host_instagram"] = ig
        updates["tem_contato"] = bool(ct.get("emails") or ct.get("instagram_links") or ct.get("instagram_handles_possiveis"))

    # muda coluna automaticamente só quando ainda está em triagem/coleta
    old_status = str(card.get("status") or "aprovado")
    if options.get("contacts", True) and old_status in {"aprovado", "coletar_dados", "sem_contato"}:
        updates["status"] = "contato_encontrado" if updates.get("tem_contato") else "sem_contato"

    _crmv2_log(f"OK {updates.get('nome') or cid}: email={bool(updates.get('email_comercial'))} ig={bool(updates.get('instagram_handle'))} videos={len(videos)} transcript={bool(updates.get('tem_transcricao_full'))}", "ok")
    return updates

def _crmv2_update_cards(ids=None, options=None, engine="api", limit=25, max_videos=10):
    cards = sincronizar_crm_com_aprovados()
    ids = [str(x).strip() for x in (ids or []) if str(x).strip()]
    selected = []
    if ids:
        wanted = set(ids)
        selected = [c for c in cards if str(c.get("id")) in wanted or str(c.get("channel_id")) in wanted]
    else:
        for c in cards:
            if c.get("archived"):
                continue
            st = str(c.get("status") or "aprovado")
            needs = not c.get("crm_enriched") or st in {"aprovado", "coletar_dados", "sem_contato"}
            if needs:
                selected.append(c)
            if len(selected) >= int(limit or 25):
                break
    if limit and not ids:
        selected = selected[:max(1, min(200, int(limit or 25)))]
    by_id = {str(c.get("id")): c for c in cards}
    results = []
    CRM_ENRICH_LAST.update({"running": True, "started_at": str(datetime.now()), "finished_at": None, "processed": 0, "ok": 0, "fail": 0})
    _crmv2_log(f"Lote iniciado: {len(selected)} card(s), engine={engine}, options={options}", "info")
    try:
        for card in selected:
            CRM_ENRICH_LAST["processed"] += 1
            cid = str(card.get("id") or card.get("channel_id") or "")
            try:
                upd = _crmv2_enrich_card(card, options=options, engine=engine, max_videos=max_videos)
                card.update(upd)
                card["updated_at"] = str(datetime.now())
                card.setdefault("historico", []).append({"data": str(datetime.now()), "evento": "CRM API Collector V2 atualizou dados"})
                results.append({"id": cid, "ok": True, "nome": card.get("nome"), "email": card.get("email_comercial", ""), "instagram": card.get("instagram_handle", ""), "status": card.get("status")})
                CRM_ENRICH_LAST["ok"] += 1
            except Exception as e:
                msg = str(e)
                _crmv2_log(f"Falha no card {card.get('nome') or cid}: {msg}", "error")
                card["crm_enrich_error"] = msg
                card["crm_enriched_at"] = str(datetime.now())
                results.append({"id": cid, "ok": False, "erro": msg, "nome": card.get("nome")})
                CRM_ENRICH_LAST["fail"] += 1
        salvar_crm(cards)
    finally:
        CRM_ENRICH_LAST["running"] = False
        CRM_ENRICH_LAST["finished_at"] = str(datetime.now())
        _crmv2_log(f"Lote finalizado: ok={CRM_ENRICH_LAST['ok']} fail={CRM_ENRICH_LAST['fail']}", "info")
    return results, cards

@app.route("/api/crm/enrich/health")
def api_crm_enrich_health():
    keys = _crmv2_keys()
    return jsonify({"ok": True, "feature": "crm_api_collector_v2", "keys_count": len(keys), "keys_masked": [_crmv2_mask_key(k) for k in keys], "last": CRM_ENRICH_LAST, "console_tail": CRM_ENRICH_CONSOLE[-20:]})

@app.route("/api/crm/enrich/console")
def api_crm_enrich_console():
    return jsonify({"ok": True, "last": CRM_ENRICH_LAST, "console": CRM_ENRICH_CONSOLE[-200:]})

@app.route("/api/crm/enrich", methods=["POST"])
def api_crm_enrich_one():
    data = request.get_json(silent=True) or {}
    ids = data.get("ids") or []
    if data.get("id"):
        ids = [data.get("id")]
    if not ids:
        return jsonify({"ok": False, "erro": "mande id ou ids"}), 400
    opts = data.get("options") or {"channel": True, "avatar": True, "contacts": True, "videos": True, "transcript": True}
    res, cards = _crmv2_update_cards(ids=ids, options=opts, engine=data.get("engine") or "api", limit=1, max_videos=data.get("max_videos") or 10)
    return jsonify({"ok": True, "results": res, "stats": crm_stats(cards), "console": CRM_ENRICH_CONSOLE[-80:]})

@app.route("/api/crm/enrich_batch", methods=["POST"])
def api_crm_enrich_batch():
    data = request.get_json(silent=True) or {}
    opts = data.get("options") or {"channel": True, "avatar": True, "contacts": True, "videos": True, "transcript": True}
    res, cards = _crmv2_update_cards(ids=data.get("ids") or [], options=opts, engine=data.get("engine") or "api", limit=data.get("limit") or 25, max_videos=data.get("max_videos") or 10)
    return jsonify({"ok": True, "processed": len(res), "results": res, "stats": crm_stats(cards), "last": CRM_ENRICH_LAST, "console": CRM_ENRICH_CONSOLE[-120:]})

@app.route("/api/transcricao", methods=["POST"])
def api_transcricao():
    """Transcrição completa de um vídeo via pytubefix (InnerTube). Recebe video_id via
    POST e devolve a lista de trechos com timestamp, igual ao formato usado no txt de
    lead ao baixar lote."""
    data = request.get_json(silent=True) or {}
    video_id = (data.get("video_id") or "").strip()
    if not video_id:
        return jsonify({"ok": False, "erro": "mande video_id"}), 400
    if _PytubeYouTube is None:
        return jsonify({"ok": False, "erro": "pytubefix não está instalado no servidor. Rode: pip install pytubefix --break-system-packages"}), 500
    url = f"https://www.youtube.com/watch?v={video_id}"
    try:
        yt = _PytubeYouTube(url)
        titulo = yt.title or ""
        captions = yt.captions
        if not captions:
            return jsonify({"ok": False, "erro": "Nenhuma legenda disponível.", "video_id": video_id, "titulo": titulo})
        caption = None
        for code in ["pt-BR", "pt"]:
            for c in captions:
                if c.code == code:
                    caption = c
                    break
            if caption:
                break
        if not caption:
            caption = next(iter(captions), None)
        if not caption:
            return jsonify({"ok": False, "erro": "Nenhuma legenda disponível.", "video_id": video_id, "titulo": titulo})
        xml_str = caption.xml_captions
        if not xml_str:
            return jsonify({"ok": False, "erro": "Legenda vazia.", "video_id": video_id, "titulo": titulo})
        trechos = _crmv2_xml_captions_to_text(xml_str)
        return jsonify({"ok": True, "video_id": video_id, "titulo": titulo, "idioma": caption.code, "total": len(trechos), "transcricao": trechos})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e), "video_id": video_id}), 500

@app.route("/api/projetos")
def api_projetos():
    return jsonify(_read_json(PROJECTS_FILE, {"projetos": []}))

@app.route("/api/projetos/create", methods=["POST"])
def api_projetos_create():
    data = request.json or {}

    def clean_part(x, fallback):
        try:
            import unicodedata
            x = unicodedata.normalize('NFKD', str(x or fallback)).encode('ascii', 'ignore').decode('ascii')
        except Exception:
            x = str(x or fallback)
        x = re.sub(r'[^A-Za-z0-9_\- ]+', '', x).strip()
        x = re.sub(r'\s+', '_', x.upper())
        return x or fallback.upper()

    cliente_raw = data.get("cliente", "CLIENTE")
    tipo_raw = data.get("tipo", "REELS")
    ano = str(data.get("ano") or datetime.now().year)

    cliente = clean_part(cliente_raw, "CLIENTE")
    tipo = clean_part(tipo_raw, "TIPO")
    nome = f"{cliente}_{tipo}_{ano}"

    default_base = os.path.join(os.path.expanduser("~"), "Downloads", "Projetos Toolkit")
    base = data.get("base_path") or default_base
    pasta = os.path.abspath(os.path.join(base, nome))

    dirs = {
        "footage": os.path.join(pasta, "01_FOOTAGE"),
        "audio": os.path.join(pasta, "02_AUDIO"),
        "images": os.path.join(pasta, "03_IMAGES"),
        "gfx": os.path.join(pasta, "04_GFX"),
        "exports": os.path.join(pasta, "05_EXPORTS"),
        "docs": os.path.join(pasta, "06_DOCS"),
        "xml": os.path.join(pasta, "07_XML_PREMIERE"),
        "project": os.path.join(pasta, "08_PROJECT"),
    }
    for p in dirs.values():
        os.makedirs(p, exist_ok=True)

    template_dir = os.path.abspath(os.path.join("templates", "premiere"))
    pr_template = os.path.join(template_dir, "CLIENTE_TIPO_ANO.prproj")
    xml_template = os.path.join(template_dir, "TEMPLATEPR.xml")

    prproj = os.path.join(dirs["project"], f"{nome}.prproj")
    if os.path.exists(pr_template):
        shutil.copyfile(pr_template, prproj)
    elif not os.path.exists(prproj):
        with open(prproj, 'w', encoding='utf-8') as f:
            f.write("THON TOOLKIT PLACEHOLDER - template Premiere ausente.\n")

    try:
        qtd = int(data.get("qtd_xml") or data.get("quantidade") or 1)
    except Exception:
        qtd = 1
    qtd = max(1, min(200, qtd))

    seq_prefix = f"{cliente}_{tipo}"

    def gerar_xml_unico():
        if not os.path.exists(xml_template):
            return None
        with open(xml_template, 'r', encoding='utf-8-sig') as f:
            tpl = f.read()

        m = re.search(r'<sequence\b.*?</sequence>', tpl, flags=re.S)
        if not m:
            return None
        seq_tpl = m.group(0)

        sequencias = []
        nomes_seq = []
        for i in range(1, qtd + 1):
            seq_name = f"{seq_prefix}_{i:02d}"
            nomes_seq.append(seq_name)
            seq = seq_tpl
            seq = re.sub(r'<sequence id="[^"]*"', f'<sequence id="sequence-{i}"', seq, count=1)
            seq = re.sub(r'<uuid>.*?</uuid>', f'<uuid>{uuid.uuid4()}</uuid>', seq, count=1, flags=re.S)
            seq = re.sub(r'<name>.*?</name>', f'<name>{seq_name}</name>', seq, count=1, flags=re.S)
            seq = seq.replace('MAINPR', seq_name)
            seq = seq.replace('SHORT 1', seq_name).replace('Short 1', seq_name).replace('short 1', seq_name)
            sequencias.append(seq)

        xml = re.sub(r'<sequence\b.*?</sequence>', '\n'.join(sequencias), tpl, count=1, flags=re.S)
        out_file = os.path.join(dirs["xml"], f"{nome}_{qtd:02d}_SEQUENCIAS.xml")
        with open(out_file, 'w', encoding='utf-8') as f:
            f.write(xml)
        return out_file, nomes_seq

    xml_result = gerar_xml_unico()
    xmls = []
    nomes_seq = []
    if xml_result:
        xmls = [xml_result[0]]
        nomes_seq = xml_result[1]

    instrucoes = os.path.join(dirs["xml"], "COMO_IMPORTAR.txt")
    with open(instrucoes, 'w', encoding='utf-8') as f:
        f.write("1. Abra o projeto .prproj no Premiere.\n")
        f.write("2. Importe o XML ÚNICO desta pasta.\n")
        f.write("3. Ele deve criar as sequências já numeradas com cliente + tipo.\n")
        f.write("4. Use o Thon Organizer CEP para organizar assets nas bins.\n")

    db = _read_json(PROJECTS_FILE, {"projetos": []})
    projeto = {
        "id": "proj_" + str(int(time.time()*1000)),
        "nome": nome,
        "cliente": cliente_raw,
        "tipo": tipo_raw,
        "ano": ano,
        "quantidade": qtd,
        "pasta": pasta,
        "premiere": prproj,
        "xml_dir": dirs["xml"],
        "xmls": xmls,
        "sequencias": nomes_seq,
        "created_at": str(datetime.now())
    }
    db.setdefault("projetos", []).insert(0, projeto)
    _write_json(PROJECTS_FILE, db)
    return jsonify({"ok": True, "projeto": projeto})

# ===== TRABALHOS ATIVOS =====
def carregar_trabalhos():
    data = _read_json(JOBS_FILE, {"trabalhos": []})
    if not isinstance(data, dict): data = {"trabalhos": []}
    data.setdefault("trabalhos", [])
    return data

def _safe_trabalhos_list(data=None):
    """Aceita JSON novo (dict) e legado (list) sem quebrar dashboard/metas."""
    if data is None:
        data = carregar_trabalhos()
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        trabalhos = data.get("trabalhos", [])
        return trabalhos if isinstance(trabalhos, list) else []
    return []

def salvar_trabalhos(trabalhos):
    _write_json(JOBS_FILE, {"trabalhos": trabalhos, "ultima_atualizacao": str(datetime.now())})

def trabalho_stats(trabalhos):
    total = len(trabalhos)
    ativos = len([t for t in trabalhos if t.get("status", "ativo") == "ativo"])
    entregues = sum(int(t.get("entregues") or 0) for t in trabalhos)
    videos = sum(int(t.get("qtd_videos") or 0) for t in trabalhos)
    receita = sum(float(t.get("valor") or 0) for t in trabalhos)
    minutos = sum(float(t.get("minutos") or 0) for t in trabalhos)
    return {"total": total, "ativos": ativos, "entregues": entregues, "videos": videos, "receita": receita, "minutos": minutos, "horas": round(minutos/60, 2)}

@app.route("/api/trabalhos")
def api_trabalhos():
    db = carregar_trabalhos(); trabalhos = _safe_trabalhos_list(db)
    return jsonify({"trabalhos": trabalhos, "stats": trabalho_stats(trabalhos)})

@app.route("/api/trabalhos/create", methods=["POST"])
def api_trabalhos_create():
    data = request.json or {}
    db = carregar_trabalhos(); trabalhos = _safe_trabalhos_list(db)
    tid = data.get("id") or ("job_" + str(int(time.time()*1000)))
    t = {
        "id": tid, "cliente": data.get("cliente", "Cliente"), "lead_id": data.get("lead_id", ""),
        "titulo": data.get("titulo") or data.get("tipo") or "Pacote de vídeos",
        "qtd_videos": int(data.get("qtd_videos") or 0), "entregues": int(data.get("entregues") or 0),
        "valor": float(data.get("valor") or 0), "prazo": data.get("prazo", ""),
        "status": data.get("status", "ativo"), "minutos": float(data.get("minutos") or 0),
        "timer_inicio": None, "observacoes": data.get("observacoes", ""),
        "cor": data.get("cor") or gerar_cor_lead(tid), "created_at": str(datetime.now()), "updated_at": str(datetime.now())
    }
    trabalhos.insert(0, t); salvar_trabalhos(trabalhos)
    return jsonify({"ok": True, "trabalho": t, "stats": trabalho_stats(trabalhos)})

@app.route("/api/trabalhos/update", methods=["POST"])
def api_trabalhos_update():
    data = request.json or {}; tid = data.get("id")
    if not tid: return jsonify({"erro":"id ausente"}), 400
    db = carregar_trabalhos(); trabalhos = _safe_trabalhos_list(db)
    found = None
    for t in trabalhos:
        if t.get("id") == tid: found = t; break
    if not found: return jsonify({"erro":"trabalho nao encontrado"}), 404
    allowed = {"cliente","titulo","qtd_videos","entregues","valor","prazo","status","minutos","observacoes","cor"}
    for k,v in data.items():
        if k in allowed: found[k] = v
    found["updated_at"] = str(datetime.now())
    salvar_trabalhos(trabalhos)
    return jsonify({"ok": True, "trabalho": found, "stats": trabalho_stats(trabalhos)})

@app.route("/api/trabalhos/timer", methods=["POST"])
def api_trabalhos_timer():
    data = request.json or {}; tid = data.get("id"); action = data.get("action")
    db = carregar_trabalhos(); trabalhos = _safe_trabalhos_list(db)
    found = next((t for t in trabalhos if t.get("id") == tid), None)
    if not found: return jsonify({"erro":"trabalho nao encontrado"}), 404
    now = time.time()
    if action == "start":
        found["timer_inicio"] = now
    elif action == "pause":
        start = found.get("timer_inicio")
        if start:
            found["minutos"] = float(found.get("minutos") or 0) + max(0, (now - float(start))/60)
        found["timer_inicio"] = None
    found["updated_at"] = str(datetime.now())
    salvar_trabalhos(trabalhos)
    return jsonify({"ok": True, "trabalho": found, "stats": trabalho_stats(trabalhos)})

@app.route("/api/trabalhos/from_crm", methods=["POST"])
def api_trabalho_from_crm():
    data = request.json or {}; lead_id = data.get("lead_id")
    cards = sincronizar_crm_com_aprovados()
    lead = next((c for c in cards if c.get("id") == lead_id), None)
    if not lead: return jsonify({"erro":"lead nao encontrado"}), 404
    payload = {
        "lead_id": lead_id, "cliente": lead.get("nome", "Cliente"), "titulo": data.get("titulo") or "Pacote de vídeos",
        "qtd_videos": data.get("qtd_videos") or lead.get("qtd_videos") or 0, "valor": data.get("valor") or lead.get("valor_mensal") or 0,
        "prazo": data.get("prazo", ""), "observacoes": data.get("observacoes", ""), "cor": lead.get("cor") or gerar_cor_lead(lead_id)
    }
    with app.test_request_context(json=payload):
        return api_trabalhos_create()

@app.route("/api/metas")
def api_metas():
    cards = sincronizar_crm_com_aprovados()
    goals = _read_json(GOALS_FILE, {"meta_mensal":10000,"meta_dms_semana":75,"meta_videos_semana":10})
    stats = crm_stats(cards)
    trabalhos = _safe_trabalhos_list()
    goals["trabalhos"] = trabalho_stats(trabalhos)
    goals["mrr_atual"] = stats["mrr"]
    goals["faltam"] = max(0, float(goals.get("meta_mensal",10000)) - stats["mrr"])
    goals["clientes_ativos"] = stats["por_status"].get("cliente_ativo",0)
    return jsonify(goals)

@app.route("/api/metas", methods=["POST"])
def api_metas_save():
    data = request.json or {}
    current = _read_json(GOALS_FILE, {})
    for k in ["meta_mensal","meta_dms_semana","meta_videos_semana","ticket_medio"]:
        if k in data:
            try: current[k] = float(data[k])
            except: current[k] = data[k]
    current["updated_at"] = str(datetime.now())
    _write_json(GOALS_FILE, current)
    return jsonify({"ok": True, **current})

@app.route("/api/dashboard")
def api_dashboard():
    with lock:
        cards = sincronizar_crm_com_aprovados()
        return jsonify({
            "ok": True,
            "version": APP_VERSION,
            "engine": ENGINE_VERSION,
            "prospector": {
                "aprovados": len(estado.get("aprovados", []) or []),
                "reprovados": len(estado.get("reprovados", []) or []),
                "vistos": len(estado.get("vistos", []) or []),
                "lote": len(estado.get("lote", []) or []),
                "status": estado.get("status", "idle"),
                "msg": estado.get("msg", ""),
                "pipeline_stage": estado.get("pipeline_stage", "idle"),
                "api_quota_used": _api_quota_used(),
                "api_quota_budget": API_DAILY_BUDGET,
                "candidatos_encontrados": estado.get("candidatos_encontrados", 0),
                "api_channels_detailed": estado.get("api_channels_detailed", 0),
                "ytdlp_verified": estado.get("ytdlp_verified", 0),
                "qualificados": estado.get("qualificados", 0),
            },
            "crm": crm_stats(cards),
            "trabalhos": trabalho_stats(_safe_trabalhos_list()),
            "brutos": len(_carregar_brutos()),
        })

# ===== DOWNLOADER =====
DEFAULT_DOWNLOADER_DIR = os.path.join(os.path.expanduser("~"), "Downloads", "THON_Downloads")

def _downloader_config():
    cfg = _read_json(DOWNLOADER_CONFIG_FILE, {})
    if not isinstance(cfg, dict):
        cfg = {}
    folder = cfg.get("folder") or DEFAULT_DOWNLOADER_DIR
    folder = os.path.abspath(os.path.expanduser(str(folder)))
    cfg["folder"] = folder
    cfg.setdefault("ultima_atualizacao", str(datetime.now()))
    return cfg

def _save_downloader_config(cfg):
    folder = cfg.get("folder") or DEFAULT_DOWNLOADER_DIR
    folder = os.path.abspath(os.path.expanduser(str(folder)))
    os.makedirs(folder, exist_ok=True)
    cfg["folder"] = folder
    cfg["ultima_atualizacao"] = str(datetime.now())
    _write_json(DOWNLOADER_CONFIG_FILE, cfg)
    return cfg

def _download_folder():
    cfg = _downloader_config()
    folder = cfg.get("folder") or DEFAULT_DOWNLOADER_DIR
    os.makedirs(folder, exist_ok=True)
    return folder

def _detectar_plataforma(url):
    u = (url or "").lower()
    if "drive.google.com" in u or "docs.google.com" in u: return "Google Drive"
    if "youtube.com" in u or "youtu.be" in u: return "YouTube"
    if "tiktok.com" in u: return "TikTok"
    if "instagram.com" in u: return "Instagram"
    if "twitter.com" in u or "x.com" in u: return "X/Twitter"
    if "facebook.com" in u or "fb.watch" in u: return "Facebook"
    if "vimeo.com" in u: return "Vimeo"
    return "Link suportado pelo yt-dlp"

_download_lock = threading.Lock()
_active_download_threads = {}

def _download_db():
    data = _read_json(DOWNLOADS_FILE, {"jobs": []})
    if not isinstance(data, dict): data = {"jobs": []}
    data.setdefault("jobs", [])
    return data

def _save_download_jobs(jobs):
    _write_json(DOWNLOADS_FILE, {"jobs": jobs[:80], "ultima_atualizacao": str(datetime.now())})

def _set_download_job(job_id, **updates):
    with _download_lock:
        db = _download_db(); jobs = db.get("jobs", [])
        for j in jobs:
            if j.get("id") == job_id:
                j.update(updates); j["updated_at"] = str(datetime.now()); break
        _save_download_jobs(jobs)

def _download_worker(job):
    url = job.get("url", "").strip()
    mode = job.get("mode", "video")
    job_id = job["id"]
    try:
        py = sys.executable or "python"
        folder = job.get("folder") or _download_folder()
        os.makedirs(folder, exist_ok=True)
        is_drive = "drive.google.com" in url or "docs.google.com" in url
        if is_drive:
            cmd = [py, "-m", "gdown", "--fuzzy", url]
            cwd = folder
        else:
            out = os.path.join(folder, "%(title).180B [%(id)s].%(ext)s")
            cmd = [py, "-m", "yt_dlp", "--no-warnings", "-o", out]
            if mode == "audio":
                cmd += ["-x", "--audio-format", "mp3"]
            elif mode == "thumbnail":
                cmd += ["--skip-download", "--write-thumbnail"]
            elif mode == "subs":
                cmd += ["--skip-download", "--write-subs", "--write-auto-subs", "--sub-lang", "pt,en,pt-BR"]
            cmd.append(url)
            cwd = None
        _set_download_job(job_id, status="running", command=" ".join(cmd), folder=folder)
        r = subprocess.run(cmd, capture_output=True, text=True, cwd=cwd, timeout=60*60)
        output = ((r.stdout or "") + "\n" + (r.stderr or "")).strip()
        if r.returncode == 0:
            _set_download_job(job_id, status="done", output=output[-5000:], finished_at=str(datetime.now()))
        else:
            _set_download_job(job_id, status="error", output=output[-5000:], error=f"Comando falhou com código {r.returncode}", finished_at=str(datetime.now()))
    except subprocess.TimeoutExpired:
        _set_download_job(job_id, status="error", error="Timeout: download demorou mais de 1 hora.", finished_at=str(datetime.now()))
    except Exception as e:
        _set_download_job(job_id, status="error", error=str(e), finished_at=str(datetime.now()))

@app.route("/api/downloader/start", methods=["POST"])
def api_downloader_start():
    data = request.json or {}
    url = (data.get("url") or "").strip()
    mode = data.get("mode") or "video"
    if not url or not re.match(r"^https?://", url):
        return jsonify({"ok": False, "erro": "Link inválido. Precisa começar com http:// ou https://"}), 400
    if mode not in ["video", "audio", "thumbnail", "subs"]:
        mode = "video"
    folder = _download_folder()
    job = {"id": "dl_" + uuid.uuid4().hex[:10], "url": url, "mode": mode, "platform": _detectar_plataforma(url), "status": "queued", "folder": folder, "created_at": str(datetime.now()), "output": "", "error": ""}
    with _download_lock:
        db = _download_db(); jobs = db.get("jobs", [])
        jobs.insert(0, job); _save_download_jobs(jobs)
    t = threading.Thread(target=_download_worker, args=(job,), daemon=True)
    _active_download_threads[job["id"]] = t
    t.start()
    return jsonify({"ok": True, "job": job})

@app.route("/api/downloader/jobs")
def api_downloader_jobs():
    return jsonify(_download_db())

@app.route("/api/downloader/check")
def api_downloader_check():
    def module_exists(mod):
        try:
            import importlib.util
            return importlib.util.find_spec(mod) is not None
        except Exception:
            return False
    setup_file = os.path.join(DATA_DIR, "setup_status.json")
    setup_cache = {}
    if os.path.exists(setup_file):
        try:
            with open(setup_file, "r", encoding="utf-8") as f:
                setup_cache = json.load(f)
        except Exception as e:
            setup_cache = {"error": str(e)}
    return jsonify({
        "ok": True,
        "python": sys.version.split()[0],
        "platform": platform.platform(),
        "folder": _download_folder(),
        "modules_now": {
            "flask": module_exists("flask"),
            "flask-cors": module_exists("flask_cors"),
            "yt-dlp": module_exists("yt_dlp"),
            "gdown": module_exists("gdown"),
        },
        "ffmpeg": {"ok": bool(shutil.which("ffmpeg")), "path": shutil.which("ffmpeg")},
        "setup_cache": setup_cache
    })

@app.route("/api/system/setup_status")
def api_system_setup_status():
    setup_file = os.path.join(DATA_DIR, "setup_status.json")
    if os.path.exists(setup_file):
        try:
            with open(setup_file, "r", encoding="utf-8") as f:
                return jsonify(json.load(f))
        except Exception as e:
            return jsonify({"ok": False, "error": str(e)}), 500
    return jsonify({"ok": False, "error": "setup_status.json ainda nao existe"})

@app.route("/api/downloader/config")
def api_downloader_config_get():
    return jsonify({"ok": True, **_downloader_config(), "default_folder": DEFAULT_DOWNLOADER_DIR})

@app.route("/api/downloader/config", methods=["POST"])
def api_downloader_config_save():
    data = request.json or {}
    folder = (data.get("folder") or "").strip()
    if not folder:
        folder = DEFAULT_DOWNLOADER_DIR
    try:
        cfg = _save_downloader_config({"folder": folder})
        return jsonify({"ok": True, **cfg})
    except Exception as e:
        return jsonify({"ok": False, "erro": f"Não consegui salvar essa pasta: {e}"}), 400

@app.route("/api/downloader/select_folder", methods=["POST"])
def api_downloader_select_folder():
    folder = _download_folder()
    os.makedirs(folder, exist_ok=True)
    return jsonify({
        "ok": False,
        "fallback": True,
        "folder": folder,
        "msg": "Por segurança, o seletor nativo foi desativado. Clique em Abrir pasta ou cole o caminho manualmente."
    }), 200

@app.route("/api/downloader/open_folder", methods=["POST"])
def api_downloader_open_folder():
    folder = _download_folder()
    os.makedirs(folder, exist_ok=True)
    try:
        if sys.platform == "darwin": subprocess.Popen(["open", folder])
        elif os.name == "nt": os.startfile(folder)
        else: subprocess.Popen(["xdg-open", folder])
        return jsonify({"ok": True, "msg": "Pasta aberta", "folder": folder})
    except Exception as e:
        return jsonify({"ok": False, "msg": str(e), "folder": folder})

@app.route("/api/downloader/clear", methods=["POST"])
def api_downloader_clear():
    _save_download_jobs([])
    return jsonify({"ok": True})

# ===== ROTAS SEGURAS DOS PROSPECTORS — SEM HTML CUSTOM / SEM MEXER NO DESIGN =====
# Estas rotas só apontam para o Prospector original com querystring. O HTML real continua sendo o seu prospector.html.
# O backend lê o Referer em /iniciar e decide o motor: api ou dlp.
@app.route('/prospectors')
def prospectors_home_page():
    return "", 302, {"Location": "/"}

@app.route('/prospector-crm')
def prospector_crm_page():
    return "", 302, {"Location": "/prospector"}

@app.route('/prospector-api')
def prospector_api_page():
    return "", 302, {"Location": "/prospector?engine_mode=api"}

@app.route('/prospector-dlp')
def prospector_dlp_page():
    return "", 302, {"Location": "/prospector?engine_mode=dlp"}

# ============================================================
# V58.26: ROTAS BLACKLIST CONCORRENTES (#4)
# ============================================================
# ============================================================
# V58.29: QUERY FACTORY V2 — status e reset
# ============================================================
@app.route("/api/query_factory/status")
def api_query_factory_status():
    """Status do query factory V2 (saturação, fila, etc)."""
    try:
        from query_factory.query_factory import get_status, load_performance
        status = get_status()
        # top 10 queries mais saturadas
        perf = load_performance()
        queries_perf = perf.get("queries") or {}
        saturadas = sorted(
            [{"query": q, "falhas": d.get("falhas",0), "sucessos": d.get("sucessos",0), "qualificados_total": d.get("qualificados_total",0), "cooldown_ate": d.get("cooldown_ate")}
             for q, d in queries_perf.items() if d.get("cooldown_ate")],
            key=lambda x: x.get("cooldown_ate") or 0, reverse=True
        )[:10]
        status["saturadas_top10"] = saturadas
        return jsonify({"ok": True, **status})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500

@app.route("/api/query_factory/reset_saturacao", methods=["POST"])
def api_query_factory_reset():
    """Limpa cooldown de todas as queries (força retry)."""
    try:
        from query_factory.query_factory import reset_saturacao
        reset_saturacao()
        return jsonify({"ok": True, "msg": "Saturação resetada. Todas as queries podem rodar de novo."})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500

@app.route("/api/query_factory/regenerar", methods=["POST"])
def api_query_factory_regenerar():
    """Força geração de novas queries e limpa fila atual."""
    try:
        from query_factory.query_factory import generate_more
        data = request.json or {}
        amount = int(data.get("amount") or 300)
        config = data.get("config") or {}
        geradas = generate_more(config=config, amount=amount)
        return jsonify({"ok": True, "geradas": geradas, "msg": f"{geradas} novas queries geradas."})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500

@app.route("/api/query_factory/reset_seen", methods=["POST"])
def api_query_factory_reset_seen():
    """V58.31: limpa TODO o seen.json pra permitir gerar todas as combinacoes de novo."""
    try:
        from query_factory.query_factory import SEEN_FILE, _save
        _save(SEEN_FILE, {"queries": {}})
        return jsonify({"ok": True, "msg": "seen.json resetado. Todas as combinações podem ser geradas de novo."})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500

@app.route("/api/query_factory/limpar_fila", methods=["POST"])
def api_query_factory_limpar_fila():
    """V58.31: limpa fila atual (queue.json) pra forçar gerar combinacoes novas."""
    try:
        from query_factory.query_factory import QUEUE_FILE, _save
        _save(QUEUE_FILE, {"queries": []})
        return jsonify({"ok": True, "msg": "Fila limpa. Próxima chamada vai gerar queries novas."})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500

# ============================================================
# V58.33: TTL BLACKLIST REPROVADOS
# ============================================================
@app.route("/api/blacklist/status")
def api_blacklist_status():
    """Status da blacklist de reprovados (total, antigos, etc)."""
    try:
        dias = int(request.args.get("dias") or REPROVADOS_TTL_DIAS)
        r = limpar_reprovados_antigos(dias_ttl=dias, dry_run=True)
        return jsonify({"ok": True, **r, "ttl_padrao_dias": REPROVADOS_TTL_DIAS})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500

@app.route("/api/blacklist/limpar_antigos", methods=["POST"])
def api_blacklist_limpar():
    """V58.33: remove reprovados com mais de N dias da blacklist."""
    try:
        data = request.json or {}
        dias = int(data.get("dias") or REPROVADOS_TTL_DIAS)
        r = limpar_reprovados_antigos(dias_ttl=dias, dry_run=False)
        print(f"[BLACKLIST TTL] limpeza manual: {r}")
        return jsonify({"ok": True, **r, "msg": f"Removidos {r['removidos']} reprovados com mais de {dias} dias. Blacklist: {r['total_antes']} → {r['total_depois']}."})
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500

@app.route("/api/concorrentes", methods=["GET"])
def api_concorrentes_list():
    """Lista IDs de canais concorrentes blacklist."""
    ids = sorted(_carregar_concorrentes_blacklist())
    return jsonify({"ok": True, "ids": ids, "total": len(ids), "file": CONCORRENTES_BLACKLIST_FILE})

@app.route("/api/concorrentes/add", methods=["POST"])
def api_concorrentes_add():
    """Adiciona 1+ IDs de canais concorrentes à blacklist."""
    data = request.json or {}
    ids_in = data.get("ids") or []
    if isinstance(ids_in, str):
        ids_in = [ids_in]
    ids_in = [str(x).strip() for x in ids_in if str(x).strip()]
    if not ids_in:
        return jsonify({"ok": False, "erro": "ids obrigatorio"}), 400
    atual = _carregar_concorrentes_blacklist()
    novos = [i for i in ids_in if i not in atual]
    atual.update(novos)
    _salvar_concorrentes_blacklist(atual)
    return jsonify({"ok": True, "adicionados": len(novos), "total": len(atual), "novos": novos})

@app.route("/api/concorrentes/remove", methods=["POST"])
def api_concorrentes_remove():
    """Remove 1+ IDs de canais concorrentes da blacklist."""
    data = request.json or {}
    ids_in = data.get("ids") or []
    if isinstance(ids_in, str):
        ids_in = [ids_in]
    ids_in = [str(x).strip() for x in ids_in if str(x).strip()]
    atual = _carregar_concorrentes_blacklist()
    removidos = [i for i in ids_in if i in atual]
    for i in removidos:
        atual.discard(i)
    _salvar_concorrentes_blacklist(atual)
    return jsonify({"ok": True, "removidos": len(removidos), "total": len(atual), "removidos_ids": removidos})

@app.route("/api/concorrentes/clear", methods=["POST"])
def api_concorrentes_clear():
    """Limpa toda a blacklist de concorrentes."""
    _salvar_concorrentes_blacklist(set())
    return jsonify({"ok": True, "total": 0})

# ============================================================
# V58.27-LOTE-XLSX: export/import planilha do lote
# ============================================================
@app.route("/api/lote/export_xlsx", methods=["POST", "GET"])
def api_lote_export_xlsx():
    """Exporta o lote de caca atual como .xlsx com colunas decisao/motivo vazias."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
    except ImportError:
        return jsonify({"ok": False, "erro": "openpyxl nao instalado no backend"}), 500

    lote = carregar_lote_ativo()
    if not lote:
        return jsonify({"ok": False, "erro": "Lote vazio. Inicie um garimpo primeiro."}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Lote de Caca"

    # Cabecalho
    headers = [
        "id", "nome", "url", "subs", "subs_fmt", "score", "nicho",
        "descricao", "email", "instagram", "instagram_link",
        "longos", "shorts", "views", "last_video_days",
        "score_tags", "sinais_monetizacao", "ja_tem_editor",
        "decisao", "motivo_reprovacao"
    ]
    header_fill = PatternFill(start_color="00C6FF", end_color="00C6FF", fill_type="solid")
    header_font = Font(color="03100F", bold=True, size=11)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Linhas
    for row_idx, c in enumerate(lote, start=2):
        ws.cell(row=row_idx, column=1, value=str(c.get("id") or c.get("channel_id") or ""))
        ws.cell(row=row_idx, column=2, value=str(c.get("nome") or c.get("title") or ""))
        ws.cell(row=row_idx, column=3, value=str(c.get("url") or ""))
        ws.cell(row=row_idx, column=4, value=int(c.get("subs") or 0))
        ws.cell(row=row_idx, column=5, value=str(c.get("subs_fmt") or ""))
        ws.cell(row=row_idx, column=6, value=int(c.get("score") or 0))
        ws.cell(row=row_idx, column=7, value=str(c.get("nicho") or ""))
        # descricao truncada em 500 chars (Excel aguenta, mas fica pesado)
        desc = str(c.get("description") or "")[:500]
        ws.cell(row=row_idx, column=8, value=desc)
        ws.cell(row=row_idx, column=9, value=str(c.get("email") or ""))
        ws.cell(row=row_idx, column=10, value=str(c.get("instagram_handle") or ""))
        ws.cell(row=row_idx, column=11, value=str(c.get("instagram_link") or ""))
        ws.cell(row=row_idx, column=12, value=int(c.get("longos") or 0))
        ws.cell(row=row_idx, column=13, value=int(c.get("shorts") or 0))
        ws.cell(row=row_idx, column=14, value=int(c.get("recent_avg_views") or 0))
        ws.cell(row=row_idx, column=15, value=int(c.get("last_video_days") or 0))
        # tags como string separada por virgula
        tags = c.get("score_tags_v26") or c.get("score_tags") or []
        if isinstance(tags, list):
            tags_str = ", ".join(str(t) for t in tags)
        else:
            tags_str = str(tags)
        ws.cell(row=row_idx, column=16, value=tags_str)
        # sinais monetizacao
        sinais = c.get("sinais_monetizacao") or []
        if isinstance(sinais, list):
            sinais_str = ", ".join(str(s) for s in sinais)
        else:
            sinais_str = str(sinais)
        ws.cell(row=row_idx, column=17, value=sinais_str)
        ws.cell(row=row_idx, column=18, value="SIM" if c.get("ja_tem_editor") else "NAO")
        # colunas decisao e motivo_reprovacao FICAM VAZIAS (usuario preenche)
        ws.cell(row=row_idx, column=19, value="")
        ws.cell(row=row_idx, column=20, value="")

    # Largura das colunas
    col_widths = [25, 30, 50, 10, 10, 8, 18, 60, 30, 25, 40, 8, 8, 10, 12, 50, 40, 12, 12, 30]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # Freeze panes (primeira linha)
    ws.freeze_panes = "A2"

    # Adiciona aba de instrucoes
    ws2 = wb.create_sheet("Instrucoes")
    instrucoes = [
        ["COMO USAR ESTA PLANILHA", ""],
        ["", ""],
        ["1.", "Nao altere as colunas A-R (id ate ja_tem_editor). Elas sao dados do canal."],
        ["2.", "Preencha a coluna 'decisao' (S) com um dos valores:"],
        ["", "  aprovar  -> lead bom, vai pra aprovados + CRM"],
        ["", "  reprovar -> lead ruim, vai pra descartados com motivo"],
        ["", "  revisar  -> mantem no lote pra voce olhar depois"],
        ["", "  (vazio)  -> mantem no lote (nao mexe)"],
        ["3.", "Se decidir 'reprovar', preencha a coluna 'motivo_reprovacao' (T)"],
        ["4.", "Salve o arquivo (mantenha .xlsx)"],
        ["5.", "Va no Prospector > Step 7 > 'Importar planilha analisada'"],
        ["6.", "Selecione este arquivo e clique importar"],
        ["", ""],
        ["DICAS DE QUALIFICACAO:", ""],
        ["- Score >= 80", "Lead excelente (audiencia engajada, monetizado, profissional)"],
        ["- Score 60-79", "Lead bom, vale revisar descricao e nicho"],
        ["- Score < 60", "Lead fraco (audiencia morta, sem longos, etc)"],
        ["- ja_tem_editor = SIM", "Lead ruim (ja tem editor, nao precisa de voce)"],
        ["- sinais_monetizacao", "Lead bom se tiver 2+ sinais (patrocinio, loja, afiliados)"],
        ["- views/subs < 0.01", "Audiencia comprada (subs falsos)"],
        ["- last_video_days > 365", "Canal parado"],
    ]
    for row in instrucoes:
        ws2.append(row)
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 80

    # Gera xlsx em memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"thon_lote_{timestamp}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )

@app.route("/api/lote/import_xlsx", methods=["POST"])
def api_lote_import_xlsx():
    """Le planilha .xlsx preenchida e dispara aprovar/descartar baseado na coluna 'decisao'."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return jsonify({"ok": False, "erro": "openpyxl nao instalado no backend"}), 500

    if "file" not in request.files:
        return jsonify({"ok": False, "erro": "Arquivo nao enviado. Use form-data com campo 'file'."}), 400

    file = request.files["file"]
    if not file.filename or not file.filename.endswith(".xlsx"):
        return jsonify({"ok": False, "erro": "Arquivo deve ser .xlsx"}), 400

    try:
        wb = load_workbook(file, data_only=True)
    except Exception as e:
        return jsonify({"ok": False, "erro": f"Falha ao ler xlsx: {e}"}), 400

    # Tenta aba "Lote de Caca" primeiro, senao primeira aba
    if "Lote de Caca" in wb.sheetnames:
        ws = wb["Lote de Caca"]
    else:
        ws = wb[wb.sheetnames[0]]

    # Le cabecalho
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            headers[str(val).strip().lower()] = col

    # Colunas obrigatorias
    required = ["id", "decisao"]
    for r in required:
        if r not in headers:
            return jsonify({"ok": False, "erro": f"Coluna '{r}' nao encontrada no cabecalho. Use a planilha exportada pelo backend."}), 400

    id_col = headers["id"]
    decisao_col = headers["decisao"]
    motivo_col = headers.get("motivo_reprovacao")

    # Processa linhas
    lote_atual = carregar_lote_ativo()
    lote_by_id = {str(c.get("id") or c.get("channel_id") or ""): c for c in lote_atual}

    aprovados_ids = []
    reprovados_ids = []
    revisar_ids = []
    ignorados_ids = []
    erros = []

    for row in range(2, ws.max_row + 1):
        cid = str(ws.cell(row=row, column=id_col).value or "").strip()
        if not cid:
            continue
        decisao = str(ws.cell(row=row, column=decisao_col).value or "").strip().lower()
        motivo = str(ws.cell(row=row, column=motivo_col).value or "").strip() if motivo_col else ""

        if decisao in ["aprovar", "ok", "sim", "1", "a"]:
            aprovados_ids.append(cid)
        elif decisao in ["reprovar", "descartar", "rejeitar", "0", "r"]:
            if not motivo:
                motivo = "Reprovado via planilha"
            reprovados_ids.append((cid, motivo))
        elif decisao in ["revisar", "manter", "m", "2"]:
            revisar_ids.append(cid)
        else:
            ignorados_ids.append(cid)

    # Dispara /aprovar e /descartar reaproveitando logica existente
    resultado = {"aprovados": 0, "reprovados": 0, "revisar": len(revisar_ids), "ignorados": len(ignorados_ids), "erros": []}

    if aprovados_ids:
        try:
            with lock:
                lote_oficial = carregar_lote_ativo()
                estado["lote"] = lote_oficial
                novos = [c for c in lote_oficial if str(c.get("id") or "") in aprovados_ids or str(c.get("channel_id") or "") in aprovados_ids]
                aprovados_existentes_ids = {str(a.get("id") or a.get("channel_id") or "") for a in estado.get("aprovados", []) or []}
                for c in novos:
                    cid = c.get("id") or c.get("channel_id")
                    if cid and str(cid) not in aprovados_existentes_ids:
                        c["id"] = cid
                        c["channel_id"] = cid
                        estado.setdefault("aprovados", []).append(c)
                        aprovados_existentes_ids.add(str(cid))
                        _atualizar_status_bruto(cid, "aprovado", {"aprovado_em": str(datetime.now()), "via": "planilha"})
                estado["vistos"].update(aprovados_ids)
                estado["lote"] = [c for c in lote_oficial if str(c.get("id") or "") not in aprovados_ids and str(c.get("channel_id") or "") not in aprovados_ids]
                salvar_lote_ativo(estado["lote"])
                salvar_memoria(estado["vistos"])
                salvar_aprovados(estado["aprovados"])
                resultado["aprovados"] = len(novos)
            # sync CRM (mesmo que /aprovar faz)
            try:
                if "sincronizar_crm_leads_incremental" in globals():
                    sincronizar_crm_leads_incremental(novos)
            except Exception as e:
                print(f"[IMPORT XLSX] aviso sync CRM: {e}")
        except Exception as e:
            resultado["erros"].append(f"Erro aprovando: {e}")

    if reprovados_ids:
        try:
            descartados = []
            with lock:
                lote_oficial = carregar_lote_ativo()
                estado["lote"] = lote_oficial
                for cid, motivo in reprovados_ids:
                    canal = next((c for c in lote_oficial if str(c.get("id") or "") == cid or str(c.get("channel_id") or "") == cid), None)
                    if canal:
                        descartados.append({
                            "id": cid,
                            "nome": canal.get("nome") or canal.get("title") or "Canal",
                            "url": canal.get("url") or (f"https://youtube.com/channel/{cid}" if cid else ""),
                            "nicho": canal.get("nicho", ""),
                            "score": canal.get("score", 0),
                            "subs_fmt": canal.get("subs_fmt", ""),
                            "motivo": motivo,
                            "data": str(datetime.now())
                        })
                        estado["vistos"].add(cid)
                        _atualizar_status_bruto(cid, "descartado", {"motivo": motivo, "via": "planilha"})
                if descartados:
                    salvar_reprovados_lista(descartados)
                    estado["reprovados"] = carregar_reprovados()
                    ids_rep = {d["id"] for d in descartados}
                    estado["lote"] = [c for c in lote_oficial if str(c.get("id") or "") not in ids_rep and str(c.get("channel_id") or "") not in ids_rep]
                    salvar_lote_ativo(estado["lote"])
                    salvar_memoria(estado["vistos"])
                resultado["reprovados"] = len(descartados)
        except Exception as e:
            resultado["erros"].append(f"Erro reprovando: {e}")

    resultado["lote_restante"] = len(carregar_lote_ativo())
    resultado["ok"] = not resultado["erros"]

    print(f"[IMPORT XLSX] aprovados={resultado['aprovados']} reprovados={resultado['reprovados']} revisar={resultado['revisar']} ignorados={resultado['ignorados']} lote_restante={resultado['lote_restante']}")
    return jsonify(resultado)

@app.route('/api/prospector_api/start', methods=['POST'])
def api_prospector_api_start():
    data = request.json or {}
    # V58.23: respeita engine_mode do body se for api_multi_source_fast; senao default api
    if not data.get('engine_mode'):
        data['engine_mode'] = 'api'
    data['ytdlp_verify_max'] = data.get('ytdlp_verify_max') or 999999
    try:
        config = normalizar_config_api(data)
        config['ytdlp_verify_max'] = int(data.get('ytdlp_verify_max') or 999999)
    except Exception as exc:
        return jsonify({'ok': False, 'erro': str(exc)}), 400
    started = iniciar_busca_api(config, auto=False)
    if not started:
        return jsonify({'ok': False, 'erro': 'Prospector já está rodando'}), 400
    return jsonify({'ok': True, 'modo': config.get('engine_mode', 'api'), 'msg': 'Engine iniciado com territorio + nicho + perfis.', 'config': config})

@app.route('/api/prospector_dlp/start', methods=['POST'])
def api_prospector_dlp_start():
    data = request.json or {}
    data['engine_mode'] = 'dlp'
    try:
        config = normalizar_config_api(data)
        config['engine_mode'] = 'dlp'
    except Exception as exc:
        return jsonify({'ok': False, 'erro': str(exc)}), 400
    started = iniciar_busca_api(config, auto=False)
    if not started:
        return jsonify({'ok': False, 'erro': 'Prospector já está rodando'}), 400
    return jsonify({'ok': True, 'modo': 'dlp_full', 'msg': 'DLP vai descobrir e verificar canais.', 'config': config})

@app.route('/api/dlp_verification_queue', methods=['GET'])
@app.route('/api/dlp_queue', methods=['GET'])
def api_dlp_verification_queue_get():
    fila = carregar_fila_dlp_verificacao()
    return jsonify({'ok': True, 'total': len(fila), 'file': DLP_VERIFY_QUEUE_FILE, 'fila': fila[:500]})

@app.route('/api/dlp_verification_queue/clear', methods=['POST'])
def api_dlp_verification_queue_clear():
    salvar_fila_dlp_verificacao([])
    return jsonify({'ok': True, 'total': 0})

@app.route('/api/dlp_verification_queue/start', methods=['POST'])
def api_dlp_verification_queue_start():
    data = request.json or {}
    data['engine_mode'] = 'api'
    try:
        config = normalizar_config_api(data)
        config['engine_mode'] = 'api'
        config['ytdlp_verify_max'] = int(data.get('ytdlp_verify_max') or 999999)
    except Exception as exc:
        return jsonify({'ok': False, 'erro': str(exc)}), 400
    with lock:
        if estado.get('rodando'):
            return jsonify({'ok': False, 'erro': 'Prospector já está rodando'}), 400
        estado['rodando'] = True
    def runner():
        lote, reprovados = [], []
        try:
            lote, reprovados = dlp_verify_queue_until_empty(config)
            if reprovados:
                salvar_reprovados_lista(reprovados)
            with lock:
                estado['lote'] = adicionar_lote_ativo(lote, base=estado.get('lote', []))
                estado['status'] = 'aguardando' if estado.get('lote') else 'fim'
                estado['msg'] = f'DLP fila finalizado: +{len(lote)} no lote | {len(reprovados)} reprovados'
                estado['pipeline_stage'] = 'aguardando_revisao' if lote else 'concluido'
        finally:
            with lock:
                estado['rodando'] = False
                salvar_lote_ativo(estado.get('lote', []))
    threading.Thread(target=runner, daemon=True, name='dlp-verify-queue-manual').start()
    return jsonify({'ok': True, 'msg': 'DLP verificando fila persistente até acabar.', 'queue_total': len(carregar_fila_dlp_verificacao())})

# ===== PÁGINAS =====
@app.route('/_base.css')
def base_css():
    return send_from_directory('.', '_base.css')

@app.route("/dashboard")
def dashboard_page():
    return send_from_directory(".", "dashboard.html")

@app.route("/crm")
def crm_page():
    return send_from_directory(".", "crm.html")

@app.route("/projetos")
def projetos_page():
    return send_from_directory(".", "projetos.html")

@app.route("/metas")
def metas_page():
    return send_from_directory(".", "metas.html")

@app.route("/trabalhos")
def trabalhos_page():
    return send_from_directory(".", "trabalhos.html")

@app.route("/downloader")
def downloader_page():
    return send_from_directory(".", "downloader.html")

@app.route("/")
def index():
    return send_from_directory(".", "index.html")

@app.route("/prospector")
def prospector_page():
    return send_from_directory(".", "prospector.html")

@app.route("/pipeline")
def pipeline_page():
    return send_from_directory(".", "pipeline_v6.html")

# ===== ROTAS DE STATUS E DIAGNÓSTICO =====
@app.route("/prospector/status")
def prospector_status():
    with lock:
        return jsonify({
            "ok": True,
            "backend": "online",
            "version": APP_VERSION,
            "engine": ENGINE_VERSION,
            "engine_mode": estado.get("engine_mode", carregar_engine_mode()),
            "engine_mode_label": _modo_label(estado.get("engine_mode", carregar_engine_mode())),
            "rodando": bool(estado.get("rodando")),
            "status": estado.get("status", "idle"),
            "msg": estado.get("msg", ""),
            "pipeline_stage": estado.get("pipeline_stage", "idle"),
            "verificados": int(estado.get("verificados", 0) or 0),
            "queries_processadas": int(estado.get("queries_processadas", 0) or 0),
            "candidatos_encontrados": int(estado.get("candidatos_encontrados", 0) or 0),
            "qualificados": int(estado.get("qualificados", 0) or 0),
            "lote_count": len(estado.get("lote", []) or []),
            "aprovados_count": len(estado.get("aprovados", []) or []),
            "reprovados_count": len(estado.get("reprovados", []) or []),
            "vistos_count": len(estado.get("vistos", []) or []),
            "pausa": bool(estado.get("pausa")),
            "auto_hunt": dict(auto_hunt),
            "api_quota_used": _api_quota_used(),
            "api_quota_budget": API_DAILY_BUDGET,
            "api_channels_detailed": int(estado.get("api_channels_detailed", 0) or 0),
            "api_videos_collected": int(estado.get("api_videos_collected", 0) or 0),
            "api_rejected": int(estado.get("api_rejected", 0) or 0),
            "api_last_error": estado.get("api_last_error", ""),
            "api_query_limit": int(estado.get("api_query_limit", 0) or 0),
            "api_candidate_target": int(estado.get("api_candidate_target", 0) or estado.get("target_candidatos", 0) or 0),
            "target_candidatos": int(estado.get("target_candidatos", 0) or estado.get("api_candidate_target", 0) or 0),
            "discovery_batch_size": int(estado.get("discovery_batch_size", 0) or 0),
            "api_verify_batch_size": int(estado.get("api_verify_batch_size", 0) or 0),
            "approval_goal": int(estado.get("approval_goal", 0) or 0),
            "multi_candidates_discovered_total": int(estado.get("multi_candidates_discovered_total", estado.get("candidatos_encontrados", 0)) or 0),
            "candidatos_descobertos_total": int(estado.get("multi_candidates_discovered_total", estado.get("candidatos_encontrados", 0)) or 0),
            "multi_candidates_by_source": dict(estado.get("multi_candidates_by_source", {}) or {}),
            "multi_chunks_processed": int(estado.get("multi_chunks_processed", 0) or 0),
            "multi_chunks_estimated": int(estado.get("multi_chunks_estimated", 0) or 0),
            "bloco_atual": int(estado.get("bloco_atual", 0) or 0),
            "bloco_tamanho": int(estado.get("bloco_tamanho", 0) or 0),
            "bloco_processado": int(estado.get("bloco_processado", 0) or 0),
            "fonte_atual": estado.get("fonte_atual", ""),
            "query_atual": estado.get("query_atual", ""),
            "pagina_atual": int(estado.get("pagina_atual", 0) or 0),
            "repetidos_total": int(estado.get("repetidos_total", 0) or 0),
            "pre_reprovados_total": int(estado.get("pre_reprovados_total", 0) or 0),
            "channels_detailed_total": int(estado.get("channels_detailed_total", estado.get("api_channels_detailed", 0)) or 0),
            "verificados_total": int(estado.get("verificados_total", estado.get("verificados", 0)) or 0),
            "qualificados_total": int(estado.get("qualificados_total", estado.get("qualificados", 0)) or 0),
            "reprovados_total": int(estado.get("reprovados_total", estado.get("api_rejected", 0)) or 0),
            "ultimo_evento": estado.get("ultimo_evento", ""),
            "stage": estado.get("stage", estado.get("pipeline_stage", "")),
            "stage_label": estado.get("stage_label", ""),
            "api_quota_pct": round((_api_quota_used() / max(1, API_DAILY_BUDGET)) * 100, 1),
            "ytdlp_verified": int(estado.get("ytdlp_verified", 0) or 0),
            "ytdlp_approved": int(estado.get("ytdlp_approved", 0) or 0),
            "ytdlp_verify_max": 160,
            "search_limit": "api-discovery",
            "query_limit_max": API_QUERY_LIMIT_MAX,
            "cookies": os.environ.get("THON_YTDLP_COOKIE_BROWSER", "none"),
            "brutos_total": len(_carregar_brutos()),
            "dlp_queue_count": len(carregar_fila_dlp_verificacao()),
            "dlp_queue_file": DLP_VERIFY_QUEUE_FILE,
        })

@app.route("/coleta_progress")
def coleta_progress():
    e = prospector_status().json
    return jsonify({
        "ok": True,
        "rodando": e["rodando"],
        "status": e["status"],
        "etapa": e["pipeline_stage"],
        "msg": e["msg"],
        "queries_processadas": e["queries_processadas"],
        "candidatos_encontrados": e["candidatos_encontrados"],
        "verificados": e["verificados"],
        "qualificados": e["qualificados"],
    })

@app.route("/logs/backend")
def logs_backend():
    linhas = _safe_int(request.args.get("linhas", 80), 80, 1, 500)
    paths = [RUNTIME_LOG_FILE, os.path.join(os.getcwd(), RUNTIME_LOG_FILE)]
    content = ""
    for path in paths:
        try:
            if os.path.exists(path):
                with open(path, "r", encoding="utf-8", errors="ignore") as f:
                    content = "".join(f.readlines()[-linhas:])
                break
        except Exception:
            pass
    return jsonify({"ok": True, "tail": content, "log": content, "linhas": linhas})

@app.route("/api/system/open_terminal", methods=["POST"])
def api_system_open_terminal():
    log_path = os.path.abspath(RUNTIME_LOG_FILE)
    try:
        os.makedirs(os.path.dirname(log_path) or APP_DIR, exist_ok=True)
        open(log_path, "a", encoding="utf-8").close()
    except Exception:
        pass
    return jsonify({"ok": True, "disabled": True, "path": log_path, "url": "/engine-console", "msg": "Terminal externo desativado; use o console web."})



# ============================================================
# API KEYS — GERENCIADOR SEGURO SEM ALTERAR HTML EXISTENTE
# ============================================================
def _api_keys_read_raw():
    try:
        if os.path.exists(API_KEYS_FILE):
            with open(API_KEYS_FILE, "r", encoding="utf-8") as f:
                return f.read()
    except Exception as e:
        print(f"[API KEYS] erro lendo {API_KEYS_FILE}: {e}")
    return ""

def _api_keys_normalize_text(raw):
    linhas = []
    vistos = set()
    for linha in str(raw or "").splitlines():
        linha = linha.strip()
        if not linha or linha.startswith("#"):
            continue
        if linha in vistos:
            continue
        vistos.add(linha)
        linhas.append(linha)
    return "\n".join(linhas) + ("\n" if linhas else "")

def _api_keys_mask(key):
    key = str(key or "").strip()
    if len(key) <= 14:
        return "*" * len(key)
    return key[:8] + "..." + key[-6:]

@app.route("/api/api_keys", methods=["GET", "POST"])
def api_api_keys_manage_safe():
    global _API_KEYS, _API_KEY_INDEX, _API_KEY_FAILED, API_DAILY_BUDGET
    if request.method == "GET":
        raw = _api_keys_read_raw()
        keys = [l.strip() for l in raw.splitlines() if l.strip() and not l.strip().startswith("#")]
        return jsonify({
            "ok": True,
            "file": API_KEYS_FILE,
            "text": raw,
            "count": len(keys),
            "active_count": len(_API_KEYS),
            "failed_today": {k: v for k, v in _API_KEY_FAILED.items() if v.get("date") == _api_today()},
            "keys_masked": [_api_keys_mask(k) for k in keys],
            "version": APP_VERSION,
            "engine": ENGINE_VERSION,
        })
    data = request.json or {}
    raw = data.get("text", "")
    normalized = _api_keys_normalize_text(raw)
    allow_empty = bool(data.get("allow_empty") or data.get("force_empty"))
    if not normalized.strip() and not allow_empty:
        return jsonify({"ok": False, "erro": "Proteção: caixa de API vazia. Para apagar todas, envie allow_empty=true."}), 400
    try:
        if os.path.exists(API_KEYS_FILE):
            try: shutil.copy2(API_KEYS_FILE, API_KEYS_FILE + ".bak")
            except Exception as be: print(f"[API KEYS] backup aviso: {be}")
        tmp = API_KEYS_FILE + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            f.write(normalized)
        os.replace(tmp, API_KEYS_FILE)
        with _API_KEY_LOCK:
            _API_KEYS = _api_carregar_chaves()
            _API_KEY_INDEX = 0
            try: _API_KEY_FAILED.clear()
            except Exception: pass
            try: API_DAILY_BUDGET = max(9000, len(_API_KEYS) * 9500)
            except Exception: pass
        return jsonify({
            "ok": True,
            "file": API_KEYS_FILE,
            "count": len(_API_KEYS),
            "keys_masked": [_api_keys_mask(k) for k in _API_KEYS],
            "msg": f"{len(_API_KEYS)} chave(s) carregada(s) em api_keys.txt sem reiniciar backend.",
        })
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500

@app.route("/api/api_keys/reload", methods=["POST", "GET"])
def api_api_keys_reload_safe():
    global _API_KEYS, _API_KEY_INDEX, _API_KEY_FAILED, API_DAILY_BUDGET
    with _API_KEY_LOCK:
        _API_KEYS = _api_carregar_chaves()
        _API_KEY_INDEX = 0
        try: _API_KEY_FAILED.clear()
        except Exception: pass
        try: API_DAILY_BUDGET = max(9000, len(_API_KEYS) * 9500)
        except Exception: pass
    return jsonify({"ok": True, "file": API_KEYS_FILE, "count": len(_API_KEYS), "keys_masked": [_api_keys_mask(k) for k in _API_KEYS], "msg": "API keys recarregadas do api_keys.txt sem reiniciar."})

_API_KEYS_SAFE_PAGE = '<!doctype html>\n<html lang="pt-BR">\n<head>\n  <meta charset="utf-8">\n  <meta name="viewport" content="width=device-width, initial-scale=1">\n  <title>THON Toolkit - API Keys</title>\n  <style>\n    :root{--bg:#07090b;--card:#10151c;--text:#eafff3;--muted:#91a99d;--green:#35f28f;--line:#23342a;--red:#ff5b5b;}\n    *{box-sizing:border-box}body{margin:0;background:#07090b;color:var(--text);font-family:Inter,system-ui,-apple-system,Segoe UI,Arial,sans-serif;min-height:100vh;padding:28px;background-image:radial-gradient(circle at 20% 0%,rgba(53,242,143,.12),transparent 34%)}\n    .wrap{max-width:980px;margin:0 auto}.top{display:flex;align-items:flex-start;justify-content:space-between;gap:16px;margin-bottom:18px}.brand{font-size:26px;font-weight:950;letter-spacing:-.04em}.muted{color:var(--muted);font-size:14px;margin-top:4px}.card{background:rgba(16,21,28,.94);border:1px solid var(--line);border-radius:18px;padding:20px;box-shadow:0 18px 60px rgba(0,0,0,.32)}\n    textarea{width:100%;min-height:320px;background:#050708;color:#dffff0;border:1px solid #263b2e;border-radius:14px;padding:16px;font:14px/1.45 ui-monospace,SFMono-Regular,Menlo,Consolas,monospace;outline:none;resize:vertical}textarea:focus{border-color:var(--green);box-shadow:0 0 0 3px rgba(53,242,143,.13)}\n    .row{display:flex;gap:10px;flex-wrap:wrap;align-items:center;margin-top:14px}.btn,button{border:0;border-radius:12px;padding:12px 16px;font-weight:850;cursor:pointer;text-decoration:none;display:inline-flex;align-items:center;gap:8px}.primary{background:var(--green);color:#021107}.ghost{background:#18231d;color:var(--text);border:1px solid #2b4033}.danger{background:#31161a;color:#ffd9d9;border:1px solid #583039}.pill{border:1px solid #294432;background:#0b120f;color:#bfffd8;border-radius:999px;padding:8px 12px;font-size:13px}.msg{margin-top:14px;white-space:pre-wrap;font-family:ui-monospace,SFMono-Regular,Menlo,Consolas,monospace;font-size:13px;color:#bdfbd3}.keys{display:grid;gap:8px;margin-top:14px}.key{padding:10px 12px;border-radius:12px;background:#080d0a;border:1px solid #1f3127;color:#bdfbd3;font-family:ui-monospace,SFMono-Regular,Menlo,Consolas,monospace;font-size:13px}.hint{margin:12px 0 0;color:var(--muted);font-size:13px;line-height:1.5}.warn{color:#ffd166}\n  </style>\n</head>\n<body>\n  <div class="wrap">\n    <div class="top">\n      <div><div class="brand">THON Toolkit - API Keys</div><div class="muted">Pagina isolada. Nao altera index, dashboard, prospector nem CRM.</div></div>\n      <div class="row" style="margin:0"><a class="btn ghost" href="/">Toolkit</a><a class="btn ghost" href="/prospector">Prospector</a></div>\n    </div>\n    <div class="card">\n      <div class="row" style="margin-top:0;margin-bottom:12px"><span class="pill" id="count">carregando...</span><span class="pill" id="active">backend...</span><span class="pill">arquivo: api_keys.txt</span></div>\n      <textarea id="keys" spellcheck="false" placeholder="Cole aqui uma chave por linha\\n\\nCHAVE_1\\nCHAVE_2\\nCHAVE_3"></textarea>\n      <div class="row"><button class="primary" onclick="saveKeys()">Salvar chaves</button><button class="ghost" onclick="loadKeys()">Recarregar</button><button class="danger" onclick="clearBox()">Limpar caixa</button></div>\n      <p class="hint"><b>Formato:</b> uma chave por linha, sem virgula e sem aspas. Ao salvar, o backend recarrega o <code>api_keys.txt</code> na hora.</p>\n      <p class="hint warn">Essa pagina nao injeta botao e nao mexe nos seus HTMLs.</p>\n      <div class="msg" id="msg"></div>\n      <div class="keys" id="masked"></div>\n    </div>\n  </div>\n<script>\nasync function loadKeys(){\n  const msg=document.getElementById(\'msg\'); msg.textContent=\'Carregando...\';\n  try{\n    const r=await fetch(\'/api/api_keys\'); const j=await r.json();\n    document.getElementById(\'keys\').value=j.text||\'\';\n    document.getElementById(\'count\').textContent=(j.count||0)+\' chave(s) no arquivo\';\n    document.getElementById(\'active\').textContent=(j.active_count||0)+\' ativa(s) no backend\';\n    document.getElementById(\'masked\').innerHTML=(j.keys_masked||[]).map((k,i)=>\'<div class="key">\'+String(i+1).padStart(2,\'0\')+\'. \'+k+\'</div>\').join(\'\');\n    msg.textContent=\'OK.\';\n  }catch(e){ msg.textContent=\'Erro carregando: \'+e; }\n}\nasync function saveKeys(){\n  const text=document.getElementById(\'keys\').value;\n  const msg=document.getElementById(\'msg\'); msg.textContent=\'Salvando...\';\n  try{\n    const r=await fetch(\'/api/api_keys\',{method:\'POST\',headers:{\'Content-Type\':\'application/json\'},body:JSON.stringify({text})});\n    const j=await r.json();\n    if(!j.ok) throw new Error(j.erro||\'falha\');\n    document.getElementById(\'count\').textContent=(j.count||0)+\' chave(s) no arquivo\';\n    document.getElementById(\'active\').textContent=(j.count||0)+\' ativa(s) no backend\';\n    document.getElementById(\'masked\').innerHTML=(j.keys_masked||[]).map((k,i)=>\'<div class="key">\'+String(i+1).padStart(2,\'0\')+\'. \'+k+\'</div>\').join(\'\');\n    msg.textContent=j.msg||\'Salvo.\';\n  }catch(e){ msg.textContent=\'Erro salvando: \'+e; }\n}\nfunction clearBox(){ document.getElementById(\'keys\').value=\'\'; }\nloadKeys();\n</script>\n</body>\n</html>'

@app.route("/api-keys")
@app.route("/config/api-keys")
def api_keys_safe_page():
    return _API_KEYS_SAFE_PAGE, 200, {"Content-Type": "text/html; charset=utf-8"}

@app.route("/global_status_widget.js")
def global_status_widget():
    return f"window.THON_BACKEND_STATUS={{ok:true,version:'{APP_VERSION}',engine:'{ENGINE_VERSION}'}};", 200, {"Content-Type": "application/javascript; charset=utf-8"}

@app.route("/thon_engine_selector.js")
def thon_engine_selector_js():
    return send_from_directory(".", "thon_engine_selector.js")

@app.route("/prospector/stop", methods=["POST"])
def prospector_stop():
    auto_hunt["enabled"] = False
    with lock:
        estado["rodando"] = False
        estado["pausa"] = False
        estado["status"] = "idle"
        estado["pipeline_stage"] = "parado"
        estado["msg"] = "Parado pelo usuário"
    return jsonify({"ok": True})

@app.route("/app/shutdown", methods=["POST"])
def app_shutdown():
    token_env = os.environ.get("THON_DESKTOP_SHUTDOWN_TOKEN", "")
    token_req = request.headers.get("X-THON-Shutdown-Token", "")
    if token_env and token_req != token_env:
        return jsonify({"ok": False, "erro": "token inválido"}), 403
    auto_hunt["enabled"] = False
    with lock:
        estado["rodando"] = False
        estado["pausa"] = False
        estado["status"] = "idle"
        estado["pipeline_stage"] = "shutdown"
        estado["msg"] = "Encerrando aplicativo"
        salvar_lote_ativo(estado.get("lote", []))
        salvar_memoria(estado.get("vistos", set()))
        salvar_aprovados(estado.get("aprovados", []))

    def _exit_later():
        time.sleep(0.25)
        os._exit(0)

    threading.Thread(target=_exit_later, daemon=True).start()
    return jsonify({"ok": True, "msg": "shutdown agendado"})

@app.route("/prospector/pause", methods=["POST"])
def prospector_pause():
    with lock:
        estado["pausa"] = True
        estado["status"] = "pausado"
        estado["msg"] = "Pausado"
    return jsonify({"ok": True})

@app.route("/prospector/resume", methods=["POST"])
def prospector_resume():
    with lock:
        estado["pausa"] = False
        if estado.get("rodando"):
            estado["status"] = "buscando"
        estado["msg"] = "Retomado"
    return jsonify({"ok": True})

@app.route("/diagnostico/v58_api")
def diagnostico_v58_api():
    return jsonify({
        "ok": True,
        "version": APP_VERSION,
        "engine": ENGINE_VERSION,
        "engine_mode": estado.get("engine_mode", carregar_engine_mode()),
        "engine_mode_label": _modo_label(estado.get("engine_mode", carregar_engine_mode())),
        "banco": diagnostico_banco(),
        "api_state": _api_load_state(),
        "chaves_ativas": len(_API_KEYS),
        "api_quota_used": _api_quota_used(),
        "api_quota_budget": API_DAILY_BUDGET,
        "status": estado.get("status", "idle"),
        "pipeline_stage": estado.get("pipeline_stage", "idle"),
        "msg": estado.get("msg", ""),
        "ultimo_erro": estado.get("api_last_error", ""),
        "brutos_total": len(_carregar_brutos()),
    })

@app.route("/api/brutos")
def api_brutos():
    """Retorna a lista de canais brutos (com status, query, etc.)"""
    page = max(1, int(request.args.get("page", 1)))
    per_page = max(1, min(500, int(request.args.get("per_page", 100))))
    filtro = request.args.get("status", "todos")  # todos, bruto, qualificado, aprovado, reprovado
    brutos = _carregar_brutos()
    if filtro != "todos":
        brutos = [c for c in brutos if c.get("status") == filtro]
    total = len(brutos)
    start = (page - 1) * per_page
    return jsonify({
        "ok": True,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": (total + per_page - 1) // per_page,
        "canais": brutos[start:start+per_page],
        "filtro": filtro,
    })

@app.route("/api/fila_pendente", methods=["GET"])
def api_fila_pendente_get():
    with lock:
        estado["lote"] = carregar_lote_ativo()
        return jsonify({
            "ok": True,
            "file": LOTE_ATIVO_FILE,
            "total": len(estado.get("lote", []) or []),
            "lote": estado.get("lote", []) or [],
            "version": APP_VERSION,
        })

@app.route("/api/fila_pendente/reconstruir", methods=["POST"])
def api_fila_pendente_reconstruir():
    data = request.json or {}
    limit = _safe_int(data.get("limit", 500), 500, 1, 5000)
    with lock:
        lote = reconstruir_lote_dos_brutos(limit=limit)
        estado["lote"] = lote
        estado["status"] = "aguardando" if lote else estado.get("status", "idle")
        estado["pausa"] = bool(lote)
    return jsonify({"ok": True, "total": len(lote), "file": LOTE_ATIVO_FILE})

@app.route("/api/fila_pendente/limpar", methods=["POST"])
def api_fila_pendente_limpar():
    # Limpeza explícita; não é chamada automaticamente por ciclo novo.
    with lock:
        estado["lote"] = salvar_lote_ativo([])
        estado["pausa"] = False
        estado["status"] = "idle"
    return jsonify({"ok": True, "cleared": True, "file": LOTE_ATIVO_FILE})




# ============================================================
# PATCH V58.13 — DLP SMART FALLBACK baseado no benchmark real
# Resultado do benchmark: pesado=40 queries, per_query=50, workers=2,
# 700 candidatos únicos, 7 qualificados, 0x 403 final.
# Regra: default_minus_websafari -> android -> no_cache -> base -> tv_simply
# ============================================================
APP_VERSION = "v58.13-dlp-smart-fallback"

DLP_YTDLP_VARIANTS = {
    "default_minus_websafari": ["--extractor-args", "youtube:player_client=default,-web_safari"],
    "android": ["--extractor-args", "youtube:player_client=android"],
    "no_cache": [],
    "base": [],
    "tv_simply": ["--extractor-args", "youtube:player_client=tv_simply"],
}
DLP_SEARCH_FALLBACK_ORDER = ["default_minus_websafari", "android", "no_cache", "base", "tv_simply"]

DLP_EXTRA_TEMPLATES = [
    "{nicho} podcast", "podcast {nicho} brasil", "{nicho} podcast brasileiro", "entrevista {nicho}",
    "{nicho} entrevista podcast", "bate papo {nicho}", "cast {nicho}", "canal {nicho} podcast",
    "{nicho} youtube", "{nicho} canal", "{nicho} criador de conteudo", "criador {nicho} brasil",
    "youtube {nicho} brasil", "especialista {nicho} youtube", "consultor {nicho} youtube", "mentor {nicho} podcast",
    "podcast pequeno {nicho}", "canal pequeno {nicho}", "podcast independente {nicho}", "bate papo {nicho} brasil",
    "entrevista sobre {nicho} brasil", "canal {nicho} brasil",
]
DLP_GENERAL_EXPANSIONS = [
    "empresario podcast brasil", "ceo podcast brasil", "fundador podcast brasil", "startup podcast brasil",
    "negocios podcast brasil", "marketing digital podcast brasil", "profissional liberal podcast",
    "medico podcast brasil", "advogado podcast brasil", "contador podcast brasil", "nutricionista podcast brasil",
    "psicologo podcast brasil", "fisioterapeuta podcast brasil", "dentista podcast brasil", "consultor podcast brasil", "mentor podcast brasil",
]

def _dlp_variant_chain(prefer=None):
    chain = []
    prefer = str(prefer or "").strip()
    if prefer and prefer in DLP_YTDLP_VARIANTS:
        chain.append(prefer)
    for v in DLP_SEARCH_FALLBACK_ORDER:
        if v not in chain:
            chain.append(v)
    return chain

def _thon_ytdlp_base_cmd(extra_variant=None):
    cmd = [sys.executable, "-m", "yt_dlp", "--no-warnings", "--quiet", "--no-cache-dir"]
    cookie_browser = os.environ.get("THON_YTDLP_COOKIE_BROWSER", "none").strip().lower()
    if cookie_browser and cookie_browser not in {"none", "off", "0", "false"}:
        cmd += ["--cookies-from-browser", cookie_browser]
    variant = extra_variant or "default_minus_websafari"
    cmd += DLP_YTDLP_VARIANTS.get(variant, [])
    return cmd

def run_cmd_raw(args, timeout=60, variant=None):
    try:
        r = subprocess.run(_thon_ytdlp_base_cmd(variant) + args,
                           capture_output=True, text=True, timeout=timeout)
        stderr = (r.stderr or "")
        out = (r.stdout or "").strip()
        is_403 = ("403" in stderr or "Forbidden" in stderr)
        if r.returncode != 0:
            if is_403:
                _registrar_403(stderr)
            elif stderr.strip():
                print(f"  [yt-dlp erro] variant={variant or 'default'} | {stderr.strip()[:160]}")
            return {"ok": False, "out": "", "err": stderr, "code": r.returncode, "is_403": is_403, "variant": variant or "default_minus_websafari"}
        if out:
            _403_state["count"] = 0
        return {"ok": True, "out": out, "err": stderr, "code": r.returncode, "is_403": False, "variant": variant or "default_minus_websafari"}
    except subprocess.TimeoutExpired:
        print(f"  [yt-dlp timeout] variant={variant or 'default'} pulado")
        return {"ok": False, "out": "", "err": "TIMEOUT", "code": 124, "is_403": False, "variant": variant or "default_minus_websafari"}
    except Exception as e:
        print(f"  [yt-dlp exception] variant={variant or 'default'} | {e}")
        return {"ok": False, "out": "", "err": str(e), "code": 999, "is_403": False, "variant": variant or "default_minus_websafari"}

def run_cmd(args, timeout=60, variant=None):
    res = run_cmd_raw(args, timeout=timeout, variant=variant)
    return res["out"] if res.get("ok") else ""

def run_cmd_fallback(args, timeout=60, prefer=None, require_output=True):
    best = {"ok": False, "out": "", "variant": prefer or "default_minus_websafari", "is_403": False, "code": 0, "err": ""}
    for variant in _dlp_variant_chain(prefer):
        res = run_cmd_raw(args, timeout=timeout, variant=variant)
        if res.get("ok") and (res.get("out") or not require_output):
            return res.get("out", ""), variant, res
        if res.get("out") and len(res.get("out")) > len(best.get("out", "")):
            best = res
        if res.get("is_403"):
            print(f"  [yt-dlp fallback] variant={variant} deu 403; tentando próxima")
    return best.get("out", ""), best.get("variant") or (prefer or "default_minus_websafari"), best

def run_lines(args, timeout=60, variant=None, fallback=False):
    if fallback:
        out, used_variant, _res = run_cmd_fallback(args, timeout=timeout, prefer=variant)
    else:
        out = run_cmd(args, timeout=timeout, variant=variant)
    return [l for l in out.splitlines() if l.strip()]

def gerar_queries_dlp_fallback(nichos, extras=None, limit=40):
    qs = []
    for q in (extras or []):
        q = _api_clean_query(q)
        if q:
            qs.append((q, "manual"))
    for nicho in nichos or ["empreendedorismo"]:
        n = str(nicho or "").strip().lower()
        if not n:
            continue
        for tpl in DLP_EXTRA_TEMPLATES:
            qs.append((_api_clean_query(tpl.format(nicho=n)), n))
        if "empreendedor" in n or "negocio" in n or "negócio" in n:
            for q in DLP_GENERAL_EXPANSIONS:
                qs.append((_api_clean_query(q), n))
    seen, out = set(), []
    for q, nicho in qs:
        if not q or q in seen:
            continue
        seen.add(q)
        out.append((q, nicho))
        if len(out) >= int(limit or 40):
            break
    return out

def _dlp_parse_search_output(out, q, nicho, variant, banco, seen):
    itens = []
    for linha in (out or "").splitlines():
        partes = linha.split("\t")
        if len(partes) < 2:
            continue
        cid = (partes[0] or "").strip()
        nome = (partes[1] or "").strip()
        if not cid or cid == "NA" or len(cid) < 8:
            continue
        if cid in seen or cid in banco:
            continue
        text = nome.lower()
        if any(k in text for k in NEGATIVOS_FORTES_API):
            continue
        itens.append({
            "id": cid, "channel_id": cid, "nome": nome, "title": nome,
            "url": f"https://youtube.com/channel/{cid}", "nicho": nicho,
            "query": q, "found_query": q, "source": "yt_dlp_search",
            "search_variant": variant, "yt_dlp_variant": variant,
        })
    return itens

def _dlp_search_query_smart(q, nicho, per_query, banco, seen):
    best_items, best_variant = [], "default_minus_websafari"
    for variant in _dlp_variant_chain("default_minus_websafari"):
        res = run_cmd_raw(["--flat-playlist", "--print", "%(channel_id)s\t%(uploader)s", f"ytsearch{per_query}:{q}"], timeout=65, variant=variant)
        itens = _dlp_parse_search_output(res.get("out", ""), q, nicho, variant, banco, seen)
        if len(itens) > len(best_items):
            best_items, best_variant = itens, variant
        if itens and not res.get("is_403"):
            if variant != "default_minus_websafari":
                print(f"    [fallback ok] {variant} salvou query | +{len(itens)}")
            return itens, variant, res
        if res.get("is_403"):
            print(f"    [403] {variant} falhou; tentando fallback")
    return best_items, best_variant, {"code": 0, "is_403": False}

def _dlp_buscar_canais(config):
    queries = obter_queries_para_engine(config, limit=config.get("query_limit") or 40, formato="tuplas", fallback_kind="dlp")
    target = int(config.get("candidate_target") or 100)
    per_query = max(5, min(50, int(config.get("dlp_search_per_query") or 50)))
    banco = ids_banco_dados()
    seen = set()
    candidatos = []
    print(f"\n[DLP] BUSCA 100% yt-dlp SMART | meta={target} canais | queries={len(queries)} | per_query={per_query}")
    for qi, (q, nicho) in enumerate(queries, start=1):
        if not estado.get("rodando") or len(candidatos) >= target:
            break
        with lock:
            estado["queries_processadas"] = qi
            estado["pipeline_stage"] = "dlp_search"
            estado["msg"] = f"DLP buscando {qi}/{len(queries)}: {q[:45]}"
        print(f"  [dlp search] {qi}/{len(queries)} {q!r}")
        before = len(candidatos)
        itens, used_variant, _res = _dlp_search_query_smart(q, nicho, per_query, banco, seen)
        for item in itens:
            cid = item.get("id")
            if not cid or cid in seen or cid in banco:
                continue
            seen.add(cid)
            candidatos.append(item)
            _adicionar_bruto({**item, "status": "bruto_dlp", "search_variant": used_variant})
            if len(candidatos) >= target:
                break
        print(f"    variant={used_variant} | +{len(candidatos)-before} | total={len(candidatos)}")
        with lock:
            estado["candidatos_encontrados"] = len(candidatos)
        if not _sleep_interrompivel(0.35):
            break
    return candidatos

def pegar_info_canal_dlp(channel_id, variant=None):
    url = f"https://www.youtube.com/channel/{channel_id}"
    used = variant or "default_minus_websafari"
    try:
        result, used, _res = run_cmd_fallback(["--dump-single-json", "--flat-playlist", "--playlist-end", "1", url], timeout=55, prefer=variant)
        if result:
            data = json.loads(result)
            subs = int(data.get("channel_follower_count") or data.get("subscriber_count") or 0)
            desc = data.get("description") or ""
            uploader = data.get("uploader") or data.get("channel") or ""
            if subs > 0:
                return {"subs": subs, "descricao": desc, "uploader": uploader, "variant": used}
    except Exception as e:
        print(f"  [dlp info canal erro] {channel_id}: {e}")

    try:
        vid_out, used2, _res2 = run_cmd_fallback(["--flat-playlist", "--print", "%(id)s", "--playlist-end", "1", url], timeout=45, prefer=used)
        vid = (vid_out.splitlines()[0].strip() if vid_out.strip() else "")
        if vid and vid != "NA":
            video_json, used3, _res3 = run_cmd_fallback(["--dump-json", f"https://youtube.com/watch?v={vid}"], timeout=65, prefer=used2)
            if video_json:
                data = json.loads(video_json)
                subs = int(data.get("channel_follower_count") or 0)
                desc = data.get("channel_description") or data.get("description") or ""
                uploader = data.get("uploader") or data.get("channel") or ""
                return {"subs": subs, "descricao": desc, "uploader": uploader, "variant": used3}
    except Exception as e:
        print(f"  [dlp info video erro] {channel_id}: {e}")
    return {"subs": 0, "descricao": "", "uploader": "", "variant": used}

def pegar_videos_canal(channel_id, variant=None):
    url = f"https://www.youtube.com/channel/{channel_id}"
    out, used_variant, _res = run_cmd_fallback(["--flat-playlist", "--print", "%(duration)s\t%(view_count)s\t%(upload_date)s",
                        "--playlist-end", "15", url], timeout=65, prefer=variant)
    linhas = [l for l in out.splitlines() if l.strip()]
    if not linhas:
        return None
    longos = medios = shorts = total_views = total_videos = 0
    datas = []
    for linha in linhas:
        partes = linha.split("\t")
        if len(partes) >= 2:
            try:
                dur = int(float(partes[0])) if partes[0] else 0
                views = int(float(partes[1])) if len(partes) > 1 and partes[1] else 0
                data = partes[2] if len(partes) > 2 else ""
                if dur > 600:
                    longos += 1
                    if views > 0:
                        total_views += views
                        total_videos += 1
                elif dur > 180:
                    medios += 1
                    if views > 0:
                        total_views += views
                        total_videos += 1
                elif 0 < dur <= 60:
                    shorts += 1
                if data and len(data) >= 8:
                    datas.append(data[:8])
            except:
                continue
    avg_views = total_views // total_videos if total_videos > 0 else 0
    dias_ultimo = 999
    if datas:
        datas.sort(reverse=True)
        try:
            dt = datetime.strptime(datas[0], "%Y%m%d")
            dias_ultimo = (datetime.now() - dt).days
        except:
            pass
    return {"longos": longos, "shorts": shorts, "avg_views": avg_views, "dias_ultimo": dias_ultimo, "variant": used_variant}

def dlp_verify_candidates(candidatos, config, source_label="api_then_dlp"):
    verify_max = int(config.get("ytdlp_verify_max") or len(candidatos) or 0)
    if bool(config.get("verify_until_empty", False)):
        verify_max = len(candidatos)
    candidatos = list(candidatos or [])[:max(0, verify_max)]
    lote = []
    reprovados = []
    print(f"\n[DLP VERIFY SMART] verificação de fila | source={source_label} | canais={len(candidatos)}")
    for i, c in enumerate(candidatos, start=1):
        if not estado.get("rodando"):
            break
        cid = c.get("id") or c.get("channel_id")
        if not cid:
            continue
        preferred_variant = c.get("search_variant") or c.get("yt_dlp_variant") or "default_minus_websafari"
        with lock:
            estado["pipeline_stage"] = "dlp_verify_queue"
            estado["msg"] = f"DLP verificando fila {i}/{len(candidatos)}: {c.get('nome','')[:35]}"
            estado["verificados"] = i
            estado["ytdlp_verified"] = int(estado.get("ytdlp_verified", 0) or 0) + 1
        info = pegar_info_canal_dlp(cid, variant=preferred_variant)
        videos = pegar_videos_canal(cid, variant=info.get("variant") or preferred_variant) or {"longos": 0, "shorts": 0, "avg_views": 0, "dias_ultimo": 999, "variant": info.get("variant") or preferred_variant}
        score, motivos = _dlp_score(c, info, videos, config)
        c.update({
            "subs": int(info.get("subs") or 0), "subscriber_count": int(info.get("subs") or 0),
            "subs_fmt": fmt_numero(info.get("subs") or 0), "description": info.get("descricao", ""),
            "longos": videos.get("longos", 0), "shorts": videos.get("shorts", 0),
            "recent_avg_views": videos.get("avg_views", 0), "ultimo": formatar_dias(videos.get("dias_ultimo", 999)),
            "last_video_days": videos.get("dias_ultimo", 999), "score": score,
            "score_reasons": source_label, "engine_mode": source_label, "source": source_label,
            "search_variant": preferred_variant, "verify_variant": videos.get("variant") or info.get("variant") or preferred_variant,
        })
        if not motivos:
            item = _canal_api_para_lote(c)
            item["source"] = source_label
            item["engine_mode"] = source_label
            item["status_fila"] = "pendente"
            item["search_variant"] = c.get("search_variant")
            item["verify_variant"] = c.get("verify_variant")
            lote.append(item)
            _atualizar_status_bruto(cid, "qualificado_dlp", {"score": score, "subs": c.get("subs"), "longos": c.get("longos"), "verify_variant": c.get("verify_variant")})
            print(f"  [DLP QUALIFICADO] score={score} | {c.get('subs_fmt')} | {c.get('nome')} | longos={c.get('longos')} avg={fmt_numero(c.get('recent_avg_views') or 0)} | search={c.get('search_variant')} verify={c.get('verify_variant')}")
        else:
            c["motivo"] = "; ".join(motivos)
            reprovados.append(_canal_api_para_reprovado(c))
            _atualizar_status_bruto(cid, "reprovado_dlp", {"motivo": c["motivo"], "score": score, "verify_variant": c.get("verify_variant")})
        with lock:
            estado["qualificados"] = len(lote)
        if not _sleep_interrompivel(0.20):
            break
    return lote, reprovados

def dlp_full_process(config):
    target = int(config.get("candidate_target") or 100)
    approval_goal = int(config.get("approval_goal") or 0)
    # V58.30: respeita query_limit do frontend (nao forca default 60)
    queries_per_cycle = int(config.get("query_limit") or config.get("dlp_queries_per_cycle") or 100)
    max_total_queries = int(config.get("max_total_queries") or 500)
    used_queries = set()
    all_lote, all_reprovados = [], []
    ciclo = 0
    result = {}
    print(f"[DLP LOOP] target={target}, q/ciclo={queries_per_cycle} (do frontend), max_total={max_total_queries}")
    temp_cfg = dict(config or {})
    temp_cfg["query_limit"] = queries_per_cycle
    while True:
        if approval_goal and len(all_lote) >= approval_goal:
            print(f"[DLP LOOP] approval_goal atingido={approval_goal}")
            break
        if len(used_queries) >= max_total_queries:
            print("[DLP LOOP] max_total_queries atingido")
            break
        ciclo += 1
        used_before = len(used_queries)
        try:
            result = _v58_15_run_external_engine(temp_cfg, candidates=None, mode='search_verify')
        except Exception as e:
            print(f"[DLP LOOP] erro ciclo {ciclo}: {e}")
            break
        # V58.25: aceita multiplas chaves (engine externa pode usar 'qualified' ou 'qualificados')
        qualificados_raw = result.get('qualificados') or result.get('qualified') or result.get('lote') or result.get('approved') or []
        reprovados_raw = result.get('reprovados') or result.get('rejected') or result.get('reproved') or []
        print(f"[DLP LOOP] ciclo={ciclo} | engine retornou: qualificados_raw={len(qualificados_raw)} reprovados_raw={len(reprovados_raw)} | chaves_disponiveis={list(result.keys())[:8]}")
        lote = [_v58_15_normalizar_lote(x, 'external_dlp_smart') for x in qualificados_raw]
        lote = [x for x in lote if x]
        reprovados = [_v58_15_normalizar_reprovado(x, 'external_dlp_smart') for x in reprovados_raw]
        reprovados = [x for x in reprovados if x]
        print(f"[DLP LOOP] ciclo={ciclo} | apos normalizar: lote={len(lote)} reprovados={len(reprovados)}")
        # V58.33: enrich de country ANTES do filtro (busca snippet.country na API)
        # Aumenta max_per_call pra cobrir todos os qualificados (antes era so 20)
        if lote and _API_KEYS:
            try:
                _enrich_country_canais(lote, max_per_call=100)
            except Exception as e:
                print(f"[DLP] aviso enrich country: {e}")
        # V58.32: FILTRO DE PAIS ANTES DO FILTRO DE IDIOMA (mais confiavel)
        lote, reprovados_pais = _filtrar_canais_por_pais(lote, temp_cfg, source_label="dlp_filtro_pais")
        for rp in reprovados_pais:
            reprovados.append(_v58_15_normalizar_reprovado(rp, 'external_dlp_smart_filtro_pais'))
        # V58.24: filtro anti-gringo por idioma (camada extra)
        lote_filtrado = []
        for c in lote:
            info_c = {"descricao": c.get("description") or "", "uploader": c.get("nome") or ""}
            gringo_reprovado, gringo_motivo, _idioma = _filtrar_gringo_por_territorio(c, info_c, temp_cfg)
            if gringo_reprovado:
                c["motivo"] = gringo_motivo
                reprovados.append(_v58_15_normalizar_reprovado(c, 'external_dlp_smart_filtro_idioma'))
                print(f"  [FILTRO IDIOMA] reprovado {c.get('nome','')[:35]}: {gringo_motivo}")
            else:
                lote_filtrado.append(c)
        lote = lote_filtrado
        all_lote.extend(lote)
        all_reprovados.extend(reprovados)
        for c in result.get('candidatos') or []:
            cid = str(c.get('id') or c.get('channel_id') or '').strip()
            if not cid:
                continue
            if '_adicionar_bruto' in globals():
                _adicionar_bruto({**c, 'status': 'bruto_dlp_loop'})
            if '_atualizar_status_bruto' in globals():
                _atualizar_status_bruto(cid, 'bruto_dlp_loop', {})
        batch_queries = result.get('queries_used') or result.get('queries') or []
        # V58.30: registra resultados das queries no query_factory V2 (saturacao)
        try:
            if query_factory_get_next_queries:
                from query_factory.query_factory import registrar_resultado_query as _v30_registrar
                # Tenta pegar metricas por query (se engine externa retornar)
                queries_metrics = result.get('queries_metrics') or {}
                sat_count = 0
                for q_item in batch_queries:
                    q_str = str(q_item.get('query') if isinstance(q_item, dict) else q_item or '').strip()
                    if not q_str:
                        continue
                    used_queries.add(q_str)
                    # Pega metricas da query se disponivel, senao estimativa
                    qm = queries_metrics.get(q_str, {}) if isinstance(queries_metrics, dict) else {}
                    novos_q = int(qm.get('novos', 0) or 0)
                    rept_q = int(qm.get('repetidos', 0) or 0)
                    quals_q = int(qm.get('qualificados', 0) or 0)
                    # Se nao tem metrica por query, usa estimativa agregada
                    if not queries_metrics:
                        # divide igualmente (estimativa)
                        n_batch = max(1, len(batch_queries))
                        novos_q = len(result.get('candidatos') or []) // n_batch
                    sat = _v30_registrar(q_str, novos_q, rept_q, quals_q)
                    if sat:
                        sat_count += 1
                        print(f"  [SATURACAO V2 DLP] {q_str!r} → {sat['motivo']} ({sat['cooldown_seg']}s)")
                if sat_count:
                    print(f"  [SATURACAO V2 DLP] {sat_count} queries saturadas neste ciclo")
        except Exception as _e_sat_dlp:
            print(f"  [SATURACAO V2 DLP] aviso: {_e_sat_dlp}")
        print(f"[DLP LOOP] ciclo={ciclo} | +{len(lote)} qtd, +{len(reprovados)} reprovados | usadas={len(used_queries)} | total={len(all_lote)}")
        if len(all_lote) >= target or len(used_queries) >= max_total_queries:
            break
        try:
            _v5818_time.sleep(0.4)
        except Exception:
            pass
    metrics = {}
    try:
        metrics = result.get('metrics') or {}
    except Exception:
        pass
    try:
        for c in all_lote:
            cid = c.get('id') or c.get('channel_id')
            if not cid:
                continue
            if '_atualizar_status_bruto' in globals():
                _atualizar_status_bruto(cid, 'qualificado_dlp_external', {'score': c.get('score'), 'subs': c.get('subs'), 'longos': c.get('longos')})
    except Exception as e:
        print(f"[DLP EXTERNAL] aviso salvando brutos: {e}")
    with lock:
        estado['candidatos_brutos_encontrados'] = int(metrics.get('raw_count') or estado.get('candidatos_brutos_encontrados', 0) or 0)
        estado['removidos_por_blacklist'] = int(metrics.get('removed_by_blacklist') or estado.get('removidos_por_blacklist', 0) or 0)
        estado['candidatos_encontrados'] = int(metrics.get('unique_after_blacklist') or estado.get('candidatos_encontrados', 0) or 0)
        estado['dlp_external_metrics'] = metrics
        estado['ytdlp_verified'] = int(estado.get('ytdlp_verified', 0) or 0)
        estado['ytdlp_approved'] = int(estado.get('ytdlp_approved', 0) or 0) + len(all_lote)
    print(f"[DLP LOOP] fim | qualificados={len(all_lote)}, reprovados={len(all_reprovados)}, queries={len(used_queries)}")
    # V58.25: garante que os qualificados vao pro lote de caca (camada extra de seguranca)
    if all_lote:
        try:
            sync_report = sync_api_qualificados_para_lote(all_lote, source="dlp_full_process_external")
            print(f"[DLP LOOP] sync_api_qualificados_para_lote | novos={sync_report['novos_adicionados']} | ja_existiam={sync_report['ja_existiam']} | aprovados_ignorados={sync_report['aprovados_ignorados']} | fila_total={sync_report['depois']}")
        except Exception as e:
            print(f"[DLP LOOP] ERRO sync_api_qualificados_para_lote: {e}")
    return all_lote, all_reprovados


# ===== V58.15 EXTERNAL DLP ENGINE + API ROTATION PATCH START =====
# Backend-only patch: DLP usa engine externa igual ao teste de terminal que funcionou.
DLP_EXTERNAL_ENGINE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "thon_dlp_engine.py")
DLP_EXTERNAL_LAST_INPUT = "thon_dlp_engine_last_input.json"
DLP_EXTERNAL_LAST_OUTPUT = "thon_dlp_engine_last_output.json"
DLP_EXTERNAL_LAST_LOG = "thon_dlp_engine_last.log"
_dlp_external_proc = None
_dlp_external_proc_lock = threading.RLock()

def _v58_15_external_dlp_config(config):
    cfg = dict(config or {})
    # V58.30: respeita query_limit do frontend (nao forca max(40))
    cfg["query_limit"] = _safe_int(cfg.get("query_limit", 100), 100, 5, 600) if '_safe_int' in globals() else int(cfg.get("query_limit") or 100)
    cfg["per_query"] = _safe_int(cfg.get("per_query", cfg.get("dlp_search_per_query", 50)), 50, 5, 100) if '_safe_int' in globals() else int(cfg.get("per_query") or cfg.get("dlp_search_per_query") or 50)
    cfg["workers"] = _safe_int(cfg.get("workers", 2), 2, 1, 4) if '_safe_int' in globals() else int(cfg.get("workers") or 2)
    cfg["verify"] = _safe_int(cfg.get("verify", cfg.get("verify_por_ciclo", cfg.get("ytdlp_verify_max", 30))), 30, 1, 500) if '_safe_int' in globals() else int(cfg.get("verify") or cfg.get("ytdlp_verify_max") or 30)
    # V58.30: removido o max(40) que forçava mínimo 40 queries no DLP
    # Agora respeita o que o frontend manda (campo "Queries máximas")
    return cfg

def _v58_15_blacklist_ids():
    try:
        ids = set()
        if 'ids_banco_dados' in globals():
            ids |= {str(x) for x in ids_banco_dados() if str(x).strip()}
        else:
            ids |= {str(x) for x in estado.get('vistos', set()) if str(x).strip()}
            ids |= {str(c.get('id')) for c in estado.get('aprovados', []) if isinstance(c, dict) and c.get('id')}
            ids |= {str(c.get('id')) for c in estado.get('reprovados', []) if isinstance(c, dict) and c.get('id')}
        return sorted(ids)
    except Exception as e:
        print(f"[DLP EXTERNAL] blacklist falhou: {e}")
        return []

def _v58_15_normalizar_lote(item, source_label="external_dlp_smart"):
    if not isinstance(item, dict):
        return None
    # V58.25: aceita multiplas chaves de ID (engine externa pode usar nomes diferentes)
    cid = str(item.get('id') or item.get('channel_id') or item.get('channelId') or item.get('canal_id') or '').strip()
    if not cid:
        # tenta extrair de URL
        url = str(item.get('url') or item.get('channel_url') or '')
        m = re.search(r"(UC[0-9A-Za-z_-]{18,})", url)
        if m:
            cid = m.group(1)
    if not cid:
        return None
    nome = item.get('nome') or item.get('title') or 'Canal'
    try:
        subs = int(item.get('subs') or item.get('subscriber_count') or 0)
    except Exception:
        subs = 0
    return {
        **item,
        'id': cid,
        'channel_id': cid,
        'nome': nome,
        'title': item.get('title') or nome,
        'url': item.get('url') or f"https://youtube.com/channel/{cid}",
        'subs': subs,
        'subscriber_count': subs,
        'subs_fmt': item.get('subs_fmt') or (fmt_numero(subs) if 'fmt_numero' in globals() else str(subs)),
        'score': int(item.get('score') or 0),
        'nicho': item.get('nicho', ''),
        'query': item.get('query') or item.get('found_query') or '',
        'found_query': item.get('found_query') or item.get('query') or '',
        'source': source_label,
        'engine_mode': source_label,
        'status_fila': 'pendente',
        'search_variant': item.get('search_variant') or 'default_minus_websafari',
        'verify_variant': item.get('verify_variant') or item.get('search_variant') or 'default_minus_websafari',
    }

def _v58_15_normalizar_reprovado(item, source_label="external_dlp_smart"):
    if not isinstance(item, dict):
        return None
    cid = str(item.get('id') or item.get('channel_id') or '').strip()
    if not cid:
        return None
    nome = item.get('nome') or item.get('title') or 'Canal'
    return {
        'id': cid,
        'channel_id': cid,
        'nome': nome,
        'title': item.get('title') or nome,
        'url': item.get('url') or f"https://youtube.com/channel/{cid}",
        'nicho': item.get('nicho', ''),
        'score': int(item.get('score') or 0),
        'subs_fmt': item.get('subs_fmt') or '',
        'motivo': item.get('motivo') or item.get('erro') or 'não qualificado DLP externo',
        'query': item.get('query') or item.get('found_query') or '',
        'source': source_label,
        'engine': ENGINE_VERSION if 'ENGINE_VERSION' in globals() else 'external_dlp_smart',
        'data': str(datetime.now()),
    }

def _v58_15_run_external_engine(config, candidates=None, mode="search_verify"):
    global _dlp_external_proc
    app_dir = os.path.dirname(os.path.abspath(__file__))
    engine_file = os.path.join(app_dir, "thon_dlp_engine.py")
    if not os.path.exists(engine_file):
        raise RuntimeError(f"Engine externa não encontrada: {engine_file}")

    cfg = _v58_15_external_dlp_config(config)

    # FIX QUERY FACTORY: passa queries geradas pelo factory para a engine externa
    # Antes era sempre [] — a engine gerava as próprias queries ignorando o factory
    factory_queries = []
    if mode in ("search_verify",):
        try:
            factory_limit = int(cfg.get("query_limit") or 60)
            raw = obter_queries_para_engine(config, limit=factory_limit, formato="tuplas", fallback_kind="dlp")
            factory_queries = [{"query": q, "nicho": nicho} for q, nicho in raw if q]
            print(f"[QUERY FACTORY → DLP EXTERNAL] {len(factory_queries)} queries injetadas na engine externa")
        except Exception as e:
            print(f"[QUERY FACTORY → DLP EXTERNAL] falhou ao gerar queries: {e} — engine vai gerar as próprias")

    payload = {
        'mode': mode,
        'config': cfg,
        'queries': factory_queries,
        'candidates': candidates or [],
        'blacklist_ids': _v58_15_blacklist_ids(),
    }
    if mode == 'verify_only':
        payload['config']['ignore_blacklist_on_verify'] = True
        payload['config']['verify_until_empty'] = True

    in_path = os.path.join(app_dir, DLP_EXTERNAL_LAST_INPUT)
    out_path = os.path.join(app_dir, DLP_EXTERNAL_LAST_OUTPUT)
    log_path = os.path.join(app_dir, DLP_EXTERNAL_LAST_LOG)
    with open(in_path, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    cmd = [sys.executable, engine_file, '--input', in_path, '--output', out_path]
    print(f"[DLP EXTERNAL] chamando engine externa | mode={mode} | config={{queries:{cfg.get('query_limit')}, per_query:{cfg.get('per_query')}, verify:{cfg.get('verify')}, workers:{cfg.get('workers')}}}")
    logs = []
    with _dlp_external_proc_lock:
        _dlp_external_proc = subprocess.Popen(cmd, cwd=app_dir, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, bufsize=1)
    try:
        with open(log_path, 'w', encoding='utf-8') as lf:
            for line in _dlp_external_proc.stdout:
                line = line.rstrip('\n')
                logs.append(line)
                lf.write(line + '\n')
                lf.flush()
                print(line)
                try:
                    with lock:
                        estado['pipeline_stage'] = 'dlp_external_engine'
                        estado['msg'] = line[-180:]
                except Exception:
                    pass
        code = _dlp_external_proc.wait()
    finally:
        with _dlp_external_proc_lock:
            _dlp_external_proc = None

    if code != 0:
        raise RuntimeError(f"Engine DLP externa falhou com código {code}. Veja {log_path}")
    if not os.path.exists(out_path):
        raise RuntimeError(f"Engine DLP externa não gerou output: {out_path}")
    with open(out_path, 'r', encoding='utf-8') as f:
        result = json.load(f)
    result.setdefault('logs_tail', logs[-80:])
    return result

def _dlp_full_process_interno(config):
    """Fallback interno quando thon_dlp_engine.py não está disponível."""
    candidatos = _dlp_buscar_canais(config)
    lote = []
    reprovados = []
    print(f"\n[DLP INTERNO] verificando {len(candidatos)} candidatos")
    for i, c in enumerate(candidatos, start=1):
        if not estado.get("rodando"):
            break
        cid = c.get("id") or c.get("channel_id")
        if not cid:
            continue
        with lock:
            estado["pipeline_stage"] = "dlp_verify"
            estado["msg"] = f"DLP verificando {i}/{len(candidatos)}: {c.get('nome','')[:35]}"
            estado["verificados"] = i
            estado["ytdlp_verified"] = int(estado.get("ytdlp_verified", 0) or 0) + 1
        info = pegar_info_canal_dlp(cid)
        videos = pegar_videos_canal(cid) or {"longos": 0, "shorts": 0, "avg_views": 0, "dias_ultimo": 999}
        score, motivos = _dlp_score(c, info, videos, config)
        c.update({
            "subs": int(info.get("subs") or 0),
            "subscriber_count": int(info.get("subs") or 0),
            "subs_fmt": fmt_numero(info.get("subs") or 0),
            "description": info.get("descricao", ""),
            "longos": videos.get("longos", 0),
            "shorts": videos.get("shorts", 0),
            "recent_avg_views": videos.get("avg_views", 0),
            "ultimo": formatar_dias(videos.get("dias_ultimo", 999)),
            "last_video_days": videos.get("dias_ultimo", 999),
            "score": score,
            "source": "yt_dlp_interno",
            "engine_mode": "dlp",
        })
        if not motivos:
            item = _canal_api_para_lote(c)
            item["source"] = "yt_dlp_interno"
            item["engine_mode"] = "dlp"
            item["status_fila"] = "pendente"
            # FIX BUG 2: garantir que id está preenchido antes de ir pro lote
            if not item.get("id") and item.get("channel_id"):
                item["id"] = item["channel_id"]
            lote.append(item)
            _atualizar_status_bruto(cid, "qualificado_dlp", {"score": score, "subs": c.get("subs"), "longos": c.get("longos")})
            print(f"  [DLP QUALIFICADO] score={score} | {c.get('subs_fmt')} | {c.get('nome')} | longos={c.get('longos')}")
        else:
            c["motivo"] = "; ".join(motivos)
            reprovados.append(_canal_api_para_reprovado(c))
            _atualizar_status_bruto(cid, "reprovado_dlp", {"motivo": c["motivo"], "score": score})
        with lock:
            estado["qualificados"] = len(lote)
        if not _sleep_interrompivel(0.25):
            break
    return lote, reprovados


def dlp_verify_candidates(candidatos, config, source_label="api_then_dlp"):
    """Override V58.15: verifica fila via engine externa; fallback interno se não existir."""
    app_dir = os.path.dirname(os.path.abspath(__file__))
    engine_file = os.path.join(app_dir, "thon_dlp_engine.py")
    if not os.path.exists(engine_file):
        print(f"[DLP] thon_dlp_engine.py não encontrado — usando verificação DLP interna")
        lote = []
        reprovados = []
        for i, c in enumerate(list(candidatos or []), start=1):
            if not estado.get("rodando"):
                break
            cid = c.get("id") or c.get("channel_id")
            if not cid:
                continue
            with lock:
                estado["pipeline_stage"] = "dlp_verify_queue"
                estado["msg"] = f"DLP verificando fila {i}/{len(candidatos)}: {c.get('nome','')[:35]}"
                estado["verificados"] = i
                estado["ytdlp_verified"] = int(estado.get("ytdlp_verified", 0) or 0) + 1
            info = pegar_info_canal_dlp(cid)
            videos = pegar_videos_canal(cid) or {"longos": 0, "shorts": 0, "avg_views": 0, "dias_ultimo": 999}
            score, motivos = _dlp_score(c, info, videos, config)
            c.update({
                "subs": int(info.get("subs") or 0),
                "subscriber_count": int(info.get("subs") or 0),
                "subs_fmt": fmt_numero(info.get("subs") or 0),
                "longos": videos.get("longos", 0),
                "shorts": videos.get("shorts", 0),
                "recent_avg_views": videos.get("avg_views", 0),
                "score": score, "source": source_label, "engine_mode": source_label,
            })
            if not motivos:
                item = _canal_api_para_lote(c)
                item["source"] = source_label
                item["status_fila"] = "pendente"
                if not item.get("id") and item.get("channel_id"):
                    item["id"] = item["channel_id"]
                lote.append(item)
                _atualizar_status_bruto(cid, "qualificado_dlp", {"score": score})
                print(f"  [DLP QUALIFICADO] score={score} | {c.get('subs_fmt')} | {c.get('nome')}")
            else:
                c["motivo"] = "; ".join(motivos)
                reprovados.append(_canal_api_para_reprovado(c))
                _atualizar_status_bruto(cid, "reprovado_dlp", {"motivo": c["motivo"]})
            with lock:
                estado["qualificados"] = len(lote)
            if not _sleep_interrompivel(0.25):
                break
        return lote, reprovados
    try:
        result = _v58_15_run_external_engine(config, candidates=list(candidatos or []), mode='verify_only')
    except Exception as e:
        print(f"[DLP EXTERNAL verify] falhou ({e}) — retornando lote vazio para não travar")
        return [], []
    metrics = result.get('metrics') or {}
    lote = [_v58_15_normalizar_lote(x, source_label) for x in (result.get('qualificados') or [])]
    lote = [x for x in lote if x]
    # FIX BUG 2: garantir id preenchido
    for item in lote:
        if not item.get("id") and item.get("channel_id"):
            item["id"] = item["channel_id"]
    reprovados = [_v58_15_normalizar_reprovado(x, source_label) for x in (result.get('reprovados') or [])]
    reprovados = [x for x in reprovados if x]
    with lock:
        estado['dlp_external_metrics'] = metrics
        estado['ytdlp_verified'] = int(estado.get('ytdlp_verified', 0) or 0) + int(metrics.get('verificados') or 0)
        estado['ytdlp_approved'] = int(estado.get('ytdlp_approved', 0) or 0) + len(lote)
    return lote, reprovados

def dlp_verify_queue_until_empty(config):
    """Override V58.15: processa fila DLP persistente com engine externa."""
    fila = carregar_fila_dlp_verificacao() if 'carregar_fila_dlp_verificacao' in globals() else []
    if not fila:
        return [], []
    lote, reprovados = dlp_verify_candidates(fila, {**dict(config or {}), 'verify_until_empty': True, 'verify': len(fila), 'ytdlp_verify_max': len(fila)}, source_label='api_then_external_dlp')
    ids_processados = {c.get('id') for c in (lote + reprovados) if c.get('id')}
    if ids_processados and 'remover_fila_dlp_verificacao_ids' in globals():
        remover_fila_dlp_verificacao_ids(ids_processados)
    return lote, reprovados

def _v58_15_kill_external_engine():
    global _dlp_external_proc
    with _dlp_external_proc_lock:
        proc = _dlp_external_proc
        _dlp_external_proc = None
    if proc and proc.poll() is None:
        try:
            proc.terminate()
            time.sleep(0.5)
            if proc.poll() is None:
                proc.kill()
            print("[DLP EXTERNAL] processo externo encerrado")
        except Exception as e:
            print(f"[DLP EXTERNAL] erro ao encerrar processo: {e}")

def _v58_15_stop_response():
    try:
        auto_hunt['enabled'] = False
    except Exception:
        pass
    _v58_15_kill_external_engine()
    with lock:
        estado['rodando'] = False
        estado['pausa'] = False
        estado['status'] = 'idle'
        estado['pipeline_stage'] = 'parado'
        estado['msg'] = 'Parado pelo usuário; engine externa encerrada'
        try:
            salvar_lote_ativo(estado.get('lote', []))
            salvar_memoria(estado.get('vistos', set()))
            salvar_aprovados(estado.get('aprovados', []))
        except Exception:
            pass
    return jsonify({'ok': True, 'msg': 'Prospector parado e engine externa encerrada', 'auto_hunt': auto_hunt if 'auto_hunt' in globals() else {}})

def parar_v58_15_external():
    return _v58_15_stop_response()

def prospector_stop_v58_15_external():
    return _v58_15_stop_response()

try:
    app.view_functions['parar'] = parar_v58_15_external
except Exception as e:
    print(f"[DLP EXTERNAL] aviso: não consegui sobrescrever /parar: {e}")
try:
    app.view_functions['prospector_stop'] = prospector_stop_v58_15_external
except Exception as e:
    print(f"[DLP EXTERNAL] aviso: não consegui sobrescrever /prospector/stop: {e}")

@app.route('/diagnostico/v58_15_dlp_external')
def diagnostico_v58_15_dlp_external():
    with lock:
        return jsonify({
            'ok': True,
            'version': APP_VERSION if 'APP_VERSION' in globals() else 'v58.14',
            'engine_file': DLP_EXTERNAL_ENGINE_FILE,
            'engine_exists': os.path.exists(DLP_EXTERNAL_ENGINE_FILE),
            'last_input': DLP_EXTERNAL_LAST_INPUT,
            'last_output': DLP_EXTERNAL_LAST_OUTPUT,
            'last_log': DLP_EXTERNAL_LAST_LOG,
            'rodando': bool(estado.get('rodando')),
            'pipeline_stage': estado.get('pipeline_stage'),
            'msg': estado.get('msg'),
            'candidatos_brutos_encontrados': estado.get('candidatos_brutos_encontrados', 0),
            'removidos_por_blacklist': estado.get('removidos_por_blacklist', 0),
            'candidatos_encontrados': estado.get('candidatos_encontrados', 0),
            'dlp_external_metrics': estado.get('dlp_external_metrics', {}),
        })


# ------------------------------------------------------------
# V58.15 API ROTATION POR CHAVE
# Corrige o bug: quota global 8913/9000 travava antes de tentar as 4 chaves.
# Agora cada key tem estado local próprio; se uma falha/quota, roda para a próxima.
# ------------------------------------------------------------
import urllib.parse, hashlib
API_KEY_ROTATION_STATE_FILE = "youtube_api_key_rotation_state.json"
API_PER_KEY_DAILY_BUDGET = int(os.environ.get("THON_API_PER_KEY_DAILY_BUDGET", "9500"))
try:
    API_DAILY_BUDGET = max(9000, API_PER_KEY_DAILY_BUDGET * max(1, len(_API_KEYS)))
except Exception:
    API_DAILY_BUDGET = max(9000, API_PER_KEY_DAILY_BUDGET)

_API_ROTATION_LOCK = threading.RLock()

def _v58_15_api_today():
    return datetime.now().strftime("%Y-%m-%d")

def _v58_15_api_key_uid(key):
    key = str(key or "").strip()
    kid = _api_key_id(key) if '_api_key_id' in globals() else (key[:8] + '...' + key[-6:])
    digest = hashlib.sha1(key.encode('utf-8')).hexdigest()[:10] if key else 'empty'
    return f"{kid}|{digest}"

def _v58_15_api_load_rotation_state():
    try:
        if os.path.exists(API_KEY_ROTATION_STATE_FILE):
            with open(API_KEY_ROTATION_STATE_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            if data.get('date') == _v58_15_api_today():
                data.setdefault('keys', {})
                data.setdefault('history', [])
                return data
    except Exception as e:
        print(f"[API ROTATION] erro lendo state: {e}")
    return {'date': _v58_15_api_today(), 'keys': {}, 'history': []}

def _v58_15_api_save_rotation_state(data):
    try:
        data['date'] = _v58_15_api_today()
        data['updated_at'] = str(datetime.now())
        hist = data.get('history') or []
        data['history'] = hist[-200:]
        with open(API_KEY_ROTATION_STATE_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"[API ROTATION] erro salvando state: {e}")

def _v58_15_api_cost(endpoint, cost=None):
    if cost is not None:
        try:
            return int(cost)
        except Exception:
            pass
    return API_SEARCH_COST if endpoint == 'search' else API_OTHER_COST

def _v58_15_api_keys_available_raw():
    global _API_KEYS
    try:
        if not _API_KEYS:
            _API_KEYS = _api_carregar_chaves()
    except Exception:
        pass
    return list(_API_KEYS or [])

def _v58_15_api_rec(state, key):
    uid = _v58_15_api_key_uid(key)
    rec = state.setdefault('keys', {}).setdefault(uid, {
        'key_id': _api_key_id(key) if '_api_key_id' in globals() else uid,
        'used': 0,
        'calls': 0,
        'ok_calls': 0,
        'failed_calls': 0,
        'exhausted_today': False,
        'cooldown_until': 0,
        'last_error': '',
        'last_status': None,
        'last_endpoint': '',
        'last_used_at': '',
    })
    rec.setdefault('key_id', _api_key_id(key) if '_api_key_id' in globals() else uid)
    rec.setdefault('used', 0)
    rec.setdefault('calls', 0)
    rec.setdefault('ok_calls', 0)
    rec.setdefault('failed_calls', 0)
    rec.setdefault('exhausted_today', False)
    rec.setdefault('cooldown_until', 0)
    return uid, rec

def _v58_15_api_key_is_available(rec, cost):
    try:
        if bool(rec.get('exhausted_today')):
            return False
        if float(rec.get('cooldown_until') or 0) > time.time():
            return False
        if int(rec.get('used') or 0) + int(cost or 0) > API_PER_KEY_DAILY_BUDGET:
            return False
        return True
    except Exception:
        return True

def _v58_15_api_pick_key(cost):
    global _API_KEY_INDEX
    keys = _v58_15_api_keys_available_raw()
    if not keys:
        raise RuntimeError(f"Nenhuma chave API carregada. Crie {API_KEYS_FILE} na pasta do app.")
    with _API_ROTATION_LOCK:
        state = _v58_15_api_load_rotation_state()
        for _ in range(len(keys)):
            idx = _API_KEY_INDEX % len(keys)
            key = keys[idx]
            _API_KEY_INDEX += 1
            uid, rec = _v58_15_api_rec(state, key)
            if _v58_15_api_key_is_available(rec, cost):
                _v58_15_api_save_rotation_state(state)
                return key, uid, rec
        _v58_15_api_save_rotation_state(state)
        return None, None, None

def _v58_15_api_mark_attempt(key, endpoint, cost, ok=False, status=None, reason='', message=''):
    with _API_ROTATION_LOCK:
        state = _v58_15_api_load_rotation_state()
        uid, rec = _v58_15_api_rec(state, key)
        rec['calls'] = int(rec.get('calls') or 0) + 1
        rec['last_status'] = status
        rec['last_endpoint'] = endpoint
        rec['last_used_at'] = str(datetime.now())
        rec['last_error'] = str(reason or message or '')[:240]
        try:
            rec['used'] = int(rec.get('used') or 0) + int(cost or 0)
        except Exception:
            pass
        if ok:
            rec['ok_calls'] = int(rec.get('ok_calls') or 0) + 1
            rec['cooldown_until'] = 0
        else:
            rec['failed_calls'] = int(rec.get('failed_calls') or 0) + 1
            rs = str(reason or '').lower()
            msg = str(message or '').lower()
            if 'quota' in rs or 'dailylimit' in rs or 'ratelimit' in rs or 'quota' in msg:
                rec['exhausted_today'] = True
            else:
                rec['cooldown_until'] = time.time() + 60
        state.setdefault('history', []).append({
            'time': str(datetime.now()),
            'key_id': rec.get('key_id'),
            'endpoint': endpoint,
            'cost': cost,
            'ok': ok,
            'status': status,
            'reason': reason,
            'used_key': rec.get('used'),
        })
        _v58_15_api_save_rotation_state(state)
        return rec

def _api_quota_used():
    """Override: soma quota local de todas as chaves do dia, sem usar o arquivo global velho como trava."""
    try:
        st = _v58_15_api_load_rotation_state()
        return sum(int((v or {}).get('used') or 0) for v in (st.get('keys') or {}).values())
    except Exception:
        return 0

def _api_consume_quota(cost, endpoint):
    """Compat: não deixa o contador global antigo travar a API antes do rodízio."""
    return True, _api_quota_used()

def youtube_api_get(endpoint, params, cost=None, quiet_404=False, max_tentativas=None):
    """V58.15: YouTube API com rotação real por chave.
    Antes: _api_consume_quota() travava em 8913/9000 antes de testar as outras keys.
    Agora: escolhe key disponível, tenta, se quota/403 troca key, e só para se todas falharem.
    """
    cost = _v58_15_api_cost(endpoint, cost)
    keys = _v58_15_api_keys_available_raw()
    total_chaves = len(keys)
    if not total_chaves:
        raise RuntimeError(f"Nenhuma chave API carregada. Crie {API_KEYS_FILE} na pasta do app.")
    tentativas = total_chaves if max_tentativas is None else max(1, min(int(max_tentativas or 1), total_chaves))
    last_error = ''
    for tentativa in range(tentativas):
        key, uid, rec = _v58_15_api_pick_key(cost)
        if not key:
            break
        params_copy = dict(params or {})
        params_copy['key'] = key
        url = 'https://www.googleapis.com/youtube/v3/' + endpoint + '?' + urllib.parse.urlencode(params_copy)
        curl_bin = shutil.which('curl') or '/usr/bin/curl'
        kid = rec.get('key_id') if rec else (_api_key_id(key) if '_api_key_id' in globals() else 'key')
        try:
            out = subprocess.check_output(
                [curl_bin, '-sS', '-L', '-w', '\n__HTTP_STATUS__:%{http_code}', url],
                stderr=subprocess.STDOUT,
                timeout=45,
            ).decode('utf-8', errors='replace')
            marker = '\n__HTTP_STATUS__:'
            if marker in out:
                body, status_raw = out.rsplit(marker, 1)
                try:
                    status = int(status_raw.strip())
                except Exception:
                    status = 0
            else:
                body, status = out, 0
            try:
                data = json.loads(body) if body.strip() else {}
            except Exception:
                data = {'raw': body[:1000]}
            if status >= 400:
                reason = ''
                msg = ''
                try:
                    err = data.get('error', {}) if isinstance(data, dict) else {}
                    msg = err.get('message', '')
                    reason = err.get('errors', [{}])[0].get('reason', '')
                except Exception:
                    pass
                reason_raw = str(reason or '')
                reason_l = reason_raw.lower()
                msg_l = str(msg or '').lower()
                last_error = f"HTTP {status} {reason} {msg}".strip()
                if status == 404 and quiet_404:
                    _v58_15_api_mark_attempt(key, endpoint, cost, ok=True, status=status, reason=reason, message=msg)
                    return {'items': [], '_http_status': 404, '_error': msg or '404'}

                # Erros de permissão/privacidade da fonte NÃO significam chave esgotada.
                # Ex.: subscriptionForbidden, activities privadas, playlist privada, canal não encontrado.
                non_key_reasons = {
                    'subscriptionforbidden','forbidden','notfound','playlistnotfound','videonotfound','channelnotfound',
                    'playlistforbidden','privateplaylist','private','autherror','insufficientpermissions'
                }
                quota_reasons = {'quotaexceeded','dailylimitexceeded','ratelimitexceeded','userratelimitexceeded','keyinvalid','accessnotconfigured'}
                is_quota = reason_l in quota_reasons or 'quota' in reason_l or 'daily limit' in msg_l or 'rate limit' in msg_l
                is_source_forbidden = reason_l in non_key_reasons or 'not allowed to access' in msg_l or 'requester is not allowed' in msg_l or 'private' in msg_l
                if is_source_forbidden and not is_quota:
                    _v58_15_api_mark_attempt(key, endpoint, cost, ok=True, status=status, reason=reason, message=msg)
                    print(f"[API SKIP] {endpoint} {reason or status}: fonte/canal sem permissão; chave preservada")
                    return {'items': [], '_http_status': status, '_error': msg or reason or f'HTTP {status}', '_non_key_error': True}

                _v58_15_api_mark_attempt(key, endpoint, cost, ok=False, status=status, reason=reason, message=msg)
                if is_quota or status == 403:
                    print(f"[API ROTATION] key {kid} falhou ({reason or 'HTTP '+str(status)}); trocando chave {tentativa+1}/{tentativas}")
                    continue
                print(f"[api erro] {endpoint} HTTP {status} | {reason} | {str(msg)[:180]}")
                return {'items': [], '_http_status': status, '_error': msg or reason or f'HTTP {status}'}
            _v58_15_api_mark_attempt(key, endpoint, cost, ok=True, status=status)
            if isinstance(data, dict):
                data['_api_key_id'] = kid
                data['_api_rotation_used'] = _api_quota_used()
                data['_api_rotation_budget'] = API_DAILY_BUDGET
            return data
        except subprocess.TimeoutExpired:
            _v58_15_api_mark_attempt(key, endpoint, cost, ok=False, status=0, reason='timeout', message='timeout')
            print(f"[api timeout] {endpoint} key={kid}; tentando próxima")
            last_error = 'timeout'
            continue
        except Exception as e:
            _v58_15_api_mark_attempt(key, endpoint, cost, ok=False, status=0, reason='exception', message=str(e))
            print(f"[api exception] {endpoint} key={kid}: {e}")
            last_error = str(e)
            continue
    raise RuntimeError(f"Todas as chaves da API falharam/esgotaram. Último erro: {last_error}")

def _v58_15_api_health_payload():
    keys = _v58_15_api_keys_available_raw()
    st = _v58_15_api_load_rotation_state()
    items = []
    for key in keys:
        uid, rec = _v58_15_api_rec(st, key)
        cooldown_left = max(0, int(float(rec.get('cooldown_until') or 0) - time.time()))
        items.append({
            'key_id': rec.get('key_id'),
            'used': int(rec.get('used') or 0),
            'budget': API_PER_KEY_DAILY_BUDGET,
            'remaining_local': max(0, API_PER_KEY_DAILY_BUDGET - int(rec.get('used') or 0)),
            'calls': int(rec.get('calls') or 0),
            'ok_calls': int(rec.get('ok_calls') or 0),
            'failed_calls': int(rec.get('failed_calls') or 0),
            'available': _v58_15_api_key_is_available(rec, 1),
            'exhausted_today': bool(rec.get('exhausted_today')),
            'cooldown_left_s': cooldown_left,
            'last_error': rec.get('last_error') or '',
            'last_status': rec.get('last_status'),
            'last_endpoint': rec.get('last_endpoint') or '',
        })
    _v58_15_api_save_rotation_state(st)
    return {
        'ok': True,
        'version': 'v58.15-api-rotation',
        'keys_total': len(keys),
        'used_total': _api_quota_used(),
        'budget_total_local': API_DAILY_BUDGET,
        'per_key_budget_local': API_PER_KEY_DAILY_BUDGET,
        'state_file': API_KEY_ROTATION_STATE_FILE,
        'keys': items,
    }

def api_health_v58_15():
    return jsonify(_v58_15_api_health_payload())

def api_rotation_reset_v58_15():
    data = {'date': _v58_15_api_today(), 'keys': {}, 'history': [], 'reset_at': str(datetime.now())}
    _v58_15_api_save_rotation_state(data)
    try:
        with open(API_STATE_FILE, 'w', encoding='utf-8') as f:
            json.dump({'date': _v58_15_api_today(), 'used': 0, 'calls': 0, 'history': [], 'reset_by': 'v58.15'}, f, ensure_ascii=False, indent=2)
    except Exception:
        pass
    return jsonify({'ok': True, 'msg': 'Estado local de rotação API resetado', **_v58_15_api_health_payload()})

try:
    app.add_url_rule('/api/api_health', 'api_health_v58_15', api_health_v58_15, methods=['GET'])
except Exception as e:
    print(f"[API ROTATION] aviso rota api_health: {e}")
try:
    app.add_url_rule('/api/api_rotation/reset', 'api_rotation_reset_v58_15', api_rotation_reset_v58_15, methods=['POST'])
except Exception as e:
    print(f"[API ROTATION] aviso rota reset: {e}")

@app.route('/diagnostico/v58_15_api_rotation')
def diagnostico_v58_15_api_rotation():
    return jsonify(_v58_15_api_health_payload())

# ===== V58.15 EXTERNAL DLP ENGINE + API ROTATION PATCH END =====



# === THON V58.17 ENGINE LIVE ROUTES ===
# Rotas somente de leitura para mostrar o que a engine externa esta fazendo.
# Nao altera HTML, CRM, Dashboard, Prospector nem regras de negocio.
try:
    from flask import Response as _thon_Response, jsonify as _thon_jsonify
except Exception:
    _thon_Response = None
    _thon_jsonify = None

import os as _thon_os_v5817
import json as _thon_json_v5817

_THON_APP_DIR_V5817 = _thon_os_v5817.path.dirname(_thon_os_v5817.path.abspath(__file__))

def _thon_v5817_read_json(name, default=None):
    if default is None:
        default = {}
    path = _thon_os_v5817.path.join(_THON_APP_DIR_V5817, name)
    try:
        if not _thon_os_v5817.path.exists(path) or _thon_os_v5817.path.getsize(path) == 0:
            return default
        with open(path, "r", encoding="utf-8") as f:
            return _thon_json_v5817.load(f)
    except Exception as e:
        return {"erro": str(e), "arquivo": name}

def _thon_v5817_tail(name="thon_engine_live.log", n=160):
    path = _thon_os_v5817.path.join(_THON_APP_DIR_V5817, name)
    try:
        if not _thon_os_v5817.path.exists(path):
            return []
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            lines = f.readlines()[-int(n):]
        return [x.rstrip("\n") for x in lines]
    except Exception as e:
        return ["erro lendo log: " + str(e)]

@app.route('/api/engine/status')
def thon_v5817_engine_status():
    status = _thon_v5817_read_json("thon_engine_status.json", {})
    last = _thon_v5817_read_json("thon_engine_last_result.json", {})
    search = _thon_v5817_read_json("thon_search_result.json", {})
    verify = _thon_v5817_read_json("thon_verify_result.json", {})
    payload = {
        "status": status,
        "last_result_stats": (last.get("stats") if isinstance(last, dict) else None),
        "search_stats": {
            "brutos_encontrados": search.get("brutos_encontrados") if isinstance(search, dict) else None,
            "blacklist_removidos": search.get("blacklist_removidos") if isinstance(search, dict) else None,
            "novos_para_verificar": search.get("novos_para_verificar") if isinstance(search, dict) else None,
        },
        "verify_stats": {
            "processados": verify.get("processados") if isinstance(verify, dict) else None,
            "qualificados_total": verify.get("qualificados_total") if isinstance(verify, dict) else None,
            "reprovados_total": verify.get("reprovados_total") if isinstance(verify, dict) else None,
        },
        "log_tail": _thon_v5817_tail(),
    }
    return _thon_jsonify(payload)

@app.route('/api/engine/log')
def thon_v5817_engine_log():
    if _thon_Response is None:
        return "flask Response indisponivel", 500
    return _thon_Response("\n".join(_thon_v5817_tail(n=300)), mimetype="text/plain; charset=utf-8")

@app.route('/engine-live')
def thon_v5817_engine_live():
    if _thon_Response is None:
        return "flask Response indisponivel", 500
    html = r"""
<!doctype html>
<html lang="pt-BR">
<head>
<meta charset="utf-8" />
<meta name="viewport" content="width=device-width,initial-scale=1" />
<title>THON Engine Live</title>
<style>
body{margin:0;background:#090d0b;color:#e8f5ec;font-family:Inter,system-ui,-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;padding:24px}.wrap{max-width:1280px;margin:0 auto}.top{display:flex;gap:12px;align-items:center;justify-content:space-between;flex-wrap:wrap;margin-bottom:18px}.brand{font-weight:900;font-size:22px}.nav a{color:#b7ffd0;text-decoration:none;margin-right:12px;font-weight:700}.grid{display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:12px}.card{background:#111913;border:1px solid #263429;border-radius:16px;padding:16px;box-shadow:0 10px 30px rgba(0,0,0,.25)}.num{font-size:34px;font-weight:900;line-height:1}.label{font-size:12px;color:#93a69a;text-transform:uppercase;letter-spacing:.08em;margin-top:5px}.wide{grid-column:1/-1}.two{display:grid;grid-template-columns:1fr 1fr;gap:12px}.bar{height:12px;border-radius:999px;background:#1e2a22;overflow:hidden}.fill{height:100%;background:#92ffb1;width:0%}.muted{color:#9bad9f}.ok{color:#92ffb1}.bad{color:#ff9d9d}.list{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:8px}.item{padding:10px;border:1px solid #223128;border-radius:12px;background:#0c120f}.log{white-space:pre-wrap;font-family:ui-monospace,SFMono-Regular,Menlo,Monaco,Consolas,"Liberation Mono",monospace;font-size:12px;line-height:1.45;max-height:460px;overflow:auto;background:#050806;border-radius:14px;padding:14px;border:1px solid #223128}button{background:#b7ffd0;color:#061009;border:0;padding:10px 14px;border-radius:12px;font-weight:900;cursor:pointer}@media(max-width:900px){.grid{grid-template-columns:1fr 1fr}.two{grid-template-columns:1fr}.list{grid-template-columns:1fr}}@media(max-width:560px){.grid{grid-template-columns:1fr}}
</style>
</head>
<body>
<div class="wrap">
  <div class="top">
    <div><div class="brand">THON Engine Live</div><div class="muted">Busca externa + verificação externa em tempo real</div></div>
    <div class="nav"><a href="/">Index</a><a href="/dashboard">Dashboard</a><a href="/prospector">Prospector</a><a href="/crm">CRM</a><a href="/api/engine/status">JSON</a></div>
  </div>
  <div class="grid">
    <div class="card"><div class="num" id="brutos">0</div><div class="label">Brutos encontrados</div></div>
    <div class="card"><div class="num" id="blacklist">0</div><div class="label">Já existiam / blacklist</div></div>
    <div class="card"><div class="num" id="novos">0</div><div class="label">Novos para verificar</div></div>
    <div class="card"><div class="num ok" id="qualificados">0</div><div class="label">Qualificados</div></div>
    <div class="card wide"><b id="msg">Carregando...</b><div class="muted" id="upd"></div><br><div class="bar"><div class="fill" id="fill"></div></div><div class="muted" id="progress"></div></div>
    <div class="card wide two">
      <div><h3>Último candidato verificado</h3><div id="current" class="item muted">...</div></div>
      <div><h3>Qualificados recentes</h3><div id="qlist" class="list"></div></div>
    </div>
    <div class="card wide two">
      <div><h3>Novos encontrados</h3><div id="newlist" class="list"></div></div>
      <div><h3>Já existiam / removidos por blacklist</h3><div id="oldlist" class="list"></div></div>
    </div>
    <div class="card wide"><h3>Log ao vivo</h3><div class="log" id="log"></div></div>
  </div>
</div>
<script>
function val(o,p,d=0){try{return p.split('.').reduce((a,k)=>a&&a[k],o)??d}catch(e){return d}}
function item(x){return `<div class="item"><b>${x.nome||x.title||x.id||'sem nome'}</b><br><span class="muted">${x.subs_fmt||''} ${x.status||''} ${x.score?('score '+x.score):''}</span><br><span class="muted">${x.motivo||x.query||''}</span></div>`}
async function tick(){
 const r=await fetch('/api/engine/status?ts='+Date.now()); const d=await r.json(); const s=d.status||{};
 const br=s.brutos_encontrados??val(d,'search_stats.brutos_encontrados',0); const bl=s.blacklist_removidos??val(d,'search_stats.blacklist_removidos',0); const nv=s.novos_para_verificar??val(d,'search_stats.novos_para_verificar',0); const q=s.verify_qualificados??val(d,'verify_stats.qualificados_total',0); const rp=s.verify_reprovados??val(d,'verify_stats.reprovados_total',0);
 document.getElementById('brutos').textContent=br||0; document.getElementById('blacklist').textContent=bl||0; document.getElementById('novos').textContent=nv||0; document.getElementById('qualificados').textContent=q||0;
 document.getElementById('msg').textContent=s.mensagem||s.etapa||'sem status'; document.getElementById('upd').textContent='Atualizado: '+(s.updated_at||'');
 const vp=s.verify_processados||0, vt=s.verify_total||0, qp=s.query_processadas||0, qt=s.query_total||0; let pct=0, txt='';
 if(vt){pct=Math.round((vp/vt)*100); txt=`Verificação ${vp}/${vt} | qualificados ${q} | reprovados ${rp}`}
 else if(qt){pct=Math.round((qp/qt)*100); txt=`Busca ${qp}/${qt}`}
 document.getElementById('fill').style.width=Math.min(100,pct)+'%'; document.getElementById('progress').textContent=txt;
 document.getElementById('current').innerHTML=s.candidato_atual?item(s.candidato_atual):'...';
 document.getElementById('qlist').innerHTML=(s.preview_qualificados||[]).slice(-8).map(item).join('')||'<span class="muted">Nenhum ainda</span>';
 document.getElementById('newlist').innerHTML=(s.preview_novos||[]).slice(0,10).map(item).join('')||'<span class="muted">...</span>';
 document.getElementById('oldlist').innerHTML=(s.preview_existentes||[]).slice(0,10).map(item).join('')||'<span class="muted">...</span>';
 document.getElementById('log').textContent=(d.log_tail||[]).join('\n');
}
setInterval(tick,2000); tick();
</script>
</body></html>
"""
    return _thon_Response(html, mimetype="text/html; charset=utf-8")
# === /THON V58.17 ENGINE LIVE ROUTES ===


# === THON V58.18 UI JSON CONSOLE SAFE START ===
# Patch de integração/visibilidade. Não altera API rotation nem algoritmo das engines.
import os as _v5818_os
import json as _v5818_json
import time as _v5818_time
import datetime as _v5818_datetime
import shutil as _v5818_shutil
from flask import jsonify as _v5818_jsonify, request as _v5818_request, Response as _v5818_Response

_THON_V5818_APP_DIR = _v5818_os.path.dirname(_v5818_os.path.abspath(__file__))
_THON_V5818_FILES = [
    "fila_pendente_api.json",
    "dlp_verification_queue.json",
    "canais_brutos_api.json",
    "thon_dlp_engine_last_output.json",
    "thon_search_result.json",
    "thon_verify_result.json",
    "winchester_aprovados.json",
    "winchester_reprovados.json",
    "winchester_vistos.json",
    "youtube_api_quota_state.json",
]

def _v5818_now():
    return _v5818_datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def _v5818_path(name):
    return _v5818_os.path.join(_THON_V5818_APP_DIR, name)

def _v5818_safe_load(name, default=None, repair=True):
    if default is None:
        default = {}
    path = _v5818_path(name)
    if not _v5818_os.path.exists(path):
        return default
    try:
        with open(path, "r", encoding="utf-8") as f:
            raw = f.read()
        if not raw.strip():
            raise ValueError("arquivo vazio")
        return _v5818_json.loads(raw)
    except Exception as e:
        # tenta tmp primeiro, caso writer tenha parado no meio
        for alt in [path + ".tmp", path + ".bak", path + ".backup"]:
            if _v5818_os.path.exists(alt):
                try:
                    with open(alt, "r", encoding="utf-8") as f:
                        raw = f.read()
                    if raw.strip():
                        data = _v5818_json.loads(raw)
                        return data
                except Exception:
                    pass
        if repair:
            try:
                bad = path + ".corrompido_" + _v5818_datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                _v5818_shutil.copy2(path, bad)
            except Exception:
                pass
        return {"_erro_json": str(e), "_arquivo": name, "_default": default}

def _v5818_atomic_write(name, data):
    path = _v5818_path(name)
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        _v5818_json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")
    _v5818_os.replace(tmp, path)
    return path

def _v5818_list_from_any(data):
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        for k in ["lote", "fila", "canais", "qualificados", "aprovados", "reprovados", "items", "resultados"]:
            v = data.get(k)
            if isinstance(v, list):
                return v
        # thon_dlp_engine_last_output normalmente tem verificacao/resultado
        for k in ["verificacao", "verify", "resultado", "data"]:
            v = data.get(k)
            if isinstance(v, dict):
                got = _v5818_list_from_any(v)
                if got:
                    return got
    return []

def _v5818_candidates_payload():
    fila = _v5818_safe_load("fila_pendente_api.json", {"lote": []}, repair=False)
    dlpq = _v5818_safe_load("dlp_verification_queue.json", {"fila": []}, repair=False)
    brutos = _v5818_safe_load("canais_brutos_api.json", {"canais": []}, repair=False)
    out = _v5818_safe_load("thon_dlp_engine_last_output.json", {}, repair=False)
    verify = _v5818_safe_load("thon_verify_result.json", {}, repair=False)
    search = _v5818_safe_load("thon_search_result.json", {}, repair=False)

    fila_list = _v5818_list_from_any(fila)
    dlpq_list = _v5818_list_from_any(dlpq)
    brutos_list = _v5818_list_from_any(brutos)
    verify_list = _v5818_list_from_any(verify)

    qualificados = []
    reprovados = []
    for src in [verify, out, fila]:
        if isinstance(src, dict):
            q = src.get("qualificados")
            r = src.get("reprovados")
            if isinstance(q, list): qualificados.extend(q)
            if isinstance(r, list): reprovados.extend(r)
            # fallback: lista geral com status
            for item in _v5818_list_from_any(src):
                if isinstance(item, dict):
                    st = str(item.get("status", "")).lower()
                    if "qual" in st and item not in qualificados: qualificados.append(item)
                    if "reprov" in st and item not in reprovados: reprovados.append(item)

    stats = {}
    if isinstance(out, dict):
        stats.update(out.get("stats") or {})
        for k in ["brutos_encontrados", "blacklist_removidos", "novos_para_verificar", "verificados", "qualificados", "reprovados"]:
            if k in out and k not in stats: stats[k] = out.get(k)
    if isinstance(search, dict):
        stats.setdefault("brutos_encontrados", search.get("brutos_encontrados"))
        stats.setdefault("blacklist_removidos", search.get("blacklist_removidos"))
        stats.setdefault("novos_para_verificar", search.get("novos_para_verificar"))
    if isinstance(verify, dict):
        stats.setdefault("verificados", verify.get("verificados") or len(verify_list))
        if isinstance(verify.get("qualificados"), list): stats.setdefault("qualificados", len(verify.get("qualificados")))
        if isinstance(verify.get("reprovados"), list): stats.setdefault("reprovados", len(verify.get("reprovados")))

    return {
        "ok": True,
        "created_at": _v5818_now(),
        "fila_total": len(fila_list),
        "fila_dlp_total": len(dlpq_list),
        "brutos_total": len(brutos_list),
        "qualificados_total": len(qualificados),
        "reprovados_total": len(reprovados),
        "stats_engine": stats,
        "fila": fila_list[:500],
        "fila_dlp": dlpq_list[:500],
        "qualificados": qualificados[:500],
        "reprovados": reprovados[:300],
        "brutos_amostra": brutos_list[-100:] if brutos_list else [],
        "arquivos": {
            "fila_pendente_api.json": isinstance(fila, dict) and not fila.get("_erro_json"),
            "dlp_verification_queue.json": isinstance(dlpq, dict) and not dlpq.get("_erro_json"),
            "canais_brutos_api.json": isinstance(brutos, dict) and not brutos.get("_erro_json"),
            "thon_dlp_engine_last_output.json": isinstance(out, dict) and not out.get("_erro_json"),
            "thon_verify_result.json": isinstance(verify, dict) and not verify.get("_erro_json"),
            "thon_search_result.json": isinstance(search, dict) and not search.get("_erro_json"),
        },
        "erros_json": [x for x in [fila, dlpq, brutos, out, verify, search] if isinstance(x, dict) and x.get("_erro_json")]
    }

def _v5818_read_tail(path, max_chars=60000):
    try:
        if not _v5818_os.path.exists(path):
            return ""
        with open(path, "rb") as f:
            f.seek(0, 2)
            size = f.tell()
            f.seek(max(0, size - max_chars))
            return f.read().decode("utf-8", "replace")
    except Exception as e:
        return f"[erro lendo {path}: {e}]"

def _v5818_log_text():
    parts = []
    for name in ["thon_engine_live.log", "thon_dlp_engine.log", "thon_search_engine.log", "thon_verify_engine.log", "backend.log"]:
        p = _v5818_path(name)
        txt = _v5818_read_tail(p, 25000)
        if txt.strip():
            parts.append(f"===== {name} =====\n" + txt)
    # se não tiver log de arquivo, cria resumo textual dos JSONs reais
    payload = _v5818_candidates_payload()
    parts.append("===== RESUMO JSON REAL =====\n" + _v5818_json.dumps({
        "fila_total": payload.get("fila_total"),
        "fila_dlp_total": payload.get("fila_dlp_total"),
        "brutos_total": payload.get("brutos_total"),
        "qualificados_total": payload.get("qualificados_total"),
        "reprovados_total": payload.get("reprovados_total"),
        "stats_engine": payload.get("stats_engine"),
        "erros_json": payload.get("erros_json"),
    }, ensure_ascii=False, indent=2))
    return "\n\n".join(parts)[-70000:]

@app.route('/engine-console')
def _v5818_engine_console():
    html = r"""
<!doctype html><html lang="pt-br"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>THON Engine Console V58.18</title>
<style>
body{margin:0;background:#020807;color:#d9fff2;font-family:Inter,system-ui,-apple-system,Segoe UI,sans-serif} .wrap{max-width:1400px;margin:0 auto;padding:22px}.top{display:flex;align-items:center;justify-content:space-between;gap:12px;margin-bottom:14px}.title{font-size:20px;font-weight:900}.sub{color:#7fcab2;font-size:12px;margin-top:4px}.btn{background:#09201b;border:1px solid rgba(0,255,170,.25);color:#c8ffec;border-radius:10px;padding:9px 12px;text-decoration:none;cursor:pointer;font-weight:800}.grid{display:grid;grid-template-columns:repeat(6,1fr);gap:10px;margin:16px 0}.card{background:#061411;border:1px solid rgba(0,255,170,.16);border-radius:14px;padding:14px}.label{font-size:10px;text-transform:uppercase;letter-spacing:.08em;color:#66ad99}.num{font-size:24px;font-weight:900;margin-top:5px}.ok{color:#8cffcf}.bad{color:#ff918c}.cols{display:grid;grid-template-columns:1.2fr .8fr;gap:14px}.panel{background:#030d0b;border:1px solid rgba(0,255,170,.16);border-radius:14px;overflow:hidden}.panel h2{font-size:14px;margin:0;padding:12px 14px;border-bottom:1px solid rgba(0,255,170,.12)}table{width:100%;border-collapse:collapse;font-size:12px}th,td{padding:9px 10px;border-bottom:1px solid rgba(0,255,170,.08);vertical-align:top}th{text-align:left;color:#7fcab2;font-size:10px;text-transform:uppercase;letter-spacing:.07em}pre{white-space:pre-wrap;word-break:break-word;margin:0;padding:14px;max-height:520px;overflow:auto;font-size:11px;color:#b8ffe5}.pill{display:inline-block;padding:3px 7px;border-radius:999px;background:#0c2a23;border:1px solid rgba(0,255,170,.2);font-size:11px}.q{color:#9dffce}.r{color:#ffaaa6}@media(max-width:900px){.grid{grid-template-columns:repeat(2,1fr)}.cols{grid-template-columns:1fr}}
</style></head><body><div class="wrap"><div class="top"><div><div class="title">THON Engine Console V58.18</div><div class="sub">Não mexe na API. Lê fila, brutos e outputs reais das engines.</div></div><div><a class="btn" href="/">Index</a> <a class="btn" href="/prospector">Prospector</a> <a class="btn" href="/api-keys">API Keys</a> <button class="btn" onclick="repair()">Reparar JSONs</button></div></div><div id="cards" class="grid"></div><div class="cols"><div class="panel"><h2>Qualificados / Fila</h2><table><thead><tr><th>Status</th><th>Canal</th><th>Score</th><th>Subs</th><th>Longos</th><th>Avg</th><th>Query</th></tr></thead><tbody id="rows"></tbody></table></div><div class="panel"><h2>Log / Estado</h2><pre id="log">carregando...</pre></div></div></div>
<script>
function val(x){return (x===undefined||x===null||x==='')?'-':x}
function pick(o,ks){for(const k of ks){if(o&&o[k]!==undefined&&o[k]!==null&&o[k]!=='')return o[k]}return '-'}
function fmt(n){ if(n===undefined||n===null||n==='')return '-'; return n; }
async function load(){
  let r=await fetch('/api/v58_18/engine',{cache:'no-store'}); let d=await r.json();
  const s=d.stats_engine||{};
  const cards=[['Fila',d.fila_total],['Fila DLP',d.fila_dlp_total],['Brutos',d.brutos_total],['Qualificados',d.qualificados_total||s.qualificados],['Reprovados',d.reprovados_total||s.reprovados],['Novos verif.',s.novos_para_verificar]];
  document.getElementById('cards').innerHTML=cards.map(c=>`<div class="card"><div class="label">${c[0]}</div><div class="num">${val(c[1])}</div></div>`).join('');
  let list=[];
  if(d.qualificados&&d.qualificados.length) list=d.qualificados;
  else if(d.fila&&d.fila.length) list=d.fila;
  else if(d.brutos_amostra&&d.brutos_amostra.length) list=d.brutos_amostra;
  document.getElementById('rows').innerHTML=list.slice(0,120).map(x=>{
    let st=String(pick(x,['status','situacao','state']));
    let cls=st.toLowerCase().includes('qual')?'q':(st.toLowerCase().includes('reprov')?'r':'');
    let url=pick(x,['url','channel_url','canal_url']);
    let nome=pick(x,['nome','title','canal','channel_title','name']);
    let link=(url&&url!=='-')?`<a class="q" target="_blank" href="${url}">${nome}</a>`:nome;
    return `<tr><td><span class="pill ${cls}">${st}</span></td><td>${link}</td><td>${pick(x,['score','lead_score'])}</td><td>${pick(x,['subs_fmt','subs','subscriber_count','inscritos'])}</td><td>${pick(x,['longos','videos_longos'])}</td><td>${pick(x,['avg_views','media_views'])}</td><td>${pick(x,['query','found_query'])}</td></tr>`
  }).join('') || '<tr><td colspan="7">Nenhum candidato encontrado nos JSONs ainda.</td></tr>';
  let lr=await fetch('/api/v58_18/engine-log',{cache:'no-store'}); let ld=await lr.json();
  document.getElementById('log').textContent=ld.log||JSON.stringify(d,null,2);
}
async function repair(){let r=await fetch('/api/v58_18/repair-jsons',{method:'POST'}); alert(JSON.stringify(await r.json(),null,2)); load();}
load(); setInterval(load,2500);
</script></body></html>
"""
    return _v5818_Response(html, mimetype='text/html')

@app.route('/api/v58_18/status')
def _v5818_status():
    return _v5818_jsonify({"ok": True, "version": "v58.18-ui-json-console-safe-no-api", "time": _v5818_now(), "note": "API rotation não foi alterada."})

@app.route('/api/v58_18/engine')
def _v5818_engine():
    return _v5818_jsonify(_v5818_candidates_payload())

@app.route('/api/v58_18/fila')
def _v5818_fila():
    return _v5818_jsonify(_v5818_candidates_payload())

@app.route('/api/v58_18/engine-log')
def _v5818_engine_log():
    return _v5818_jsonify({"ok": True, "created_at": _v5818_now(), "log": _v5818_log_text()})

@app.route('/api/v58_18/api_state')
def _v5818_api_state():
    keys_path = _v5818_path("api_keys.txt")
    keys=[]
    try:
        if _v5818_os.path.exists(keys_path):
            for line in open(keys_path,"r",encoding="utf-8",errors="replace"):
                k=line.strip()
                if k and not k.startswith('#'):
                    keys.append(k[:8]+'...'+k[-6:])
    except Exception: pass
    state=_v5818_safe_load("youtube_api_quota_state.json", {}, repair=False)
    return _v5818_jsonify({"ok": True, "keys_count": len(keys), "keys_masked": keys, "quota_state": state, "note":"rota de leitura; não altera API rotation"})

@app.route('/api/v58_18/repair-jsons', methods=['POST','GET'])
def _v5818_repair_jsons():
    report=[]
    for name in _THON_V5818_FILES:
        path=_v5818_path(name)
        if not _v5818_os.path.exists(path):
            report.append({"arquivo":name,"status":"nao_existe"}); continue
        data=_v5818_safe_load(name, {}, repair=True)
        if isinstance(data, dict) and data.get('_erro_json'):
            # recria estrutura mínima só se arquivo estiver ilegível; corrupto já foi backupado
            if name == "fila_pendente_api.json": empty=[]
            elif name == "dlp_verification_queue.json": empty={"fila":[],"total":0,"updated_at":_v5818_now(),"nota":"recriado v58.18 apos JSON invalido"}
            elif name == "canais_brutos_api.json": empty={"canais":[],"total":0,"updated_at":_v5818_now(),"nota":"recriado v58.18 apos JSON invalido"}
            elif name == "youtube_api_quota_state.json": empty={"date":_v5818_datetime.date.today().isoformat(),"used":0,"calls":0,"history":[],"updated_at":_v5818_now(),"nota":"recriado v58.18 apos JSON invalido"}
            else: empty={"updated_at":_v5818_now(),"nota":"recriado v58.18 apos JSON invalido"}
            _v5818_atomic_write(name, empty)
            report.append({"arquivo":name,"status":"corrompido_backupado_e_recriado","erro":data.get('_erro_json')})
        else:
            report.append({"arquivo":name,"status":"ok"})
    return _v5818_jsonify({"ok": True, "report": report})

# aliases para telas antigas não quebrarem se esperarem JSON
@app.route('/api/engine-live/status')
@app.route('/api/engine/status')
@app.route('/api/prospector/fila')
@app.route('/api/fila-pendente')
@app.route('/api/fila_pendente')
@app.route('/api/coleta_v93/status')
@app.route('/api/coleta-v93/status')
@app.route('/coleta_v93/status')
@app.route('/coleta-v93/status')
@app.route('/coleta_progress')
@app.route('/api/coleta_progress')
def _v5818_alias_json():
    return _v5818_jsonify(_v5818_candidates_payload())

# Evita erro no browser: Unexpected token '<', '<!doctype'... is not valid JSON
# Para chamadas de API/coleta que ainda estejam apontando para rota antiga.
@app.errorhandler(404)
def _v5818_404(e):
    p = getattr(_v5818_request, 'path', '') or ''
    wants_json = p.startswith('/api') or ('coleta' in p.lower()) or ('v93' in p.lower()) or ('json' in p.lower())
    if wants_json:
        return _v5818_jsonify({"ok": False, "error": "rota_nao_encontrada", "path": p, "hint": "V58.18 retornou JSON para evitar <!doctype no frontend."}), 404
    try:
        return e
    except Exception:
        return _v5818_jsonify({"ok": False, "error": "not_found", "path": p}), 404

@app.errorhandler(500)
def _v5818_500(e):
    p = getattr(_v5818_request, 'path', '') or ''
    if p.startswith('/api') or ('coleta' in p.lower()) or ('v93' in p.lower()):
        return _v5818_jsonify({"ok": False, "error": "erro_interno", "path": p, "detail": str(e)}), 500
    raise e

print("[V58.18] UI/JSON/Console Safe carregado | /engine-console | /api/v58_18/engine | API rotation preservada")
# === THON V58.18 UI JSON CONSOLE SAFE END ===

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print(f"\n{'='*55}")
    print(f" THON TOOLKIT {APP_VERSION} - FILA PERSISTENTE + API DISCOVERY + YT-DLP")
    print(f"{'='*55}")
    print(f" Local: http://localhost:{port}")
    print(f" Nichos: 12 | Lote: {LOTE_PADRAO} | Prospectors: CRM separado | API descobre -> DLP verifica | DLP Auto Hunt")
    print(f" Regra: {fmt_numero(MIN_SUBS)}-{fmt_numero(MAX_SUBS)} inscritos")
    print(f" Reprovados: {len(carregar_reprovados())} totais (paginação 100)")
    print(f" API budget: {API_DAILY_BUDGET} unidades/dia | multi-key")
    print(f" Depósito de brutos: {CANAL_BRUTO_FILE} (total atual: {len(_carregar_brutos())})")
    print(f" Fila pendente: {LOTE_ATIVO_FILE} (total atual: {len(carregar_lote_ativo())})")
    print(f"{'='*55}\n")
    app.run(host="0.0.0.0", port=port, debug=False)
